#!/usr/bin/env python3
"""
importar_abastecimentos.py
--------------------------
Importa abastecimentos de "Relatório Abastecimento.xlsx" para o banco SQLite
usado pelo servidor gestao-abastecimentos.

Uso:
    python3 importar_abastecimentos.py                         # usa DB padrão
    python3 importar_abastecimentos.py caminho/planilha.xlsx   # arquivo custom
    python3 importar_abastecimentos.py planilha.xlsx meu.db    # arquivo + DB custom
"""

import sys, os, sqlite3

# ── Auto-instala openpyxl se necessário ──────────────────────────────────────
try:
    import openpyxl
except ImportError:
    print("⚙  Instalando openpyxl…")
    import subprocess
    subprocess.check_call([sys.executable, '-m', 'pip', 'install',
                           'openpyxl', '--break-system-packages', '-q'])
    import openpyxl

# ── Paths padrão ─────────────────────────────────────────────────────────────
SCRIPT_DIR   = os.path.dirname(os.path.abspath(__file__))
DEFAULT_XLS  = os.path.join(SCRIPT_DIR, 'Relatório Abastecimento.xlsx')
DEFAULT_DB   = os.path.join(SCRIPT_DIR, 'gestao_frota.db')

xlsx_path = sys.argv[1] if len(sys.argv) > 1 else DEFAULT_XLS
db_path   = sys.argv[2] if len(sys.argv) > 2 else DEFAULT_DB

# ── Helpers ───────────────────────────────────────────────────────────────────
def fmt_date(val):
    """Converte DD/MM/YYYY ou objeto date para YYYY-MM-DD."""
    if val is None:
        return ''
    if hasattr(val, 'strftime'):          # objeto datetime / date do openpyxl
        return val.strftime('%Y-%m-%d')
    s = str(val).strip()
    if '/' in s:                           # DD/MM/YYYY
        parts = s.split('/')
        if len(parts) == 3:
            return f'{parts[2]}-{parts[1]}-{parts[0]}'
    return s

def fmt_hora(val):
    """Converte timedelta / time / string para HH:MM:SS."""
    if val is None:
        return '00:00:00'
    if hasattr(val, 'strftime'):           # time object
        return val.strftime('%H:%M:%S')
    if hasattr(val, 'seconds'):            # timedelta (openpyxl devolve assim)
        total = int(val.seconds)
        h, rem = divmod(total, 3600)
        m, s   = divmod(rem, 60)
        return f'{h:02d}:{m:02d}:{s:02d}'
    s = str(val).strip()
    # já está como HH:MM:SS
    return s

def clean_cnpj(val):
    """Remove espaços do CNPJ."""
    return str(val).strip() if val else ''

def clean_str(val):
    return str(val).strip() if val else ''

def clean_float(val):
    if val is None:
        return 0.0
    try:
        return float(val)
    except (ValueError, TypeError):
        return 0.0

def clean_int(val):
    if val is None:
        return 0
    try:
        return int(float(str(val)))
    except (ValueError, TypeError):
        return 0

# ── Lê o Excel ───────────────────────────────────────────────────────────────
print(f"📂 Abrindo: {xlsx_path}")
if not os.path.exists(xlsx_path):
    print(f"❌ Arquivo não encontrado: {xlsx_path}")
    sys.exit(1)

wb = openpyxl.load_workbook(xlsx_path, data_only=True)
ws = wb.worksheets[0]
print(f"ℹ  Aba: '{ws.title}' | {ws.max_row - 1} linhas de dados")

# Monta mapa de colunas pelo cabeçalho (linha 1)
col_map = {}
for c in range(1, ws.max_column + 1):
    val = ws.cell(1, c).value
    if val is not None:
        col_map[str(val).strip().upper()] = c

def col(name):
    """Retorna índice da coluna (case-insensitive), ou None se não existir."""
    return col_map.get(name.upper())

# Verifica colunas obrigatórias
required = ['DATA', 'HORA', 'PLACA', 'MOTORISTA', 'POSTO', 'COMBUSTÍVEL']
missing  = [r for r in required if col(r) is None]
if missing:
    print(f"⚠  Colunas não encontradas: {missing}")
    print(f"   Colunas disponíveis: {list(col_map.keys())}")
    sys.exit(1)

def cell(row, name):
    idx = col(name)
    return ws.cell(row, idx).value if idx else None

# ── Conecta ao banco ──────────────────────────────────────────────────────────
print(f"🗄  Banco: {db_path}")
conn = sqlite3.connect(db_path)
conn.row_factory = sqlite3.Row

# Garante que a tabela existe
conn.execute('''CREATE TABLE IF NOT EXISTS abastecimentos (
    id              INTEGER PRIMARY KEY AUTOINCREMENT,
    data            TEXT    NOT NULL,
    hora            TEXT    NOT NULL,
    placa           TEXT    NOT NULL,
    motorista       TEXT    NOT NULL,
    cpf_motorista   TEXT    DEFAULT '',
    hodometro       REAL    DEFAULT 0,
    posto           TEXT    NOT NULL,
    cnpj_posto      TEXT    DEFAULT '',
    cidade_posto    TEXT    DEFAULT '',
    uf_posto        TEXT    DEFAULT '',
    combustivel     TEXT    NOT NULL,
    volume          REAL    DEFAULT 0,
    preco_unitario  REAL    DEFAULT 0,
    valor_total     REAL    DEFAULT 0,
    created_at      TEXT    DEFAULT (datetime('now','localtime'))
)''')
conn.commit()

# Busca registros existentes para verificar duplicatas (data+hora+placa)
existing = set()
for row in conn.execute("SELECT data, hora, placa FROM abastecimentos"):
    existing.add((row['data'], row['hora'], row['placa']))

# ── Processa e insere ─────────────────────────────────────────────────────────
INSERT_SQL = '''INSERT INTO abastecimentos
    (data, hora, placa, motorista, cpf_motorista, hodometro,
     posto, cnpj_posto, cidade_posto, uf_posto,
     combustivel, volume, preco_unitario, valor_total)
    VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?)'''

inserted = 0
skipped  = 0
errors   = 0

for r in range(2, ws.max_row + 1):
    # Ignora linhas totalmente vazias
    if all(ws.cell(r, c).value is None for c in range(1, ws.max_column + 1)):
        continue
    try:
        data       = fmt_date(cell(r, 'DATA'))
        hora       = fmt_hora(cell(r, 'HORA'))
        placa      = clean_str(cell(r, 'PLACA')).upper()
        motorista  = clean_str(cell(r, 'MOTORISTA'))
        cpf        = clean_str(cell(r, 'CPF MOTORISTA'))
        hodometro  = clean_int(cell(r, 'HODÔMETRO') or cell(r, 'HODOMETRO'))
        posto      = clean_str(cell(r, 'POSTO'))
        cnpj_posto = clean_cnpj(cell(r, 'CNPJ POSTO') or cell(r, 'CNPJ'))
        cidade     = clean_str(cell(r, 'CIDADE'))
        uf         = clean_str(cell(r, 'UF'))
        combustivel= clean_str(cell(r, 'COMBUSTÍVEL') or cell(r, 'COMBUSTIVEL'))
        volume     = clean_float(cell(r, 'VOLUME (L)') or cell(r, 'VOLUME'))
        preco      = clean_float(cell(r, 'R$/L') or cell(r, 'PRECO') or cell(r, 'PREÇO'))
        valor      = clean_float(cell(r, 'VALOR TOTAL') or cell(r, 'TOTAL'))

        # Validação básica
        if not placa or not motorista or not posto or not combustivel or not data:
            print(f"  ⚠ Linha {r} ignorada: campos obrigatórios ausentes")
            errors += 1
            continue

        # Duplicata
        key = (data, hora, placa)
        if key in existing:
            skipped += 1
            continue

        conn.execute(INSERT_SQL, [
            data, hora, placa, motorista, cpf, hodometro,
            posto, cnpj_posto, cidade, uf,
            combustivel, volume, preco, valor
        ])
        existing.add(key)
        inserted += 1

    except Exception as e:
        print(f"  ⚠ Linha {r} ignorada: {e}")
        errors += 1

conn.commit()
conn.close()

# ── Resumo ────────────────────────────────────────────────────────────────────
print()
print(f"✅ Importação concluída!")
print(f"   ➕ Inseridos : {inserted}")
print(f"   ⏭  Duplicatas: {skipped}")
if errors:
    print(f"   ⚠  Erros    : {errors}")
