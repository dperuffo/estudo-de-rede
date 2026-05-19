# ═══════════════════════════════════════════════════════════════════
#  Estudo de Rede – Gestão de Frotas
#  Versão 5.1  |  Plotly WebGL map + NumPy vetorizado + cache 24h
# ═══════════════════════════════════════════════════════════════════

import base64
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import datetime
import hashlib
import io
import json as _json_mod
import math
import os
import re
import time
import unicodedata
import requests
import numpy as np
import pandas as pd
import plotly.graph_objects as go
import streamlit as st
import folium
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import matplotlib.patches as mpatches
from matplotlib.lines import Line2D

# Diretório onde este script está — usado para localizar arquivos do repo
_DIR = os.path.dirname(os.path.abspath(__file__))

# ═══════════════════════════════════════════════════════════════════
#  SUPABASE — Banco de Dados para Persistência
# ═══════════════════════════════════════════════════════════════════

def _db_client():
    """Retorna o cliente Supabase. Cria uma vez por sessão via session_state."""
    if "_supabase_client" not in st.session_state:
        try:
            from supabase import create_client
            _url = st.secrets.get("supabase", {}).get("url") or os.environ.get("SUPABASE_URL", "")
            _key = st.secrets.get("supabase", {}).get("key") or os.environ.get("SUPABASE_KEY", "")
            if _url and _key:
                st.session_state["_supabase_client"] = create_client(_url, _key)
            else:
                st.session_state["_supabase_client"] = None
        except Exception:
            st.session_state["_supabase_client"] = None
    return st.session_state["_supabase_client"]


def _db_email() -> str:
    """E-mail do usuário logado, usado como identificador no banco."""
    return (st.session_state.get("_auth_user") or {}).get("email", "anonimo")


# ── Rotas ──────────────────────────────────────────────────────────

def _db_carregar_rotas() -> list:
    """Carrega rotas do Supabase. Fallback para JSON local."""
    db = _db_client()
    if db:
        try:
            res = db.table("rotas_salvas") \
                    .select("*") \
                    .eq("usuario_email", _db_email()) \
                    .order("criado_em", desc=True) \
                    .execute()
            return [{"id": r["id"], "nome": r["nome"], "tipo": r["tipo"],
                     "criado_em": r["criado_em"], **r.get("dados", {})}
                    for r in (res.data or [])]
        except Exception:
            pass
    return _carregar_rotas_salvas_local()


def _db_salvar_rota(nome: str, tipo: str, dados: dict) -> bool:
    """Salva rota no Supabase. Fallback para JSON local."""
    db = _db_client()
    _id = f"{int(time.time())}_{nome[:8]}"
    # Garante que dados é JSON puro (converte numpy, datetime, etc.)
    try:
        _dados_json = _json_mod.loads(_json_mod.dumps(dados, default=str))
    except Exception:
        _dados_json = {}
    if db:
        try:
            db.table("rotas_salvas").insert({
                "id":            _id,
                "usuario_email": _db_email(),
                "nome":          nome.strip() or "Rota",
                "tipo":          tipo,
                "criado_em":     _agora(),
                "dados":         _dados_json,
            }).execute()
            return True
        except Exception as _e:
            st.warning(f"⚠️ Banco indisponível ({_e}), salvando localmente.", icon="💾")
    return _salvar_rota_nova_local(nome, tipo, dados)


def _db_deletar_rota(rota_id: str) -> bool:
    """Remove rota do Supabase. Fallback para JSON local."""
    db = _db_client()
    if db:
        try:
            db.table("rotas_salvas") \
              .delete() \
              .eq("id", rota_id) \
              .eq("usuario_email", _db_email()) \
              .execute()
            return True
        except Exception:
            pass
    return _deletar_rota_local(rota_id)


# ── Preferências ───────────────────────────────────────────────────

def _db_salvar_preferencias(placa: str = "", combustivel: str = "",
                             autonomia: float = 0.0, capacidade: float = 0.0,
                             extras: dict = None) -> bool:
    """Grava preferências do usuário no Supabase."""
    db = _db_client()
    if not db:
        return False
    try:
        db.table("preferencias").upsert({
            "usuario_email": _db_email(),
            "placa":         placa,
            "combustivel":   combustivel,
            "autonomia":     autonomia,
            "capacidade":    capacidade,
            "extras":        extras or {},
            "atualizado_em": datetime.now().isoformat(),
        }).execute()
        return True
    except Exception:
        return False


def _db_carregar_preferencias() -> dict:
    """Carrega preferências do usuário do Supabase."""
    db = _db_client()
    if not db:
        return {}
    try:
        res = db.table("preferencias") \
                .select("*") \
                .eq("usuario_email", _db_email()) \
                .limit(1) \
                .execute()
        if res.data:
            return res.data[0]
    except Exception:
        pass
    return {}


# ── Postos Favoritos ───────────────────────────────────────────────

def _db_favoritos() -> list:
    """Lista postos favoritos do usuário."""
    db = _db_client()
    if not db:
        return []
    try:
        res = db.table("postos_favoritos") \
                .select("*") \
                .eq("usuario_email", _db_email()) \
                .execute()
        return res.data or []
    except Exception:
        return []


def _db_add_favorito(cnpj: str, razao_social: str, municipio: str,
                     uf: str, lat: float = None, lon: float = None) -> bool:
    """Adiciona posto aos favoritos."""
    db = _db_client()
    if not db:
        return False
    try:
        db.table("postos_favoritos").upsert({
            "usuario_email": _db_email(),
            "cnpj":          cnpj,
            "razao_social":  razao_social,
            "municipio":     municipio,
            "uf":            uf,
            "lat":           lat,
            "lon":           lon,
        }).execute()
        return True
    except Exception:
        return False


def _db_remove_favorito(cnpj: str) -> bool:
    """Remove posto dos favoritos."""
    db = _db_client()
    if not db:
        return False
    try:
        db.table("postos_favoritos") \
          .delete() \
          .eq("usuario_email", _db_email()) \
          .eq("cnpj", cnpj) \
          .execute()
        return True
    except Exception:
        return False


# ── Notas por Posto ───────────────────────────────────────────────

def _db_nota_posto(cnpj: str) -> str:
    """Retorna a nota interna salva para um posto (por CNPJ)."""
    db = _db_client()
    if not db or not cnpj:
        return ""
    try:
        res = db.table("notas_posto") \
                .select("nota") \
                .eq("usuario_email", _db_email()) \
                .eq("cnpj", cnpj) \
                .limit(1) \
                .execute()
        return res.data[0]["nota"] if res.data else ""
    except Exception:
        return ""


def _db_salvar_nota_posto(cnpj: str, nota: str) -> bool:
    """Salva/atualiza nota interna para um posto."""
    db = _db_client()
    if not db or not cnpj:
        return False
    try:
        db.table("notas_posto").upsert({
            "usuario_email": _db_email(),
            "cnpj":          cnpj,
            "nota":          nota.strip(),
            "atualizado_em": datetime.now().isoformat(),
        }, on_conflict="usuario_email,cnpj").execute()
        return True
    except Exception:
        return False


# ── Perfis de Veículo ─────────────────────────────────────────────

def _db_perfis_veiculo() -> list:
    """Lista perfis de veículo salvos pelo usuário."""
    db = _db_client()
    if not db:
        return []
    try:
        res = db.table("perfis_veiculo") \
                .select("*") \
                .eq("usuario_email", _db_email()) \
                .order("criado_em", desc=False) \
                .execute()
        return res.data or []
    except Exception:
        return []


def _db_salvar_perfil_veiculo(nome: str, placa: str, combustivel: str,
                               tanque: float, autonomia: float) -> bool:
    """Salva um novo perfil de veículo."""
    db = _db_client()
    if not db:
        return False
    try:
        db.table("perfis_veiculo").insert({
            "usuario_email": _db_email(),
            "nome":          nome.strip() or placa or "Veículo",
            "placa":         placa.strip().upper(),
            "combustivel":   combustivel,
            "tanque":        tanque,
            "autonomia":     autonomia,
            "criado_em":     datetime.now().isoformat(),
        }).execute()
        return True
    except Exception:
        return False


def _db_deletar_perfil_veiculo(perfil_id) -> bool:
    """Remove um perfil de veículo."""
    db = _db_client()
    if not db:
        return False
    try:
        db.table("perfis_veiculo") \
          .delete() \
          .eq("id", perfil_id) \
          .eq("usuario_email", _db_email()) \
          .execute()
        return True
    except Exception:
        return False


# ── Histórico de Preços ────────────────────────────────────────────

def _db_gravar_preco(cnpj: str, razao_social: str, municipio: str, uf: str,
                     combustivel: str, preco: float, fonte: str = "ANP",
                     lat: float = None, lon: float = None) -> bool:
    """Grava snapshot de preço no histórico. Ignora duplicatas do mesmo dia."""
    db = _db_client()
    if not db or not cnpj or not preco:
        return False
    try:
        db.table("historico_precos").upsert({
            "cnpj":         cnpj,
            "razao_social": razao_social,
            "municipio":    municipio,
            "uf":           uf,
            "combustivel":  combustivel,
            "preco":        round(float(preco), 3),
            "fonte":        fonte,
            "data_ref":     datetime.now().strftime("%Y-%m-%d"),
            "lat":          lat,
            "lon":          lon,
        }, on_conflict="cnpj,combustivel,data_ref").execute()
        return True
    except Exception:
        return False


def _db_historico_preco(cnpj: str, combustivel: str = None, dias: int = 90) -> list:
    """Retorna histórico de preços de um posto nos últimos N dias."""
    db = _db_client()
    if not db:
        return []
    try:
        from datetime import timedelta
        data_ini = (datetime.now() - timedelta(days=dias)).strftime("%Y-%m-%d")
        q = db.table("historico_precos") \
              .select("data_ref,preco,combustivel,fonte") \
              .eq("cnpj", cnpj) \
              .gte("data_ref", data_ini) \
              .order("data_ref")
        if combustivel:
            q = q.eq("combustivel", combustivel)
        res = q.execute()
        return res.data or []
    except Exception:
        return []


# ── Controle de Acesso (Allowlist / Blacklist) ────────────────────

# E-mail do administrador — único com acesso ao painel de gestão
_ADMIN_EMAIL = "d.peruffo@gmail.com"


def _db_verificar_acesso(email: str) -> tuple[bool, str]:
    """
    Verifica se o e-mail tem permissão para acessar o app.
    Retorna (permitido: bool, motivo: str).

    Modo 'blacklist' (padrão): todos entram, exceto bloqueados.
    Modo 'allowlist':          só entra quem está com status='permitido'.
    """
    if not email or email == "anonimo":
        return False, "E-mail não identificado."

    # Admin sempre tem acesso
    if email.lower() == _ADMIN_EMAIL.lower():
        return True, "admin"

    db = _db_client()
    if not db:
        return True, "banco indisponível — acesso liberado"

    try:
        # Lê modo de acesso configurado
        _cfg = db.table("configuracoes") \
                 .select("valor") \
                 .eq("chave", "modo_acesso") \
                 .limit(1) \
                 .execute()
        _modo = (_cfg.data[0]["valor"] if _cfg.data else "blacklist")

        # Busca registro do e-mail
        _res = db.table("controle_acesso") \
                 .select("status,motivo") \
                 .eq("email", email.lower()) \
                 .limit(1) \
                 .execute()
        _registro = _res.data[0] if _res.data else None

        if _registro:
            if _registro["status"] == "bloqueado":
                return False, _registro.get("motivo") or "Acesso bloqueado pelo administrador."
            if _registro["status"] == "permitido":
                # Atualiza último acesso
                db.table("controle_acesso") \
                  .update({"ultimo_acesso": datetime.now().isoformat()}) \
                  .eq("email", email.lower()) \
                  .execute()
                return True, "permitido"

        # Sem registro: depende do modo
        if _modo == "allowlist":
            # Registra como pendente para o admin revisar
            db.table("controle_acesso").upsert({
                "email":  email.lower(),
                "status": "pendente",
                "nome":   (st.session_state.get("_auth_user") or {}).get("name", ""),
            }).execute()
            return False, "Seu acesso está pendente de aprovação pelo administrador."

        # Modo blacklist: não está bloqueado → pode entrar
        return True, "blacklist-livre"

    except Exception:
        return True, "erro ao verificar — acesso liberado"


def _db_atualizar_status_acesso(email: str, status: str,
                                 motivo: str = "", admin: str = "") -> bool:
    """Admin altera status de um e-mail: 'permitido' | 'bloqueado' | 'pendente'."""
    db = _db_client()
    if not db:
        return False
    try:
        db.table("controle_acesso").upsert({
            "email":         email.lower(),
            "status":        status,
            "motivo":        motivo,
            "adicionado_por": admin,
            "adicionado_em": datetime.now().isoformat(),
        }).execute()
        return True
    except Exception:
        return False


def _db_listar_controle_acesso() -> list:
    """Retorna todos os registros de controle de acesso."""
    db = _db_client()
    if not db:
        return []
    try:
        res = db.table("controle_acesso") \
                .select("*") \
                .order("adicionado_em", desc=True) \
                .execute()
        return res.data or []
    except Exception:
        return []


def _db_modo_acesso() -> str:
    """Retorna o modo atual: 'blacklist' ou 'allowlist'."""
    db = _db_client()
    if not db:
        return "blacklist"
    try:
        res = db.table("configuracoes") \
                .select("valor") \
                .eq("chave", "modo_acesso") \
                .limit(1) \
                .execute()
        return res.data[0]["valor"] if res.data else "blacklist"
    except Exception:
        return "blacklist"


def _db_set_modo_acesso(modo: str) -> bool:
    """Admin altera o modo de acesso global."""
    db = _db_client()
    if not db:
        return False
    try:
        db.table("configuracoes").upsert({
            "chave": "modo_acesso",
            "valor": modo,
            "atualizado_em": datetime.now().isoformat(),
        }).execute()
        return True
    except Exception:
        return False


# ── Logs de Acesso ─────────────────────────────────────────────────

def _db_gravar_log(entry: dict) -> None:
    """Grava um evento de log no Supabase. Fire-and-forget (falha silenciosa)."""
    db = _db_client()
    if not db:
        return
    try:
        db.table("logs_acesso").insert({
            "timestamp":     entry.get("timestamp"),
            "data":          entry.get("data"),
            "hora":          entry.get("hora"),
            "ip":            entry.get("ip"),
            "session_id":    entry.get("session_id"),
            "user_email":    entry.get("user_email"),
            "user_name":     entry.get("user_name"),
            "auth_provider": entry.get("auth_provider"),
            "modo":          entry.get("modo"),
            "uf":            entry.get("uf"),
            "municipio":     entry.get("municipio"),
            "acao":          entry.get("acao"),
            "detalhe":       entry.get("detalhe"),
            "user_agent":    entry.get("user_agent"),
        }).execute()
    except Exception:
        pass  # log nunca deve travar o app


def _db_ler_logs(limite: int = 2000) -> list:
    """Lê logs do Supabase. Retorna lista de dicts ordenada por mais recente."""
    db = _db_client()
    if not db:
        return []
    try:
        res = db.table("logs_acesso") \
                .select("*") \
                .order("criado_em", desc=True) \
                .limit(limite) \
                .execute()
        return res.data or []
    except Exception:
        return []

# ─── Imagem banner do sidebar (header) ───────────────────────────
# Designer.jpg tem prioridade máxima; fallback para versões anteriores
for _menu_nome in ["Designer.jpg", "designer.jpg", "Designer.png", "designer.png",
                   "Menu.jpg", "menu.jpg", "Menu.png", "menu.png",
                   "Menu_rincipal.jpg", "Menu_principal.jpg", "menu_principal.jpg",
                   "Menu_rincipal.png", "Menu_principal.png"]:
    _menu_candidato = os.path.join(_DIR, _menu_nome)
    if os.path.exists(_menu_candidato):
        _MENU_PATH = _menu_candidato
        break
else:
    _MENU_PATH = ""

if os.path.exists(_MENU_PATH):
    with open(_MENU_PATH, "rb") as _f:
        _menu_bytes = _f.read()
    _menu_mime = "image/jpeg" if _MENU_PATH.lower().endswith(".jpg") else "image/png"
    _MENU_B64  = base64.b64encode(_menu_bytes).decode()
    _MENU_IMG  = (
        f'<img src="data:{_menu_mime};base64,{_MENU_B64}" '
        f'style="width:100%;display:block;object-fit:cover;">'
    )
else:
    _MENU_B64  = None
    _MENU_IMG  = ""

# ─── Logo Gestão de Frotas (topbar + fallback sidebar) ──────────────────
for _logo_nome in ["Logo_profrotas.jpg", "logo_profrotas.jpg",
                   "Logo_profrotas.png", "logo_profrotas.png"]:
    _logo_candidato = os.path.join(_DIR, _logo_nome)
    if os.path.exists(_logo_candidato):
        _LOGO_PATH = _logo_candidato
        break
else:
    _LOGO_PATH = ""

if os.path.exists(_LOGO_PATH):
    with open(_LOGO_PATH, "rb") as _f:
        _logo_bytes = _f.read()
    _logo_mime = "image/jpeg" if _LOGO_PATH.lower().endswith(".jpg") else "image/png"
    _LOGO_B64 = base64.b64encode(_logo_bytes).decode()
    # Topbar azul: mix-blend-mode screen funde o fundo azul da logo com o gradiente
    _LOGO_TOPBAR  = (
        f'<img src="data:{_logo_mime};base64,{_LOGO_B64}" '
        f'style="height:46px;object-fit:contain;mix-blend-mode:screen;'
        f'filter:brightness(1.15) contrast(1.05)">'
    )
    # Sidebar: logo natural sobre fundo branco (fallback quando Menu.jpg ausente)
    _LOGO_SIDEBAR = (
        f'<img src="data:{_logo_mime};base64,{_LOGO_B64}" '
        f'style="height:78px;object-fit:contain;display:block;margin:0 auto">'
    )
    _LOGO_PAGE_ICON = _LOGO_PATH
else:
    _LOGO_B64       = None
    _LOGO_TOPBAR    = '<span style="font-size:36px">⛽</span>'
    _LOGO_SIDEBAR   = '<span style="font-size:32px">⛽</span>'
    _LOGO_PAGE_ICON = "⛽"

# ─── Configuração da página ────────────────────────────────────────
st.set_page_config(
    page_title="Estudo de Rede – Gestão de Frotas",
    page_icon=_LOGO_PAGE_ICON,
    layout="wide",
    initial_sidebar_state="expanded",  # sempre aberta
)

# ─── CSS Global + Responsivo ───────────────────────────────────────
st.markdown("""
<style>
/* ══ OCULTAR ELEMENTOS STREAMLIT ══════════════════════════════════ */
#MainMenu                                         { display: none !important; }
footer                                            { display: none !important; }
.stDeployButton                                   { display: none !important; }
[data-testid="manage-app-button"]                 { display: none !important; }
[data-testid="stStatusWidget"]                    { display: none !important; }
[class*="viewerBadge"]                            { display: none !important; }
[class*="ViewerBadge"]                            { display: none !important; }
[data-testid="stDecoration"]                      { display: none !important; }
a[href*="streamlit.io"]                           { display: none !important; }
a[href*="github.com"]                             { display: none !important; }
svg[data-icon="mark-github"]                      { display: none !important; }
img[alt*="github" i]                              { display: none !important; }
img[src*="github" i]                              { display: none !important; }

/* ── Seta recolher/expandir — OCULTA no desktop ── */
[data-testid="collapsedControl"]                  { display: none !important; }
[data-testid="stSidebarCollapseButton"]           { display: none !important; }
[data-testid="stSidebarNavCollapseButton"]        { display: none !important; }
button[data-testid="baseButton-headerNoPadding"]  { display: none !important; }

/* ── Botões ocultos do Tour — invisíveis mas clicáveis via JS ── */
.st-key-btn_tour_done_hidden,
.st-key-btn_tour_open_hidden {
    position: fixed !important;
    left: -9999px !important;
    top: -9999px !important;
    width: 1px !important;
    height: 1px !important;
    overflow: hidden !important;
}

/* Header transparente */
header[data-testid="stHeader"] {
    background: transparent !important;
    box-shadow: none !important;
}

/* ══ LAYOUT GERAL ══════════════════════════════════════════════════ */
.main .block-container {
    padding: 0.25rem 1rem 0.5rem !important;
    max-width: 100% !important;
}

/* ══ SIDEBAR REDESIGN ══════════════════════════════════════════════ */
[data-testid="stSidebar"] .sb-label {
    font-size: 10px;
    font-weight: 700;
    letter-spacing: 1.2px;
    text-transform: uppercase;
    color: #90a4ae;
    margin: 14px 0 6px;
}
[data-testid="stSidebar"] .modo-toggle [data-testid="stButton"] > button {
    height: 62px !important;
    font-size: 13px !important;
    line-height: 1.35 !important;
    white-space: pre-wrap !important;
    padding: 6px 4px !important;
}
[data-testid="stSidebar"] .pf-badge {
    border-radius: 8px;
    padding: 7px 11px;
    font-size: 11px;
    margin-bottom: 8px;
    line-height: 1.5;
}

/* ══ TOPBAR ════════════════════════════════════════════════════════ */
.topbar {
    background: linear-gradient(100deg, #0D47A1 0%, #1565C0 35%, #1976D2 55%, #E65100 100%);
    color: white;
    padding: 12px 28px;
    border-radius: 0 0 14px 14px;
    margin-bottom: 12px;
    display: flex;
    align-items: center;
    gap: 16px;
    box-shadow: 0 4px 18px rgba(13,71,161,0.35), 0 2px 6px rgba(230,81,0,0.20);
    flex-wrap: nowrap;
    position: relative;
    overflow: hidden;
}
.topbar::before {
    content: "";
    position: absolute;
    top: 0; left: 0; right: 0;
    height: 40%;
    background: linear-gradient(180deg, rgba(255,255,255,0.10) 0%, transparent 100%);
    pointer-events: none;
}
.topbar-title {
    font-size: 23px;
    font-weight: 900;
    letter-spacing: 0.3px;
    text-shadow: 0 1px 4px rgba(0,0,0,0.30);
    white-space: nowrap;
}
.topbar-sub {
    font-size: 12px;
    font-weight: 600;
    opacity: 0.88;
    margin-top: 2px;
    letter-spacing: 0.8px;
    text-transform: uppercase;
    text-shadow: 0 1px 3px rgba(0,0,0,0.25);
}
.topbar-badge {
    margin-left: auto;
    background: rgba(255,255,255,0.18);
    border: 1.5px solid rgba(255,255,255,0.45);
    border-radius: 20px;
    padding: 5px 14px;
    font-size: 12px;
    font-weight: 700;
    white-space: nowrap;
    backdrop-filter: blur(4px);
    text-shadow: 0 1px 2px rgba(0,0,0,0.20);
    flex-shrink: 0;
}

/* ── Botão hambúrguer (visível só no mobile) ── */
.topbar-menu-btn {
    display: none;
    flex-shrink: 0;
    background: rgba(255,255,255,0.20);
    border: 1.5px solid rgba(255,255,255,0.45);
    border-radius: 10px;
    width: 42px;
    height: 42px;
    align-items: center;
    justify-content: center;
    font-size: 22px;
    color: white;
    cursor: pointer;
    transition: background 0.18s ease, transform 0.18s ease;
    -webkit-tap-highlight-color: transparent;
    user-select: none;
    line-height: 1;
}
.topbar-menu-btn:active {
    background: rgba(255,255,255,0.38);
    transform: scale(0.93);
}

/* ══ OVERLAY (fecha drawer ao tocar fora) ══════════════════════════ */
.gf-overlay {
    display: none;
    position: fixed;
    inset: 0;
    background: rgba(0,0,0,0.48);
    z-index: 1099;
    -webkit-tap-highlight-color: transparent;
}
.gf-overlay.ativo { display: block; }

/* ══ SIDEBAR — desktop (sempre visível) ═══════════════════════════ */
section[data-testid="stSidebar"] > div:first-child {
    background: linear-gradient(180deg, #e6edf8 0%, #f2f6fc 60%, #f7f9fd 100%);
    border-right: 1px solid #c2d3e8;
}
section[data-testid="stSidebar"] .stButton > button {
    border-radius: 8px;
    font-weight: 600;
    min-height: 44px;
}

/* ══ MÉTRICAS ══════════════════════════════════════════════════════ */
[data-testid="stMetric"] {
    background: white;
    border-radius: 10px;
    padding: 14px 18px !important;
    box-shadow: 0 2px 6px rgba(0,0,0,0.08);
    border-left: 4px solid #1565c0;
}
[data-testid="stMetricLabel"] { font-size: 12px !important; color: #555 !important; }
[data-testid="stMetricValue"] { font-size: 24px !important; font-weight: 700 !important; }

/* ══ TABS ══════════════════════════════════════════════════════════ */
button[data-baseweb="tab"] { font-weight: 600; font-size: 13px; }

/* ══ EXPANDER ══════════════════════════════════════════════════════ */
details summary { font-weight: 700; font-size: 14px; }

/* ══ ALERTS / SEPARADORES ══════════════════════════════════════════ */
.stAlert { border-radius: 8px !important; }
hr { margin: 10px 0 !important; border-color: #c8d8e8 !important; }

/* ══ BOTÕES ════════════════════════════════════════════════════════ */
.stButton > button {
    min-height: 44px;
    border-radius: 8px;
    font-weight: 600;
    display: flex !important;
    align-items: center !important;
    justify-content: center !important;
    text-align: center !important;
}
.stButton > button[kind="primary"] {
    background: linear-gradient(135deg, #1565c0, #0d47a1);
    border: none;
    font-size: 14px;
    padding: 10px 0;
}
.stButton > button[kind="primary"]:hover {
    background: linear-gradient(135deg, #1976d2, #1565c0);
    box-shadow: 0 4px 12px rgba(21,101,192,0.4);
}

/* ══ LOADING / ANIMAÇÕES ═══════════════════════════════════════════ */
iframe { animation: mapFadeIn 0.45s ease-in; }
@keyframes mapFadeIn {
    from { opacity: 0; transform: translateY(6px); }
    to   { opacity: 1; transform: translateY(0);   }
}
[data-testid="stSpinner"] {
    padding: 18px 0 !important;
    text-align: center !important;
}
[data-testid="stSpinner"] > div {
    justify-content: center !important;
    font-size: 15px !important;
    color: #1565c0 !important;
    font-weight: 600 !important;
    gap: 10px !important;
}
[data-testid="stProgress"] > div > div {
    background: linear-gradient(90deg, #0d1b4b, #1565c0, #0288d1) !important;
    border-radius: 4px !important;
}

/* ══ EMPTY STATE ═══════════════════════════════════════════════════ */
.empty-state {
    text-align: center;
    padding: 60px 40px;
    color: #90a4b0;
}
.empty-state-icon  { font-size: 64px; margin-bottom: 16px; }
.empty-state-title { font-size: 20px; font-weight: 700; color: #546e7a; margin-bottom: 8px; }
.empty-state-desc  { font-size: 14px; line-height: 1.6; }

/* ══ TABELA SCROLL HORIZONTAL ══════════════════════════════════════ */
[data-testid="stDataFrame"] { overflow-x: auto !important; }

/* ══ RESPONSIVO — TABLET (≤ 1024px) ═══════════════════════════════ */
@media (max-width: 1024px) {
    .main .block-container { padding: 0.5rem 1rem 2rem !important; }
    .topbar { padding: 12px 18px; }
    .topbar-title { font-size: 18px; }
    [data-testid="stMetricValue"] { font-size: 20px !important; }
}

/* ══ RESPONSIVO — MOBILE (≤ 768px) ════════════════════════════════ */
@media (max-width: 768px) {
    /* Layout principal — sem deslocamento pelo sidebar */
    .main .block-container { padding: 0.25rem 0.5rem 2rem !important; }
    .main, [data-testid="stMain"] {
        margin-left: 0 !important;
        width: 100% !important;
    }

    /* Topbar: recua à esquerda para dar espaço ao FAB (54px) */
    .topbar {
        padding: 9px 10px 9px 60px;
        border-radius: 0 0 10px 10px;
        margin-bottom: 10px;
        gap: 8px;
        align-items: center;
        flex-wrap: nowrap;
    }
    .topbar-menu-btn { display: none !important; } /* visual desativado; FAB cuida de tudo */
    .topbar-title { font-size: 15px; letter-spacing: 0; }
    .topbar-sub   { font-size: 9px; margin-top: 1px; }
    .topbar-badge { font-size: 10px; padding: 3px 9px; }

    /* ── Sidebar: drawer fixo fora da tela por padrão ── */
    section[data-testid="stSidebar"] {
        position: fixed !important;
        top: 0 !important;
        left: 0 !important;
        height: 100% !important;
        height: 100dvh !important;
        width: 83vw !important;
        min-width: 260px !important;
        max-width: 320px !important;
        z-index: 9000 !important;
        overflow-y: auto !important;
        overflow-x: hidden !important;
        transition: transform 0.30s cubic-bezier(0.4,0,0.2,1),
                    box-shadow  0.30s ease !important;
    }
    /* Conteúdo interno do drawer ocupa altura toda */
    section[data-testid="stSidebar"] > div:first-child {
        min-height: 100% !important;
        padding-bottom: 48px !important;
    }
    /* Ocultar seta nativa de fechar sidebar */
    [data-testid="stSidebarCollapseButton"],
    [data-testid="stSidebarNavCollapseButton"]  { display: none !important; }
    /* Ocultar >> nativo (FAB substitui) */
    [data-testid="collapsedControl"]             { display: none !important; }

    /* Métricas: 2 colunas */
    [data-testid="stHorizontalBlock"] { flex-wrap: wrap !important; }
    [data-testid="stHorizontalBlock"] > [data-testid="column"] {
        flex: 0 0 48% !important;
        min-width: 48% !important;
        max-width: 48% !important;
    }
    [data-testid="stMetric"] { padding: 10px 12px !important; }
    [data-testid="stMetricValue"] { font-size: 18px !important; }
    [data-testid="stMetricLabel"] { font-size: 10px !important; }

    /* Tabs menores */
    button[data-baseweb="tab"] { font-size: 11px !important; padding: 8px 5px !important; }

    /* Inputs — font 16px evita zoom iOS */
    input, select, textarea { font-size: 16px !important; }
    [data-testid="stTextInput"] input  { min-height: 44px !important; }
    [data-testid="stSelectbox"] select { min-height: 44px !important; }

    /* Botões touch */
    .stButton > button { min-height: 48px !important; font-size: 14px !important; }

    /* Mapa */
    iframe { max-height: 380px !important; }

    /* Empty state */
    .empty-state       { padding: 30px 16px; }
    .empty-state-icon  { font-size: 40px; }
    .empty-state-title { font-size: 16px; }
    .empty-state-desc  { font-size: 12px; }

    [data-testid="stSlider"] { padding: 12px 0 !important; }
    .stAlert { font-size: 13px !important; }
}

/* ══ RESPONSIVO — SMARTPHONE PEQUENO (≤ 480px) ════════════════════ */
@media (max-width: 480px) {
    .topbar { padding-left: 56px; }
    .topbar-title { font-size: 13px; }
    .topbar-sub   { display: none; }
    .topbar-badge { display: none; }
    [data-testid="stMetricValue"] { font-size: 16px !important; }
    button[data-baseweb="tab"]    { font-size: 10px !important; padding: 7px 4px !important; }
    iframe                        { max-height: 300px !important; }
}

/* ══ RESPONSIVO — TELA MUITO PEQUENA (≤ 360px) ═══════════════════ */
@media (max-width: 360px) {
    .topbar-title { font-size: 12px; }
    section[data-testid="stSidebar"] { width: 90vw !important; max-width: 300px !important; }
}
</style>
<script>
(function () {
    /* ═══════════════════════════════════════════════════════════════════
       ESTRATÉGIA DEFINITIVA
       ─ Botão FAB criado no <body> via JS (fora do React)
       ─ Click/touch direto no elemento real (sem delegation)
       ─ Sidebar controlada por inline style com !important
         (mais forte que qualquer CSS externo ou do Streamlit)
       ─ Overlay escuro fecha o drawer ao toque
       ═══════════════════════════════════════════════════════════════════ */

    /* ── 1. Ocultar elementos Streamlit ─────────────────────────────── */
    var HIDE_SELS = [
        'button[title*="GitHub"]','button[title*="github"]',
        'a[href*="github.com"]','svg[data-icon="mark-github"]',
        '[data-testid="collapsedControl"]',
        '[data-testid="stSidebarCollapseButton"]',
        '[data-testid="stSidebarNavCollapseButton"]',
        'button[data-testid="baseButton-headerNoPadding"]',
        '.stDeployButton','#MainMenu',
    ];
    function ocultarEls() {
        HIDE_SELS.forEach(function(s) {
            document.querySelectorAll(s).forEach(function(el) {
                el.style.setProperty('display','none','important');
            });
        });
    }

    /* ── 2. Helpers ─────────────────────────────────────────────────── */
    var _aberto   = false;
    var _fab      = null;
    var _overlay  = null;

    function isMobile() { return window.innerWidth <= 768; }

    function getSidebar() {
        return document.querySelector('section[data-testid="stSidebar"]');
    }

    /* Inline style com !important sobrepõe qualquer regra CSS */
    function setSidebarTransform(val) {
        var sb = getSidebar();
        if (!sb) return;
        sb.style.setProperty('transform', val, 'important');
        sb.style.setProperty('display',   'block', 'important');
        sb.style.setProperty('visibility','visible','important');
        if (val === 'translateX(0px)' || val === 'translateX(0)') {
            sb.style.setProperty('box-shadow','6px 0 32px rgba(0,0,0,0.32)','important');
        } else {
            sb.style.setProperty('box-shadow','none','important');
        }
    }

    /* ── 3. Abrir / fechar ──────────────────────────────────────────── */
    function abrir() {
        if (!isMobile()) return;
        setSidebarTransform('translateX(0)');
        if (_overlay) _overlay.style.display = 'block';
        document.body.style.overflow = 'hidden';
        if (_fab) _fab.textContent = '✕';
        _aberto = true;
    }

    function fechar() {
        setSidebarTransform('translateX(-110%)');
        if (_overlay) _overlay.style.display = 'none';
        document.body.style.overflow = '';
        if (_fab) _fab.textContent = '☰';
        _aberto = false;
    }

    function toggle(e) {
        if (e) { e.preventDefault(); e.stopPropagation(); }
        if (_aberto) fechar(); else abrir();
    }

    /* ── 4. Criar FAB (botão flutuante real no <body>) ──────────────── */
    function criarFAB() {
        if (document.getElementById('gf-fab')) {
            _fab = document.getElementById('gf-fab');
            return;
        }
        _fab = document.createElement('button');
        _fab.id = 'gf-fab';
        _fab.textContent = '☰';
        _fab.setAttribute('aria-label','Abrir menu de navegação');
        _fab.setAttribute('type','button');
        /* Inline styles — nunca sobrescritos por React */
        _fab.style.cssText = [
            'position:fixed','top:8px','left:10px','z-index:99999',
            'width:44px','height:44px','border-radius:12px',
            'background:linear-gradient(135deg,#0D47A1,#1565C0)',
            'color:white','border:none','font-size:22px',
            'cursor:pointer','touch-action:manipulation',
            '-webkit-tap-highlight-color:transparent',
            'display:none',                /* começa oculto; tick() decide */
            'align-items:center','justify-content:center',
            'box-shadow:0 3px 14px rgba(13,71,161,0.55)',
            'line-height:1','padding:0',
        ].join(';');

        _fab.addEventListener('click',   toggle);
        _fab.addEventListener('touchend',toggle);
        document.body.appendChild(_fab);
    }

    /* ── 5. Criar Overlay ───────────────────────────────────────────── */
    function criarOverlay() {
        if (document.getElementById('gf-overlay')) {
            _overlay = document.getElementById('gf-overlay');
            return;
        }
        _overlay = document.createElement('div');
        _overlay.id = 'gf-overlay';
        _overlay.style.cssText = [
            'display:none','position:fixed','inset:0',
            'background:rgba(0,0,0,0.48)','z-index:9500',
        ].join(';');
        _overlay.addEventListener('click',   fechar);
        _overlay.addEventListener('touchend',function(e){ e.preventDefault(); fechar(); });
        /* swipe esquerda no overlay fecha */
        var _tsX = 0;
        _overlay.addEventListener('touchstart',function(e){ _tsX=e.touches[0].clientX; },{passive:true});
        _overlay.addEventListener('touchend',function(e){
            if (_tsX - e.changedTouches[0].clientX > 40) fechar();
        },{passive:true});
        document.body.appendChild(_overlay);
    }

    /* ── 6. Swipe p/ fechar dentro do sidebar ───────────────────────── */
    function vincularSwipe() {
        var sb = getSidebar();
        if (!sb || sb._gfSwipe) return;
        sb._gfSwipe = true;
        var _tX = 0;
        sb.addEventListener('touchstart',function(e){ _tX=e.touches[0].clientX; },{passive:true});
        sb.addEventListener('touchend',function(e){
            if (_tX - e.changedTouches[0].clientX > 55) fechar();
        },{passive:true});
    }

    /* ── 7. Tick: chamado no carregamento e a cada mutação do DOM ───── */
    function tick() {
        ocultarEls();
        criarFAB();
        criarOverlay();
        vincularSwipe();
        /* mostrar/ocultar FAB conforme largura da tela */
        if (_fab) _fab.style.display = isMobile() ? 'flex' : 'none';
        /* garantir sidebar off-screen no mobile quando fechado */
        if (isMobile() && !_aberto) setSidebarTransform('translateX(-110%)');
    }

    /* ── 8. Boot ────────────────────────────────────────────────────── */
    if (document.readyState === 'loading') {
        document.addEventListener('DOMContentLoaded', tick);
    } else {
        tick();
    }

    /* Re-aplica a cada re-render do Streamlit */
    new MutationObserver(function() {
        tick();
        /* Se estava aberto, manter aberto após re-render */
        if (_aberto) setSidebarTransform('translateX(0)');
    }).observe(document.body, { childList:true, subtree:true });

    /* Fechar ao girar p/ landscape */
    window.addEventListener('resize', function() {
        if (_fab) _fab.style.display = isMobile() ? 'flex' : 'none';
        if (!isMobile() && _aberto) fechar();
    });
})();
</script>
""", unsafe_allow_html=True)

# ─── Bottom Navigation + Mobile Map Enhancements ─────────────────
st.markdown("""
<style>
/* ══ BOTTOM NAV BAR (mobile only) ════════════════════════════════ */
#gf-bottom-nav {
    display: none;
    position: fixed;
    bottom: 0; left: 0; right: 0;
    height: 62px;
    background: #fff;
    border-top: 1.5px solid #dde3ee;
    box-shadow: 0 -3px 20px rgba(13,71,161,0.10);
    z-index: 8990;
    align-items: stretch;
    justify-content: space-around;
    padding: 0;
}
.gf-nav-item {
    flex: 1;
    display: flex;
    flex-direction: column;
    align-items: center;
    justify-content: center;
    gap: 2px;
    cursor: pointer;
    -webkit-tap-highlight-color: transparent;
    user-select: none;
    padding: 6px 1px 4px;
    border: none;
    background: transparent;
    transition: background 0.12s;
    position: relative;
    min-width: 0;
}
.gf-nav-item:active { background: #f0f4ff; border-radius: 10px; }
.gf-nav-item.ativo::before {
    content: '';
    position: absolute;
    top: 0; left: 15%; right: 15%;
    height: 3px;
    border-radius: 0 0 4px 4px;
    background: linear-gradient(90deg,#0d47a1,#1565c0);
}
.gf-nav-icon {
    font-size: 20px;
    line-height: 1;
    transition: transform 0.12s;
}
.gf-nav-item.ativo .gf-nav-icon { transform: scale(1.12); }
.gf-nav-label {
    font-size: 9px;
    font-weight: 700;
    color: #9aa5b8;
    letter-spacing: 0.2px;
    text-transform: uppercase;
    white-space: nowrap;
    overflow: hidden;
    max-width: 54px;
    text-overflow: ellipsis;
}
.gf-nav-item.ativo .gf-nav-label { color: #1565c0; }

/* ── Padding-bottom para conteúdo não ficar atrás da nav ── */
@media (max-width: 768px) {
    .main .block-container,
    [data-testid="stMain"] .block-container {
        padding-bottom: 74px !important;
    }

    /* Plotly map: altura máxima adaptada + touch */
    [data-testid="stPlotlyChart"] {
        max-height: 370px !important;
        overflow: hidden;
    }
    [data-testid="stPlotlyChart"] > div { height: 370px !important; }
    .js-plotly-plot, .plot-container   { max-height: 370px !important; }
    .js-plotly-plot, .plot-container, .plotly-graph-div {
        touch-action: pan-x pan-y !important;
    }
    /* Toolbar do Plotly — botões maiores para toque */
    .modebar-btn { min-width: 28px !important; min-height: 28px !important; }

    /* Tabelas com scroll horizontal */
    [data-testid="stDataFrame"] table {
        font-size: 11px !important;
    }

    /* Cards ANP menores */
    .fc { padding: 10px 12px 8px !important; }
    .fc-preco { font-size: 20px !important; }

    /* Cards cc (calc custo) */
    .cc-preco  { font-size: 17px !important; }
    .cc-custo  { font-size: 16px !important; }

    /* Expanders — mais altura de toque */
    summary { min-height: 44px; display: flex; align-items: center; }

    /* Popup do posto selecionado — responsivo */
    .main-posto-card { font-size: 12px !important; }
}

@media (max-width: 480px) {
    .gf-nav-icon  { font-size: 18px; }
    .gf-nav-label { font-size: 8px; max-width: 44px; }
    #gf-bottom-nav { height: 56px; }
    .main .block-container { padding-bottom: 62px !important; }
    [data-testid="stPlotlyChart"]      { max-height: 300px !important; }
    [data-testid="stPlotlyChart"] > div{ height: 300px !important; }
    .js-plotly-plot, .plot-container   { max-height: 300px !important; }
}
</style>

<script>
(function() {
    'use strict';

    var NAV = [
        { e:'📍', l:'Mapa',    k:'btn_modo_estado' },
        { e:'🗺️', l:'Rota',   k:'btn_modo_rota' },
        { e:'🔍', l:'Busca',  k:'btn_modo_consulta' },
        { e:'🛣️', l:'Roteiro',k:'btn_modo_roteirizacao' },
        { e:'📋', l:'Salvas', k:'btn_rotas_salvas' },
        { e:'📊', l:'Dash',   k:'btn_dashboard' },
        { e:'🧠', l:'Intel',  k:'btn_inteligencia' },
    ];

    var _nav = null;

    function isMobile() { return window.innerWidth <= 768; }

    /* ── Clicar no botão sidebar correspondente ── */
    function clickSidebarBtn(key) {
        var c = document.querySelector('.st-key-' + key);
        if (!c) return;
        var b = c.querySelector('button');
        if (b) {
            b.click();
            /* Feedback háptico se disponível */
            if (window.navigator && window.navigator.vibrate) {
                window.navigator.vibrate(8);
            }
        }
    }

    /* ── Detectar modo ativo (botão primary no sidebar) ── */
    function getActiveKey() {
        for (var i = 0; i < NAV.length; i++) {
            var c = document.querySelector('.st-key-' + NAV[i].k);
            if (!c) continue;
            if (c.querySelector('[data-testid="stBaseButton-primary"]')) {
                return NAV[i].k;
            }
        }
        return null;
    }

    /* ── Criar a barra ── */
    function criarNav() {
        if (document.getElementById('gf-bottom-nav')) {
            _nav = document.getElementById('gf-bottom-nav');
            return;
        }
        _nav = document.createElement('nav');
        _nav.id = 'gf-bottom-nav';
        _nav.setAttribute('role', 'navigation');
        _nav.setAttribute('aria-label', 'Navegação');

        NAV.forEach(function(item) {
            var btn = document.createElement('button');
            btn.className = 'gf-nav-item';
            btn.setAttribute('data-navkey', item.k);
            btn.setAttribute('aria-label', item.l);
            btn.setAttribute('type', 'button');
            btn.innerHTML =
                '<span class="gf-nav-icon" aria-hidden="true">' + item.e + '</span>' +
                '<span class="gf-nav-label">' + item.l + '</span>';

            /* Touch — preferir touchend para resposta imediata */
            var _tapping = false;
            btn.addEventListener('touchstart', function() { _tapping = true; }, { passive: true });
            btn.addEventListener('touchend', function(e) {
                if (!_tapping) return;
                _tapping = false;
                e.preventDefault();
                clickSidebarBtn(item.k);
                /* Feedback visual imediato */
                _nav.querySelectorAll('.gf-nav-item').forEach(function(el) {
                    el.classList.remove('ativo');
                });
                btn.classList.add('ativo');
            });
            btn.addEventListener('click', function() { clickSidebarBtn(item.k); });

            _nav.appendChild(btn);
        });

        document.body.appendChild(_nav);
    }

    /* ── Atualizar item ativo ── */
    function atualizarAtivo() {
        if (!_nav) return;
        var activeKey = getActiveKey();
        _nav.querySelectorAll('.gf-nav-item').forEach(function(el) {
            var key = el.getAttribute('data-navkey');
            if (key === activeKey) {
                el.classList.add('ativo');
            } else {
                el.classList.remove('ativo');
            }
        });
    }

    /* ── Aplicar touch-action ao mapa Plotly (sem bloquear pan) ── */
    function melhorarMapaToque() {
        document.querySelectorAll('.js-plotly-plot').forEach(function(el) {
            el.style.touchAction = 'pan-x pan-y';
            /* Garantir que o Plotly não bloqueie scroll da página */
            var dragLayer = el.querySelector('.dragcover,.drag,.nsewdrag');
            if (dragLayer) dragLayer.style.touchAction = 'none';
        });
    }

    /* ── Tick principal ── */
    function tick() {
        criarNav();
        atualizarAtivo();
        melhorarMapaToque();
        if (_nav) {
            _nav.style.display = isMobile() ? 'flex' : 'none';
        }
    }

    /* Boot */
    if (document.readyState === 'loading') {
        document.addEventListener('DOMContentLoaded', tick);
    } else {
        setTimeout(tick, 50);
    }

    /* Re-aplica após re-renders do Streamlit */
    var _mutTimer = null;
    new MutationObserver(function() {
        clearTimeout(_mutTimer);
        _mutTimer = setTimeout(tick, 60);
    }).observe(document.body, { childList: true, subtree: true });

    window.addEventListener('resize', function() {
        if (_nav) _nav.style.display = isMobile() ? 'flex' : 'none';
    });

    /* iOS safe-area (notch) — padding extra para iPhones */
    if (CSS && CSS.supports && CSS.supports('padding-bottom', 'env(safe-area-inset-bottom)')) {
        var style = document.createElement('style');
        style.textContent = (
            '#gf-bottom-nav { padding-bottom: env(safe-area-inset-bottom); ' +
            'height: calc(62px + env(safe-area-inset-bottom)); }'
        );
        document.head.appendChild(style);
    }
})();
</script>
""", unsafe_allow_html=True)

# ═══════════════════════════════════════════════════════════════════
#  AUTENTICAÇÃO OAuth2 — Google + Microsoft
# ═══════════════════════════════════════════════════════════════════

# ── Importa a biblioteca (opcional — sem ela o auth fica desativado) ──
try:
    from streamlit_oauth import OAuth2Component
    _OAUTH_LIB_OK = True
except ImportError:
    OAuth2Component = None          # type: ignore
    _OAUTH_LIB_OK   = False


def _oauth_cfg(provider: str) -> bool:
    """True se as credenciais do provider existem em st.secrets."""
    try:
        return (provider in st.secrets
                and bool(st.secrets[provider].get("client_id")))
    except Exception:
        return False


_OAUTH_GOOGLE_OK = _OAUTH_LIB_OK and _oauth_cfg("oauth_google")
_OAUTH_MS_OK     = _OAUTH_LIB_OK and _oauth_cfg("oauth_microsoft")
_OAUTH_ATIVO     = _OAUTH_GOOGLE_OK or _OAUTH_MS_OK


# ── Decodifica payload de JWT sem verificar assinatura ──
def _auth_decode_jwt(token_str: str) -> dict:
    try:
        import base64 as _b64, json as _j
        parts   = token_str.split(".")
        if len(parts) < 2:
            return {}
        payload = parts[1]
        payload += "=" * (4 - len(payload) % 4)
        return _j.loads(_b64.urlsafe_b64decode(payload))
    except Exception:
        return {}


def _auth_user_from_token(token_result: dict, provider: str) -> dict:
    """Extrai nome, e-mail e foto a partir do token OAuth2 retornado."""
    id_token = token_result.get("id_token", "")
    claims   = _auth_decode_jwt(id_token) if id_token else {}
    return {
        "name"    : (claims.get("name")
                     or claims.get("preferred_username")
                     or claims.get("email", "Usuário")),
        "email"   : claims.get("email", ""),
        "picture" : claims.get("picture", ""),
        "provider": provider,
    }


def _auth_login_page():
    """Página de login — design moderno com fundo animado e glassmorphism."""

    st.markdown("""
    <style>
    /* ── Esconde elementos padrão do Streamlit na tela de login ── */
    #MainMenu, header[data-testid="stHeader"], footer,
    [data-testid="stSidebar"], [data-testid="stToolbar"],
    [data-testid="collapsedControl"] { display: none !important; }

    /* ── Fundo animado full-screen ── */
    [data-testid="stAppViewContainer"] {
        background: linear-gradient(135deg, #0a0e27 0%, #0d1b4b 35%, #0a2a6e 65%, #061840 100%);
        min-height: 100vh;
        position: relative;
        overflow: hidden;
    }
    [data-testid="stAppViewContainer"]::before {
        content: "";
        position: fixed;
        inset: 0;
        background:
            radial-gradient(ellipse 80% 60% at 20% 30%, rgba(25,118,210,0.18) 0%, transparent 60%),
            radial-gradient(ellipse 60% 50% at 80% 70%, rgba(0,200,150,0.10) 0%, transparent 55%),
            radial-gradient(ellipse 50% 40% at 60% 10%, rgba(100,181,246,0.12) 0%, transparent 50%);
        animation: bgPulse 8s ease-in-out infinite alternate;
        pointer-events: none;
        z-index: 0;
    }
    @keyframes bgPulse {
        0%   { opacity: 0.7; transform: scale(1); }
        100% { opacity: 1;   transform: scale(1.04); }
    }

    /* ── Partículas decorativas ── */
    [data-testid="stAppViewContainer"]::after {
        content: "";
        position: fixed;
        inset: 0;
        background-image:
            radial-gradient(circle 1.5px at 15% 20%, rgba(255,255,255,0.25) 0%, transparent 100%),
            radial-gradient(circle 1px at 35% 55%, rgba(255,255,255,0.18) 0%, transparent 100%),
            radial-gradient(circle 2px at 70% 15%, rgba(100,181,246,0.35) 0%, transparent 100%),
            radial-gradient(circle 1px at 85% 40%, rgba(255,255,255,0.2) 0%, transparent 100%),
            radial-gradient(circle 1.5px at 50% 80%, rgba(255,255,255,0.15) 0%, transparent 100%),
            radial-gradient(circle 1px at 90% 75%, rgba(100,181,246,0.28) 0%, transparent 100%),
            radial-gradient(circle 1px at 25% 90%, rgba(255,255,255,0.18) 0%, transparent 100%),
            radial-gradient(circle 2px at 60% 45%, rgba(255,255,255,0.12) 0%, transparent 100%);
        pointer-events: none;
        z-index: 0;
        animation: particlesDrift 12s ease-in-out infinite alternate;
    }
    @keyframes particlesDrift {
        0%   { transform: translateY(0px); }
        100% { transform: translateY(-12px); }
    }

    /* ── Centralização do conteúdo ── */
    section[data-testid="stMain"] > div:first-child {
        display: flex !important;
        flex-direction: column !important;
        align-items: center !important;
        justify-content: center !important;
        min-height: 100vh !important;
        padding: 2rem 1rem !important;
        position: relative;
        z-index: 1;
    }
    .block-container {
        padding: 0 !important;
        max-width: 100% !important;
    }

    /* ── Card principal — glassmorphism ── */
    .login-card {
        background: rgba(255, 255, 255, 0.07);
        backdrop-filter: blur(24px);
        -webkit-backdrop-filter: blur(24px);
        border: 1px solid rgba(255, 255, 255, 0.14);
        border-radius: 28px;
        padding: 3rem 2.8rem 2.4rem;
        text-align: center;
        box-shadow:
            0 24px 64px rgba(0, 0, 0, 0.45),
            0 4px 20px rgba(0, 0, 0, 0.3),
            inset 0 1px 0 rgba(255,255,255,0.12);
        animation: cardAppear 0.6s cubic-bezier(0.16,1,0.3,1) both;
        position: relative;
        overflow: hidden;
        max-width: 440px;
        width: 100%;
        margin: 0 auto;
    }
    .login-card::before {
        content: "";
        position: absolute;
        top: -60px; left: -60px;
        width: 180px; height: 180px;
        background: radial-gradient(circle, rgba(100,181,246,0.15) 0%, transparent 70%);
        pointer-events: none;
    }
    @keyframes cardAppear {
        from { opacity: 0; transform: translateY(28px) scale(0.97); }
        to   { opacity: 1; transform: translateY(0) scale(1); }
    }

    /* ── Logo ── */
    .login-logo-ring {
        display: inline-flex;
        align-items: center;
        justify-content: center;
        width: 120px; height: 120px;
        background: linear-gradient(135deg, rgba(13,71,161,0.7), rgba(25,118,210,0.5), rgba(0,200,150,0.25));
        border: 2px solid rgba(100,181,246,0.5);
        border-radius: 30px;
        margin-bottom: 1.4rem;
        box-shadow:
            0 12px 40px rgba(25,118,210,0.45),
            0 4px 16px rgba(0,0,0,0.4),
            inset 0 1px 0 rgba(255,255,255,0.2);
        animation: logoPop 0.7s cubic-bezier(0.16,1,0.3,1) 0.2s both;
        position: relative;
        overflow: hidden;
    }
    .login-logo-ring::before {
        content: "";
        position: absolute;
        inset: 0;
        background: radial-gradient(circle at 30% 30%, rgba(100,181,246,0.3) 0%, transparent 60%);
        pointer-events: none;
    }
    .login-logo-svg {
        width: 72px; height: 72px;
        position: relative; z-index: 1;
        filter: drop-shadow(0 4px 12px rgba(25,118,210,0.6));
    }
    @keyframes logoPop {
        from { opacity: 0; transform: scale(0.6) rotate(-10deg); }
        to   { opacity: 1; transform: scale(1) rotate(0deg); }
    }

    /* ── Título e subtítulo ── */
    .login-title {
        font-size: 1.85rem;
        font-weight: 800;
        background: linear-gradient(135deg, #ffffff 0%, #90caf9 100%);
        -webkit-background-clip: text;
        -webkit-text-fill-color: transparent;
        background-clip: text;
        margin: 0 0 0.25rem;
        letter-spacing: -0.5px;
        animation: fadeUp 0.5s ease 0.3s both;
    }
    .login-badge {
        display: inline-block;
        background: linear-gradient(90deg, rgba(25,118,210,0.35), rgba(0,200,150,0.25));
        border: 1px solid rgba(100,181,246,0.3);
        border-radius: 20px;
        padding: 3px 14px;
        font-size: 0.75rem;
        color: #90caf9;
        font-weight: 600;
        letter-spacing: 0.5px;
        margin-bottom: 1rem;
        animation: fadeUp 0.5s ease 0.35s both;
    }
    .login-sub {
        font-size: 0.875rem;
        color: rgba(255,255,255,0.5);
        margin-bottom: 2rem;
        line-height: 1.55;
        animation: fadeUp 0.5s ease 0.4s both;
    }
    @keyframes fadeUp {
        from { opacity: 0; transform: translateY(14px); }
        to   { opacity: 1; transform: translateY(0); }
    }

    /* ── Divisor ── */
    .login-divider {
        display: flex;
        align-items: center;
        gap: 12px;
        margin: 0.5rem 0 1.4rem;
        animation: fadeUp 0.5s ease 0.45s both;
    }
    .login-divider::before, .login-divider::after {
        content: "";
        flex: 1;
        height: 1px;
        background: linear-gradient(90deg, transparent, rgba(255,255,255,0.12), transparent);
    }
    .login-divider span {
        font-size: 0.72rem;
        color: rgba(255,255,255,0.3);
        font-weight: 500;
        text-transform: uppercase;
        letter-spacing: 1px;
    }

    /* ── Features ── */
    .login-features {
        display: flex;
        justify-content: center;
        gap: 1.4rem;
        margin-bottom: 1.8rem;
        animation: fadeUp 0.5s ease 0.45s both;
    }
    .login-feature {
        display: flex;
        flex-direction: column;
        align-items: center;
        gap: 5px;
    }
    .login-feature-icon {
        width: 38px; height: 38px;
        background: rgba(255,255,255,0.07);
        border: 1px solid rgba(255,255,255,0.1);
        border-radius: 11px;
        display: flex; align-items: center; justify-content: center;
        font-size: 18px;
        transition: transform 0.2s, background 0.2s;
    }
    .login-feature-icon:hover {
        transform: translateY(-3px);
        background: rgba(100,181,246,0.15);
    }
    .login-feature-label {
        font-size: 0.65rem;
        color: rgba(255,255,255,0.4);
        font-weight: 500;
        text-align: center;
        max-width: 60px;
        line-height: 1.3;
    }

    /* ── Botões OAuth (override streamlit) — discreto ── */
    div[data-testid="stButton"] > button,
    div.stButton > button {
        background: rgba(255,255,255,0.05) !important;
        border: 1px solid rgba(255,255,255,0.10) !important;
        color: rgba(255,255,255,0.6) !important;
        border-radius: 10px !important;
        font-weight: 500 !important;
        font-size: 0.78rem !important;
        padding: 0.45rem 1rem !important;
        transition: all 0.22s cubic-bezier(0.4,0,0.2,1) !important;
        backdrop-filter: blur(8px) !important;
        box-shadow: none !important;
        max-width: 220px !important;
        margin: 0 auto !important;
    }
    div[data-testid="stButton"] > button:hover,
    div.stButton > button:hover {
        background: rgba(255,255,255,0.10) !important;
        border-color: rgba(255,255,255,0.20) !important;
        color: rgba(255,255,255,0.85) !important;
        transform: translateY(-1px) !important;
        box-shadow: 0 3px 12px rgba(0,0,0,0.2) !important;
    }
    /* Centraliza o botão OAuth */
    div[data-testid="stButton"],
    div.stButton {
        display: flex !important;
        justify-content: center !important;
    }

    /* ── Rodapé ── */
    .login-footer {
        font-size: 0.7rem;
        color: rgba(255,255,255,0.25);
        margin-top: 1.6rem;
        line-height: 1.6;
        animation: fadeUp 0.5s ease 0.6s both;
    }
    .login-footer a { color: rgba(144,202,249,0.5); text-decoration: none; }

    /* ── Selo de versão ── */
    .login-version {
        position: fixed;
        bottom: 1.2rem; right: 1.4rem;
        font-size: 0.65rem;
        color: rgba(255,255,255,0.2);
        font-weight: 500;
        z-index: 10;
    }
    </style>
    """, unsafe_allow_html=True)

    # ── Layout: coluna central ──
    _, _c, _ = st.columns([1, 2.4, 1])
    with _c:
        # ── Card principal ──
        st.markdown("""
        <div class='login-card'>
          <div class='login-logo-ring'>
            <svg class='login-logo-svg' viewBox="0 0 72 72" fill="none" xmlns="http://www.w3.org/2000/svg">
              <!-- Fundo do mapa -->
              <rect x="6" y="10" width="60" height="52" rx="6" fill="rgba(13,71,161,0.6)" stroke="rgba(100,181,246,0.4)" stroke-width="1.2"/>
              <!-- Grade do mapa -->
              <line x1="6" y1="27" x2="66" y2="27" stroke="rgba(100,181,246,0.2)" stroke-width="0.8"/>
              <line x1="6" y1="44" x2="66" y2="44" stroke="rgba(100,181,246,0.2)" stroke-width="0.8"/>
              <line x1="24" y1="10" x2="24" y2="62" stroke="rgba(100,181,246,0.2)" stroke-width="0.8"/>
              <line x1="48" y1="10" x2="48" y2="62" stroke="rgba(100,181,246,0.2)" stroke-width="0.8"/>
              <!-- Rota principal (estrada) -->
              <path d="M12 52 Q22 44 30 36 Q40 26 52 20" stroke="#64B5F6" stroke-width="2.8" stroke-linecap="round" stroke-linejoin="round"/>
              <path d="M12 52 Q22 44 30 36 Q40 26 52 20" stroke="rgba(255,255,255,0.15)" stroke-width="5" stroke-linecap="round"/>
              <!-- Rota secundária -->
              <path d="M30 36 Q38 42 52 48" stroke="rgba(100,181,246,0.5)" stroke-width="1.6" stroke-linecap="round" stroke-dasharray="3 3"/>
              <!-- Marcador origem (verde) -->
              <circle cx="12" cy="52" r="5" fill="#2E7D32" stroke="white" stroke-width="1.5"/>
              <circle cx="12" cy="52" r="2.5" fill="white"/>
              <!-- Marcador destino (vermelho) -->
              <circle cx="52" cy="20" r="5" fill="#C62828" stroke="white" stroke-width="1.5"/>
              <circle cx="52" cy="20" r="2.5" fill="white"/>
              <!-- Posto intermediário (laranja) -->
              <polygon points="30,31 33,38 27,38" fill="#FF8F00" stroke="white" stroke-width="1.2"/>
              <!-- Legenda mini -->
              <rect x="38" y="48" width="22" height="10" rx="3" fill="rgba(0,0,0,0.35)" stroke="rgba(255,255,255,0.1)" stroke-width="0.5"/>
              <circle cx="43" cy="53" r="2" fill="#2E7D32"/>
              <circle cx="53" cy="53" r="2" fill="#C62828"/>
            </svg>
          </div>
          <div class='login-title'>Estudo de Rede</div>
          <div class='login-badge'>Gestão de Frotas</div>
          <div class='login-sub'>
            Plataforma inteligente para análise de postos,<br>
            roteirização e monitoramento de frota.
          </div>

          <div class='login-features'>
            <div class='login-feature'>
              <div class='login-feature-icon'>⛽</div>
              <div class='login-feature-label'>Postos ANP</div>
            </div>
            <div class='login-feature'>
              <div class='login-feature-icon'>🗺️</div>
              <div class='login-feature-label'>Roteirização</div>
            </div>
            <div class='login-feature'>
              <div class='login-feature-icon'>📊</div>
              <div class='login-feature-label'>Dashboard</div>
            </div>
            <div class='login-feature'>
              <div class='login-feature-icon'>🚛</div>
              <div class='login-feature-label'>GF</div>
            </div>
          </div>

          <div class='login-divider'><span>Acesso seguro</span></div>
        </div>
        """, unsafe_allow_html=True)

        # ── Redirect URI ──
        try:
            _redir = st.secrets.get("redirect_uri", "http://localhost:8501")
        except Exception:
            _redir = "http://localhost:8501"

        # ── Botão Google ──
        if _OAUTH_GOOGLE_OK:
            _g_oauth = OAuth2Component(
                client_id=st.secrets["oauth_google"]["client_id"],
                client_secret=st.secrets["oauth_google"]["client_secret"],
                authorize_endpoint="https://accounts.google.com/o/oauth2/v2/auth",
                token_endpoint="https://oauth2.googleapis.com/token",
                refresh_token_endpoint="https://oauth2.googleapis.com/token",
                revoke_token_endpoint="https://oauth2.googleapis.com/revoke",
            )
            _res_g = _g_oauth.authorize_button(
                name="  Continuar com Google",
                redirect_uri=_redir,
                scope="openid email profile",
                use_container_width=True,
                pkce="S256",
                icon="https://www.google.com/favicon.ico",
                key="oauth_btn_google",
            )
            if _res_g and "token" in _res_g:
                st.session_state["_auth_user"] = _auth_user_from_token(
                    _res_g["token"], "google"
                )
                st.rerun()

        # ── Espaço entre botões ──
        if _OAUTH_GOOGLE_OK and _OAUTH_MS_OK:
            st.markdown(
                "<div style='text-align:center;font-size:11px;color:rgba(255,255,255,0.25);"
                "margin:8px 0;letter-spacing:1px;text-transform:uppercase'>ou</div>",
                unsafe_allow_html=True,
            )

        # ── Botão Microsoft ──
        if _OAUTH_MS_OK:
            _ms_oauth = OAuth2Component(
                client_id=st.secrets["oauth_microsoft"]["client_id"],
                client_secret=st.secrets["oauth_microsoft"]["client_secret"],
                authorize_endpoint=(
                    "https://login.microsoftonline.com/common/oauth2/v2.0/authorize"
                ),
                token_endpoint=(
                    "https://login.microsoftonline.com/common/oauth2/v2.0/token"
                ),
                refresh_token_endpoint=(
                    "https://login.microsoftonline.com/common/oauth2/v2.0/token"
                ),
            )
            _res_ms = _ms_oauth.authorize_button(
                name="  Continuar com Microsoft",
                redirect_uri=_redir,
                scope="openid email profile User.Read",
                use_container_width=True,
                pkce="S256",
                key="oauth_btn_microsoft",
            )
            if _res_ms and "token" in _res_ms:
                st.session_state["_auth_user"] = _auth_user_from_token(
                    _res_ms["token"], "microsoft"
                )
                st.rerun()

        st.markdown(
            "<p class='login-footer'>🔒 Acesso restrito a colaboradores autorizados.<br>"
            "Em caso de dúvidas, contate o administrador do sistema.</p>",
            unsafe_allow_html=True,
        )
        st.markdown("<div class='login-version'>v2.0 · Pró-Frotas</div>", unsafe_allow_html=True)


# ── Inicializar estado de autenticação ──────────────────────────────
if "_auth_user" not in st.session_state:
    st.session_state["_auth_user"] = None

# ── Verificar autenticação — bloqueia o app se não autenticado ──────
if _OAUTH_ATIVO and st.session_state["_auth_user"] is None:
    _auth_login_page()
    st.stop()

# ── Verificar controle de acesso (allowlist / blacklist) ────────────
if _OAUTH_ATIVO and st.session_state.get("_auth_user"):
    _email_logado = (st.session_state["_auth_user"] or {}).get("email", "")
    if _email_logado and not st.session_state.get("_acesso_verificado"):
        _acesso_ok, _acesso_motivo = _db_verificar_acesso(_email_logado)
        st.session_state["_acesso_verificado"] = True
        if not _acesso_ok:
            # Limpa sessão e mostra tela de bloqueio
            _nome_bloq = (st.session_state["_auth_user"] or {}).get("name", _email_logado)
            st.session_state["_auth_user"] = None
            st.session_state["_acesso_verificado"] = False
            st.markdown("""
            <style>
            #MainMenu, header, footer, [data-testid="stSidebar"],
            [data-testid="stToolbar"] { display: none !important; }
            [data-testid="stAppViewContainer"] {
                background: linear-gradient(135deg, #0a0e27 0%, #0d1b4b 100%);
                display: flex; align-items: center; justify-content: center;
                min-height: 100vh;
            }
            </style>
            """, unsafe_allow_html=True)
            _, _cc, _ = st.columns([1, 2, 1])
            with _cc:
                st.markdown(f"""
                <div style="background:rgba(255,255,255,0.07);backdrop-filter:blur(20px);
                            border:1px solid rgba(255,255,255,0.12);border-radius:20px;
                            padding:2.5rem 2rem;text-align:center;margin-top:15vh;">
                  <div style="font-size:56px;margin-bottom:1rem;">🚫</div>
                  <div style="font-size:1.4rem;font-weight:700;color:#fff;margin-bottom:.5rem;">
                    Acesso Negado
                  </div>
                  <div style="font-size:.9rem;color:rgba(255,255,255,.5);margin-bottom:1.5rem;">
                    {_acesso_motivo}
                  </div>
                  <div style="font-size:.75rem;color:rgba(255,255,255,.3);">
                    {_email_logado}<br>Entre em contato com o administrador do sistema.
                  </div>
                </div>
                """, unsafe_allow_html=True)
            st.stop()


# ─── Constantes ───────────────────────────────────────────────────
API_BASE_URL = "https://revendedoresapi.anp.gov.br"
ENDPOINT     = "/v1/combustivel"
NOMINATIM    = "https://nominatim.openstreetmap.org/search"

# Headers que simulam um navegador — evita bloqueio 403 da API ANP
HEADERS_ANP = {
    "User-Agent": (
        "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/124.0.0.0 Safari/537.36"
    ),
    "Accept":          "application/json, text/plain, */*",
    "Accept-Language": "pt-BR,pt;q=0.9,en-US;q=0.8,en;q=0.7",
    "Referer":         "https://postos.anp.gov.br/",
    "Origin":          "https://postos.anp.gov.br",
}

UFS = [
    "AC","AL","AM","AP","BA","CE","DF","ES","GO","MA","MG","MS",
    "MT","PA","PB","PE","PI","PR","RJ","RN","RO","RR","RS","SC",
    "SE","SP","TO",
]

# Mapeamento UF -> nome completo (padrão ANP nas planilhas)
UF_NOME = {
    "AC":"Acre","AL":"Alagoas","AM":"Amazonas","AP":"Amapá",
    "BA":"Bahia","CE":"Ceará","DF":"Distrito Federal","ES":"Espírito Santo",
    "GO":"Goiás","MA":"Maranhão","MG":"Minas Gerais","MS":"Mato Grosso do Sul",
    "MT":"Mato Grosso","PA":"Pará","PB":"Paraíba","PE":"Pernambuco",
    "PI":"Piauí","PR":"Paraná","RJ":"Rio de Janeiro","RN":"Rio Grande do Norte",
    "RO":"Rondônia","RR":"Roraima","RS":"Rio Grande do Sul","SC":"Santa Catarina",
    "SE":"Sergipe","SP":"São Paulo","TO":"Tocantins",
}

# URL da página de levantamento de preços semanais da ANP
ANP_PRECOS_URL = (
    "https://www.gov.br/anp/pt-br/assuntos/precos-e-defesa-da-concorrencia"
    "/precos/levantamento-de-precos-de-combustiveis-ultimas-semanas-pesquisadas"
)

# Produtos-chave exibidos (ordem de exibição)
PRODUTOS_CHAVE = [
    "GASOLINA COMUM",
    "GASOLINA ADITIVADA",
    "ETANOL HIDRATADO COMBUSTÍVEL",
    "ÓLEO DIESEL",
    "ÓLEO DIESEL S10",
    "GNV",
    "GLP",
]
# Chaves já normalizadas (sem acento, uppercase) — igual ao _anp_norm()
PRODUTO_CURTO = {
    # ANP canônicos
    "GASOLINA COMUM":                    "⛽ Gasolina",
    "GASOLINA ADITIVADA":                "⛽ Gasolina Aditivada",
    "ETANOL HIDRATADO COMBUSTIVEL":      "🌿 Etanol",
    "ETANOL HIDRATADO":                  "🌿 Etanol",
    "OLEO DIESEL":                       "🛢️ Diesel Comum",
    "OLEO DIESEL S10":                   "🛢️ Diesel S10",
    "GNV":                               "💨 GNV",
    "GLP":                               "🔵 GLP",
    "GAS NATURAL COMPRIMIDO":            "💨 GNV",
    "GAS NATURAL VEICULAR":              "💨 GNV",
    "GAS LIQUEFEITO DE PETROLEO":        "🔵 GLP",
    "GAS LIQUEFEITO DO PETROLEO":        "🔵 GLP",
    # Variantes comuns em planilhas de clientes (Preço Posto)
    "GASOLINA":                          "⛽ Gasolina",
    "GASOLINA C":                        "⛽ Gasolina",
    "GASOLINA ALTA OCTANAGEM":           "⛽ Gasolina Aditivada",
    "GASOLINA PREMIUM":                  "⛽ Gasolina Aditivada",
    "DIESEL":                            "🛢️ Diesel Comum",
    "DIESEL COMUM":                      "🛢️ Diesel Comum",
    "DIESEL S500":                       "🛢️ Diesel Comum",
    "DIESEL S-500":                      "🛢️ Diesel Comum",
    "DIESEL S-500 COMUM":               "🛢️ Diesel Comum",
    "DIESEL S-500 ADITIVADO":           "🛢️ Diesel Comum",
    "DIESEL S10":                        "🛢️ Diesel S10",
    "DIESEL S-10":                       "🛢️ Diesel S10",
    "DIESEL S-10 COMUM":                "🛢️ Diesel S10",
    "DIESEL S-10 ADITIVADO":            "🛢️ Diesel S10",
    "OLEO DIESEL S-10":                  "🛢️ Diesel S10",
    "ETANOL":                            "🌿 Etanol",
    "ETANOL COMUM":                      "🌿 Etanol",
    "ETANOL ADITIVADO":                  "🌿 Etanol",
    "ALCOOL":                            "🌿 Etanol",
    "ALCOOL HIDRATADO":                  "🌿 Etanol",
}

# Mapeamento PK da planilha "Preço Posto" → PK canônico ANP
# Resolve casos onde o nome do cliente difere completamente do nome ANP.
# Variantes do mesmo tipo são consolidadas no mesmo PK canônico para agrupamento.
_PP_PARA_ANP_PK: dict = {
    # ── Gasolina ──────────────────────────────────────────────────────
    "GASOLINA":                          "GASOLINA COMUM",
    "GASOLINA COMUM":                    "GASOLINA COMUM",
    "GASOLINA C":                        "GASOLINA COMUM",
    "GASOLINA ADITIVADA":                "GASOLINA ADITIVADA",
    "GASOLINA PREMIUM":                  "GASOLINA ADITIVADA",
    "GASOLINA ALTA OCTANAGEM":           "GASOLINA ADITIVADA",
    "GASOLINA PODIUM":                   "GASOLINA ADITIVADA",
    "GASOLINA FORMULA":                  "GASOLINA ADITIVADA",
    # ── Etanol / Álcool ───────────────────────────────────────────────
    "ETANOL":                            "ETANOL HIDRATADO COMBUSTIVEL",
    "ETANOL COMUM":                      "ETANOL HIDRATADO COMBUSTIVEL",
    "ETANOL ADITIVADO":                  "ETANOL HIDRATADO COMBUSTIVEL",
    "ETANOL HIDRATADO":                  "ETANOL HIDRATADO COMBUSTIVEL",
    "ETANOL HIDRATADO COMBUSTIVEL":      "ETANOL HIDRATADO COMBUSTIVEL",
    "ALCOOL":                            "ETANOL HIDRATADO COMBUSTIVEL",
    "ALCOOL HIDRATADO":                  "ETANOL HIDRATADO COMBUSTIVEL",
    # ── Diesel S500 (= Óleo Diesel / Diesel Comum) ────────────────────
    "DIESEL":                            "OLEO DIESEL",
    "DIESEL COMUM":                      "OLEO DIESEL",
    "OLEO DIESEL":                       "OLEO DIESEL",
    "DIESEL S500":                       "OLEO DIESEL",
    "DIESEL S-500":                      "OLEO DIESEL",
    "DIESEL S 500":                      "OLEO DIESEL",
    "DIESEL S-500 COMUM":               "OLEO DIESEL",
    "DIESEL S-500 ADITIVADO":           "OLEO DIESEL",
    "DIESEL S500 COMUM":                "OLEO DIESEL",
    "DIESEL S500 ADITIVADO":            "OLEO DIESEL",
    "OLEO DIESEL S500":                  "OLEO DIESEL",
    "OLEO DIESEL S-500":                 "OLEO DIESEL",
    # ── Diesel S10 ────────────────────────────────────────────────────
    "DIESEL S10":                        "OLEO DIESEL S10",
    "DIESEL S-10":                       "OLEO DIESEL S10",
    "DIESEL S 10":                       "OLEO DIESEL S10",
    "DIESEL S-10 COMUM":                "OLEO DIESEL S10",
    "DIESEL S-10 ADITIVADO":            "OLEO DIESEL S10",
    "DIESEL S10 COMUM":                  "OLEO DIESEL S10",
    "DIESEL S10 ADITIVADO":              "OLEO DIESEL S10",
    "OLEO DIESEL S10":                   "OLEO DIESEL S10",
    "OLEO DIESEL S-10":                  "OLEO DIESEL S10",
    "OLEO DIESEL S 10":                  "OLEO DIESEL S10",
    # ── GNV ───────────────────────────────────────────────────────────
    "GNV":                               "GNV",
    "GAS NATURAL VEICULAR":              "GNV",
    "GAS NATURAL COMPRIMIDO":            "GNV",
    # ── GLP ───────────────────────────────────────────────────────────
    "GLP":                               "GLP",
    "GAS LIQUEFEITO DE PETROLEO":        "GLP",
    "GAS LIQUEFEITO DO PETROLEO":        "GLP",
}

BBOX_UFS = {
    "AC": (-11.15,-73.99,-7.11,-66.63), "AL": (-10.50,-38.24,-8.81,-35.15),
    "AM": ( -9.82,-73.80, 2.25,-56.10), "AP": ( -1.24,-52.00, 4.44,-49.87),
    "BA": (-18.36,-46.62,-8.53,-37.35), "CE": ( -7.86,-41.44,-2.77,-37.25),
    "DF": (-16.06,-48.28,-15.50,-47.31),"ES": (-21.31,-41.88,-17.87,-39.69),
    "GO": (-19.49,-53.24,-12.40,-45.92),"MA": (-10.25,-48.75,-1.02,-41.81),
    "MG": (-22.92,-51.05,-14.24,-39.86),"MS": (-24.06,-58.16,-17.16,-50.92),
    "MT": (-18.05,-61.00,-7.35,-50.22), "PA": ( -9.84,-58.90, 2.59,-46.02),
    "PB": ( -8.27,-38.82,-6.03,-34.79), "PE": ( -9.49,-41.36,-7.29,-32.41),
    "PI": (-10.95,-45.99,-2.74,-40.37), "PR": (-26.72,-54.62,-22.52,-48.02),
    "RJ": (-23.37,-44.89,-20.76,-40.96),"RN": ( -6.99,-38.60,-4.83,-34.97),
    "RO": (-13.69,-66.09,-7.96,-59.77), "RR": ( -1.55,-64.82, 5.27,-58.90),
    "RS": (-33.75,-53.70,-27.10,-49.69),"SC": (-29.36,-53.84,-25.96,-48.37),
    "SE": (-11.57,-38.25,-9.52,-36.39), "SP": (-25.31,-53.11,-19.78,-44.16),
    "TO": (-13.46,-50.74,-5.18,-45.74),
}

# Paleta de fallback — usada apenas para distribuidoras não mapeadas em CORES_MARCAS.
# Cores vibrantes e distintas entre si para fácil diferenciação visual.
CORES_FALLBACK = [
    "#00ACC1","#FB8C00","#8D6E63","#546E7A","#EC407A","#66BB6A",
    "#AB47BC","#EF5350","#26A69A","#7E57C2","#D4E157","#FF7043",
    "#29B6F6","#9CCC65","#FFA726","#26C6DA","#EF9A9A","#80CBC4",
    "#CE93D8","#A5D6A7","#FFCC02","#90CAF9",
]

# ── Cores padronizadas por marca ──────────────────────────────────────────────
# Identificadas por substring no nome da distribuidora (case-insensitive).
# A ordem importa: substrings mais específicas devem vir antes das genéricas.
CORES_MARCAS = {
    # ── Grandes redes nacionais ──────────────────────────────────
    "IPIRANGA":          "#FFB300",  # amarelo âmbar — cor da marca Ipiranga
    "ULTRAPAR":          "#FFB300",  # grupo controlador da Ipiranga
    "VIBRA":             "#43A047",  # verde — cor da marca Vibra Energy
    "BR DISTRIBUIDORA":  "#43A047",  # nome comercial anterior da Vibra
    "PETROBRAS DIST":    "#43A047",  # razão social anterior da Vibra
    "RAIZEN":            "#E53935",  # vermelho — Raízen Combustíveis/Shell
    "RAÍZEN":            "#E53935",
    "SHELL":             "#E53935",  # Raízen opera sob a marca Shell no Brasil
    "BANDEIRA BRANCA":   "#7B1FA2",  # roxo — postos independentes
    "SEM BANDEIRA":      "#7B1FA2",  # variação do nome no cadastro ANP
    # ── Redes regionais e demais ─────────────────────────────────
    "ALESAT":            "#F57C00",  # laranja — Alesat Combustíveis
    "ALE COMB":          "#F57C00",  # ALE Combustíveis
    "SABBA":             "#00838F",  # teal — Sabbá (região Norte/AM)
    "SABBÁ":             "#00838F",
    "DISLUB":            "#455A64",  # cinza ardósia — Dislub Equador
    "PETRONIT":          "#37474F",  # cinza chumbo — Petronit
    "PLURAL":            "#1976D2",  # azul — Plural Distribuidora
    "REPSOL":            "#003087",  # azul escuro — Repsol Sinopec
    "TEXACO":            "#D32F2F",  # vermelho escuro — Texaco
    "NACIONAL GAS":      "#5D4037",  # marrom — Nacional Gás
    "NACIONAL GÁS":      "#5D4037",
    "COSAN":             "#F44336",  # vermelho — Cosan (holding Raízen)
    "PETRO RIO":         "#0288D1",  # azul claro — PetroRio
    "PETROIL":           "#0097A7",  # ciano — Petroil
    "COMFORGAS":         "#558B2F",  # verde musgo — Comforgas
    "TERPASTOS":         "#795548",  # marrom — Terpastos
    "GLP":               "#FDD835",  # amarelo vivo — distribuidoras de GLP
    "LIQUIGAS":          "#1565C0",  # azul — Liquigás
    "ULTRAGAZ":          "#E91E63",  # rosa — Ultragaz
    "COPAGAZ":           "#FF6F00",  # âmbar escuro — Copagaz
    "SUPERGASB":         "#6A1B9A",  # roxo escuro — Supergasbras
    "SUPERGASB":         "#6A1B9A",
    "NACIONAL":          "#4E342E",  # marrom escuro — Nacional (genérico)
}

# Cor e estilo do marcador Gestão de Frotas
COR_PF_FILL  = "#1565C0"   # azul — identificação visual do credenciamento
COR_PF_BORDA = "#0D47A1"   # azul escuro

# Cor e estilo do marcador Rodo Rede (perfil de venda especial)
# Deep Orange — distinto do amarelo-âmbar #FFB300 (Ipiranga) e do #FF8F00 (Cercados)
COR_RR_FILL  = "#E64A19"   # deep orange — identidade visual Rodo Rede
COR_RR_BORDA = "#BF360C"   # deep orange escuro — borda do marcador
PERFIL_RODO_REDE = "RODO REDE"  # valor normalizado para comparação

# Helper: verifica se a distribuidora é Ipiranga (ou grupo Ultrapar)
def _is_ipiranga(distribuidora: str) -> bool:
    d = str(distribuidora).upper().strip()
    return "IPIRANGA" in d or "ULTRAPAR" in d


def _cor_marca(distribuidora: str) -> str:
    """Retorna a cor do pin para qualquer distribuidora.

    Ordem de resolução:
    1. Marcas conhecidas em CORES_MARCAS → cor da identidade da marca.
    2. Desconhecidas → cor determinística via MD5 do nome (sempre igual,
       independente do estado ou ordem de carregamento).
    """
    d = str(distribuidora).upper().strip()
    for marca, cor in CORES_MARCAS.items():
        if marca in d:
            return cor
    # Fallback determinístico: mesma distribuidora → mesma cor em qualquer consulta
    h = int(hashlib.md5(d.encode()).hexdigest(), 16)
    return CORES_FALLBACK[h % len(CORES_FALLBACK)]


# Limite de marcadores no mapa. Acima disso os postos são amostrados
# (Gestão de Frotas têm prioridade) e o popup é simplificado.
# → evita serializar 5-6 MB de HTML para estados como SP (4 000+ postos).
MAX_MAPA_POSTOS = 5000  # Plotly WebGL suporta 10 000+ marcadores sem travar


# ═══════════════════════════════════════════════════════════════════
#  SISTEMA DE LOGS DE USO
# ═══════════════════════════════════════════════════════════════════

import csv as _csv_mod
import os as _os_mod

_LOG_PATH = _os_mod.path.join(
    _os_mod.path.dirname(_os_mod.path.abspath(__file__)),
    "_usage_logs.csv"
)
_LOG_FIELDS = [
    "timestamp", "data", "hora", "ip", "session_id",
    "modo", "uf", "municipio", "acao", "detalhe", "user_agent",
    "user_email", "user_name", "auth_provider",
]


def _get_session_id() -> str:
    """ID de sessão baseado no id() do session_state — estável dentro da sessão."""
    if "_session_id" not in st.session_state:
        import hashlib as _hl
        st.session_state["_session_id"] = _hl.md5(
            str(id(st.session_state)).encode()
        ).hexdigest()[:10]
    return st.session_state["_session_id"]


def _get_client_ip() -> str:
    """Tenta obter IP real do cliente via headers Streamlit (≥1.31)."""
    try:
        _h = st.context.headers
        for _hk in ["X-Forwarded-For", "X-Real-Ip", "Cf-Connecting-Ip", "True-Client-Ip"]:
            _v = _h.get(_hk, "")
            if _v:
                return _v.split(",")[0].strip()
        return _h.get("Remote-Addr", "—")
    except Exception:
        return "—"


def _get_user_agent() -> str:
    """User-Agent do navegador."""
    try:
        return st.context.headers.get("User-Agent", "—")[:200]
    except Exception:
        return "—"


def _log_acesso(acao: str, detalhe: str = "", modo_override: str = None):
    """
    Registra um evento de uso.
    Grava em st.session_state['_uso_logs'] (sempre) e no arquivo CSV (se possível).
    Inclui dados do usuário autenticado via OAuth2 quando disponível.
    """
    _now   = datetime.now()
    _auth  = st.session_state.get("_auth_user") or {}
    _entry = {
        "timestamp":     _now.strftime("%Y-%m-%d %H:%M:%S"),
        "data":          _now.strftime("%d/%m/%Y"),
        "hora":          _now.strftime("%H:%M:%S"),
        "ip":            _get_client_ip(),
        "session_id":    _get_session_id(),
        "user_email":    _auth.get("email",    "—"),
        "user_name":     _auth.get("name",     "—"),
        "auth_provider": _auth.get("provider", "—"),
        "modo":          modo_override or st.session_state.get("modo_selecionado", "—"),
        "uf":            str(st.session_state.get("uf_input") or "—"),
        "municipio":     str(st.session_state.get("municipio_input") or "—"),
        "acao":          acao,
        "detalhe":       str(detalhe)[:200],
        "user_agent":    _get_user_agent(),
    }

    # ── Guarda em memória (sessão atual) ──
    if "_uso_logs" not in st.session_state:
        st.session_state["_uso_logs"] = []
    st.session_state["_uso_logs"].append(_entry)

    # ── Persiste no Supabase ──
    _db_gravar_log(_entry)

    # ── Persiste em arquivo CSV (fallback local) ──
    try:
        _existe = _os_mod.path.exists(_LOG_PATH)
        if _existe:
            try:
                with open(_LOG_PATH, "r", encoding="utf-8") as _chk:
                    _hdr = next(_csv_mod.reader(_chk), [])
                if _hdr != _LOG_FIELDS:
                    _os_mod.remove(_LOG_PATH)
                    _existe = False
            except Exception:
                _existe = False
        with open(_LOG_PATH, "a", newline="", encoding="utf-8") as _f:
            _w = _csv_mod.DictWriter(_f, fieldnames=_LOG_FIELDS,
                                     extrasaction="ignore", quoting=_csv_mod.QUOTE_ALL)
            if not _existe:
                _w.writeheader()
            _w.writerow(_entry)
    except Exception:
        pass  # Falha silenciosa em ambientes read-only (Streamlit Cloud)


def _log_ler_arquivo() -> list:
    """Lê logs do Supabase (prioritário) + arquivo local + sessão atual."""
    _rows = []

    # ── 1. Tenta ler do Supabase ──
    _rows = _db_ler_logs(limite=2000)

    # ── 2. Fallback: arquivo CSV local (se banco vazio) ──
    if not _rows:
        try:
            if _os_mod.path.exists(_LOG_PATH):
                with open(_LOG_PATH, "r", encoding="utf-8") as _f:
                    _rows = list(_csv_mod.DictReader(_f))
        except Exception:
            pass

    # ── 3. Merge com sessão atual (evita duplicatas) ──
    _sess_logs = st.session_state.get("_uso_logs", [])
    _seen = {(r.get("timestamp"), r.get("session_id")) for r in _rows}
    for _r in _sess_logs:
        _key = (_r.get("timestamp"), _r.get("session_id"))
        if _key not in _seen:
            _rows.append(_r)
            _seen.add(_key)
    return _rows


# ── Registra acesso inicial da sessão (uma vez por sessão) ──────────
if not st.session_state.get("_log_inicio_ok", False):
    _auth_u = st.session_state.get("_auth_user") or {}
    if _auth_u:
        # Usuário autenticado: registra LOGIN com dados do provider
        _log_acesso(
            "LOGIN",
            detalhe=(
                f"provider={_auth_u.get('provider','—')} | "
                f"email={_auth_u.get('email','—')} | "
                f"name={_auth_u.get('name','—')}"
            ),
        )
    else:
        _log_acesso("SESSÃO_INÍCIO", detalhe="App carregado")
    st.session_state["_log_inicio_ok"] = True


# ═══════════════════════════════════════════════════════════════════
#  INTELIGÊNCIA DE DADOS — Histórico de Preços · Score · Alertas
# ═══════════════════════════════════════════════════════════════════

import json  as _json_mod
import math  as _math_mod

_INTEL_PATH = _os_mod.path.join(
    _os_mod.path.dirname(_os_mod.path.abspath(__file__)),
    "_intel_data.json",
)

# ── Persistência ────────────────────────────────────────────────────

def _intel_load() -> dict:
    """Carrega dados de inteligência do JSON (cache por sessão)."""
    if st.session_state.get("_intel_loaded"):
        return st.session_state.get("_intel_data", {})
    try:
        if _os_mod.path.exists(_INTEL_PATH):
            with open(_INTEL_PATH, "r", encoding="utf-8") as _f:
                _data = _json_mod.load(_f)
        else:
            _data = {}
    except Exception:
        _data = {}
    _data.setdefault("historico", {})   # {cnpj14: [{data, preco, combustivel, nome, municipio, uf}]}
    _data.setdefault("limiar",    {})   # {combustivel: float}
    _data.setdefault("last_report", None)
    st.session_state["_intel_data"]   = _data
    st.session_state["_intel_loaded"] = True
    return _data


def _intel_save(data: dict) -> bool:
    """Persiste dados de inteligência no JSON."""
    try:
        with open(_INTEL_PATH, "w", encoding="utf-8") as _f:
            _json_mod.dump(data, _f, ensure_ascii=False, separators=(",", ":"))
        st.session_state["_intel_data"]   = data
        st.session_state["_intel_loaded"] = True
        return True
    except Exception:
        return False


# ── Registro de observações de preço ───────────────────────────────

def _hist_record_lote(
    df: "pd.DataFrame",
    combustivel: str,
    data_str: str = None,
) -> int:
    """
    Registra preços de um DataFrame de postos no histórico.
    df precisa de: cnpj (ou _cnpj_norm), razaoSocial/nome, coluna de preço.
    Retorna quantidade de novos registros adicionados.
    """
    if df is None or df.empty:
        return 0
    if data_str is None:
        data_str = datetime.now().strftime("%Y-%m-%d")

    # detecta coluna de preço
    _preco_col = None
    for _c in df.columns:
        _cl = _c.lower()
        if "preco" in _cl or "preço" in _cl or _cl.startswith("_preco"):
            _preco_col = _c
            break
    if _preco_col is None:
        return 0

    # detecta coluna de CNPJ
    _cnpj_col = ("_cnpj_norm" if "_cnpj_norm" in df.columns
                 else ("cnpj" if "cnpj" in df.columns else None))
    if _cnpj_col is None:
        return 0

    _nome_col = next((c for c in ["razaoSocial","nome","nomeFantasia"] if c in df.columns), None)
    _mun_col  = "municipio" if "municipio" in df.columns else None
    _uf_col   = "uf"        if "uf"        in df.columns else None

    intel = _intel_load()
    hist  = intel["historico"]
    novos = 0

    for _, row in df.iterrows():
        _cnpj = re.sub(r"\D", "", str(row.get(_cnpj_col, "") or ""))
        if len(_cnpj) != 14:
            continue
        _preco = pd.to_numeric(row.get(_preco_col), errors="coerce")
        if pd.isna(_preco) or _preco <= 0:
            continue

        _entry: dict = {"data": data_str, "preco": round(float(_preco), 3),
                        "combustivel": combustivel}
        if _nome_col: _entry["nome"]      = str(row.get(_nome_col, ""))[:50]
        if _mun_col:  _entry["municipio"] = str(row.get(_mun_col,  ""))
        if _uf_col:   _entry["uf"]        = str(row.get(_uf_col,   ""))

        _lista = hist.setdefault(_cnpj, [])
        _ja_tem = any(
            e.get("data") == data_str and e.get("combustivel") == combustivel
            for e in _lista
        )
        if not _ja_tem:
            _lista.append(_entry)
            if len(_lista) > 52:
                _lista[:] = sorted(_lista, key=lambda e: e.get("data", ""))[-52:]
            novos += 1

    if novos > 0:
        _intel_save(intel)
    return novos


def _hist_record_pp_df(pp_df: "pd.DataFrame") -> int:
    """
    Registra preços do _pp_df normalizado no histórico.

    _pp_df tem colunas: cnpj_norm, combustivel_pk, combustivel_label,
                        preco, data_atualizacao.
    Retorna quantidade de novos registros adicionados.
    """
    if pp_df is None or pp_df.empty:
        return 0

    # Data de referência: usa data_atualizacao se disponível, senão hoje
    _hoje = datetime.now().strftime("%Y-%m-%d")

    intel = _intel_load()
    hist  = intel["historico"]
    novos = 0

    for _, row in pp_df.iterrows():
        _cnpj = str(row.get("cnpj_norm", "") or "").strip()
        if len(_cnpj) != 14:
            continue
        _preco = pd.to_numeric(row.get("preco"), errors="coerce")
        if pd.isna(_preco) or _preco <= 0:
            continue
        _comb = str(
            row.get("combustivel_pk") or row.get("combustivel_label") or ""
        ).upper().strip()
        if not _comb:
            continue

        # Data: tenta usar data_atualizacao da planilha (YYYY-MM-DD)
        _data_raw = str(row.get("data_atualizacao", "") or "").strip()
        try:
            _data = pd.to_datetime(_data_raw, dayfirst=True).strftime("%Y-%m-%d")
        except Exception:
            _data = _hoje

        _lista = hist.setdefault(_cnpj, [])
        _ja_tem = any(
            e.get("data") == _data and e.get("combustivel") == _comb
            for e in _lista
        )
        if not _ja_tem:
            _lista.append({
                "data":        _data,
                "preco":       round(float(_preco), 3),
                "combustivel": _comb,
            })
            if len(_lista) > 52:
                _lista[:] = sorted(_lista, key=lambda e: e.get("data", ""))[-52:]
            novos += 1

    if novos > 0:
        _intel_save(intel)
        # Invalida o cache da sessão para que os KPIs sejam atualizados
        st.session_state.pop("_intel_loaded", None)
    return novos


def _hist_get_posto(cnpj: str, combustivel: str = None) -> list:
    """Retorna histórico de preços de um posto (lista de dicts)."""
    _cnpj_n   = re.sub(r"\D", "", str(cnpj or ""))
    intel     = _intel_load()
    registros = intel.get("historico", {}).get(_cnpj_n, [])
    if combustivel:
        _ck = combustivel.upper()
        registros = [r for r in registros if r.get("combustivel", "").upper() == _ck]
    return sorted(registros, key=lambda r: r.get("data", ""))


def _hist_chart_posto(cnpj: str, nome: str, combustivel: str = None):
    """Retorna figura Plotly com evolução de preço de um posto."""
    import plotly.graph_objects as _pgo
    from collections import defaultdict as _dd
    registros = _hist_get_posto(cnpj, combustivel)
    if not registros:
        return None

    _CORES_COMB = {
        "GASOLINA COMUM":      "#EF5350",
        "GASOLINA ADITIVADA":  "#FF7043",
        "ETANOL HIDRATADO":    "#66BB6A",
        "ETANOL COMUM":        "#81C784",
        "DIESEL S10":          "#42A5F5",
        "DIESEL S500":         "#1E88E5",
        "DIESEL S-500 COMUM":  "#1565C0",
        "DIESEL S-500 ADITIVADO": "#0D47A1",
        "GNV":                 "#AB47BC",
        "GLP":                 "#FFA726",
    }
    por_comb: dict = _dd(list)
    for r in registros:
        por_comb[r.get("combustivel", "Combustível")].append(
            (r["data"], r["preco"])
        )

    fig = _pgo.Figure()
    for comb, pts in sorted(por_comb.items()):
        pts_s  = sorted(pts, key=lambda x: x[0])
        # Converte "YYYY-MM-DD" → "DD/MM/YYYY" para exibir no eixo X como categoria
        datas  = []
        for _d in [p[0] for p in pts_s]:
            try:
                _parts = str(_d).split("-")
                datas.append(f"{_parts[2]}/{_parts[1]}/{_parts[0]}" if len(_parts) == 3 else str(_d))
            except Exception:
                datas.append(str(_d))
        precos = [p[1] for p in pts_s]
        fig.add_trace(_pgo.Scatter(
            x=datas, y=precos,
            mode="lines+markers" if len(precos) > 1 else "markers",
            name=comb,
            line=dict(color=_CORES_COMB.get(comb, "#90CAF9"), width=2.5),
            marker=dict(size=8, line=dict(width=1.5, color="white")),
            hovertemplate=(
                "<b>%{x}</b><br>"
                f"<span style='color:{_CORES_COMB.get(comb,'#90CAF9')}'>{comb}</span><br>"
                "R$ <b>%{y:.3f}</b>/L<extra></extra>"
            ),
        ))

    fig.update_layout(
        # Sem title interno — será renderizado via st.markdown antes do gráfico
        xaxis_title=None,
        yaxis_title="R$/L",
        legend=dict(
            orientation="h",
            yanchor="top", y=-0.18,       # legenda abaixo do gráfico
            xanchor="center", x=0.5,
            font=dict(size=11),
            bgcolor="rgba(255,255,255,0.85)",
            bordercolor="rgba(0,0,0,0.08)",
            borderwidth=1,
        ),
        plot_bgcolor="rgba(0,0,0,0)",
        paper_bgcolor="rgba(0,0,0,0)",
        margin=dict(t=16, b=110, l=55, r=16),   # margem inferior para a legenda
        height=360,
        hovermode="x unified",
        xaxis=dict(
            type="category",                     # trata datas como texto — sem timezone
            gridcolor="rgba(0,0,0,0.07)",
            tickangle=-35,
            tickfont=dict(size=11),
            title=dict(text="Data", font=dict(size=12)),
        ),
        yaxis=dict(
            gridcolor="rgba(0,0,0,0.07)",
            tickprefix="R$ ",
            tickformat=".3f",
            tickfont=dict(size=11),
        ),
    )
    return fig


# ── Score de posto ──────────────────────────────────────────────────

_SCORE_ICONES = {"A": "🟢", "B": "🔵", "C": "🟡", "D": "🔴"}
_SCORE_CORES  = {
    "A": ("#e8f5e9", "#2e7d32", "#a5d6a7"),
    "B": ("#e3f2fd", "#1565c0", "#90caf9"),
    "C": ("#fff8e1", "#f57f17", "#ffe082"),
    "D": ("#ffebee", "#c62828", "#ef9a9a"),
}


def _calcular_score_posto(
    row:                  dict,
    preco_ref_anp:        float = None,
    lat_ref:              float = None,
    lon_ref:              float = None,
    servicos_keys:        list  = None,
    n_servicos_max:       int   = 10,
) -> dict:
    """
    Score composto 0-100 para um posto.
    Pesos: preço vs ANP 50% · serviços 30% · distância 20%.
    Graus: A≥75 · B≥55 · C≥35 · D<35
    """
    # ── Preço (50%) ─────────────────────────────────────────────────
    _s_preco = 50.0
    _det_preco = "Sem referência ANP"
    if preco_ref_anp and preco_ref_anp > 0:
        _p = pd.to_numeric(
            row.get("_preco_posto") or row.get("preco") or row.get("_preco"),
            errors="coerce")
        if pd.notna(_p) and _p > 0:
            _diff = (_p - preco_ref_anp) / preco_ref_anp
            _s_preco = max(0.0, min(100.0, 50.0 - _diff * 500.0))
            _det_preco = f"{_diff:+.1%} vs ANP ({preco_ref_anp:.3f})"

    # ── Serviços (30%) ──────────────────────────────────────────────
    _s_svc = 0.0
    _det_svc = "Sem dados de serviços"
    if servicos_keys and n_servicos_max > 0:
        _n = sum(1 for s in servicos_keys if row.get(s))
        _s_svc = min(100.0, _n / n_servicos_max * 100.0)
        _det_svc = f"{_n}/{n_servicos_max} serviços"

    # ── Distância (20%) ─────────────────────────────────────────────
    _s_dist = 50.0
    _det_dist = "Sem ponto de referência"
    if lat_ref is not None and lon_ref is not None:
        _lat = pd.to_numeric(row.get("_lat") or row.get("lat"), errors="coerce")
        _lon = pd.to_numeric(row.get("_lon") or row.get("lon"), errors="coerce")
        if pd.notna(_lat) and pd.notna(_lon):
            _dlat = _math_mod.radians(_lat - lat_ref)
            _dlon = _math_mod.radians(_lon - lon_ref)
            _a    = (_math_mod.sin(_dlat/2)**2 +
                     _math_mod.cos(_math_mod.radians(lat_ref)) *
                     _math_mod.cos(_math_mod.radians(_lat)) *
                     _math_mod.sin(_dlon/2)**2)
            _d_km = 6371 * 2 * _math_mod.atan2(
                _math_mod.sqrt(_a), _math_mod.sqrt(1 - _a))
            _s_dist   = max(0.0, min(100.0, 100.0 - _d_km))
            _det_dist = f"{_d_km:.1f} km do ponto de busca"

    _score = 0.50 * _s_preco + 0.30 * _s_svc + 0.20 * _s_dist
    _grade = ("A" if _score >= 75 else
              "B" if _score >= 55 else
              "C" if _score >= 35 else "D")
    return {
        "score":           round(_score, 1),
        "grade":           _grade,
        "score_preco":     round(_s_preco, 1),
        "score_servicos":  round(_s_svc,   1),
        "score_distancia": round(_s_dist,  1),
        "detalhe_preco":   _det_preco,
        "detalhe_svc":     _det_svc,
        "detalhe_dist":    _det_dist,
    }


def _score_badge_html(score: float, grade: str, tooltip: str = "",
                      size: str = "normal") -> str:
    """Badge colorido com score e grau (HTML)."""
    bg, txt, brd = _SCORE_CORES.get(grade, ("#f5f5f5", "#424242", "#e0e0e0"))
    _ic   = _SCORE_ICONES.get(grade, "⚪")
    _font = "13px" if size == "normal" else "11px"
    _pad  = "4px 10px" if size == "normal" else "2px 7px"
    _tt   = f" title='{tooltip}'" if tooltip else ""
    return (
        f"<span{_tt} style='display:inline-flex;align-items:center;gap:5px;"
        f"background:{bg};border:1px solid {brd};border-radius:20px;"
        f"padding:{_pad};font-size:{_font};font-weight:700;color:{txt};cursor:default'>"
        f"{_ic} Score {score:.0f}"
        f"<span style='opacity:.65;font-weight:600'>({grade})</span>"
        f"</span>"
    )


def _calcular_score_df(
    df: "pd.DataFrame",
    preco_ref_anp: float = None,
    lat_ref: float = None,
    lon_ref: float = None,
) -> "pd.DataFrame":
    """
    Adiciona coluna '⭐ Score' ao DataFrame de postos.
    Retorna df com a coluna inserida na posição 0.
    """
    if df is None or df.empty:
        return df
    _svc_keys = list(st.session_state.get("_servicos_pf_labels", {}).keys())
    _n_max    = max(len(_svc_keys), 1)
    _scores   = []
    for _, row in df.iterrows():
        _res = _calcular_score_posto(
            row.to_dict(),
            preco_ref_anp=preco_ref_anp,
            lat_ref=lat_ref,
            lon_ref=lon_ref,
            servicos_keys=_svc_keys,
            n_servicos_max=_n_max,
        )
        _scores.append(f"{_res['grade']} {_res['score']:.0f}")
    _df = df.copy()
    _df.insert(0, "⭐ Score", _scores)
    return _df


# ── Relatório semanal de alertas ────────────────────────────────────

def _gerar_relatorio_alertas_xlsx(
    df_pp:  "pd.DataFrame",
    limiar: dict,
    semana: str = None,
) -> tuple:
    """
    Gera relatório Excel de postos em alerta de preço.

    df_pp:  DataFrame de preços dos postos GF.
    limiar: {combustivel_norm: preco_max_float}.
    semana: label de referência (ex: "2024-W05").

    Retorna (bytes_xlsx, filename, error_msg).
    """
    if df_pp is None or df_pp.empty:
        return None, None, "Nenhuma planilha de preços carregada."
    if not limiar:
        return None, None, "Defina ao menos um limiar de preço."

    try:
        import io as _io
        import openpyxl as _opxl
        from openpyxl.styles import (PatternFill, Font, Alignment,
                                     Border, Side, numbers)
        from openpyxl.utils import get_column_letter

        # ── Enriquece df_pp com dados cadastrais do pf_coords_df ──────
        _pf = st.session_state.get("pf_coords_df", pd.DataFrame())
        if not _pf.empty and "cnpj_norm" in _pf.columns:
            _info_cols = [c for c in ["cnpj_norm","razaoSocial","municipio","uf","distribuidora"]
                          if c in _pf.columns]
            _pf_info = _pf[_info_cols].drop_duplicates("cnpj_norm")
            # df_pp usa cnpj_norm; garante coluna antes do merge
            _key = "cnpj_norm" if "cnpj_norm" in df_pp.columns else None
            if _key:
                df_pp = df_pp.merge(_pf_info, on="cnpj_norm", how="left")

        _data_rel = semana or datetime.now().strftime("%Y-W%V")
        _wb  = _opxl.Workbook()
        _ws  = _wb.active
        _ws.title = "Alertas de Preço"

        # ── Cabeçalho do relatório ──────────────────────────────────
        _titulo_fill = PatternFill("solid", fgColor="0D47A1")
        _titulo_font = Font(bold=True, color="FFFFFF", size=13)
        _ws.merge_cells("A1:H1")
        _ws["A1"] = f"Relatório de Alertas de Preço — {_data_rel}"
        _ws["A1"].fill = _titulo_fill
        _ws["A1"].font = _titulo_font
        _ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
        _ws.row_dimensions[1].height = 26

        _ws.merge_cells("A2:H2")
        _ws["A2"] = f"Gerado em {datetime.now().strftime('%d/%m/%Y %H:%M')} · Pró-Frotas"
        _ws["A2"].font = Font(italic=True, color="666666", size=10)
        _ws["A2"].alignment = Alignment(horizontal="center")

        _row = 4
        _alert_total = 0

        # ── Para cada combustível com limiar ────────────────────────
        for _comb_key, _lim in limiar.items():
            if _lim <= 0:
                continue

            # Detecta coluna de preço no df_pp
            _col_preco = None
            for _c in df_pp.columns:
                _cn = re.sub(r"[^a-z]", "", _c.lower())
                if _cn in _comb_key.lower().replace(" ", "") or \
                   _comb_key.lower().replace(" ", "") in _cn:
                    _col_preco = _c
                    break
            if _col_preco is None:
                # fallback: primeira coluna numérica
                for _c in df_pp.columns:
                    if pd.api.types.is_numeric_dtype(df_pp[_c]):
                        _col_preco = _c
                        break
            if _col_preco is None:
                continue

            _df_c = df_pp.copy()
            _df_c["_preco_num"] = pd.to_numeric(_df_c[_col_preco], errors="coerce")
            _df_alert = _df_c[
                _df_c["_preco_num"].notna() &
                (_df_c["_preco_num"] > _lim)
            ].sort_values("_preco_num", ascending=False)
            if _df_alert.empty:
                continue

            _alert_total += len(_df_alert)

            # Título do combustível
            _ws.merge_cells(f"A{_row}:H{_row}")
            _ws[f"A{_row}"] = f"⚠️ {_comb_key.title()} — Limiar: R$ {_lim:.3f}/L · {len(_df_alert)} postos em alerta"
            _ws[f"A{_row}"].fill = PatternFill("solid", fgColor="FFF9C4")
            _ws[f"A{_row}"].font = Font(bold=True, color="F57F17", size=11)
            _ws[f"A{_row}"].alignment = Alignment(horizontal="left", indent=1)
            _row += 1

            # Cabeçalho da tabela
            _hdrs = ["CNPJ","Razão Social","Município","UF",
                     f"Preço ({_comb_key.title()})","Desvio vs Limiar","% Acima","Bandeira"]
            _hdr_fill = PatternFill("solid", fgColor="E3F2FD")
            _hdr_font = Font(bold=True, color="1565C0", size=10)
            for _ci, _h in enumerate(_hdrs, 1):
                _cell = _ws.cell(row=_row, column=_ci, value=_h)
                _cell.fill = _hdr_fill
                _cell.font = _hdr_font
                _cell.alignment = Alignment(horizontal="center")
                _cell.border = Border(
                    bottom=Side(style="thin", color="90CAF9"),
                    top=Side(style="thin", color="90CAF9"),
                )
            _row += 1

            # Linhas de dados
            _alt_fill_r = PatternFill("solid", fgColor="FFEBEE")
            _alt_fill_w = PatternFill("solid", fgColor="FFFFFF")
            for _ri, (_, _ar) in enumerate(_df_alert.iterrows()):
                _preco_p  = _ar["_preco_num"]
                _desvio   = _preco_p - _lim
                _pct      = _desvio / _lim * 100 if _lim else 0
                _fill     = _alt_fill_r if _ri % 2 == 0 else _alt_fill_w

                _cnpj_raw = _ar.get("cnpj_norm", _ar.get("cnpj", _ar.get("_cnpj_norm", "")))
                # Formata CNPJ: 00.000.000/0000-00
                try:
                    _d = re.sub(r"\D", "", str(_cnpj_raw))
                    _cnpj_c = f"{_d[:2]}.{_d[2:5]}.{_d[5:8]}/{_d[8:12]}-{_d[12:14]}" if len(_d)==14 else _cnpj_raw
                except Exception:
                    _cnpj_c = _cnpj_raw
                _nome_c   = _ar.get("razaoSocial", _ar.get("nome", _ar.get("nomeFantasia", "")))
                _mun_c    = _ar.get("municipio", "")
                _uf_c     = _ar.get("uf", "")
                _band_c   = _ar.get("distribuidora", "")

                _vals = [
                    _cnpj_c, _nome_c, _mun_c, _uf_c,
                    _preco_p, _desvio, _pct / 100, _band_c,
                ]
                for _ci, _v in enumerate(_vals, 1):
                    _cell = _ws.cell(row=_row, column=_ci, value=_v)
                    _cell.fill = _fill
                    _cell.font = Font(size=9)
                    if _ci == 5:  # preço
                        _cell.number_format = 'R$ #,##0.000'
                        _cell.font = Font(size=9, bold=True,
                                          color=("C62828" if _preco_p > _lim * 1.1 else "E65100"))
                    elif _ci == 6:
                        _cell.number_format = 'R$ #,##0.000'
                    elif _ci == 7:
                        _cell.number_format = '0.0%'
                _row += 1
            _row += 2  # espaço entre combustíveis

        if _alert_total == 0:
            _ws[f"A{_row}"] = "✅ Nenhum posto acima do limiar configurado."
            _ws[f"A{_row}"].font = Font(color="2E7D32", italic=True, size=10)
            _row += 1

        # Resumo final
        _row += 1
        _ws.merge_cells(f"A{_row}:H{_row}")
        _ws[f"A{_row}"] = f"Total de postos em alerta: {_alert_total}"
        _ws[f"A{_row}"].font = Font(bold=True, size=11,
                                     color=("C62828" if _alert_total > 0 else "2E7D32"))

        # Ajusta largura das colunas
        _larguras = [18, 40, 22, 5, 12, 12, 8, 18]
        for _ci, _larg in enumerate(_larguras, 1):
            _ws.column_dimensions[get_column_letter(_ci)].width = _larg

        _buf = _io.BytesIO()
        _wb.save(_buf)
        _buf.seek(0)
        _fname = f"alertas_preco_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"

        # Atualiza data do último relatório
        _intel = _intel_load()
        _intel["last_report"] = datetime.now().isoformat()
        _intel_save(_intel)

        return _buf.getvalue(), _fname, None

    except Exception as _ex:
        return None, None, str(_ex)


# ═══════════════════════════════════════════════════════════════════
#  GESTÃO DE FROTAS — Upload e comparação de CNPJs
# ═══════════════════════════════════════════════════════════════════

# Nome do arquivo fixo esperado na raiz do repositório
ARQUIVO_PF_REPO       = "pro_frotas.xlsx"
ARQUIVO_CERCADOS_REPO = "Postos Cercados.xlsx"
COR_CERCADO_FILL      = "#FF8F00"   # laranja âmbar — alerta visual
COR_CERCADO_BORDA     = "#E65100"   # laranja escuro
ARQUIVO_PP_REPO       = "preco_posto.xlsx"   # planilha de preços por posto
_PP_PARSER_VERSION    = "v5"                 # incrementar aqui força re-parse automático
ARQUIVO_DOC_PDF       = "documentacao_gestao_frotas.pdf"   # documentação da aplicação

# Candidatos de nome para o PDF de documentação (ordem de prioridade)
_DOC_PDF_CANDIDATOS = [
    "Gestao de Frotas.pdf",
    "Gestão de Frotas.pdf",
    "gestao_de_frotas.pdf",
    "documentacao_gestao_frotas.pdf",
    "documentacao_gestao_frotas.pdf",
    "Documentacao_Gestao_Frotas.pdf",
    "gestao de frotas.pdf",
]


@st.cache_data(show_spinner=False, ttl=86400)   # 24 h — relê o PDF do repo uma vez por dia
def _carregar_doc_pdf():
    """
    Carrega o PDF de documentação do repositório.
    Tenta vários nomes possíveis para o arquivo.
    Retorna (bytes, nome_arquivo) ou (None, None) se não encontrado.
    """
    for _nome in _DOC_PDF_CANDIDATOS:
        _caminho_doc = os.path.join(_DIR, _nome)
        if os.path.exists(_caminho_doc):
            try:
                with open(_caminho_doc, "rb") as _f:
                    return _f.read(), _nome
            except Exception:
                continue
    return None, None


def normalizar_cnpj(valor):
    if pd.isna(valor):
        return ""
    return "".join(c for c in str(valor) if c.isdigit())


def detectar_coluna_cnpj(df: pd.DataFrame):
    for col in df.columns:
        if "cnpj" in col.lower():
            return col
    for col in df.columns:
        amostra = df[col].dropna().astype(str).head(5)
        if amostra.apply(lambda x: len("".join(c for c in x if c.isdigit())) >= 14).any():
            return col
    return None


def _detectar_col(df: pd.DataFrame, termos: list) -> str | None:
    """Retorna o nome da coluna que melhor bate com os termos.

    Estratégia em 2 passos:
      1. Match EXATO por ordem de termos (prioridade): varre todos os termos em
         sequência e, para cada um, percorre as colunas buscando correspondência
         exata. Isso garante que, p.ex., 'BANDEIRA' seja encontrado antes que
         'REDE' (substring de 'CREDENCIADO') cause uma falsa detecção.
      2. Match SUBSTRING por ordem de colunas (fallback): mantém comportamento
         original para planilhas com nomes de coluna compostos.
    """
    # 1ª passagem — match exato, prioridade determinada pela ordem dos termos
    for _t in termos:
        for _c in df.columns:
            if _anp_norm(_c) == _t:
                return _c
    # 2ª passagem — substring como fallback
    for _c in df.columns:
        _cn = _anp_norm(_c)
        if any(t in _cn for t in termos):
            return _c
    return None


# ── Mapeamento de colunas de serviço "Possui X?" / "Oferece X?" ─────────────
# Chave interna → label para exibição no filtro avançado.
# A detecção usa substring normalizada (sem acentos, uppercase).
_PF_SVC_LABELS_MASTER: dict[str, str] = {
    "svc_conveniencia":   "🛒 Conveniência",
    "svc_ampm":           "🏪 Conv. AM/PM",
    "svc_restaurante":    "🍽️ Restaurante",
    "svc_banheiro":       "🚻 Banheiro",
    "svc_banheiro_pago":  "💰 Banheiro Pago",
    "svc_estacionamento": "🅿️ Estacionamento",
    "svc_troca_oleo":     "🔧 Troca de Óleo",
    "svc_oleo_granel":    "🛢️ Óleo a Granel",
    "svc_arla32":         "🧪 ARLA 32",
    "svc_tipo_arla":      "🏷️ Tipo de ARLA",
}

# Colunas legadas que são "substituídas" por colunas dinâmicas da planilha.
# { col_key_legada: {col_keys_dinamicas_que_cobrem_o_conceito} }
_SVC_LEGADO_SUPERSEDE: dict[str, set] = {
    "conveniencia": {"svc_conveniencia"},
    "arla":         {"svc_arla32"},
}


def _atualizar_servicos_pf(df_coords) -> None:
    """
    Reconstrói st.session_state["_servicos_pf_labels"] e complementa
    "_servicos_cols_disponiveis" com base nas colunas presentes em df_coords.

    Deve ser chamado sempre que pf_coords_df for atualizado no session_state.
    """
    if df_coords is None or (hasattr(df_coords, "empty") and df_coords.empty):
        return
    _labels: dict[str, str] = {}
    _cols_disp: list = list(st.session_state.get("_servicos_cols_disponiveis") or [])
    for _key, _lbl in _PF_SVC_LABELS_MASTER.items():
        if _key in df_coords.columns:
            _labels[_key] = _lbl
            if _key not in _cols_disp:
                _cols_disp.append(_key)
    if _labels:
        st.session_state["_servicos_pf_labels"]       = _labels
        st.session_state["_servicos_cols_disponiveis"] = _cols_disp


def _processar_bytes_pro_frotas(nome: str, conteudo: bytes):
    """
    Núcleo de leitura da planilha Gestão de Frotas.
    Aceita o nome do arquivo e seus bytes brutos.
    Retorna (set_cnpjs, msg, df_preview, perfil_map, df_coords) ou (None, msg, None, None, None).

    df_coords: DataFrame com colunas [cnpj_norm, _lat, _lon, razaoSocial, distribuidora,
               municipio, uf] para postos que possuem coordenadas na planilha.
    perfil_map: dict {cnpj_norm: "Perfil de Venda"}.
    """
    buf = io.BytesIO(conteudo)
    nome_l = nome.lower()

    if nome_l.endswith(".csv"):
        try:
            df = pd.read_csv(buf, dtype=str, encoding="utf-8")
        except UnicodeDecodeError:
            buf.seek(0)
            df = pd.read_csv(buf, dtype=str, encoding="latin-1")
    elif nome_l.endswith(".xls"):
        df = pd.read_excel(buf, dtype=str, engine="xlrd")
    else:
        df = pd.read_excel(buf, dtype=str, engine="openpyxl")

    if df.empty:
        return None, "A planilha está vazia.", None, None, None

    col = detectar_coluna_cnpj(df)
    if col is None:
        colunas = ", ".join(df.columns.tolist())
        return None, (
            f"Coluna CNPJ não encontrada. "
            f"Colunas disponíveis: **{colunas}**. "
            "Renomeie a coluna de CNPJ para 'CNPJ'."
        ), None, None, None

    cnpjs = {c for c in df[col].dropna().apply(normalizar_cnpj) if len(c) == 14}
    if not cnpjs:
        return None, "Nenhum CNPJ válido (14 dígitos) encontrado na coluna detectada.", None, None, None

    # ── Detecta coluna "Perfil de Venda" ───────────────────────────
    col_perfil = _detectar_col(df, ["PERFIL DE VENDA", "PERFIL VENDA", "PERFILDEVENDA", "PERFIL"])

    perfil_map: dict = {}
    if col_perfil:
        for _, row in df.iterrows():
            cnpj_n = normalizar_cnpj(row.get(col, ""))
            if len(cnpj_n) == 14:
                perfil = str(row.get(col_perfil, "")).strip()
                if perfil and perfil.upper() not in ("NAN", "NONE", ""):
                    perfil_map[cnpj_n] = perfil

    # ── Detecta colunas de coordenadas e dados complementares ──────
    col_lat  = _detectar_col(df, ["LATITUDE", "LAT"])
    col_lon  = _detectar_col(df, ["LONGITUDE", "LON", "LNG", "LONG"])
    col_nome = _detectar_col(df, ["RAZAO SOCIAL", "NOME FANTASIA", "RAZAO", "NOME"])
    col_dist = _detectar_col(df, ["BANDEIRA", "DISTRIBUIDORA", "REDE"])   # BANDEIRA primeiro — prioridade exata
    col_mun  = _detectar_col(df, ["MUNICIPIO", "CIDADE"])
    col_uf   = _detectar_col(df, ["UF", "ESTADO"])

    # ── Colunas opcionais de serviços / horário ─────────────────
    col_horario   = _detectar_col(df, ["HORARIO FUNCIONAMENTO", "HORARIO", "FUNCIONAMENTO", "HORA FUNC"])
    col_24h       = _detectar_col(df, ["FUNCIONA 24H", "24H", "24 HORAS", "ABERTO 24H", "FUNCIONAMENTO 24H"])
    col_caminhao  = _detectar_col(df, ["PISTA CAMINHAO", "CAMINHAO", "PISTA TRUCK", "TRUCK", "PISTA CAM"])
    col_arla      = _detectar_col(df, ["ARLA", "ARLA 32", "ARLA32"])
    col_conv      = _detectar_col(df, ["CONVENIENCIA", "LOJA CONVENIENCIA", "LOJA", "CONVENIENCE"])

    # ── Detecção genérica de colunas "Possui X?" / "Oferece X?" ──────────────
    # Normaliza nome de coluna: remove acentos, pontuação → uppercase limpo.
    import unicodedata as _ud_svc
    def _norm_svc_col(s: str) -> str:
        s = _ud_svc.normalize("NFD", str(s).upper())
        s = "".join(c for c in s if _ud_svc.category(c) != "Mn")
        return re.sub(r"[^A-Z0-9 /]+", " ", s).strip()

    # Lista de (keyword_normalizada, col_key_interna, ignorar_se_col_exata_ja_detectada)
    _PF_SVC_DETECT = [
        ("POSSUI CONVENIENCIA",   "svc_conveniencia",   col_conv),
        ("CONVENIENCIA E AM",     "svc_ampm",           None),
        ("CONVENIENCIA AM PM",    "svc_ampm",           None),
        ("POSSUI RESTAURANTE",    "svc_restaurante",    None),
        ("POSSUI BANHEIRO",       "svc_banheiro",       None),
        ("COBRANCA PARA USAR",    "svc_banheiro_pago",  None),
        ("POSSUI ESTACIONAMENTO", "svc_estacionamento", None),
        ("TROCA DE OLEO",         "svc_troca_oleo",     None),
        ("OLEO A GRANEL",         "svc_oleo_granel",    None),
        ("OFERECE PRODUTO ARLA",  "svc_arla32",         col_arla),
        ("QUAL TIPO DE ARLA",     "svc_tipo_arla",      None),
    ]

    # Mapeia col_real_no_df → col_key_interna (evita duplicatas de chave)
    _pf_svc_detected: dict[str, str] = {}
    _seen_svc_keys: set = set()
    for _raw_col in df.columns:
        _n = _norm_svc_col(str(_raw_col))
        for _kw, _ckey, _skip_col in _PF_SVC_DETECT:
            if _kw in _n and _ckey not in _seen_svc_keys:
                # Se este conceito já está coberto por uma coluna legada detectada,
                # ainda registramos a coluna planilha (mais rica) para sobrescrever.
                _pf_svc_detected[str(_raw_col)] = _ckey
                _seen_svc_keys.add(_ckey)
                break

    def _bool_col(val):
        """Converte 'SIM','S','1','YES','TRUE','X' → True; demais → False."""
        return str(val).strip().upper() in ("SIM", "S", "1", "YES", "TRUE", "X", "VERDADEIRO")

    df_coords = pd.DataFrame()
    if col_lat and col_lon:
        rows = []
        for _, row in df.iterrows():
            cnpj_n = normalizar_cnpj(row.get(col, ""))
            if len(cnpj_n) != 14:
                continue
            try:
                lat = float(str(row[col_lat]).replace(",", "."))
                lon = float(str(row[col_lon]).replace(",", "."))
            except (ValueError, TypeError):
                continue
            if not (-33.8 <= lat <= 5.3 and -73.9 <= lon <= -34.7):
                continue
            rec = {
                "cnpj":         cnpj_n,
                "_lat":         lat,
                "_lon":         lon,
                "razaoSocial":  str(row.get(col_nome, "")).strip() if col_nome else "",
                "distribuidora":str(row.get(col_dist, "")).strip() if col_dist else "",
                "municipio":    str(row.get(col_mun,  "")).strip() if col_mun  else "",
                "uf":           _normalizar_uf(str(row.get(col_uf, ""))) if col_uf else "",
                # Serviços / horário legados — None quando coluna ausente
                "horario":       str(row.get(col_horario, "")).strip() if col_horario else None,
                "funciona_24h":  _bool_col(row.get(col_24h, ""))      if col_24h     else None,
                "pista_caminhao":_bool_col(row.get(col_caminhao, "")) if col_caminhao else None,
                "arla":          _bool_col(row.get(col_arla, ""))     if col_arla    else None,
                "conveniencia":  _bool_col(row.get(col_conv, ""))     if col_conv    else None,
            }
            # ── Adiciona serviços detectados dinamicamente ("Possui X?") ──
            for _rcol, _ckey in _pf_svc_detected.items():
                rec[_ckey] = _bool_col(row.get(_rcol, ""))
            rows.append(rec)
        if rows:
            df_coords = pd.DataFrame(rows)
            # Limpa valores "nan" / "None" que vieram como string
            for _c in ["razaoSocial","distribuidora","municipio","uf"]:
                df_coords[_c] = df_coords[_c].replace(
                    {"nan": "", "None": "", "NaN": ""})
            # Normaliza distribuidora para Title Case uniforme
            df_coords["distribuidora"] = _normalizar_distribuidora(df_coords["distribuidora"])
            # Informa quais colunas de serviço foram detectadas (para feedback no sidebar)
            _servicos_detectados = [s for s, c in [
                ("horario", col_horario), ("funciona_24h", col_24h),
                ("pista_caminhao", col_caminhao), ("arla", col_arla),
                ("conveniencia", col_conv),
            ] if c]
            # Adiciona chaves das colunas dinâmicas
            for _ckey_d in _pf_svc_detected.values():
                if _ckey_d not in _servicos_detectados:
                    _servicos_detectados.append(_ckey_d)
            if _servicos_detectados:
                st.session_state["_servicos_cols_disponiveis"] = _servicos_detectados

    preview = df[[col]].rename(columns={col: "CNPJ (original)"}).head(10)
    perfil_info  = f" · {len(set(perfil_map.values()))} perfis" if perfil_map else ""
    coords_info  = f" · {len(df_coords)} coords" if not df_coords.empty else ""
    return (cnpjs,
            f"{len(cnpjs)} CNPJs carregados (coluna: **{col}**){perfil_info}{coords_info}",
            preview,
            perfil_map,
            df_coords)


def _processar_bytes_anp_postos(nome: str, conteudo: bytes):
    """
    Lê o arquivo XLSX de postos ANP baixado manualmente do site da ANP.
    Detecta automaticamente as colunas de CNPJ, lat/lon, nome, bandeira, município e UF.
    Retorna (DataFrame, msg) ou (None, msg_erro).
    O DataFrame terá colunas: cnpj, _lat, _lon, razaoSocial, distribuidora, municipio, uf
    """
    try:
        buf = io.BytesIO(conteudo)
        nome_l = nome.lower()
        if nome_l.endswith(".xls"):
            df = pd.read_excel(buf, dtype=str, engine="xlrd")
        else:
            df = pd.read_excel(buf, dtype=str, engine="openpyxl")
    except Exception as e:
        return None, f"Erro ao ler o arquivo: {type(e).__name__} — {e}"

    if df.empty:
        return None, "O arquivo está vazio."

    # Normaliza nomes de colunas para detecção
    df.columns = [str(c).strip().upper() for c in df.columns]

    col_cnpj = _detectar_col(df, [
        "CNPJ DA REVENDA", "CNPJ REVENDA", "CNPJ_REVENDA",
        "CNPJ DO ESTABELECIMENTO", "CNPJ ESTABELECIMENTO", "CNPJ",
    ])
    col_lat  = _detectar_col(df, ["LATITUDE", "LAT"])
    col_lon  = _detectar_col(df, ["LONGITUDE", "LON", "LNG", "LONG"])
    col_nome = _detectar_col(df, [
        "RAZÃO SOCIAL", "RAZAO SOCIAL", "RAZÃO", "RAZAO",
        "NOME FANTASIA", "NOMEFANTASIA", "NOME",
    ])
    col_dist = _detectar_col(df, ["BANDEIRA", "DISTRIBUIDORA", "REDE", "BRAND"])
    col_mun  = _detectar_col(df, ["MUNICÍPIO", "MUNICIPIO", "CIDADE", "CITY"])
    col_uf   = _detectar_col(df, ["UF", "ESTADO", "ESTADO (UF)"])

    if col_cnpj is None:
        colunas = ", ".join(df.columns.tolist()[:20])
        return None, (
            f"Coluna CNPJ não encontrada. Colunas disponíveis: {colunas}. "
            "Verifique se o arquivo correto (postos ANP) foi selecionado."
        )

    rows = []
    for _, row in df.iterrows():
        cnpj_n = re.sub(r"\D", "", str(row.get(col_cnpj, "") or ""))
        if len(cnpj_n) != 14:
            continue
        try:
            lat = float(str(row[col_lat]).replace(",", ".")) if col_lat else float("nan")
            lon = float(str(row[col_lon]).replace(",", ".")) if col_lon else float("nan")
        except (ValueError, TypeError):
            lat, lon = float("nan"), float("nan")

        # Rejeita coordenadas fora do Brasil
        if col_lat and col_lon and not (math.isnan(lat) or math.isnan(lon)):
            if not (-33.8 <= lat <= 5.3 and -73.9 <= lon <= -34.7):
                continue

        rows.append({
            "cnpj":          cnpj_n,
            "_lat":          lat if col_lat else None,
            "_lon":          lon if col_lon else None,
            "razaoSocial":   str(row.get(col_nome, "")).strip() if col_nome else "",
            "distribuidora": str(row.get(col_dist, "")).strip() if col_dist else "",
            "municipio":     str(row.get(col_mun,  "")).strip() if col_mun  else "",
            "uf":            str(row.get(col_uf,   "")).strip().upper() if col_uf else "",
        })

    if not rows:
        return None, "Nenhum CNPJ válido (14 dígitos) encontrado no arquivo."

    result_df = pd.DataFrame(rows)
    for _c in ["razaoSocial", "distribuidora", "municipio", "uf"]:
        result_df[_c] = result_df[_c].replace({"nan": "", "None": "", "NaN": ""})
    # Normaliza distribuidora para Title Case uniforme
    result_df["distribuidora"] = _normalizar_distribuidora(result_df["distribuidora"])

    # Remove linhas sem coordenadas válidas (necessário para o mapa)
    _n_total = len(result_df)
    if col_lat and col_lon:
        result_df = result_df.dropna(subset=["_lat", "_lon"]).reset_index(drop=True)

    _n_com_coord = len(result_df)
    _sem_coord   = _n_total - _n_com_coord
    sem_msg = f" ({_sem_coord} sem coordenadas ignorados)" if _sem_coord else ""
    return result_df, f"{_n_com_coord} postos ANP carregados{sem_msg}"


def ler_planilha_pro_frotas(arquivo):
    """Lê UploadedFile do Streamlit. Sem @st.cache_data (upload não é cacheável)."""
    try:
        return _processar_bytes_pro_frotas(arquivo.name, arquivo.read())
    except ImportError as e:
        return None, f"Biblioteca ausente no servidor: **{e}**.", None, None, None
    except Exception as e:
        return None, f"Erro ao processar arquivo: **{type(e).__name__}** — {e}", None, None, None


@st.cache_data(show_spinner=False, ttl=86400)   # 24 horas — lê o arquivo do repo uma vez por dia
def _auto_carregar_pro_frotas_repo():
    """
    Tenta carregar automaticamente a planilha Gestão de Frotas do repositório.
    Aceita: pro_frotas.xlsx / pro_frotas.xls / pro_frotas.csv
    Retorna (set_cnpjs, msg, df_preview, perfil_map, df_coords) ou (None, msg, None, None, None).
    """
    for nome in [ARQUIVO_PF_REPO, "pro_frotas.xls", "pro_frotas.csv"]:
        caminho = os.path.join(_DIR, nome)
        if os.path.exists(caminho):
            try:
                with open(caminho, "rb") as f:
                    conteudo = f.read()
                cnpjs, msg, preview, perfil_map, df_coords = _processar_bytes_pro_frotas(nome, conteudo)
                if cnpjs:
                    return cnpjs, msg, preview, perfil_map, df_coords
            except Exception as e:
                return None, f"Erro ao ler {nome} do repositório: {e}", None, None, None
    return None, f"Arquivo `{ARQUIVO_PF_REPO}` não encontrado em: {_DIR}", None, None, None


# ── Postos Cercados ─────────────────────────────────────────────────

def _processar_bytes_cercados(nome: str, conteudo: bytes):
    """
    Lê planilha de Postos Cercados (CNPJ).
    Retorna (set_cnpjs, msg, df_preview) ou (None, msg_erro, None).
    """
    buf = io.BytesIO(conteudo)
    nome_l = nome.lower()
    if nome_l.endswith(".csv"):
        try:
            df = pd.read_csv(buf, dtype=str, encoding="utf-8")
        except UnicodeDecodeError:
            buf.seek(0)
            df = pd.read_csv(buf, dtype=str, encoding="latin-1")
    elif nome_l.endswith(".xls"):
        df = pd.read_excel(buf, dtype=str, engine="xlrd")
    else:
        df = pd.read_excel(buf, dtype=str, engine="openpyxl")

    if df.empty:
        return None, "A planilha está vazia.", None

    col = detectar_coluna_cnpj(df)
    if col is None:
        colunas = ", ".join(df.columns.tolist())
        return None, (
            f"Coluna CNPJ não encontrada. "
            f"Colunas disponíveis: **{colunas}**. "
            "Renomeie a coluna de CNPJ para 'CNPJ'."
        ), None

    cnpjs = {c for c in df[col].dropna().apply(normalizar_cnpj) if len(c) == 14}
    if not cnpjs:
        return None, "Nenhum CNPJ válido (14 dígitos) encontrado na coluna detectada.", None

    preview = df[[col]].rename(columns={col: "CNPJ (original)"}).head(10)
    return cnpjs, f"{len(cnpjs)} postos cercados carregados (coluna: **{col}**)", preview


def ler_planilha_cercados(arquivo):
    """Lê UploadedFile do Streamlit. Sem @st.cache_data (upload não é cacheável)."""
    try:
        return _processar_bytes_cercados(arquivo.name, arquivo.read())
    except ImportError as e:
        return None, f"Biblioteca ausente no servidor: **{e}**.", None
    except Exception as e:
        return None, f"Erro ao processar arquivo: **{type(e).__name__}** — {e}", None


@st.cache_data(show_spinner=False, ttl=86400)   # 24 h — re-lê o arquivo do repo uma vez por dia
def _auto_carregar_cercados_repo():
    """
    Tenta carregar automaticamente 'Postos Cercados.xlsx' do repositório.
    Aceita variações de nome: com/sem espaço, maiúsculo/minúsculo.
    Retorna (set_cnpjs, msg, df_preview) ou (None, msg_erro, None).
    """
    candidatos = [
        ARQUIVO_CERCADOS_REPO,          # "Postos Cercados.xlsx"
        "postos_cercados.xlsx",
        "postos_cercados.xls",
        "Postos_Cercados.xlsx",
        "postos cercados.xlsx",
    ]
    for nome in candidatos:
        caminho = os.path.join(_DIR, nome)
        if os.path.exists(caminho):
            try:
                with open(caminho, "rb") as f:
                    conteudo = f.read()
                cnpjs, msg, preview = _processar_bytes_cercados(nome, conteudo)
                if cnpjs:
                    return cnpjs, msg, preview
            except Exception as e:
                return None, f"Erro ao ler {nome} do repositório: {e}", None
    return None, f"Arquivo `{ARQUIVO_CERCADOS_REPO}` não encontrado em: {_DIR}", None


def marcar_cercados(df: pd.DataFrame, cnpjs_cercados: set) -> pd.DataFrame:
    """Adiciona coluna '_cercado' ao DataFrame com base nos CNPJs."""
    df = df.copy()
    if cnpjs_cercados and "cnpj" in df.columns:
        if "_cnpj_norm" not in df.columns:
            df["_cnpj_norm"] = df["cnpj"].fillna("").str.replace(r'\D', '', regex=True)
        df["_cercado"] = df["_cnpj_norm"].isin(cnpjs_cercados)
    else:
        df["_cercado"] = False
    return df


# ── Preço Posto (preços por CNPJ com data de atualização) ───────────

def _detectar_col_combustivel_pp(df):
    for col in df.columns:
        if any(t in _anp_norm(col) for t in ["PRODUTO","COMBUSTIVEL","FUEL","PRODUCT"]):
            return col
    return None

def _detectar_col_preco_pp(df):
    for col in df.columns:
        if any(t in _anp_norm(col) for t in ["PRECO","VALOR","PRICE","VLR","CUSTO"]):
            return col
    return None

def _detectar_col_data_pp(df):
    for col in df.columns:
        if any(t in _anp_norm(col) for t in ["DATA","DATE","ATUALIZ","UPDATE","VIGENCIA"]):
            return col
    return None


def _colunas_combustivel_wide(df):
    """
    Detecta colunas que sejam nomes de combustíveis (formato wide).
    Retorna lista de (col_original, pk_normalizado) ou [] se não for wide.
    """
    # Conjunto de tokens que indicam que a coluna É um combustível
    tokens_comb = {
        "GASOLINA", "DIESEL", "ETANOL", "ALCOOL", "GNV", "GLP",
        "OLEO", "GAS NATURAL", "GAS LIQUEFEITO",
    }
    cols_comb = []
    for col in df.columns:
        pk = _anp_norm(str(col))
        if any(tok in pk for tok in tokens_comb):
            cols_comb.append((col, pk))
    return cols_comb


def _processar_bytes_precos_postos(nome: str, conteudo: bytes):
    """
    Lê planilha de Preços por Posto.
    Aceita dois formatos:
      • Long  — colunas: CNPJ | Produto/Combustível | Preço | Data
      • Wide  — colunas: CNPJ | Gasolina | Diesel | Diesel S10 | ... | Data
                (uma coluna por combustível, valor = preço)

    Retorna (df_normalizado, msg, None) ou (None, msg_erro, None).
    df tem colunas: cnpj_norm, combustivel_pk, combustivel_label, preco, data_atualizacao
    """
    buf = io.BytesIO(conteudo)
    nome_l = nome.lower()
    if nome_l.endswith(".csv"):
        try:
            df = pd.read_csv(buf, dtype=str, encoding="utf-8")
        except UnicodeDecodeError:
            buf.seek(0)
            df = pd.read_csv(buf, dtype=str, encoding="latin-1")
    elif nome_l.endswith(".xls"):
        df = pd.read_excel(buf, dtype=str, engine="xlrd")
    else:
        df = pd.read_excel(buf, dtype=str, engine="openpyxl")

    if df.empty:
        return None, "A planilha está vazia.", None

    col_cnpj = detectar_coluna_cnpj(df)
    if col_cnpj is None:
        return None, (
            f"Coluna CNPJ não encontrada. Disponíveis: {', '.join(df.columns.tolist())}"
        ), None

    col_data = _detectar_col_data_pp(df)

    # ── Detecção automática de formato ────────────────────────────────
    col_comb  = _detectar_col_combustivel_pp(df)
    col_preco = _detectar_col_preco_pp(df)

    # Formato WIDE: sem coluna "Produto/Combustível" mas com colunas de combustíveis
    cols_wide = _colunas_combustivel_wide(df)
    usar_wide = (col_comb is None or col_preco is None) and len(cols_wide) >= 2

    result = []

    if usar_wide:
        # ── Formato Wide ──────────────────────────────────────────────
        for _, row in df.iterrows():
            cnpj_n = normalizar_cnpj(row.get(col_cnpj, ""))
            if len(cnpj_n) != 14:
                continue
            data_str = str(row.get(col_data, "")).strip() if col_data else ""
            for col_orig, pk in cols_wide:
                raw_val = str(row.get(col_orig, "")).strip()
                if not raw_val or raw_val.upper() in ("NAN", "NONE", "", "-", "N/A"):
                    continue
                try:
                    preco = float(raw_val.replace(",", "."))
                    if preco <= 0:
                        continue
                except (ValueError, TypeError):
                    continue
                result.append({
                    "cnpj_norm":         cnpj_n,
                    "combustivel_pk":    pk,
                    "combustivel_label": str(col_orig).strip(),
                    "preco":             preco,
                    "data_atualizacao":  data_str,
                })
    else:
        # ── Formato Long ──────────────────────────────────────────────
        if col_comb is None or col_preco is None:
            faltando = []
            if col_comb  is None: faltando.append("Combustível/Produto")
            if col_preco is None: faltando.append("Preço")
            return None, (
                f"Colunas não encontradas: **{', '.join(faltando)}**. "
                f"Disponíveis: {', '.join(df.columns.tolist())}"
            ), None

        for _, row in df.iterrows():
            cnpj_n = normalizar_cnpj(row.get(col_cnpj, ""))
            if len(cnpj_n) != 14:
                continue
            comb_raw = str(row.get(col_comb, "")).strip()
            comb_pk  = _anp_norm(comb_raw)
            if not comb_pk:
                continue
            try:
                preco_str = str(row.get(col_preco, "")).replace(",", ".").strip()
                preco = float(preco_str)
                if preco <= 0:
                    continue
            except (ValueError, TypeError):
                continue
            data_str = str(row.get(col_data, "")).strip() if col_data else ""
            result.append({
                "cnpj_norm":         cnpj_n,
                "combustivel_pk":    comb_pk,
                "combustivel_label": comb_raw,
                "preco":             preco,
                "data_atualizacao":  data_str,
            })

    if not result:
        return None, "Nenhuma linha válida encontrada (verifique CNPJ e preço).", None

    df_out = pd.DataFrame(result)
    fmt = "wide" if usar_wide else "long"
    msg = (f"{len(df_out)} registros · "
           f"{df_out['cnpj_norm'].nunique()} postos · "
           f"{df_out['combustivel_pk'].nunique()} combustíveis "
           f"[formato {fmt}]")
    return df_out, msg, None


def ler_planilha_precos_postos(arquivo):
    """Lê UploadedFile do Streamlit (sem cache — upload não é cacheável)."""
    try:
        return _processar_bytes_precos_postos(arquivo.name, arquivo.read())
    except Exception as e:
        return None, f"Erro: {type(e).__name__} — {e}", None


def _normalizar_nome_arquivo(nome: str) -> str:
    """Remove acentos e normaliza nome de arquivo para comparação tolerante."""
    sem_acento = unicodedata.normalize("NFD", nome)
    sem_acento = "".join(c for c in sem_acento
                         if unicodedata.category(c) != "Mn")
    return sem_acento.lower().replace(" ", "_").replace("-", "_")


@st.cache_data(show_spinner=False, ttl=3600)   # 1 h — preços atualizam com frequência
def _auto_carregar_precos_postos_repo():
    """
    Tenta carregar a planilha de Preços Posto do diretório do repositório.
    Usa comparação tolerante a acentos e maiúsculas/minúsculas para encontrar
    o arquivo mesmo quando o nome contém caracteres especiais (ex: ç em 'Preço').
    """
    # Fragmentos-chave que o arquivo deve conter (normalizados)
    fragmentos_chave = ["preco", "posto"]

    try:
        arquivos_dir = os.listdir(_DIR)
    except Exception:
        arquivos_dir = []

    # 1ª tentativa: comparação exata (rápida)
    candidatos_exatos = [
        "preco_posto.xlsx",      # nome padrão do repositório
        ARQUIVO_PP_REPO,         # alias configurado (atualmente = "preco_posto.xlsx")
        "preco_postos.xlsx",
        "Preco Posto.xlsx",
        "Preço Posto.xlsx",
        "precos_postos.xlsx",
        "precos_posto.xlsx",
        "Preco_Posto.xlsx",
        "Preços Posto.xlsx",
        "Precos Posto.xlsx",
    ]
    for nome in candidatos_exatos:
        caminho = os.path.join(_DIR, nome)
        if os.path.exists(caminho):
            try:
                with open(caminho, "rb") as f:
                    conteudo = f.read()
                df, msg, _ = _processar_bytes_precos_postos(nome, conteudo)
                if df is not None:
                    return df, msg, None
            except Exception as e:
                return None, f"Erro ao ler {nome}: {e}", None

    # 2ª tentativa: varredura tolerante a acentos sobre os arquivos reais do diretório
    for arq in arquivos_dir:
        if not arq.lower().endswith((".xlsx", ".xls")):
            continue
        arq_norm = _normalizar_nome_arquivo(arq)
        if all(frag in arq_norm for frag in fragmentos_chave):
            caminho = os.path.join(_DIR, arq)
            try:
                with open(caminho, "rb") as f:
                    conteudo = f.read()
                df, msg, _ = _processar_bytes_precos_postos(arq, conteudo)
                if df is not None:
                    return df, msg, None
            except Exception as e:
                return None, f"Erro ao ler {arq}: {e}", None

    return None, f"Arquivo `{ARQUIVO_PP_REPO}` não encontrado em: {_DIR}", None


# ── Gestão de Frotas ───────────────────────────────────────────────────────

def marcar_pro_frotas(df: pd.DataFrame, cnpjs_pf: set) -> pd.DataFrame:
    df = df.copy()
    if cnpjs_pf and "cnpj" in df.columns:
        df["_cnpj_norm"] = df["cnpj"].fillna("").str.replace(r'\D', '', regex=True)
        df["_pro_frotas"] = df["_cnpj_norm"].isin(cnpjs_pf)
    else:
        df["_pro_frotas"] = False
    return df


# ═══════════════════════════════════════════════════════════════════
#  AUTOCOMPLETE — Nominatim
# ═══════════════════════════════════════════════════════════════════

@st.cache_data(show_spinner=False, ttl=3600)
def sugestoes_nominatim(texto: str):
    texto = texto.strip()
    if len(texto) < 3:
        return []
    try:
        r = requests.get(NOMINATIM, params={
            "q": f"{texto}, Brasil", "format": "json", "limit": 6,
            "countrycodes": "br", "addressdetails": 1,
        }, headers={"User-Agent": "EstudoDeRedeANP/4.0"}, timeout=8)
        opcoes, vistos = [], set()
        for item in r.json():
            addr   = item.get("address", {})
            cidade = addr.get("city") or addr.get("town") or addr.get("village") \
                     or addr.get("municipality") or addr.get("county") or ""
            estado = addr.get("state", "")
            label  = f"{cidade} – {estado}" if cidade and estado else (
                estado or ", ".join(item["display_name"].split(", ")[:2]))
            if label not in vistos:
                vistos.add(label)
                opcoes.append({"label": label, "lat": float(item["lat"]), "lon": float(item["lon"])})
        return opcoes
    except Exception:
        return []


def _formatar_cnpj(cnpj_str: str) -> str:
    """Formata string de dígitos como CNPJ: XX.XXX.XXX/XXXX-XX."""
    d = "".join(c for c in str(cnpj_str) if c.isdigit())
    if len(d) == 14:
        return f"{d[:2]}.{d[2:5]}.{d[5:8]}/{d[8:12]}-{d[12:]}"
    return cnpj_str


def buscar_posto_por_texto(texto: str, max_results: int = 6) -> list:
    """
    Busca postos por razão social (parcial) ou CNPJ nos estados já em cache.
    Usa apenas dados que já estão na memória — sem chamadas extras à API ANP.
    Retorna lista de dicts: {label, lat, lon, tipo='posto'}.
    """
    texto_clean = texto.strip()
    if len(texto_clean) < 3:
        return []

    cnpj_digits = "".join(c for c in texto_clean if c.isdigit())
    # É CNPJ se a maioria dos caracteres são dígitos (ex: "12.345" ou "123456")
    is_cnpj = (
        len(cnpj_digits) >= 6
        and len(cnpj_digits) / max(len(texto_clean.replace(" ", "")), 1) > 0.65
    )

    estados = st.session_state.get("_estados_precarregados", [])
    if not estados:
        return []

    resultados = []
    for uf in estados:
        try:
            df = buscar_postos(uf=uf)   # instantâneo se estiver em cache
            if df.empty or "razaoSocial" not in df.columns:
                continue

            if is_cnpj:
                mask = df["cnpj"].fillna("").apply(
                    lambda x: cnpj_digits in "".join(c for c in str(x) if c.isdigit())
                )
            else:
                mask = df["razaoSocial"].fillna("").str.upper().str.contains(
                    texto_clean.upper(), regex=False, na=False
                )

            for _, row in df[mask].head(3).iterrows():
                nome   = str(row.get("razaoSocial", "?"))
                cidade = str(row.get("municipio", ""))
                uf_r   = str(row.get("uf", uf))
                cnpj   = _formatar_cnpj(str(row.get("cnpj", "")))
                label  = f"⛽ {nome} — {cidade}/{uf_r} | CNPJ: {cnpj}"
                resultados.append({
                    "label": label,
                    "lat":   float(row["_lat"]),
                    "lon":   float(row["_lon"]),
                    "tipo":  "posto",
                })
                if len(resultados) >= max_results:
                    break
        except Exception:
            continue
        if len(resultados) >= max_results:
            break

    return resultados


def campo_autocomplete(titulo, placeholder, key_texto, key_estado):
    """Campo inteligente que aceita 4 tipos de entrada:
    • UF/Estado  — ex: SP, RJ, MG  (2 letras → centro do estado)
    • Cidade     — ex: Ribeirao Preto  (busca no cache ANP, sem acento)
    • Razão Social — ex: Rudnick  (busca por nome do posto)
    • CNPJ       — ex: 12.345.678/0001-99  (busca por CNPJ)
    """
    fk = st.session_state.get("_form_key", 0)
    key_txt_widget = f"{key_texto}_{fk}"
    key_sel_widget = f"_sel_{key_estado}_{fk}"

    st.markdown(f"<div style='font-weight:700;font-size:13px;margin-bottom:4px'>{titulo}</div>",
                unsafe_allow_html=True)
    texto = st.text_input(titulo, placeholder=placeholder,
                          key=key_txt_widget, label_visibility="collapsed")

    ultimo = st.session_state.get(f"_{key_estado}_txt_ant", "")
    if texto != ultimo:
        st.session_state[f"_{key_estado}_txt_ant"] = texto
        if len(texto) < 2:
            st.session_state.pop(key_estado, None)

    texto_strip = texto.strip()
    texto_up    = texto_strip.upper()
    sugestoes   = []

    if len(texto_strip) >= 2:
        cnpj_digits = "".join(c for c in texto_strip if c.isdigit())
        is_cnpj = (
            len(cnpj_digits) >= 6
            and len(cnpj_digits) / max(len(texto_strip.replace(" ", "")), 1) > 0.65
        )

        # ── 1. UF/Estado (2 letras exatas) ──────────────────────────
        if texto_up in UFS:
            bbox   = BBOX_UFS.get(texto_up, (-15.8, -47.9, -15.7, -47.8))
            lat_c  = (bbox[0] + bbox[2]) / 2
            lon_c  = (bbox[1] + bbox[3]) / 2
            sugestoes = [{
                "label": f"🗺️ Estado {texto_up}",
                "lat": lat_c, "lon": lon_c, "tipo": "estado",
            }]

        # ── 2. CNPJ (maioria dígitos) ────────────────────────────────
        elif is_cnpj:
            sugestoes = buscar_posto_por_texto(texto_strip)
            if not sugestoes:
                n_est = len(st.session_state.get("_estados_precarregados", []))
                msg = ("⚠️ Base ainda carregando — tente novamente em instantes."
                       if n_est == 0 else
                       f"⚠️ CNPJ não encontrado nos {n_est} estado(s) carregado(s).")
                st.markdown(f"<small style='color:#e65100'>{msg}</small>",
                            unsafe_allow_html=True)

        # ── 3. Cidade ou Razão Social (mín. 3 letras) ───────────────
        elif len(texto_strip) >= 3:
            # Cidade: busca no cache ANP com normalização de acentos
            sug_cidades = _buscar_cidades_cache(texto_strip)
            # Fallback Nominatim se base ainda não tiver sido carregada
            if not sug_cidades:
                sug_cidades = [
                    dict(s, tipo="cidade") for s in sugestoes_nominatim(texto_strip)
                ]
            # Razão social: busca por nome do posto
            sug_postos = buscar_posto_por_texto(texto_strip)
            sugestoes  = sug_cidades[:4] + sug_postos[:4]

        else:
            st.markdown(
                "<small style='color:#888'>"
                "Digite UF (ex: SP), cidade, nome do posto ou CNPJ…"
                "</small>",
                unsafe_allow_html=True,
            )

    elif len(texto_strip) == 1:
        st.markdown(
            "<small style='color:#888'>"
            "Digite UF (ex: SP), cidade, nome do posto ou CNPJ…"
            "</small>",
            unsafe_allow_html=True,
        )

    if sugestoes:
        labels = [s["label"] for s in sugestoes]
        idx = st.selectbox("Sugestões:", range(len(labels)),
                           format_func=lambda i: labels[i], key=key_sel_widget)
        sel = sugestoes[idx]
        st.session_state[key_estado] = sel
        tipo  = sel.get("tipo", "")
        icone = {"estado": "🗺️", "cidade": "📍", "posto": "⛽"}.get(tipo, "📍")
        st.markdown(
            f"<small style='color:#1565c0'>{icone} {sel['label']}</small>",
            unsafe_allow_html=True,
        )
        return sel
    elif len(texto_strip) >= 3 and texto_up not in UFS:
        st.markdown(
            "<small style='color:#e65100'>⚠️ Nenhuma sugestão encontrada. "
            "Tente outra grafia ou verifique se a base foi carregada.</small>",
            unsafe_allow_html=True,
        )
        return st.session_state.get(key_estado)

    return st.session_state.get(key_estado)


def _campo_rota_compacto(
    placeholder: str,
    key_texto: str,
    key_estado: str,
    icon_bg: str = "#2E7D32",
    icon_number: str = "",
    always_show_action: bool = False,
    action_key: str | None = None,
    action_help: str = "Limpar seleção",
) -> bool:
    """Campo compacto para seleção de ponto de rota.

    Confirmado → linha única: [ícone] [chip ✓ Local] [✕]
    Digitando  → linha única: [ícone] [text_input] + sugestões abaixo

    Parâmetros
    ----------
    always_show_action : bool
        True → mostra botão ✕ mesmo sem seleção (para paradas intermediárias,
        onde ✕ remove a parada inteira).
    action_key : str | None
        Chave Streamlit do botão ✕ (gerada automaticamente se None).
    action_help : str
        Tooltip do botão ✕.

    Retorna
    -------
    bool – True se o botão ✕ foi clicado (chamador decide o que fazer).
    """
    fk             = st.session_state.get("_form_key", 0)
    key_txt_widget = f"{key_texto}_{fk}"
    key_sel_widget = f"_sel_{key_estado}_{fk}"
    _act_key       = action_key or f"_act_{key_estado}_{fk}"

    # ── ícone circular ─────────────────────────────────────────────
    _inner = (
        f"<span style='font-size:9px;font-weight:800;color:#fff'>{icon_number}</span>"
        if icon_number else ""
    )
    _icon_html = (
        f"<div style='min-width:18px;height:18px;border-radius:50%;background:{icon_bg};"
        f"display:flex;align-items:center;justify-content:center;flex-shrink:0;"
        f"margin-top:6px'>{_inner}</div>"
    )

    _sel = st.session_state.get(key_estado)

    # ══ Estado CONFIRMADO ══════════════════════════════════════════
    if _sel:
        _lbl  = str(_sel.get("label", ""))[:40]
        _tipo = _sel.get("tipo", "")
        _ico  = {"estado": "🗺️", "cidade": "📍", "posto": "⛽"}.get(_tipo, "📍")

        _c_ic, _c_chip, _c_act = st.columns([1, 8, 1])
        with _c_ic:
            st.markdown(_icon_html, unsafe_allow_html=True)
        with _c_chip:
            st.markdown(
                f"<div style='background:#e8f5e9;border:1px solid #a5d6a7;"
                f"border-radius:6px;padding:5px 8px;font-size:11px;color:#2e7d32;"
                f"overflow:hidden;white-space:nowrap;text-overflow:ellipsis;"
                f"line-height:1.5'>{_ico} {_lbl}</div>",
                unsafe_allow_html=True,
            )
        with _c_act:
            if st.button("✕", key=_act_key, help=action_help,
                         use_container_width=True):
                # Limpa seleção interna (caller pode fazer cleanup adicional)
                st.session_state.pop(key_estado, None)
                st.session_state.pop(f"_{key_estado}_txt_ant", None)
                st.session_state[key_txt_widget] = ""
                return True
        return False

    # ══ Estado DIGITANDO ═══════════════════════════════════════════
    if always_show_action:
        _c_ic2, _c_inp, _c_act2 = st.columns([1, 8, 1])
    else:
        _c_ic2, _c_inp = st.columns([1, 9])

    with _c_ic2:
        st.markdown(_icon_html, unsafe_allow_html=True)
    with _c_inp:
        texto = st.text_input(
            "", placeholder=placeholder,
            key=key_txt_widget, label_visibility="collapsed",
        )

    if always_show_action:
        with _c_act2:
            if st.button("✕", key=_act_key, help=action_help,
                         use_container_width=True):
                return True  # caller deleta a parada

    # ── Lógica de sugestões (idêntica a campo_autocomplete) ────────
    _ultimo = st.session_state.get(f"_{key_estado}_txt_ant", "")
    if texto != _ultimo:
        st.session_state[f"_{key_estado}_txt_ant"] = texto
        if len(texto) < 2:
            st.session_state.pop(key_estado, None)

    _ts = texto.strip()
    _tu = _ts.upper()
    _sugestoes: list = []

    if len(_ts) >= 2:
        _digits = "".join(c for c in _ts if c.isdigit())
        _is_cnpj = (
            len(_digits) >= 6
            and len(_digits) / max(len(_ts.replace(" ", "")), 1) > 0.65
        )
        if _tu in UFS:
            _bbox  = BBOX_UFS.get(_tu, (-15.8, -47.9, -15.7, -47.8))
            _lat_c = (_bbox[0] + _bbox[2]) / 2
            _lon_c = (_bbox[1] + _bbox[3]) / 2
            _sugestoes = [{"label": f"🗺️ Estado {_tu}",
                           "lat": _lat_c, "lon": _lon_c, "tipo": "estado"}]
        elif _is_cnpj:
            _sugestoes = buscar_posto_por_texto(_ts)
            if not _sugestoes:
                _n_est = len(st.session_state.get("_estados_precarregados", []))
                _msg = (
                    "⚠️ Base ainda carregando — tente novamente em instantes."
                    if _n_est == 0 else
                    f"⚠️ CNPJ não encontrado nos {_n_est} estado(s) carregado(s)."
                )
                st.markdown(f"<small style='color:#e65100'>{_msg}</small>",
                            unsafe_allow_html=True)
        elif len(_ts) >= 3:
            _sug_cid = _buscar_cidades_cache(_ts)
            if not _sug_cid:
                _sug_cid = [dict(s, tipo="cidade") for s in sugestoes_nominatim(_ts)]
            _sug_pos   = buscar_posto_por_texto(_ts)
            _sugestoes = _sug_cid[:4] + _sug_pos[:4]

    if _sugestoes:
        _labels = [s["label"] for s in _sugestoes]
        _idx = st.selectbox(
            "", range(len(_labels)),
            format_func=lambda i: _labels[i],
            key=key_sel_widget,
            index=None,
            placeholder="↑ selecione uma sugestão…",
            label_visibility="collapsed",
        )
        if _idx is not None:
            st.session_state[key_estado] = _sugestoes[_idx]
            st.rerun()

    return False


# ═══════════════════════════════════════════════════════════════════
#  DISTÂNCIA / GEOMETRIA
# ═══════════════════════════════════════════════════════════════════

def _haversine(lat1, lon1, lat2, lon2):
    R = 6_371_000
    φ1, φ2 = math.radians(lat1), math.radians(lat2)
    dφ, dλ = math.radians(lat2-lat1), math.radians(lon2-lon1)
    a = math.sin(dφ/2)**2 + math.cos(φ1)*math.cos(φ2)*math.sin(dλ/2)**2
    return R * 2 * math.asin(math.sqrt(a))


def dist_minima_rota_np(lats_arr, lons_arr, coords_rota, chunk=4000):
    """
    Calcula distância mínima de TODOS os postos à rota (NumPy vetorizado).
    lats_arr, lons_arr : arrays 1-D com as coordenadas dos postos (M elementos)
    coords_rota        : lista de [lat, lon] do trajeto  (N pontos)
    chunk              : processa postos em lotes para limitar uso de RAM
    Retorna            : array 1-D com distância em metros para cada posto

    Limite de memória: chunk × 150 segmentos × 6 arrays × 8 bytes ≈ 28 MB/lote
    → seguro para o limite de 1 GB do Streamlit Community Cloud.
    """
    if not coords_rota:
        return np.full(len(lats_arr), np.inf)

    # Limita a rota a 150 pontos — suficiente para precisão de ±50 m
    coords_ds = _downsample(coords_rota, 150)
    rota = np.array(coords_ds, dtype=np.float64)            # (N, 2)
    lats = np.asarray(lats_arr, dtype=np.float64)           # (M,)
    lons = np.asarray(lons_arr, dtype=np.float64)           # (M,)

    # Projeção plana local (erro < 0,1 % para distâncias até 500 km)
    R       = 6_371_000.0
    lat0    = rota[0, 0];  lon0 = rota[0, 1]
    cos_lat = np.cos(np.radians(rota[:, 0].mean()))

    # Converte para metros (eixo X = leste, Y = norte)
    rx = np.radians(rota[:, 1] - lon0) * cos_lat * R        # (N,)
    ry = np.radians(rota[:, 0] - lat0) * R                  # (N,)
    px = np.radians(lons - lon0)        * cos_lat * R        # (M,)
    py = np.radians(lats - lat0)        * R                  # (M,)

    # Vetores de cada segmento A→B
    ax = rx[:-1];  ay = ry[:-1]                              # (N-1,)
    dx = rx[1:] - ax;  dy = ry[1:] - ay                     # (N-1,)
    ab2 = dx*dx + dy*dy
    ab2 = np.where(ab2 < 1e-10, 1e-10, ab2)                 # evita /0

    # Processa em lotes para não estourar RAM (cada lote ≈ 28 MB)
    M = len(lats)
    result = np.empty(M, dtype=np.float64)
    for start in range(0, M, chunk):
        end = min(start + chunk, M)
        apx = px[start:end, None] - ax[None, :]             # (chunk, N-1)
        apy = py[start:end, None] - ay[None, :]
        t   = np.clip((apx * dx + apy * dy) / ab2, 0.0, 1.0)
        ex  = apx - t * dx
        ey  = apy - t * dy
        result[start:end] = np.sqrt((ex*ex + ey*ey).min(axis=1))

    return result                                            # (M,)


def _downsample(coords, max_pts=300):
    """Reduz o número de pontos da polyline sem perder o traçado geral."""
    if len(coords) <= max_pts:
        return coords
    step = max(1, len(coords) // max_pts)
    result = coords[::step]
    # Garante que o último ponto está incluído
    if result[-1] != coords[-1]:
        result = result + [coords[-1]]
    return result


def ufs_ao_longo_rota(coords_rota):
    ufs = set()
    passo = max(1, len(coords_rota)//80)
    for lat, lon in coords_rota[::passo]:
        for uf, (la_min, lo_min, la_max, lo_max) in BBOX_UFS.items():
            if la_min <= lat <= la_max and lo_min <= lon <= lo_max:
                ufs.add(uf)
    return sorted(ufs)


# ═══════════════════════════════════════════════════════════════════
#  PDF — RELATÓRIO DE ROTEIRIZAÇÃO
# ═══════════════════════════════════════════════════════════════════

def _gerar_mapa_rota_png(coords_rota, orig, dest, paradas, sugest):
    """
    Renderiza o mapa da rota e retorna bytes PNG.
    Estratégia:
      1. staticmap com tiles OpenStreetMap (mapa real com estradas/cidades).
         Funciona quando o servidor tem acesso à internet (Streamlit Cloud).
      2. Fallback: matplotlib limpo sem eixos, visual de cartografia.
    coords_rota: [[lat, lon], ...]   orig/dest: dict com lat/lon/label
    paradas: list de dict           sugest: list de dict com lat/lon
    """

    # ── Tentativa 1: staticmap + tiles OSM ───────────────────────────────────
    try:
        from staticmap import StaticMap, Line, CircleMarker
        from PIL import Image as _PILImage, ImageDraw as _PILDraw, ImageFont as _PILFont

        _W, _H = 900, 480

        _sm = StaticMap(
            _W, _H,
            url_template="https://tile.openstreetmap.org/{z}/{x}/{y}.png",
            headers={"User-Agent": "EstudoDeRede-ProFrotas/2.0"},
        )

        # Linha da rota
        _route_xy = [(float(c[1]), float(c[0])) for c in coords_rota]
        _sm.add_line(Line(_route_xy, "#1565C0", 4))

        # Postos sugeridos (⛽ laranja)
        for _s in sugest:
            _sm.add_marker(CircleMarker((float(_s["lon"]), float(_s["lat"])),
                                        "#FF6F00", 16))
            _sm.add_marker(CircleMarker((float(_s["lon"]), float(_s["lat"])),
                                        "#ffffff", 8))

        # Paradas intermediárias (laranja escuro)
        for _p in paradas:
            _sm.add_marker(CircleMarker((float(_p["lon"]), float(_p["lat"])),
                                        "#E65100", 16))

        # Destino (vermelho)
        _sm.add_marker(CircleMarker((float(dest["lon"]), float(dest["lat"])),
                                    "#C62828", 20))
        _sm.add_marker(CircleMarker((float(dest["lon"]), float(dest["lat"])),
                                    "#ffffff", 8))

        # Origem (verde) — desenhada por cima para ficar visível
        _sm.add_marker(CircleMarker((float(orig["lon"]), float(orig["lat"])),
                                    "#2E7D32", 20))
        _sm.add_marker(CircleMarker((float(orig["lon"]), float(orig["lat"])),
                                    "#ffffff", 8))

        _img = _sm.render()
        _draw = _PILDraw.Draw(_img)

        # Função auxiliar: converte lon/lat → pixel no resultado renderizado
        import math as _math
        def _ll_to_px(lat, lon, zoom, w, h):
            _n = 2 ** zoom
            _x_tile = (lon + 180) / 360 * _n
            _y_tile = (1 - _math.log(_math.tan(_math.radians(lat)) +
                       1 / _math.cos(_math.radians(lat))) / _math.pi) / 2 * _n
            # centro do mapa
            _lats = [c[0] for c in coords_rota]
            _lons = [c[1] for c in coords_rota]
            _lat_c = (max(_lats) + min(_lats)) / 2
            _lon_c = (max(_lons) + min(_lons)) / 2
            _xc = ((_lon_c + 180) / 360 * _n)
            _yc = ((1 - _math.log(_math.tan(_math.radians(_lat_c)) +
                    1 / _math.cos(_math.radians(_lat_c))) / _math.pi) / 2 * _n)
            _px = int((_x_tile - _xc) * 256 + w / 2)
            _py = int((_y_tile - _yc) * 256 + h / 2)
            return _px, _py

        # Determina zoom usado pelo staticmap para calcular posição dos labels
        _lats_r = [c[0] for c in coords_rota]
        _lons_r = [c[1] for c in coords_rota]
        _span_lat = max(_lats_r) - min(_lats_r)
        _span_lon = max(_lons_r) - min(_lons_r)
        _span_max = max(_span_lat, _span_lon * 0.65)
        _zoom_est = max(4, min(13, int(7 - _math.log2(max(_span_max, 0.01)))))

        def _label(draw, text, lon, lat, color, bg):
            _px, _py = _ll_to_px(lat, lon, _zoom_est, _W, _H)
            _txt = str(text)[:28]
            _bbox = draw.textbbox((_px + 10, _py - 8), _txt)
            _pad = 3
            draw.rounded_rectangle(
                (_bbox[0]-_pad, _bbox[1]-_pad, _bbox[2]+_pad, _bbox[3]+_pad),
                radius=4, fill=bg + "DD"
            )
            draw.text((_px + 10, _py - 8), _txt, fill=color)

        try:
            _label(_draw, orig.get("label", "Origem"),
                   float(orig["lon"]), float(orig["lat"]), "#1B5E20", "#ffffff")
            _label(_draw, dest.get("label", "Destino"),
                   float(dest["lon"]), float(dest["lat"]), "#B71C1C", "#ffffff")
            for _i, _p in enumerate(paradas, 1):
                _label(_draw, _p.get("label", f"Parada {_i}"),
                       float(_p["lon"]), float(_p["lat"]), "#BF360C", "#fff3e0")
            for _s in sugest:
                _label(_draw, _s.get("razaoSocial", _s.get("nome", "Posto"))[:20],
                       float(_s["lon"]), float(_s["lat"]), "#E65100", "#fff8e1")
        except Exception:
            pass  # labels opcionais

        # Legenda simples no canto inferior direito
        _leg_items = [
            ("#1565C0", "Rota"),
            ("#2E7D32", "Origem"),
            ("#C62828", "Destino"),
        ]
        if paradas:  _leg_items.append(("#E65100", f"{len(paradas)} Parada(s)"))
        if sugest:   _leg_items.append(("#FF6F00", f"{len(sugest)} Posto(s) sugerido(s)"))

        _lx, _ly = _W - 180, _H - len(_leg_items) * 22 - 12
        _draw.rounded_rectangle((_lx - 8, _ly - 8, _W - 6, _H - 6),
                                  radius=6, fill="#ffffffCC")
        for _j, (_cor, _txt) in enumerate(_leg_items):
            _yy = _ly + _j * 22
            _draw.ellipse((_lx, _yy + 3, _lx + 14, _yy + 17), fill=_cor)
            _draw.text((_lx + 20, _yy), _txt, fill="#333333")

        _buf_sm = io.BytesIO()
        _img.save(_buf_sm, format="PNG")
        _buf_sm.seek(0)
        return _buf_sm.read()

    except Exception:
        pass  # cai no fallback matplotlib

    # ── Fallback: matplotlib sem eixos ────────────────────────────────────────
    lats = [c[0] for c in coords_rota]
    lons = [c[1] for c in coords_rota]

    fig, ax = plt.subplots(figsize=(10, 5.5), dpi=130)

    # Fundo estilo cartográfico claro
    ax.set_facecolor("#EEF2F7")
    fig.patch.set_facecolor("#EEF2F7")

    # Rota com sombra
    ax.plot(lons, lats, color="#90CAF9", linewidth=5.0, zorder=2,
            solid_capstyle="round", alpha=0.5)
    ax.plot(lons, lats, color="#1565C0", linewidth=2.5, zorder=3,
            solid_capstyle="round")

    # Postos sugeridos
    for s in sugest:
        ax.scatter(float(s["lon"]), float(s["lat"]),
                   s=120, c="#FF6F00", marker="^", zorder=5,
                   linewidths=1.0, edgecolors="#fff")

    # Paradas intermediárias
    for p in paradas:
        ax.scatter(float(p["lon"]), float(p["lat"]),
                   s=110, c="#E65100", marker="D", zorder=6,
                   linewidths=1.0, edgecolors="#fff")

    # Origem e Destino
    ax.scatter(float(orig["lon"]), float(orig["lat"]),
               s=200, c="#2E7D32", marker="o", zorder=7,
               linewidths=1.2, edgecolors="#fff")
    ax.scatter(float(dest["lon"]), float(dest["lat"]),
               s=200, c="#C62828", marker="s", zorder=7,
               linewidths=1.2, edgecolors="#fff")

    # Labels origem/destino
    for _pt, _cor_txt, _cor_bg in [
        (orig, "#1B5E20", "#E8F5E9"),
        (dest, "#B71C1C", "#FFEBEE"),
    ]:
        ax.annotate(
            str(_pt.get("label", ""))[:25],
            xy=(float(_pt["lon"]), float(_pt["lat"])),
            xytext=(8, 8), textcoords="offset points",
            fontsize=7.5, color=_cor_txt, fontweight="bold",
            bbox=dict(boxstyle="round,pad=0.3", fc=_cor_bg, alpha=0.9, ec="none"),
        )

    # Margens sem eixos visíveis
    _pad_lat = max((max(lats) - min(lats)) * 0.13, 0.35)
    _pad_lon = max((max(lons) - min(lons)) * 0.13, 0.35)
    ax.set_xlim(min(lons) - _pad_lon, max(lons) + _pad_lon)
    ax.set_ylim(min(lats) - _pad_lat, max(lats) + _pad_lat)

    # Remove todos os eixos e ticks
    ax.set_xticks([])
    ax.set_yticks([])
    ax.set_xlabel("")
    ax.set_ylabel("")
    for spine in ax.spines.values():
        spine.set_visible(False)

    # Grade ultra-suave
    ax.grid(True, linestyle=":", linewidth=0.3, color="#B0BEC5", alpha=0.5)

    # Legenda
    _legenda = [
        Line2D([0], [0], color="#1565C0", linewidth=2.5, label="Rota"),
        mpatches.Patch(color="#2E7D32", label="Origem"),
        mpatches.Patch(color="#C62828", label="Destino"),
    ]
    if paradas:
        _legenda.append(mpatches.Patch(color="#E65100", label="Parada(s)"))
    if sugest:
        _legenda.append(Line2D([0], [0], marker="^", color="w",
                               markerfacecolor="#FF6F00", markersize=9,
                               label="Posto sugerido"))
    ax.legend(handles=_legenda, fontsize=7, loc="lower right",
              framealpha=0.9, edgecolor="#CFD8DC",
              facecolor="white", labelcolor="#333333")

    plt.tight_layout(pad=0.4)
    _buf = io.BytesIO()
    fig.savefig(_buf, format="png", bbox_inches="tight", dpi=130)
    plt.close(fig)
    _buf.seek(0)
    return _buf.read()


# ═══════════════════════════════════════════════════════════════════
#  CARD PNG PARA COMPARTILHAMENTO DE ROTA
# ═══════════════════════════════════════════════════════════════════

def _gerar_card_rota_png(rot_res: dict, sugest: list) -> bytes:
    """
    Gera um card PNG visual para compartilhamento da roteirização
    (WhatsApp, e-mail, etc.).  Retorna bytes PNG.

    Layout:
      ┌──────────────────────────────────────┐
      │  Header azul: branding + placa       │
      ├──────────────────┬───────────────────┤
      │   Mini-mapa      │  Stats da rota    │
      ├──────────────────┴───────────────────┤
      │  Lista de postos sugeridos           │
      └──────────────────────────────────────┘
    """
    import matplotlib.gridspec as _mgs
    import textwrap as _tw

    _AZ    = "#0D47A1"
    _AZ2   = "#1565C0"
    _AZ3   = "#1976D2"
    _VERDE = "#2E7D32"
    _VERD2 = "#E8F5E9"
    _LAR   = "#E65100"
    _LAR2  = "#FFF3E0"
    _CINZ  = "#ECEFF1"
    _BRNC  = "#FFFFFF"
    _DARK  = "#212121"
    _MED   = "#546E7A"

    DPI   = 140
    W_IN  = 10.0
    H_IN  = 6.5

    fig = plt.figure(figsize=(W_IN, H_IN), dpi=DPI, facecolor=_AZ)

    # ── GridSpec: header | corpo | footer ──────────────────────────
    gs_outer = _mgs.GridSpec(3, 1, figure=fig,
                              height_ratios=[0.13, 0.72, 0.15],
                              hspace=0, left=0, right=1, top=1, bottom=0)

    # ── HEADER ─────────────────────────────────────────────────────
    ax_hdr = fig.add_subplot(gs_outer[0])
    ax_hdr.set_facecolor(_AZ)
    ax_hdr.set_xlim(0, 1); ax_hdr.set_ylim(0, 1)
    ax_hdr.axis("off")
    _placa = rot_res.get("placa","") or ""
    _hdr_txt = "🚛  Estudo de Rede — Gestão de Frotas"
    if _placa:
        _hdr_txt += f"  ·  Placa {_placa}"
    ax_hdr.text(0.5, 0.5, _hdr_txt,
                ha="center", va="center", fontsize=11, fontweight="bold",
                color=_BRNC, transform=ax_hdr.transAxes)

    # ── CORPO ──────────────────────────────────────────────────────
    ax_body = fig.add_subplot(gs_outer[1])
    ax_body.set_facecolor(_BRNC)
    ax_body.axis("off")

    # Divide corpo: mapa (esq) + info (dir)
    gs_body = _mgs.GridSpecFromSubplotSpec(
        1, 2, subplot_spec=gs_outer[1],
        width_ratios=[1.1, 0.9], wspace=0.0
    )

    # ── Mapa (coluna esquerda) ──────────────────────────────────────
    ax_map = fig.add_subplot(gs_body[0])
    ax_map.set_facecolor(_CINZ)
    ax_map.axis("off")

    _rc  = rot_res.get("coords",  [])
    _ro  = rot_res.get("orig",    {})
    _rt  = rot_res.get("dest",    {})
    _rp  = rot_res.get("paradas", [])

    if _rc and len(_rc) >= 2:
        try:
            _map_bytes = _gerar_mapa_rota_png(_rc, _ro, _rt, _rp, sugest)
            from PIL import Image as _PILI
            _map_img = _PILI.open(io.BytesIO(_map_bytes))
            ax_map.imshow(_map_img, aspect="auto")
        except Exception:
            # Fallback matplotlib inline
            _lts = [c[0] for c in _rc]; _lns = [c[1] for c in _rc]
            ax_map.set_facecolor("#EEF2F7")
            ax_map.plot(_lns, _lts, color=_AZ2, linewidth=2.5)
            if _ro: ax_map.scatter(float(_ro["lon"]), float(_ro["lat"]),
                                   s=100, c=_VERDE, zorder=5)
            if _rt: ax_map.scatter(float(_rt["lon"]), float(_rt["lat"]),
                                   s=100, c="#C62828", zorder=5)
            for _s in sugest:
                ax_map.scatter(float(_s["lon"]), float(_s["lat"]),
                               s=70, c=_LAR, marker="^", zorder=6)
            _pad = 0.3
            ax_map.set_xlim(min(_lns)-_pad, max(_lns)+_pad)
            ax_map.set_ylim(min(_lts)-_pad, max(_lts)+_pad)
            ax_map.set_xticks([]); ax_map.set_yticks([])
    else:
        ax_map.text(0.5, 0.5, "Mapa indisponível\n(sem coordenadas de rota)",
                    ha="center", va="center", fontsize=9, color=_MED)

    # ── Info (coluna direita) ───────────────────────────────────────
    ax_info = fig.add_subplot(gs_body[1])
    ax_info.set_facecolor(_BRNC)
    ax_info.set_xlim(0, 1); ax_info.set_ylim(0, 1)
    ax_info.axis("off")

    _rd  = float(rot_res.get("dist_km", 0) or 0)
    _rm  = float(rot_res.get("dur_min", 0) or 0)
    _cap = float(rot_res.get("capacidade", 0) or 0)
    _aut = float(rot_res.get("autonomia",  0) or 0)
    _comb = rot_res.get("combustivel", "—") or "—"

    _orig_lbl = (_ro.get("label","Origem") or "Origem")[:32]
    _dest_lbl = (_rt.get("label","Destino") or "Destino")[:32]

    _y = 0.94

    # Título Origem → Destino
    ax_info.text(0.5, _y, _orig_lbl, ha="center", va="top", fontsize=9,
                 color=_VERDE, fontweight="bold",
                 transform=ax_info.transAxes)
    _y -= 0.07
    ax_info.text(0.5, _y, "▼", ha="center", va="top", fontsize=12,
                 color=_AZ3, transform=ax_info.transAxes)
    _y -= 0.09
    ax_info.text(0.5, _y, _dest_lbl, ha="center", va="top", fontsize=9,
                 color="#C62828", fontweight="bold",
                 transform=ax_info.transAxes)
    _y -= 0.06

    # Linha divisória
    ax_info.plot([0.05, 0.95], [_y + 0.01, _y + 0.01],
                 color="#B0BEC5", linewidth=0.8, transform=ax_info.transAxes)
    _y -= 0.04

    # Stats
    def _stat(y, icon, label, value, vcolor=_DARK):
        ax_info.text(0.08, y, icon, ha="left", va="top", fontsize=9,
                     transform=ax_info.transAxes)
        ax_info.text(0.24, y, label, ha="left", va="top", fontsize=7.5,
                     color=_MED, transform=ax_info.transAxes)
        ax_info.text(0.92, y, value, ha="right", va="top", fontsize=8.5,
                     fontweight="bold", color=vcolor, transform=ax_info.transAxes)
        return y - 0.085

    _dist_fmt = f"{_rd:,.0f} km".replace(",",".")
    _h, _min  = int(_rm//60), int(_rm%60)
    _time_fmt = f"{_h}h {_min:02d}min"
    _cons_fmt = (f"{_rd/_aut:.0f} L ({_comb})" if _aut else f"— ({_comb})")
    _para_fmt = str(len(sugest)) if sugest else "Nenhuma"

    _y = _stat(_y, "📏", "Distância",    _dist_fmt)
    _y = _stat(_y, "⏱️",  "Tempo est.",   _time_fmt)
    _y = _stat(_y, "⛽",  "Consumo",     _cons_fmt)
    _y = _stat(_y, "🚦",  "Paradas abast.", _para_fmt,
               vcolor=(_LAR if sugest else _VERDE))

    if _cap and _aut:
        _y = _stat(_y, "🛢️", "Tanque/Aut.",
                   f"{_cap:.0f} L / {_aut:.0f} km")

    n_par = len(rot_res.get("paradas", []))
    if n_par:
        _y = _stat(_y, "📍", "Paradas interm.", str(n_par))

    # ── Lista de postos sugeridos ─────────────────────────────────
    ax_stops = fig.add_subplot(gs_outer[2] if len(sugest) == 0 else gs_outer[1])

    # Usar um ax extra no rodapé para a lista de postos
    ax_body2 = fig.add_axes([0.0, 0.01, 1.0, 0.17])
    ax_body2.set_facecolor(_LAR2)
    ax_body2.set_xlim(0, 1); ax_body2.set_ylim(0, 1)
    ax_body2.axis("off")

    _stops_txt = "⛽  Postos de abastecimento sugeridos:  "
    _max_show  = min(len(sugest), 4)
    for _i, _s in enumerate(sugest[:_max_show]):
        _nm = str(_s.get("razaoSocial") or _s.get("nome") or "Posto")[:28]
        _uf_s = _s.get("uf", "")
        _mun  = _s.get("municipio", "")
        _loc  = f"{_mun}/{_uf_s}" if _mun else _uf_s
        _pr   = _s.get("preco", None)
        _pr_s = f"  R$ {_pr:.3f}/L" if _pr else ""
        _stops_txt += f"  [{_i+1}] {_nm} ({_loc}){_pr_s}    "
    if len(sugest) > _max_show:
        _stops_txt += f"  +{len(sugest)-_max_show} mais…"
    if not sugest:
        _stops_txt = "✅  Nenhuma parada de abastecimento necessária na rota."

    ax_body2.text(0.01, 0.55, _stops_txt, ha="left", va="center",
                  fontsize=7.5, color=_LAR if sugest else _VERDE,
                  fontweight="bold", wrap=True,
                  transform=ax_body2.transAxes)

    # ── FOOTER ─────────────────────────────────────────────────────
    ax_ftr = fig.add_axes([0.0, 0.0, 1.0, 0.055])
    ax_ftr.set_facecolor(_AZ)
    ax_ftr.set_xlim(0, 1); ax_ftr.set_ylim(0, 1)
    ax_ftr.axis("off")
    _now_str = datetime.now().strftime("%d/%m/%Y  %H:%M")
    ax_ftr.text(0.5, 0.5, f"Gerado em {_now_str}  ·  Estudo de Rede — Pró-Frotas",
                ha="center", va="center", fontsize=7.5, color="#BBDEFB",
                transform=ax_ftr.transAxes)

    # Salva
    _buf_card = io.BytesIO()
    fig.savefig(_buf_card, format="png", bbox_inches="tight",
                dpi=DPI, facecolor=_AZ)
    plt.close(fig)
    _buf_card.seek(0)
    return _buf_card.read()


# ═══════════════════════════════════════════════════════════════════
#  EXPORTAÇÃO DO MAPA DE POSTOS COMO PNG
# ═══════════════════════════════════════════════════════════════════

def _exportar_mapa_postos_png(df: "pd.DataFrame", titulo: str = "",
                               subtitulo: str = "") -> bytes:
    """
    Gera imagem PNG estática dos postos presentes no df.
    Usa matplotlib com scatter plot georreferenciado.
    Retorna bytes PNG.
    """
    if df is None or df.empty or "_lat" not in df.columns:
        # Mapa vazio
        fig, ax = plt.subplots(figsize=(8, 5), dpi=120)
        ax.set_facecolor("#EEF2F7")
        ax.text(0.5, 0.5, "Sem postos para exibir", ha="center", va="center",
                fontsize=12, color="#90A4AE", transform=ax.transAxes)
        ax.axis("off")
        _b = io.BytesIO(); fig.savefig(_b, format="png", dpi=120)
        plt.close(fig); _b.seek(0); return _b.read()

    _df = df.dropna(subset=["_lat","_lon"]).copy()
    if _df.empty:
        return _exportar_mapa_postos_png(None)

    lats = _df["_lat"].astype(float)
    lons = _df["_lon"].astype(float)

    fig, ax = plt.subplots(figsize=(11, 6.5), dpi=130)
    ax.set_facecolor("#EEF2F7")
    fig.patch.set_facecolor("#FFFFFF")

    # ── Camadas de pontos ──────────────────────────────────────────
    # Postos comuns (ANP)
    _mask_reg = ~(
        (_df.get("_pro_frotas", pd.Series(False, index=_df.index)).fillna(False)) |
        (_df.get("_cercado",    pd.Series(False, index=_df.index)).fillna(False)) |
        (_df.get("_rodo_rede",  pd.Series(False, index=_df.index)).fillna(False))
    )
    _col_reg = "#42A5F5"
    if "_pro_frotas" in _df.columns:
        _mask_reg = (~_df["_pro_frotas"].fillna(False) &
                     ~_df.get("_cercado",   pd.Series(False, index=_df.index)).fillna(False) &
                     ~_df.get("_rodo_rede", pd.Series(False, index=_df.index)).fillna(False))

    if _mask_reg.any():
        ax.scatter(lons[_mask_reg], lats[_mask_reg],
                   s=22, c=_col_reg, alpha=0.7, zorder=3,
                   linewidths=0.3, edgecolors="#1565C0", label="Postos ANP")

    # Rodo Rede
    if "_rodo_rede" in _df.columns:
        _mask_rr = _df["_rodo_rede"].fillna(False) & ~_df.get("_cercado", pd.Series(False, index=_df.index)).fillna(False)
        if _mask_rr.any():
            ax.scatter(lons[_mask_rr], lats[_mask_rr],
                       s=35, c="#9C27B0", alpha=0.85, zorder=4,
                       linewidths=0.4, edgecolors="#4A148C", label="Rodo Rede")

    # Cercados
    if "_cercado" in _df.columns:
        _mask_cer = _df["_cercado"].fillna(False)
        if _mask_cer.any():
            ax.scatter(lons[_mask_cer], lats[_mask_cer],
                       s=45, c="#FF8F00", alpha=0.9, zorder=5,
                       linewidths=0.5, edgecolors="#E65100", label="Cercados ⚠️",
                       marker="D")

    # Gestão de Frotas
    if "_pro_frotas" in _df.columns:
        _mask_pf = _df["_pro_frotas"].fillna(False)
        if _mask_pf.any():
            ax.scatter(lons[_mask_pf], lats[_mask_pf],
                       s=55, c="#FFD700", alpha=0.95, zorder=6,
                       linewidths=0.6, edgecolors="#F57F17", label="Gestão de Frotas ⭐",
                       marker="*")

    # Margens
    _pad_lat = max((lats.max() - lats.min()) * 0.08, 0.4)
    _pad_lon = max((lons.max() - lons.min()) * 0.08, 0.4)
    ax.set_xlim(lons.min() - _pad_lon, lons.max() + _pad_lon)
    ax.set_ylim(lats.min() - _pad_lat, lats.max() + _pad_lat)

    # Eixos simplificados
    ax.set_xticks([]); ax.set_yticks([])
    for sp in ax.spines.values(): sp.set_visible(False)
    ax.grid(True, linestyle=":", linewidth=0.3, color="#B0BEC5", alpha=0.5)

    # Título
    _titulo_final = titulo or "Postos de Combustível"
    ax.set_title(_titulo_final, fontsize=12, fontweight="bold",
                 color="#0D47A1", pad=8)
    if subtitulo:
        ax.text(0.5, 1.01, subtitulo, ha="center", va="bottom",
                fontsize=8, color="#546E7A", transform=ax.transAxes)

    # Legenda
    ax.legend(fontsize=7.5, loc="lower right", framealpha=0.92,
              edgecolor="#CFD8DC", facecolor="white")

    # Contador
    ax.text(0.01, 0.01, f"{len(_df)} postos  ·  {datetime.now().strftime('%d/%m/%Y %H:%M')}",
            ha="left", va="bottom", fontsize=7, color="#90A4AE",
            transform=ax.transAxes)

    plt.tight_layout(pad=0.5)
    _buf_map = io.BytesIO()
    fig.savefig(_buf_map, format="png", bbox_inches="tight", dpi=130)
    plt.close(fig)
    _buf_map.seek(0)
    return _buf_map.read()


def gerar_pdf_roteirizacao(rot_res: dict, sugest: list, logo_b64: str = None) -> bytes:
    """
    Gera PDF completo do relatório de roteirização.
    rot_res: dicionário com todos os campos do resultado.
    sugest:  lista de postos sugeridos.
    Retorna bytes do PDF.
    """
    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.lib import colors
        from reportlab.lib.units import cm
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
        from reportlab.platypus import (
            SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle,
            Image as RLImage, HRFlowable, KeepTogether
        )
    except ImportError:
        return b""

    _buf = io.BytesIO()
    doc = SimpleDocTemplate(
        _buf, pagesize=A4,
        leftMargin=1.8*cm, rightMargin=1.8*cm,
        topMargin=1.5*cm, bottomMargin=1.5*cm,
    )

    W, H = A4
    _w = W - 3.6*cm   # largura útil

    # ── Estilos ───────────────────────────────────────────────────────
    styles = getSampleStyleSheet()
    _azul   = colors.HexColor("#0D47A1")
    _azul2  = colors.HexColor("#1565C0")
    _laranja= colors.HexColor("#E65100")
    _verde  = colors.HexColor("#2E7D32")
    _cinza  = colors.HexColor("#607D8B")
    _cinzacl= colors.HexColor("#ECEFF1")
    _branco = colors.white

    _sT = ParagraphStyle("sT", parent=styles["Normal"],
                         fontSize=18, textColor=_branco,
                         fontName="Helvetica-Bold", spaceAfter=0)
    _sS = ParagraphStyle("sS", parent=styles["Normal"],
                         fontSize=8, textColor=colors.HexColor("#BBDEFB"),
                         fontName="Helvetica", spaceAfter=0)
    _sH2 = ParagraphStyle("sH2", parent=styles["Normal"],
                          fontSize=10, textColor=_azul,
                          fontName="Helvetica-Bold", spaceBefore=10, spaceAfter=4)
    _sN  = ParagraphStyle("sN",  parent=styles["Normal"],
                          fontSize=8.5, textColor=colors.HexColor("#37474F"),
                          fontName="Helvetica", leading=13)
    _sB  = ParagraphStyle("sB",  parent=styles["Normal"],
                          fontSize=8.5, textColor=colors.HexColor("#212121"),
                          fontName="Helvetica-Bold", leading=13)
    _sC  = ParagraphStyle("sC",  parent=styles["Normal"],
                          fontSize=7, textColor=_cinza,
                          fontName="Helvetica", alignment=TA_CENTER)

    story = []

    # ═══════════════════════════════════════════
    # CABEÇALHO AZUL
    # ═══════════════════════════════════════════
    _hdr_cells = []
    if logo_b64:
        try:
            _logo_bytes = base64.b64decode(logo_b64)
            _logo_io    = io.BytesIO(_logo_bytes)
            _logo_img   = RLImage(_logo_io, width=3*cm, height=1.4*cm)
            _logo_img.hAlign = "LEFT"
            _hdr_cells.append([_logo_img])
        except Exception:
            _hdr_cells = None
    else:
        _hdr_cells = None

    _titulo_txt = [
        Paragraph("Relatório de Roteirização", _sT),
        Paragraph("Gestão de Frotas – Estudo de Rede", _sS),
    ]

    _data_txt = Paragraph(
        datetime.now().strftime("%d/%m/%Y  %H:%M"),
        ParagraphStyle("dt", parent=_sS, alignment=TA_RIGHT)
    )

    if _hdr_cells:
        _hdr_tbl = Table(
            [[_logo_img, _titulo_txt, _data_txt]],
            colWidths=[3.2*cm, _w - 5.5*cm, 2.3*cm],
        )
    else:
        _hdr_tbl = Table(
            [[_titulo_txt, _data_txt]],
            colWidths=[_w - 2.5*cm, 2.5*cm],
        )

    _hdr_tbl.setStyle(TableStyle([
        ("BACKGROUND",   (0,0), (-1,-1), _azul),
        ("VALIGN",       (0,0), (-1,-1), "MIDDLE"),
        ("LEFTPADDING",  (0,0), (-1,-1), 10),
        ("RIGHTPADDING", (0,0), (-1,-1), 10),
        ("TOPPADDING",   (0,0), (-1,-1), 10),
        ("BOTTOMPADDING",(0,0), (-1,-1), 10),
        ("ROUNDEDCORNERS", (0,0), (-1,-1), [6,6,6,6]),
    ]))
    story.append(_hdr_tbl)
    story.append(Spacer(1, 0.35*cm))

    # ═══════════════════════════════════════════
    # PARÂMETROS DO VEÍCULO / ROTA
    # ═══════════════════════════════════════════
    _placa  = rot_res.get("placa","—") or "—"
    _comb   = rot_res.get("combustivel","—") or "—"
    _cap    = float(rot_res.get("capacidade", 0) or 0)
    _aut    = float(rot_res.get("autonomia",  0) or 0)
    _rmin_v = _cap * 0.25
    _range_v= (_cap - _rmin_v) * _aut if _aut else 0
    _rd_v   = float(rot_res.get("dist_km", 0))
    _rm_v   = float(rot_res.get("dur_min", 0))
    _n_pts  = len(rot_res.get("paradas", [])) + 2

    _param_rows = [
        [Paragraph("<b>Parâmetro</b>", _sB), Paragraph("<b>Valor</b>", _sB),
         Paragraph("<b>Parâmetro</b>", _sB), Paragraph("<b>Valor</b>", _sB)],
        [Paragraph("🚛 Placa",    _sN), Paragraph(_placa,          _sB),
         Paragraph("⛽ Combustível",_sN),Paragraph(_comb,           _sB)],
        [Paragraph("🛢 Tanque",   _sN), Paragraph(f"{_cap:.0f} L",  _sB),
         Paragraph("📏 Autonomia",_sN), Paragraph(f"{_aut:.0f} km/tanque" if _aut else "—", _sB)],
        [Paragraph("⚠️ Nível mín.",_sN),Paragraph(f"{_rmin_v:.0f} L ({25:.0f}%)", _sB),
         Paragraph("🎯 Alcance ef.",_sN),Paragraph(f"{_range_v:.0f} km" if _range_v else "—", _sB)],
    ]
    _cw = _w / 4
    _tbl_param = Table(_param_rows, colWidths=[_cw]*4, hAlign="LEFT")
    _tbl_param.setStyle(TableStyle([
        ("BACKGROUND",   (0,0), (-1,0), _azul2),
        ("TEXTCOLOR",    (0,0), (-1,0), _branco),
        ("BACKGROUND",   (0,1), (-1,-1), _cinzacl),
        ("ROWBACKGROUNDS",(0,1),(-1,-1), [_branco, _cinzacl]),
        ("GRID",         (0,0), (-1,-1), 0.4, colors.HexColor("#B0BEC5")),
        ("FONTSIZE",     (0,0), (-1,-1), 8),
        ("TOPPADDING",   (0,0), (-1,-1), 5),
        ("BOTTOMPADDING",(0,0), (-1,-1), 5),
        ("LEFTPADDING",  (0,0), (-1,-1), 7),
        ("VALIGN",       (0,0), (-1,-1), "MIDDLE"),
    ]))
    story.append(Paragraph("Parâmetros da Roteirização", _sH2))
    story.append(_tbl_param)
    story.append(Spacer(1, 0.3*cm))

    # ═══════════════════════════════════════════
    # KPIs — INDICADORES
    # ═══════════════════════════════════════════
    _consumo_v = f"{_rd_v/_aut:.0f} L" if _aut else "—"
    _kpis = [
        ("📏 Distância",    f"{_rd_v:,.0f} km".replace(",",".")),
        ("⏱️ Tempo est.",   f"{int(_rm_v//60)}h {int(_rm_v%60):02d}min"),
        ("🛢 Consumo",      _consumo_v),
        ("📍 Pontos na rota", str(_n_pts)),
        ("⛽ Paradas abast.",str(len(sugest)) if sugest else "Nenhuma"),
    ]
    _kpi_w = _w / len(_kpis)
    _kpi_row_lbl = [Paragraph(f"<b>{k}</b>", ParagraphStyle(
        "kl", parent=_sC, fontSize=7.5, textColor=_branco)) for k,v in _kpis]
    _kpi_row_val = [Paragraph(f"<b>{v}</b>", ParagraphStyle(
        "kv", parent=_sC, fontSize=13, textColor=_branco, fontName="Helvetica-Bold")) for k,v in _kpis]

    _tbl_kpi = Table([_kpi_row_lbl, _kpi_row_val],
                     colWidths=[_kpi_w]*len(_kpis), hAlign="LEFT")
    _tbl_kpi.setStyle(TableStyle([
        ("BACKGROUND",   (0,0), (-1,-1), _azul2),
        ("GRID",         (0,0), (-1,-1), 0.5, colors.HexColor("#1976D2")),
        ("TOPPADDING",   (0,0), (-1,0),  5),
        ("BOTTOMPADDING",(0,0), (-1,0),  2),
        ("TOPPADDING",   (0,1), (-1,1),  2),
        ("BOTTOMPADDING",(0,1), (-1,1),  8),
        ("ALIGN",        (0,0), (-1,-1), "CENTER"),
        ("VALIGN",       (0,0), (-1,-1), "MIDDLE"),
        ("ROUNDEDCORNERS",(0,0),(-1,-1), [4,4,4,4]),
    ]))
    story.append(Paragraph("Indicadores da Rota", _sH2))
    story.append(_tbl_kpi)
    story.append(Spacer(1, 0.35*cm))

    # ═══════════════════════════════════════════
    # MAPA DA ROTA
    # ═══════════════════════════════════════════
    _rc = rot_res.get("coords", [])
    _ro = rot_res.get("orig",  {})
    _rt = rot_res.get("dest",  {})
    _rp = rot_res.get("paradas", [])

    if _rc and len(_rc) >= 2:
        story.append(Paragraph("Mapa da Rota", _sH2))
        try:
            _map_png = _gerar_mapa_rota_png(_rc, _ro, _rt, _rp, sugest)
            _map_io  = io.BytesIO(_map_png)
            _map_img = RLImage(_map_io, width=_w, height=_w * 0.52)
            _map_img.hAlign = "CENTER"
            story.append(_map_img)
        except Exception as _em:
            story.append(Paragraph(f"⚠️ Mapa indisponível: {_em}", _sN))
        story.append(Spacer(1, 0.35*cm))

    # ═══════════════════════════════════════════
    # POSTOS DE ABASTECIMENTO SUGERIDOS
    # ═══════════════════════════════════════════
    story.append(Paragraph("Postos de Abastecimento Sugeridos", _sH2))
    if not sugest:
        story.append(Paragraph(
            "✅ Nenhuma parada necessária — o alcance efetivo cobre toda a rota.",
            _sN))
    else:
        _ab_header = [
            Paragraph("<b>#</b>",           _sB),
            Paragraph("<b>Posto</b>",       _sB),
            Paragraph("<b>Município/UF</b>",_sB),
            Paragraph("<b>Km orig.</b>",    _sB),
            Paragraph("<b>Chega</b>",       _sB),
            Paragraph("<b>Preço/L</b>",     _sB),
            Paragraph("<b>Litros</b>",      _sB),
            Paragraph("<b>Custo</b>",       _sB),
            Paragraph("<b>Sai com</b>",     _sB),
        ]
        _ab_rows = [_ab_header]
        for _i, _s in enumerate(sugest, 1):
            _preco   = float(_s.get("preco", 0))
            _litros  = int(_s.get("litros_sugeridos", 0))
            _custo_s = float(_s.get("custo_abast", 0))
            _f_ch    = float(_s.get("fuel_chegada", 0))
            _p_ch    = float(_s.get("pct_chegada", 0))
            _f_ap    = float(_s.get("fuel_apos", 0))
            _p_ap    = float(_s.get("pct_apos", 0))
            _motivo  = _s.get("motivo", "mais_barato")
            _cor_m   = _verde if _motivo == "mais_barato" else colors.HexColor("#B71C1C")
            _ab_rows.append([
                Paragraph(str(_i), _sN),
                Paragraph(str(_s.get("label",""))[:35], _sN),
                Paragraph(f"{_s.get('municipio','')} / {_s.get('uf','')}", _sN),
                Paragraph(f"{_s.get('_km',0):.0f} km", _sN),
                Paragraph(f"{_f_ch:.0f} L\n({_p_ch:.0f}%)", _sN),
                Paragraph(f"R$ {_preco:.3f}".replace(".",","), _sN),
                Paragraph(str(_litros) + " L",
                          ParagraphStyle("lf", parent=_sN, textColor=_cor_m,
                                         fontName="Helvetica-Bold")),
                Paragraph(f"R$ {_custo_s:.2f}".replace(".",","),
                          ParagraphStyle("cf", parent=_sN, textColor=_cor_m,
                                         fontName="Helvetica-Bold")),
                Paragraph(f"{_f_ap:.0f} L\n({_p_ap:.0f}%)", _sN),
            ])
        _cws_ab = [0.5*cm, 4.0*cm, 2.6*cm, 1.5*cm, 1.5*cm, 1.6*cm, 1.4*cm, 1.7*cm, 1.6*cm]
        _tbl_ab = Table(_ab_rows, colWidths=_cws_ab, hAlign="LEFT", repeatRows=1)
        _tbl_ab.setStyle(TableStyle([
            ("BACKGROUND",    (0,0), (-1,0), _azul2),
            ("TEXTCOLOR",     (0,0), (-1,0), _branco),
            ("ROWBACKGROUNDS",(0,1),(-1,-1), [_branco, _cinzacl]),
            ("GRID",          (0,0), (-1,-1), 0.4, colors.HexColor("#B0BEC5")),
            ("FONTSIZE",      (0,0), (-1,-1), 8),
            ("TOPPADDING",    (0,0), (-1,-1), 4),
            ("BOTTOMPADDING", (0,0), (-1,-1), 4),
            ("LEFTPADDING",   (0,0), (-1,-1), 5),
            ("VALIGN",        (0,0), (-1,-1), "MIDDLE"),
        ]))
        story.append(_tbl_ab)

    story.append(Spacer(1, 0.35*cm))

    # ═══════════════════════════════════════════
    # RESUMO DA ROTA — PONTOS
    # ═══════════════════════════════════════════
    story.append(Paragraph("Resumo da Rota — Pontos de Passagem", _sH2))
    _pontos_pdf = [_ro] + _rp + [_rt]
    _tipos_pdf  = (["Origem"] + [f"Parada {i+1}" for i in range(len(_rp))] + ["Destino"])
    _cores_pdf  = ([_verde] + [_laranja]*len(_rp) + [colors.HexColor("#C62828")])

    _pt_header = [
        Paragraph("<b>#</b>",   _sB),
        Paragraph("<b>Tipo</b>",_sB),
        Paragraph("<b>Local</b>",_sB),
        Paragraph("<b>Lat</b>", _sB),
        Paragraph("<b>Lon</b>", _sB),
    ]
    _pt_rows = [_pt_header]
    for _i, (_pt, _tp, _cr_pt) in enumerate(zip(_pontos_pdf, _tipos_pdf, _cores_pdf), 1):
        _pt_rows.append([
            Paragraph(str(_i), _sN),
            Paragraph(_tp, ParagraphStyle("tp", parent=_sN, textColor=_cr_pt,
                                          fontName="Helvetica-Bold")),
            Paragraph(str(_pt.get("label",""))[:50], _sN),
            Paragraph(f"{float(_pt.get('lat',0)):.5f}", _sN),
            Paragraph(f"{float(_pt.get('lon',0)):.5f}", _sN),
        ])
    _cws_pt = [0.6*cm, 2.2*cm, 8.0*cm, 2.2*cm, 2.2*cm]
    _tbl_pt = Table(_pt_rows, colWidths=_cws_pt, hAlign="LEFT", repeatRows=1)
    _tbl_pt.setStyle(TableStyle([
        ("BACKGROUND",    (0,0), (-1,0), _azul2),
        ("TEXTCOLOR",     (0,0), (-1,0), _branco),
        ("ROWBACKGROUNDS",(0,1),(-1,-1), [_branco, _cinzacl]),
        ("GRID",          (0,0), (-1,-1), 0.4, colors.HexColor("#B0BEC5")),
        ("FONTSIZE",      (0,0), (-1,-1), 8),
        ("TOPPADDING",    (0,0), (-1,-1), 4),
        ("BOTTOMPADDING", (0,0), (-1,-1), 4),
        ("LEFTPADDING",   (0,0), (-1,-1), 5),
        ("VALIGN",        (0,0), (-1,-1), "MIDDLE"),
    ]))
    story.append(_tbl_pt)
    story.append(Spacer(1, 0.5*cm))

    # ── Rodapé ────────────────────────────────────────────────────────
    story.append(HRFlowable(width=_w, thickness=0.5, color=_cinza))
    story.append(Spacer(1, 0.15*cm))
    story.append(Paragraph(
        f"Gerado em {datetime.now().strftime('%d/%m/%Y às %H:%M')} · "
        "Gestão de Frotas – Estudo de Rede · Documento gerado automaticamente",
        _sC))

    doc.build(story)
    _buf.seek(0)
    return _buf.read()


# ═══════════════════════════════════════════════════════════════════
#  GPX — EXPORTAÇÃO DE ROTA PARA GPS / GOOGLE MAPS
# ═══════════════════════════════════════════════════════════════════

def gerar_gpx_roteirizacao(rot_res: dict, sugest: list) -> bytes:
    """
    Gera arquivo GPX 1.1 com waypoints e trilha da rota.
    Compatível com GPS Garmin, Google Maps, OsmAnd, Waze etc.
    """
    import xml.etree.ElementTree as ET

    _ts   = datetime.utcnow().strftime("%Y-%m-%dT%H:%M:%SZ")
    _orig = rot_res.get("orig", {})
    _dest = rot_res.get("dest", {})
    _rp   = rot_res.get("paradas", [])
    _rc   = rot_res.get("coords", [])
    _nome = (f"{_orig.get('label','Origem')[:25]} → "
             f"{_dest.get('label','Destino')[:25]}")

    # Namespace GPX 1.1
    ET.register_namespace("", "http://www.topografix.com/GPX/1/1")
    ET.register_namespace("xsi", "http://www.w3.org/2001/XMLSchema-instance")

    gpx = ET.Element("gpx", {
        "version": "1.1",
        "creator": "Estudo de Rede – Gestão de Frotas",
        "xmlns":   "http://www.topografix.com/GPX/1/1",
        "xmlns:xsi": "http://www.w3.org/2001/XMLSchema-instance",
        "xsi:schemaLocation": (
            "http://www.topografix.com/GPX/1/1 "
            "http://www.topografix.com/GPX/1/1/gpx.xsd"
        ),
    })

    # ── Metadata ────────────────────────────────────────────────────
    meta = ET.SubElement(gpx, "metadata")
    ET.SubElement(meta, "name").text   = _nome
    ET.SubElement(meta, "desc").text   = (
        f"Rota: {rot_res.get('dist_km',0):.0f} km · "
        f"{len(sugest)} parada(s) de abastecimento · "
        f"Placa: {rot_res.get('placa','—')}"
    )
    ET.SubElement(meta, "time").text   = _ts

    # ── Waypoints ────────────────────────────────────────────────────
    def _wpt(lat, lon, name, desc="", sym="Waypoint"):
        w = ET.SubElement(gpx, "wpt", {"lat": f"{lat:.6f}", "lon": f"{lon:.6f}"})
        ET.SubElement(w, "name").text = name[:50]
        if desc:
            ET.SubElement(w, "desc").text = desc[:120]
        ET.SubElement(w, "sym").text  = sym
        return w

    # Origem
    if _orig.get("lat") and _orig.get("lon"):
        _wpt(float(_orig["lat"]), float(_orig["lon"]),
             f"🟢 Origem: {_orig.get('label','Origem')[:40]}",
             desc="Ponto de partida da rota", sym="Flag, Green")

    # Paradas intermediárias
    for _wi, _p in enumerate(_rp, 1):
        if _p.get("lat") and _p.get("lon"):
            _wpt(float(_p["lat"]), float(_p["lon"]),
                 f"🟠 Parada {_wi}: {_p.get('label','')[:35]}",
                 desc=f"Parada intermediária {_wi}", sym="Flag, Blue")

    # Postos de abastecimento sugeridos
    for _si, _s in enumerate(sugest, 1):
        if _s.get("lat") and _s.get("lon"):
            _wpt(
                float(_s["lat"]), float(_s["lon"]),
                f"⛽ Posto {_si}: {_s.get('label','')[:35]}",
                desc=(
                    f"Abastecer {_s.get('litros_sugeridos',0)} L · "
                    f"R$ {_s.get('preco',0):.3f}/L · "
                    f"{_s.get('municipio','')} / {_s.get('uf','')}"
                ),
                sym="Gas Station",
            )

    # Destino
    if _dest.get("lat") and _dest.get("lon"):
        _wpt(float(_dest["lat"]), float(_dest["lon"]),
             f"🔴 Destino: {_dest.get('label','Destino')[:40]}",
             desc="Ponto de chegada da rota", sym="Flag, Red")

    # ── Track (trilha da rota) ────────────────────────────────────────
    if _rc and len(_rc) >= 2:
        trk = ET.SubElement(gpx, "trk")
        ET.SubElement(trk, "name").text = _nome
        ET.SubElement(trk, "desc").text = f"Rota calculada · {rot_res.get('dist_km',0):.0f} km"
        trkseg = ET.SubElement(trk, "trkseg")
        # Limita a 2000 pontos para manter o arquivo leve
        _step = max(1, len(_rc) // 2000)
        for _i, (_lat, _lon) in enumerate(_rc):
            if _i % _step == 0 or _i == len(_rc) - 1:
                ET.SubElement(trkseg, "trkpt",
                              {"lat": f"{_lat:.6f}", "lon": f"{_lon:.6f}"})

    # Serializa para bytes UTF-8
    # encoding="unicode" escreve str → usa StringIO; depois codifica para bytes
    tree = ET.ElementTree(gpx)
    _sbuf = io.StringIO()
    tree.write(_sbuf, encoding="unicode", xml_declaration=False)
    _xml_str = '<?xml version="1.0" encoding="UTF-8"?>\n' + _sbuf.getvalue()
    return _xml_str.encode("utf-8")


# ═══════════════════════════════════════════════════════════════════
#  ROTEAMENTO — OSRM + fallback linha reta
# ═══════════════════════════════════════════════════════════════════

_OSRM_SERVIDORES = [
    "http://router.project-osrm.org/route/v1/driving",
    "https://routing.openstreetmap.de/routed-car/route/v1/driving",
]


def _tentar_osrm(srv, pontos: list):
    """pontos = [[lat, lon], ...] — suporta origem, N paradas e destino."""
    coords_str = ";".join(f"{lon},{lat}" for lat, lon in pontos)
    r = requests.get(f"{srv}/{coords_str}",
                     params={"overview": "full", "geometries": "geojson"}, timeout=6)
    d = r.json()
    if d.get("code") == "Ok":
        geo = d["routes"][0]["geometry"]["coordinates"]
        return (
            [[c[1], c[0]] for c in geo],
            d["routes"][0]["distance"] / 1000,
            d["routes"][0]["duration"] / 60,
        )
    return None


def calcular_rota(lat1, lon1, lat2, lon2, waypoints=None):
    """
    Calcula rota entre origem e destino com paradas intermediárias opcionais.

    waypoints: list of [lat, lon] — paradas na ordem desejada (até 10).
    Retorna (coords_rota, dist_km, dur_min, linha_reta).
    Tenta os servidores OSRM em paralelo para reduzir latência.
    """
    pontos = [[lat1, lon1]] + (waypoints or []) + [[lat2, lon2]]
    # Tenta servidores em paralelo — usa o primeiro que responder com sucesso
    with ThreadPoolExecutor(max_workers=len(_OSRM_SERVIDORES)) as _ex:
        _futs = {_ex.submit(_tentar_osrm, srv, pontos): srv
                 for srv in _OSRM_SERVIDORES}
        for _fut in as_completed(_futs):
            try:
                res = _fut.result()
                if res:
                    return res[0], res[1], res[2], False
            except Exception:
                continue
    # Fallback: segmentos de linha reta entre todos os pontos
    coords, n_seg = [], max(20, 15 * len(pontos))
    segs_por_trecho = max(10, n_seg // (len(pontos) - 1))
    for i in range(len(pontos) - 1):
        la1, lo1 = pontos[i]; la2, lo2 = pontos[i + 1]
        for j in range(segs_por_trecho):
            t = j / segs_por_trecho
            coords.append([la1 + (la2 - la1) * t, lo1 + (lo2 - lo1) * t])
    coords.append(pontos[-1])
    d = sum(
        _haversine(pontos[i][0], pontos[i][1], pontos[i+1][0], pontos[i+1][1])
        for i in range(len(pontos) - 1)
    ) / 1000
    return coords, d, (d / 80) * 60, True


# ═══════════════════════════════════════════════════════════════════
#  API ANP
# ═══════════════════════════════════════════════════════════════════

def _get(url, params, tentativas=4):
    """Requisição GET com retry e backoff exponencial.

    Trata 429 (rate-limit) e 5xx com espera maior.
    Levanta a exceção original após esgotar as tentativas.
    """
    for i in range(tentativas):
        try:
            r = requests.get(url, params=params, headers=HEADERS_ANP, timeout=45)
            r.raise_for_status()
            return r
        except requests.exceptions.HTTPError as e:
            status = e.response.status_code if e.response is not None else 0
            if i == tentativas - 1:
                raise
            # 429 = rate limit; 5xx = servidor sobrecarregado → espera maior
            espera = 10 if status in (429, 503, 504) else (2 ** i)
            time.sleep(espera)
        except Exception:
            if i == tentativas - 1:
                raise
            time.sleep(2 ** i)


@st.cache_data(show_spinner=False, ttl=86400)   # 24 horas
def buscar_postos(uf=None):
    params = {"numeropagina": 1}
    if uf: params["uf"] = uf
    resp = _get(f"{API_BASE_URL}{ENDPOINT}", params)
    data = resp.json()
    registros = data["data"] if isinstance(data, dict) and "data" in data else data
    if not registros: return pd.DataFrame()
    df = pd.DataFrame(registros)
    df["_lat"] = pd.to_numeric(df.get("latitude"),  errors="coerce")
    df["_lon"] = pd.to_numeric(df.get("longitude"), errors="coerce")
    df = df.dropna(subset=["_lat","_lon"])
    df = df[df["_lat"].between(-33.8,5.3) & df["_lon"].between(-73.9,-34.7)]
    return df.reset_index(drop=True)


@st.cache_data(show_spinner=False, ttl=86400)   # 24 horas — resultado por CNPJ
def buscar_posto_por_cnpj(cnpj_norm: str) -> pd.DataFrame:
    """
    Consulta a API ANP pelo CNPJ do posto individual (sem filtro de UF).
    Tenta dois parâmetros: 'cnpj' (estabelecimento) e 'cnpjRevenda' (bandeira).
    Aplica pós-filtro para garantir que só o CNPJ exato seja retornado.
    Retorna DataFrame com o posto (ou vazio se não encontrado / erro).
    """
    if len(cnpj_norm) != 14:
        return pd.DataFrame()
    cf = cnpj_norm
    cnpj_fmt = f"{cf[:2]}.{cf[2:5]}.{cf[5:8]}/{cf[8:12]}-{cf[12:]}"

    def _filtrar_cnpj(df: pd.DataFrame) -> pd.DataFrame:
        """Mantém apenas linhas com o CNPJ exato pesquisado."""
        if df.empty or "cnpj" not in df.columns:
            return df
        mask = df["cnpj"].fillna("").str.replace(r"\D", "", regex=True) == cnpj_norm
        return df[mask].reset_index(drop=True)

    def _normalizar(lst) -> pd.DataFrame:
        if not lst:
            return pd.DataFrame()
        df = pd.DataFrame(lst if isinstance(lst, list) else [lst])
        df["_lat"] = pd.to_numeric(df.get("latitude"),  errors="coerce")
        df["_lon"] = pd.to_numeric(df.get("longitude"), errors="coerce")
        df = df.dropna(subset=["_lat", "_lon"])
        df = df[df["_lat"].between(-33.8, 5.3) & df["_lon"].between(-73.9, -34.7)]
        return df.reset_index(drop=True)

    # Tentativa 1: parâmetro 'cnpj' (CNPJ do estabelecimento/posto)
    try:
        resp = _get(f"{API_BASE_URL}{ENDPOINT}", {"numeropagina": 1, "cnpj": cnpj_fmt})
        data = resp.json()
        registros = data.get("data", data) if isinstance(data, dict) else data
        df = _filtrar_cnpj(_normalizar(registros))
        if not df.empty:
            return df
    except Exception:
        pass

    # Tentativa 2: parâmetro 'cnpjRevenda' + pós-filtro obrigatório
    try:
        resp = _get(f"{API_BASE_URL}{ENDPOINT}",
                    {"numeropagina": 1, "cnpjRevenda": cnpj_fmt})
        data = resp.json()
        registros = data.get("data", data) if isinstance(data, dict) else data
        df = _filtrar_cnpj(_normalizar(registros))
        if not df.empty:
            return df
    except Exception:
        pass

    return pd.DataFrame()


@st.cache_data(show_spinner=False, ttl=3600)
def buscar_postos_por_nome(termo: str, uf: str = "") -> pd.DataFrame:
    """Busca postos na API ANP por razão social (parcial) + UF opcional.
    Aplica pós-filtro de UF pois a API nem sempre respeita o parâmetro."""
    params: dict = {"numeropagina": 1, "razaoSocial": termo}
    if uf:
        params["uf"] = uf
    try:
        resp = _get(f"{API_BASE_URL}{ENDPOINT}", params)
        data = resp.json()
        registros = data.get("data", data) if isinstance(data, dict) else data
        if not registros:
            return pd.DataFrame()
        lst = registros if isinstance(registros, list) else [registros]
        df = pd.DataFrame(lst)
        df["_lat"] = pd.to_numeric(df.get("latitude"),  errors="coerce")
        df["_lon"] = pd.to_numeric(df.get("longitude"), errors="coerce")
        df = df.dropna(subset=["_lat", "_lon"])
        df = df[df["_lat"].between(-33.8, 5.3) & df["_lon"].between(-73.9, -34.7)]
        # ── Pós-filtro de UF (a API pode ignorar o parâmetro) ────────
        if uf and not df.empty and "uf" in df.columns:
            df = df[
                df["uf"].fillna("").str.upper().str.strip() == uf.upper()
            ].reset_index(drop=True)
        return df.reset_index(drop=True)
    except Exception:
        return pd.DataFrame()


def _buscar_posto_completo(termo: str, uf: str = "") -> tuple[pd.DataFrame, str]:
    """
    Ponto único de busca para o Modo 3 — Consulta por Posto.

    Estratégia:
      1. Detecta CNPJ (14 dígitos) → busca direta na API ANP por CNPJ.
      2. Texto curto / nome → busca por razão social na API ANP (com UF opcional).
      3. Enriquece com flag _pro_frotas / _cercado / _rodo_rede das listas locais.

    Retorna (DataFrame, mensagem_fonte).
    """
    termo = termo.strip()
    if not termo:
        return pd.DataFrame(), ""

    cnpjs_pf      = st.session_state.get("cnpjs_pro_frotas", set())
    cnpjs_cercados = st.session_state.get("cnpjs_cercados",  set())
    perfil_map    = st.session_state.get("perfil_venda_map", {})

    # ── Detecta se parece CNPJ ──────────────────────────────────────
    _digits = re.sub(r"\D", "", termo)
    is_cnpj = len(_digits) == 14
    cnpj_fmt_label = (f"CNPJ {_digits[:2]}.{_digits[2:5]}.{_digits[5:8]}"
                      f"/{_digits[8:12]}-{_digits[12:]}" if is_cnpj else "")

    pf_df = st.session_state.get("pf_coords_df", pd.DataFrame())

    if is_cnpj:
        # ── Busca por CNPJ ──────────────────────────────────────────────
        fonte = cnpj_fmt_label
        df = pd.DataFrame()
        # Prioridade 1: planilha local Gestão de Frotas (mais precisa, sem API)
        if not pf_df.empty and "cnpj" in pf_df.columns:
            _mask_cnpj = pf_df["cnpj"].fillna("").str.replace(r"\D", "", regex=True) == _digits
            df = pf_df[_mask_cnpj].copy()
            if not df.empty:
                fonte += " (planilha Gestão de Frotas)"
        # Prioridade 2: API ANP (com pós-filtro de CNPJ exato)
        if df.empty:
            df = buscar_posto_por_cnpj(_digits)
            if not df.empty:
                fonte += " (API ANP)"
    else:
        # ── Busca por nome / razão social ────────────────────────────────
        # Prioridade 1: planilha local Gestão de Frotas (sem chamada à API ANP)
        fonte = f'Razão social "{termo}"' + (f" · UF {uf}" if uf else "")
        df = pd.DataFrame()
        if not pf_df.empty:
            _t_norm = _anp_norm(termo)
            _mask = pd.Series(False, index=pf_df.index)
            for _col in ["razaoSocial", "nome", "nomeFantasia"]:
                if _col in pf_df.columns:
                    _mask |= (
                        pf_df[_col].fillna("")
                        .apply(_anp_norm)
                        .str.contains(_t_norm, regex=False, na=False)
                    )
            if uf:
                _mask &= (
                    pf_df["uf"].fillna("").str.upper().str.strip() == uf.upper()
                )
            df = pf_df[_mask].copy()
            if not df.empty:
                fonte += " (planilha Gestão de Frotas)"

        # Prioridade 2: API ANP como fallback — filtra somente postos GF
        if df.empty:
            _df_api = buscar_postos_por_nome(termo, uf=uf)
            if not _df_api.empty:
                # ── Pós-filtro UF (defesa em profundidade) ────────────
                if uf and "uf" in _df_api.columns:
                    _df_api = _df_api[
                        _df_api["uf"].fillna("").str.upper().str.strip() == uf.upper()
                    ].reset_index(drop=True)
                if cnpjs_pf and not _df_api.empty and "cnpj" in _df_api.columns:
                    _cnpjs_api = _df_api["cnpj"].fillna("").str.replace(r"\D", "", regex=True)
                    _df_api = _df_api[_cnpjs_api.isin(cnpjs_pf)].reset_index(drop=True)
                if not _df_api.empty:
                    df = _df_api
                    fonte += " (API ANP)"

    if df.empty:
        return pd.DataFrame(), fonte

    # ── Pós-filtro de UF (defesa final — garante mesmo se API ignorou) ──
    if uf and not is_cnpj and "uf" in df.columns:
        _mask_uf_final = df["uf"].fillna("").str.upper().str.strip() == uf.upper()
        # Só aplica se ao menos alguns registros têm UF preenchida;
        # senão mantém tudo (evita apagar resultados de planilhas sem UF).
        if _mask_uf_final.any():
            df = df[_mask_uf_final].reset_index(drop=True)

    if df.empty:
        return pd.DataFrame(), fonte

    # ── Normaliza CNPJ e injeta flags ──────────────────────────────
    if "cnpj" in df.columns:
        df["_cnpj_norm"] = df["cnpj"].fillna("").str.replace(r"\D", "", regex=True)
    else:
        df["_cnpj_norm"] = ""

    df["_pro_frotas"] = df["_cnpj_norm"].isin(cnpjs_pf)
    df["_cercado"]    = df["_cnpj_norm"].isin(cnpjs_cercados)

    if perfil_map and "_cnpj_norm" in df.columns:
        df["_perfil_venda"] = df["_cnpj_norm"].map(perfil_map).fillna("")
        df["_rodo_rede"] = df["_perfil_venda"].str.upper().str.strip() == PERFIL_RODO_REDE
    else:
        df["_perfil_venda"] = ""
        df["_rodo_rede"] = False

    # Garante colunas de coordenadas
    if "_lat" not in df.columns and "latitude" in df.columns:
        df["_lat"] = pd.to_numeric(df["latitude"], errors="coerce")
    if "_lon" not in df.columns and "longitude" in df.columns:
        df["_lon"] = pd.to_numeric(df["longitude"], errors="coerce")

    df = df.dropna(subset=["_lat", "_lon"])
    return df.reset_index(drop=True), fonte


def _injetar_pf_ausentes(df_raw: pd.DataFrame, cnpjs_pf: set,
                         uf_atual: str = "", ufs_permitidas: set = None) -> pd.DataFrame:
    """
    ARQUITETURA EM CAMADAS:
    ─ Camada 1 (primária):  Postos Gestão de Frotas — planilha pro_frotas.xlsx do GitHub.
    ─ Camada 2 (complementar): API ANP — enriquece brand/dados dos postos GF em comum
                                e acrescenta postos não-GF da região selecionada.

    Esta função garante que postos GF da Camada 1 que não aparecem na API ANP
    (Camada 2) sejam injetados no dataset, usando as coordenadas da planilha.

    Resultado final (df_raw + injetados):
    • Postos GF ∩ ANP  → dados ANP (mais completos) + flag _pro_frotas=True
    • Postos GF - ANP  → dados da planilha + flag _pro_frotas=True  [esta função]
    • Postos ANP - GF  → dados ANP + flag _pro_frotas=False

    Filtro de UF:
      - uf_atual:      UF única (Modo 1 — Por Estado).
      - ufs_permitidas: conjunto de UFs (Modo 2 — Rota).
    """
    df_coords = st.session_state.get("pf_coords_df", pd.DataFrame())
    if df_coords.empty or not cnpjs_pf:
        return df_raw

    # CNPJs já presentes no dataset (vetorizado)
    if not df_raw.empty and "cnpj" in df_raw.columns:
        cnpjs_presentes = set(df_raw["cnpj"].fillna("").str.replace(r'\D', '', regex=True))
    else:
        cnpjs_presentes = set()

    ausentes = cnpjs_pf - cnpjs_presentes
    if not ausentes:
        return df_raw

    df_novos = df_coords[df_coords["cnpj"].isin(ausentes)].copy()
    if df_novos.empty:
        return df_raw

    # ── FILTRO POR UF ────────────────────────────────────────────────
    # Evita injetar postos de estados que não fazem parte da consulta.
    # SEGURANÇA: se uf_atual ou ufs_permitidas foram informados mas a planilha
    # não tem coluna UF preenchida, NÃO injeta (evita flood de todos os GF em 1 estado).
    _tem_uf_col = "uf" in df_novos.columns and df_novos["uf"].fillna("").str.strip().ne("").any()

    if uf_atual or ufs_permitidas:
        if not _tem_uf_col:
            # Planilha sem UF → não é possível filtrar → não injeta para evitar dados errados
            return df_raw
        if ufs_permitidas:
            _ufs_upper = {u.upper().strip() for u in ufs_permitidas}
            df_novos = df_novos[
                df_novos["uf"].fillna("").str.upper().str.strip().isin(_ufs_upper)
            ]
        else:
            df_novos = df_novos[
                df_novos["uf"].fillna("").str.upper().str.strip() == uf_atual.upper().strip()
            ]
        if df_novos.empty:
            return df_raw

    # Garante compatibilidade de colunas com df_raw
    for _col in (df_raw.columns if not df_raw.empty else []):
        if _col not in df_novos.columns:
            df_novos[_col] = pd.NA

    if df_raw.empty:
        return df_novos.reset_index(drop=True)

    return pd.concat([df_raw, df_novos], ignore_index=True)


# ═══════════════════════════════════════════════════════════════════
#  MAPA FOLIUM
# ═══════════════════════════════════════════════════════════════════

def _cor(distribuidora, mapa_cores):
    """Retorna cor do pin — delega a _cor_marca para garantir consistência."""
    return _cor_marca(distribuidora)


def _popup(row):
    pf_badge = ""
    if row.get("_pro_frotas"):
        pf_badge = (
            "<div style='background:#FFF9C4;border:2px solid #FFD700;border-radius:6px;"
            "padding:5px 10px;margin-bottom:8px;font-weight:700;color:#7B5E00;"
            "display:flex;align-items:center;gap:6px;font-size:12px'>"
            "⭐ CREDENCIADO GESTÃO DE FROTAS</div>"
        )
    produtos_html = ""
    try:
        prods = row.get("produtos", [])
        if isinstance(prods, list) and prods:
            linhas = "".join(
                f"<tr><td>{p.get('produto','')}</td>"
                f"<td style='text-align:center'>{p.get('tancagem','')}&nbsp;{p.get('unidMedidaTancagem','')}</td>"
                f"<td style='text-align:center'>{p.get('qtdeBicos','')}</td></tr>"
                for p in prods
            )
            produtos_html = (
                f"<details style='margin-top:6px'>"
                f"<summary style='cursor:pointer;font-weight:bold'>🛢️ Produtos ({len(prods)})</summary>"
                f"<table style='font-size:11px;width:100%;margin-top:4px'>"
                f"<tr style='background:#f5f5f5'><th>Produto</th><th>Tanc.</th><th>Bicos</th></tr>"
                f"{linhas}</table></details>"
            )
    except Exception: pass

    def v(c):
        s = str(row.get(c,"")).strip()
        return s if s and s not in ("nan","None") else "—"

    dist_txt = ""
    try:
        d = row.get("_dist_rota")
        if d is not None and pd.notna(d):
            dist_txt = f"<b>Dist. da rota:</b> {int(d)} m<br>"
    except Exception: pass

    lat = row.get("_lat", "")
    lon = row.get("_lon", "")
    sv_url   = f"https://www.google.com/maps?q=&layer=c&cbll={lat},{lon}"
    maps_url = f"https://www.google.com/maps?q={lat},{lon}"
    botoes_html = (
        f"<div style='display:flex;gap:6px;margin-top:10px'>"
        f"<a href='{sv_url}' target='_blank' style='"
        f"flex:1;background:#1a73e8;color:#fff;text-decoration:none;"
        f"border-radius:6px;padding:6px 0;text-align:center;"
        f"font-size:11px;font-weight:700;display:block'>"
        f"📷 Street View</a>"
        f"<a href='{maps_url}' target='_blank' style='"
        f"flex:1;background:#34a853;color:#fff;text-decoration:none;"
        f"border-radius:6px;padding:6px 0;text-align:center;"
        f"font-size:11px;font-weight:700;display:block'>"
        f"🗺️ Google Maps</a>"
        f"</div>"
    )

    # Marcador oculto para captura de clique no Streamlit
    nome_safe = v("razaoSocial").replace(";", ",")[:80]
    coord_tag = f"<!-- POSTO_SEL:{row.get('_lat', '')};{row.get('_lon', '')};{nome_safe} -->"

    # ── Preço do posto via _pp_df + tendência vs ANP semana anterior ─
    _preco_posto_html = ""
    try:
        _pp_df_popup = st.session_state.get("_pp_df")
        _cnpj_n_popup = str(row.get("_cnpj_norm", row.get("cnpj", ""))).replace(r"\D", "")
        if _pp_df_popup is not None and _cnpj_n_popup and "cnpj_norm" in _pp_df_popup.columns:
            _pp_row = _pp_df_popup[_pp_df_popup["cnpj_norm"] == _cnpj_n_popup]
            if not _pp_row.empty:
                _uf_popup = str(row.get("uf", "")).strip().upper()
                _cache_ant_popup  = st.session_state.get("_precos_anp_cache_anterior", {})
                _sheets_ant_popup = _cache_ant_popup.get("sheets", {})
                _semana_ant_popup = _cache_ant_popup.get("semana", "")
                _linhas_popup = []
                for _, _pp_r in _pp_row.iterrows():
                    _comb_label = str(_pp_r.get("combustivel_label", "")).strip()
                    _pk_popup   = _anp_norm(_pp_r.get("combustivel_pk", _comb_label))
                    _preco_p    = float(_pp_r["preco"])
                    # Preço ANP semana anterior para este UF/combustível
                    _anp_ant = None
                    if _sheets_ant_popup and _uf_popup:
                        _anp_ant = _anp_preco_uf(_sheets_ant_popup, _pk_popup, _uf_popup)
                        if _anp_ant is None:
                            _anp_ant = _anp_preco_brasil(_sheets_ant_popup, _pk_popup)
                    _tend_html = _tendencia_badge(_preco_p, _anp_ant, inline=True) if _anp_ant else ""
                    _data_p    = str(_pp_r.get("data_atualizacao", "")).strip()
                    _data_lbl  = f" · {_data_p}" if _data_p and _data_p not in ("nan","None","") else ""
                    _linhas_popup.append(
                        f"<tr>"
                        f"<td style='padding:2px 6px;font-size:11px;color:#555'>{_comb_label}</td>"
                        f"<td style='padding:2px 6px;font-size:12px;font-weight:700;color:#0d1b4b'>"
                        f"R$ {_brl(_preco_p, 3)}</td>"
                        f"<td style='padding:2px 6px'>{_tend_html}</td>"
                        f"</tr>"
                    )
                if _linhas_popup:
                    _ant_nota = (f"<div style='font-size:9px;color:#aaa;margin-top:2px'>"
                                 f"↑↓ vs ANP sem. {_semana_ant_popup}</div>"
                                 if _semana_ant_popup else "")
                    _preco_posto_html = (
                        f"<hr style='margin:5px 0'>"
                        f"<b style='font-size:11px'>💰 Preços cadastrados:</b>"
                        f"<table style='width:100%;border-collapse:collapse;margin-top:3px'>"
                        + "".join(_linhas_popup) +
                        f"</table>{_ant_nota}"
                    )
    except Exception:
        pass

    # ── Serviços disponíveis (colunas opcionais da planilha) ──────────
    _svc_badges = []
    if row.get("funciona_24h") is True:
        _svc_badges.append(
            "<span style='background:#1565c0;color:#fff;border-radius:4px;"
            "padding:1px 5px;font-size:10px;font-weight:700'>🕐 24h</span>"
        )
    if row.get("pista_caminhao") is True:
        _svc_badges.append(
            "<span style='background:#4e342e;color:#fff;border-radius:4px;"
            "padding:1px 5px;font-size:10px;font-weight:700'>🚛 Pista</span>"
        )
    if row.get("arla") is True:
        _svc_badges.append(
            "<span style='background:#1b5e20;color:#fff;border-radius:4px;"
            "padding:1px 5px;font-size:10px;font-weight:700'>🧪 ARLA</span>"
        )
    if row.get("conveniencia") is True:
        _svc_badges.append(
            "<span style='background:#4a148c;color:#fff;border-radius:4px;"
            "padding:1px 5px;font-size:10px;font-weight:700'>🛒 Conv.</span>"
        )
    _svc_html = ""
    if _svc_badges:
        _svc_html = (
            "<div style='display:flex;flex-wrap:wrap;gap:4px;margin:5px 0'>"
            + " ".join(_svc_badges)
            + "</div>"
        )
    _horario_val = str(row.get("horario", "") or "").strip()
    _horario_html = ""
    if _horario_val and _horario_val not in ("nan", "None", "—"):
        _horario_html = f"<b>🕐 Horário:</b> {_horario_val}<br>"

    return folium.Popup(
        f"<div style='font-family:sans-serif;font-size:12px;min-width:260px;max-width:320px'>"
        f"{pf_badge}"
        f"<b style='font-size:13px'>⛽ {v('razaoSocial')}</b><br>"
        f"<span style='color:#555;font-size:11px'>{v('distribuidora')}</span><br>"
        f"<hr style='margin:5px 0'>"
        f"<b>CNPJ:</b> {v('cnpj')}<br>"
        f"<b>Endereço:</b> {v('endereco')}{(', '+v('complemento')) if v('complemento')!='—' else ''}<br>"
        f"<b>Bairro:</b> {v('bairro')} — {v('municipio')}/{v('uf')}<br>"
        f"<b>CEP:</b> {v('cep')}<br>"
        f"<b>Autorização:</b> {v('autorizacao')}<br>"
        f"<b>Situação:</b> {v('situacaoConstatada')} | <b>SIGAF:</b> {v('statusSIGAF')}<br>"
        f"{_horario_html}"
        f"{dist_txt}"
        f"{_preco_posto_html}"
        f"{_svc_html}"
        f"{produtos_html}"
        f"{botoes_html}"
        f"{coord_tag}"
        f"</div>",
        max_width=360
    )


def _popup_simples(row):
    """Popup compacto (~250 bytes × marcador vs ~1.2 KB do popup completo).
    Usado automaticamente quando o mapa exibe muitos postos (> MAX_MAPA_POSTOS)
    para reduzir o HTML serializado e manter o mapa responsivo.
    """
    def v(c):
        s = str(row.get(c, "")).strip()
        return s if s and s not in ("nan", "None") else "—"

    lat = row.get("_lat", ""); lon = row.get("_lon", "")
    maps_url = f"https://www.google.com/maps?q={lat},{lon}"
    pf_txt = ("<div style='color:#7B5E00;font-weight:700;font-size:11px;"
              "margin-bottom:4px'>⭐ GESTÃO DE FROTAS</div>"
              if row.get("_pro_frotas") else "")

    return folium.Popup(
        f"<div style='font-family:sans-serif;font-size:12px;min-width:200px;max-width:260px'>"
        f"{pf_txt}"
        f"<b>⛽ {v('razaoSocial')}</b><br>"
        f"<span style='color:#777;font-size:11px'>{v('distribuidora')}</span>"
        f"<hr style='margin:5px 0'>"
        f"<b>CNPJ:</b> {v('cnpj')}<br>"
        f"<b>📍</b> {v('municipio')} / {v('uf')}<br>"
        f"<a href='{maps_url}' target='_blank' "
        f"style='font-size:11px;color:#1a73e8'>🗺️ Ver no Google Maps</a>"
        f"</div>",
        max_width=280,
    )


def _extrair_posto_do_popup(popup_html: str):
    """Extrai lat, lon e nome do marcador oculto no HTML do popup."""
    if not popup_html:
        return None
    try:
        m = re.search(r"<!-- POSTO_SEL:([-\d.]+);([-\d.]+);(.+?) -->", popup_html)
        if not m:
            return None
        return {"lat": float(m.group(1)), "lon": float(m.group(2)),
                "label": m.group(3).strip()}
    except Exception:
        return None


def _marcador_cercado(lat, lon, popup, tooltip):
    """Marcador laranja com ⚠️ para postos cercados — destaca a situação competitiva."""
    icon = folium.DivIcon(
        html=(
            f"<div style='"
            f"background:{COR_CERCADO_FILL};"
            f"border:2.5px solid {COR_CERCADO_BORDA};"
            f"border-radius:50%;"
            f"width:26px;height:26px;"
            f"display:flex;align-items:center;justify-content:center;"
            f"font-size:14px;line-height:1;"
            f"box-shadow:0 2px 6px rgba(0,0,0,.5);"
            f"'>⚠️</div>"
        ),
        icon_size=(26, 26),
        icon_anchor=(13, 13),
    )
    return folium.Marker(
        location=[lat, lon],
        icon=icon,
        popup=popup,
        tooltip=tooltip,
    )


@st.cache_data(show_spinner=False)
def _img_pro_frotas_b64() -> str:
    """
    Carrega logo_profrotas.jpg do repositório e retorna data-URL base64.
    Tenta variações de nome e extensão. Retorna "" se não encontrado.
    """
    for nome in ["logo_profrotas.jpg", "logo_profrotas.jpeg", "logo_profrotas.png",
                 "Logo_profrotas.jpg", "Logo_Profrotas.jpg", "logo_ProFrotas.jpg",
                 "profrotas.jpg", "profrotas.png"]:
        caminho = os.path.join(_DIR, nome)
        if os.path.exists(caminho):
            ext = nome.rsplit(".", 1)[-1].lower()
            mime = "image/jpeg" if ext in ("jpg", "jpeg") else "image/png"
            with open(caminho, "rb") as f:
                dados = base64.b64encode(f.read()).decode()
            return f"data:{mime};base64,{dados}"
    return ""


def _marcador_pf(lat, lon, popup, tooltip):
    """Marcador Gestão de Frotas: usa logo_profrotas.jpg se disponível, senão círculo azul."""
    img_b64 = _img_pro_frotas_b64()

    if img_b64:
        html_icon = (
            f"<div style='"
            f"width:36px;height:36px;"
            f"border-radius:50%;"
            f"border:3px solid {COR_PF_BORDA};"
            f"box-shadow:0 2px 8px rgba(0,0,0,.5);"
            f"overflow:hidden;"
            f"background:#fff;'>"
            f"<img src='{img_b64}' "
            f"style='width:100%;height:100%;object-fit:cover;display:block;'/>"
            f"</div>"
        )
        icon = folium.DivIcon(html=html_icon, icon_size=(36, 36), icon_anchor=(18, 18))
        return folium.Marker(
            location=[lat, lon],
            icon=icon,
            popup=popup,
            tooltip=tooltip,
        )
    else:
        # Fallback: círculo azul original
        return folium.CircleMarker(
            location=[lat, lon],
            radius=14,
            color=COR_PF_BORDA,
            weight=2.5,
            fill=True,
            fill_color=COR_PF_FILL,
            fill_opacity=0.92,
            popup=popup,
            tooltip=tooltip,
        )


@st.cache_data(show_spinner=False)
def _img_rodo_rede_b64() -> str:
    """
    Carrega RodoRede.jpg do repositório e retorna data-URL base64.
    Tenta extensões .jpg, .jpeg e .png. Retorna "" se não encontrado.
    """
    for nome in ["RodoRede.jpg", "RodoRede.jpeg", "RodoRede.png",
                 "rodorede.jpg", "rodorede.jpeg", "rodorede.png"]:
        caminho = os.path.join(_DIR, nome)
        if os.path.exists(caminho):
            ext = nome.rsplit(".", 1)[-1].lower()
            mime = "image/jpeg" if ext in ("jpg", "jpeg") else "image/png"
            with open(caminho, "rb") as f:
                dados = base64.b64encode(f.read()).decode()
            return f"data:{mime};base64,{dados}"
    return ""


@st.cache_data(show_spinner=False)
def _logos_bandeiras_b64() -> dict:
    """
    Carrega logos de bandeiras do repositório e retorna dict
    {SUBSTRING_UPPER: data_url_base64}.
    Adicione mais entradas para suportar outras bandeiras no futuro.
    Ex: {"IPIRANGA": "data:image/jpeg;base64,..."}
    """
    # Mapeamento: substring da distribuidora → lista de nomes de arquivo candidatos
    MAPA = {
        "IPIRANGA": ["Ipiranga.jpg", "ipiranga.jpg", "Ipiranga.jpeg",
                     "ipiranga.jpeg", "Ipiranga.png", "ipiranga.png"],
    }
    resultado = {}
    for marca, candidatos in MAPA.items():
        for nome in candidatos:
            caminho = os.path.join(_DIR, nome)
            if os.path.exists(caminho):
                ext  = nome.rsplit(".", 1)[-1].lower()
                mime = "image/jpeg" if ext in ("jpg", "jpeg") else "image/png"
                with open(caminho, "rb") as f:
                    dados = base64.b64encode(f.read()).decode()
                resultado[marca] = f"data:{mime};base64,{dados}"
                break
    return resultado


def _logo_para_distribuidora(distribuidora: str, logos: dict) -> str:
    """Retorna a data-URL da logo se a distribuidora tiver logo cadastrada, senão ''."""
    d = str(distribuidora).upper().strip()
    for marca, url in logos.items():
        if marca in d:
            return url
    return ""


def _marcador_logo_bandeira(lat, lon, popup, tooltip, img_b64: str, cor_borda: str):
    """Marcador circular com a logo da bandeira para postos regulares."""
    html_icon = (
        f"<div style='"
        f"width:28px;height:28px;"
        f"border-radius:50%;"
        f"border:2px solid {cor_borda};"
        f"box-shadow:0 2px 6px rgba(0,0,0,.5);"
        f"overflow:hidden;"
        f"background:#fff;'>"
        f"<img src='{img_b64}' "
        f"style='width:100%;height:100%;object-fit:cover;display:block;'/>"
        f"</div>"
    )
    icon = folium.DivIcon(html=html_icon, icon_size=(28, 28), icon_anchor=(14, 14))
    return folium.Marker(location=[lat, lon], icon=icon, popup=popup, tooltip=tooltip)


def _marcador_pf_bandeira(lat, lon, popup, tooltip, img_b64: str):
    """
    Marcador para posto Gestão de Frotas que tem logo de bandeira reconhecida (ex: Ipiranga).
    Usa a logo da bandeira com borda azul GF + anel dourado externo para diferenciar
    de postos regulares da mesma bandeira.
    Tamanho maior (36px) que o pin regular (28px) para destacar o credenciamento GF.
    """
    html_icon = (
        f"<div style='"
        f"width:36px;height:36px;"
        f"border-radius:50%;"
        # Borda interna azul GF + sombra dourada = indicação visual de credenciado
        f"border:3px solid {COR_PF_BORDA};"
        f"box-shadow:0 0 0 2px #FFD700, 0 3px 8px rgba(0,0,0,.55);"
        f"overflow:hidden;"
        f"background:#fff;"
        f"'>"
        f"<img src='{img_b64}' "
        f"style='width:100%;height:100%;object-fit:cover;display:block;'/>"
        f"</div>"
    )
    icon = folium.DivIcon(html=html_icon, icon_size=(36, 36), icon_anchor=(18, 18))
    return folium.Marker(location=[lat, lon], icon=icon, popup=popup, tooltip=tooltip)


def _marcador_rodo_rede(lat, lon, popup, tooltip):
    """Marcador com logo Rodo Rede para postos Gestão de Frotas com Perfil de Venda = Rodo Rede."""
    img_b64 = _img_rodo_rede_b64()

    if img_b64:
        # Pin com a imagem RodoRede.jpg — circular com borda roxa
        html_icon = (
            f"<div style='"
            f"width:40px;height:40px;"
            f"border-radius:50%;"
            f"border:3px solid {COR_RR_BORDA};"
            f"box-shadow:0 2px 8px rgba(0,0,0,.6);"
            f"overflow:hidden;"
            f"background:#fff;'>"
            f"<img src='{img_b64}' "
            f"style='width:100%;height:100%;object-fit:cover;display:block;'/>"
            f"</div>"
        )
        icon_size   = (40, 40)
        icon_anchor = (20, 20)
    else:
        # Fallback: emoji 🚛 roxo se imagem não encontrada
        html_icon = (
            f"<div style='"
            f"background:{COR_RR_FILL};"
            f"border:3px solid {COR_RR_BORDA};"
            f"border-radius:50%;"
            f"width:34px;height:34px;"
            f"display:flex;align-items:center;justify-content:center;"
            f"font-size:17px;line-height:1;"
            f"box-shadow:0 2px 8px rgba(0,0,0,.55);'>"
            f"🚛"
            f"</div>"
        )
        icon_size   = (34, 34)
        icon_anchor = (17, 17)

    icon = folium.DivIcon(html=html_icon, icon_size=icon_size, icon_anchor=icon_anchor)
    return folium.Marker(
        location=[lat, lon],
        icon=icon,
        popup=popup,
        tooltip=tooltip,
    )


def marcar_perfil_venda(df: pd.DataFrame, perfil_map: dict) -> pd.DataFrame:
    """Adiciona colunas '_perfil_venda' e '_rodo_rede' ao DataFrame."""
    df = df.copy()
    # Garante que _cnpj_norm existe mesmo que marcar_pro_frotas não tenha criado
    if "_cnpj_norm" not in df.columns and "cnpj" in df.columns:
        df["_cnpj_norm"] = df["cnpj"].fillna("").str.replace(r'\D', '', regex=True)
    if perfil_map and "_cnpj_norm" in df.columns:
        df["_perfil_venda"] = df["_cnpj_norm"].map(perfil_map).fillna("")
        df["_rodo_rede"] = df["_perfil_venda"].str.upper().str.strip() == PERFIL_RODO_REDE
    else:
        df["_perfil_venda"] = ""
        df["_rodo_rede"] = False
    return df


def criar_mapa(df, coords_rota=None, lat_orig=None, lon_orig=None,
               lat_dest=None, lon_dest=None, label_orig="Origem", label_dest="Destino",
               waypoints=None):
    """Retorna go.Figure (Plotly Scattermapbox) — WebGL, suporta 10 000+ marcadores."""
    # ── Cap de marcadores — Gestão de Frotas sempre priorizados ──────────────────────
    MAX_PF_MAPA = 5000
    n_total = len(df)
    foi_limitado = False
    if not df.empty and n_total > MAX_MAPA_POSTOS:
        foi_limitado = True
        tem_pf_col  = "_pro_frotas" in df.columns
        tem_cer_col = "_cercado"    in df.columns
        tem_rr_col  = "_rodo_rede"  in df.columns

        _mask_prio = pd.Series(False, index=df.index)
        if tem_pf_col:
            _mask_prio |= df["_pro_frotas"].fillna(False)
        if tem_cer_col:
            _mask_prio |= df["_cercado"].fillna(False)
        if tem_rr_col:
            _mask_prio |= df["_rodo_rede"].fillna(False)

        df_prio = df[_mask_prio]
        df_reg  = df[~_mask_prio]

        if len(df_prio) > MAX_PF_MAPA:
            df_prio = df_prio.sample(n=MAX_PF_MAPA, random_state=42)

        n_reg_max = max(0, MAX_MAPA_POSTOS - len(df_prio))
        if len(df_reg) > n_reg_max:
            df_reg = df_reg.sample(n=n_reg_max, random_state=42)

        df = pd.concat([df_prio, df_reg], ignore_index=True)

    # ── Centro e zoom do mapa ──────────────────────────────────────────────────
    if not df.empty:
        clat = float(df["_lat"].mean())
        clon = float(df["_lon"].mean())
        zoom = 6
    elif coords_rota:
        lats = [c[0] for c in coords_rota]
        lons = [c[1] for c in coords_rota]
        clat = (min(lats) + max(lats)) / 2
        clon = (min(lons) + max(lons)) / 2
        zoom = 5
    else:
        clat, clon, zoom = -15.0, -47.0, 4

    traces = []

    # ── Linha da rota ──────────────────────────────────────────────────────────
    if coords_rota and len(coords_rota) >= 2:
        coords_poly = _downsample(coords_rota, 500)
        traces.append(go.Scattermapbox(
            lat=[c[0] for c in coords_poly],
            lon=[c[1] for c in coords_poly],
            mode="lines",
            line=dict(width=4, color="#1565C0"),
            hoverinfo="skip",
            name="Rota",
            showlegend=False,
        ))

    # ── Marcadores origem / paradas / destino ─────────────────────────────────
    if lat_orig is not None:
        traces.append(go.Scattermapbox(
            lat=[lat_orig], lon=[lon_orig],
            mode="markers",
            marker=dict(size=16, color="#2E7D32"),
            text=[f"🟢 Origem: {label_orig}"],
            hoverinfo="text",
            name="Origem",
            showlegend=False,
        ))
    # Paradas intermediárias — PIN laranja grande com número da parada
    if waypoints:
        for idx_wp, wp in enumerate(waypoints, start=1):
            traces.append(go.Scattermapbox(
                lat=[wp["lat"]], lon=[wp["lon"]],
                mode="markers+text",
                marker=dict(size=22, color="#FF8F00"),
                text=[str(idx_wp)],
                textfont=dict(size=12, color="#fff", family="Arial Black, sans-serif"),
                textposition="middle center",
                customdata=[[wp["label"], idx_wp]],
                hovertemplate=(
                    f"<b>📍 Parada {idx_wp}</b><br>"
                    f"{wp['label']}<extra></extra>"
                ),
                name=f"Parada {idx_wp}",
                showlegend=False,
            ))
    if lat_dest is not None:
        traces.append(go.Scattermapbox(
            lat=[lat_dest], lon=[lon_dest],
            mode="markers",
            marker=dict(size=16, color="#C62828"),
            text=[f"🔴 Destino: {label_dest}"],
            hoverinfo="text",
            name="Destino",
            showlegend=False,
        ))

    # ── Postos ─────────────────────────────────────────────────────────────────
    if not df.empty:
        distribuidoras = sorted(df["distribuidora"].dropna().unique())
        mapa_cores = {d.upper().strip(): _cor_marca(d) for d in distribuidoras}

        tem_pf  = "_pro_frotas" in df.columns
        tem_cer = "_cercado"    in df.columns
        tem_rr  = "_rodo_rede"  in df.columns

        # Máscaras de grupo (mutuamente exclusivas; cercado > rr > pf > regular)
        mask_cer     = df["_cercado"].fillna(False)    if tem_cer else pd.Series(False, index=df.index)
        mask_rr      = (df["_rodo_rede"].fillna(False) if tem_rr  else pd.Series(False, index=df.index)) & ~mask_cer
        _pf_base     = (df["_pro_frotas"].fillna(False) if tem_pf else pd.Series(False, index=df.index)) & ~mask_cer & ~mask_rr
        mask_pf_ipi  = _pf_base & df["distribuidora"].fillna("").apply(_is_ipiranga)
        mask_pf_out  = _pf_base & ~mask_pf_ipi
        mask_reg     = ~mask_cer & ~mask_rr & ~_pf_base

        def _hover_txt(row):
            """Texto do tooltip: nome, bandeira, cidade/UF, CNPJ."""
            nome = str(row.get("razaoSocial", "?"))[:40]
            dist = str(row.get("distribuidora", "?"))
            mun  = str(row.get("municipio", ""))
            uf_  = str(row.get("uf", ""))
            cnpj = str(row.get("cnpj", ""))
            geo  = f"{mun}/{uf_}" if mun and uf_ else mun or uf_
            pf_  = " ⭐" if row.get("_pro_frotas") else ""
            rr_  = " · Rodo Rede" if row.get("_rodo_rede")  else ""
            cer_ = " ⚠️" if row.get("_cercado")    else ""
            cnpj_str = f"<br>📋 {cnpj}" if cnpj else ""
            return f"<b>{nome}</b>{pf_}{rr_}{cer_}<br>{dist}<br>{geo}{cnpj_str}"

        def _customdata(sub_df: pd.DataFrame) -> list:
            """customdata por ponto: [cnpj, nome, distribuidora, municipio/uf, lat, lon]"""
            rows = []
            for _, r in sub_df.iterrows():
                mun = str(r.get("municipio", ""))
                uf_ = str(r.get("uf", ""))
                geo = f"{mun}/{uf_}" if mun and uf_ else mun or uf_
                rows.append([
                    str(r.get("cnpj", "")),
                    str(r.get("razaoSocial", "")),
                    str(r.get("distribuidora", "")),
                    geo,
                    float(r["_lat"]),
                    float(r["_lon"]),
                    "1" if r.get("_pro_frotas") else "",
                    "1" if r.get("_cercado")    else "",
                    "1" if r.get("_rodo_rede")  else "",
                ])
            return rows

        # Postos Cercados — laranja
        if mask_cer.any():
            dfc = df[mask_cer]
            traces.append(go.Scattermapbox(
                lat=dfc["_lat"].tolist(), lon=dfc["_lon"].tolist(),
                mode="markers",
                marker=dict(size=13, color=COR_CERCADO_FILL, opacity=0.92),
                text=dfc.apply(_hover_txt, axis=1).tolist(),
                customdata=_customdata(dfc),
                hoverinfo="text",
                name="⚠️ Postos Cercados",
            ))

        # Rodo Rede — amarelo com destaque
        if mask_rr.any():
            dfr = df[mask_rr]
            traces.append(go.Scattermapbox(
                lat=dfr["_lat"].tolist(), lon=dfr["_lon"].tolist(),
                mode="markers",
                marker=dict(size=15, color=COR_RR_FILL, opacity=0.95),
                text=dfr.apply(_hover_txt, axis=1).tolist(),
                customdata=_customdata(dfr),
                hoverinfo="text",
                name="⭐ Ipiranga RodoRede",
            ))

        # Gestão de Frotas Ipiranga — amarelo
        if mask_pf_ipi.any():
            dfi = df[mask_pf_ipi]
            traces.append(go.Scattermapbox(
                lat=dfi["_lat"].tolist(), lon=dfi["_lon"].tolist(),
                mode="markers",
                marker=dict(size=14, color="#FFB300", opacity=0.95),
                text=dfi.apply(_hover_txt, axis=1).tolist(),
                customdata=_customdata(dfi),
                hoverinfo="text",
                name="⭐ GF Ipiranga",
            ))

        # Gestão de Frotas demais bandeiras — azul
        if mask_pf_out.any():
            dfp = df[mask_pf_out]
            traces.append(go.Scattermapbox(
                lat=dfp["_lat"].tolist(), lon=dfp["_lon"].tolist(),
                mode="markers",
                marker=dict(size=14, color=COR_PF_FILL, opacity=0.95),
                text=dfp.apply(_hover_txt, axis=1).tolist(),
                customdata=_customdata(dfp),
                hoverinfo="text",
                name="⭐ Gestão de Frotas",
            ))

        # Postos regulares ANP — cor por marca
        if mask_reg.any():
            dfg = df[mask_reg].copy()
            dfg["_cor_plot"] = dfg["distribuidora"].apply(lambda d: _cor(d, mapa_cores))
            traces.append(go.Scattermapbox(
                lat=dfg["_lat"].tolist(), lon=dfg["_lon"].tolist(),
                mode="markers",
                marker=dict(size=8, color=dfg["_cor_plot"].tolist(), opacity=0.85),
                text=dfg.apply(_hover_txt, axis=1).tolist(),
                customdata=_customdata(dfg),
                hoverinfo="text",
                name="⛽ Postos ANP",
            ))

    # ── Trace Top 5 Mais Baratos — estrelas douradas sobrepostas ─────────────────
    if not df.empty and "_rank_barato" in df.columns:
        _df_top5 = df[df["_rank_barato"] > 0].sort_values("_rank_barato")
        if not _df_top5.empty:
            _top5_hover = []
            for _, _tr in _df_top5.iterrows():
                _rank_n  = int(_tr["_rank_barato"])
                _emoji_r = _RANK_EMOJI.get(_rank_n, str(_rank_n))
                _nome_r  = str(_tr.get("razaoSocial", "?"))[:35]
                _preco_r = _tr.get("_preco_barato")
                _comb_r  = str(_tr.get("_comb_barato", ""))
                _mun_r   = str(_tr.get("municipio", ""))
                _uf_r    = str(_tr.get("uf", ""))
                _geo_r   = f"{_mun_r}/{_uf_r}" if _mun_r and _uf_r else _mun_r or _uf_r
                _preco_str = f"R$ {_preco_r:.3f}/L" if _preco_r is not None else "—"
                _top5_hover.append(
                    f"<b>{_emoji_r} #{_rank_n} Mais Barato</b><br>"
                    f"{_nome_r}<br>"
                    f"💰 {_preco_str} — {_comb_r}<br>"
                    f"📍 {_geo_r}<extra></extra>"
                )
            traces.append(go.Scattermapbox(
                lat=_df_top5["_lat"].tolist(),
                lon=_df_top5["_lon"].tolist(),
                mode="markers+text",
                marker=dict(size=22, color="#FFD700", opacity=1.0),
                text=[_RANK_EMOJI.get(int(r), "★") for r in _df_top5["_rank_barato"]],
                textfont=dict(size=11, color="#7B3F00", family="Arial Black, sans-serif"),
                textposition="middle center",
                hovertemplate=_top5_hover,
                name="💰 Top 5 Mais Baratos",
                showlegend=True,
            ))

    # ── Garante ao menos 1 trace Scattermapbox para ativar tiles mesmo sem dados ──
    # Sem isso, Plotly cai no modo cartesiano quando traces=[] e exibe eixos vazios.
    if not traces:
        traces.append(go.Scattermapbox(
            lat=[], lon=[],
            mode="markers",
            marker=dict(size=1, opacity=0),
            hoverinfo="skip",
            showlegend=False,
        ))

    # ── Anotação de limitação ──────────────────────────────────────────────────
    layout_annotations = []
    if foi_limitado:
        layout_annotations.append(dict(
            text=(f"⚠️ Exibindo {_fmt_int(MAX_MAPA_POSTOS)} de {_fmt_int(n_total)} postos "
                  f"(Gestão de Frotas priorizados). Veja todos na aba Dados Tabulares."),
            x=0.5, y=0.02, xref="paper", yref="paper",
            showarrow=False, align="center",
            font=dict(size=11, color="#E65100"),
            bgcolor="rgba(255,243,224,0.92)",
            bordercolor="#FF9800", borderwidth=1,
        ))

    fig = go.Figure(
        data=traces,
        layout=go.Layout(
            mapbox=dict(
                style="carto-positron",
                center=dict(lat=clat, lon=clon),
                zoom=zoom,
            ),
            margin=dict(l=0, r=0, t=0, b=0),
            legend=dict(
                bgcolor="rgba(255,255,255,0.92)",
                bordercolor="#ccc",
                borderwidth=1,
                font=dict(size=11),
                x=0.01, y=0.99,
                xanchor="left", yanchor="top",
            ),
            annotations=layout_annotations,
            uirevision="mapa_estudo_rede",
        ),
    )
    return fig


def _renderizar_mapa(fig: go.Figure, height: int = 660, key: str = "mapa_plot") -> None:
    """
    Renderiza o mapa Plotly e exibe painel de detalhe ao clicar num posto.

    — Seleção ativa: banner amarelo acima do mapa + card destacado abaixo.
    — Botão "✕ Limpar seleção" força re-render limpo (sem seleção).
    — CNPJ + localização aparecem no tooltip (hover) nativo do Plotly.
    """
    # ── Chaves de sessão para esta instância do mapa ────────────────────────────
    _sel_key = f"_msel_{key}"   # armazena dados do posto selecionado
    _ver_key = f"_mver_{key}"   # versão para forçar re-render ao limpar

    _ver      = st.session_state.get(_ver_key, 0)
    _chart_key = f"{key}_v{_ver}"
    _sel       = st.session_state.get(_sel_key)  # dict com dados do posto ou None

    # ── Banner de seleção ativa (aparece ACIMA do mapa) ─────────────────────────
    if _sel:
        _ban_nome = _sel.get("nome", "Posto")
        _ban_geo  = _sel.get("geo", "")
        _ban_lat  = _sel.get("lat")
        _ban_lon  = _sel.get("lon")
        _ban_maps = (f"https://maps.google.com/?q={float(_ban_lat):.6f},{float(_ban_lon):.6f}"
                     if _ban_lat is not None and _ban_lon is not None else None)

        c_info, c_maps, c_limpar = st.columns([5, 1, 1])
        with c_info:
            st.markdown(
                f"<div style='"
                f"background:linear-gradient(90deg,#fff8e1,#fffde7);"
                f"border:1.5px solid #f9a825;"
                f"border-radius:8px;padding:8px 14px;"
                f"display:flex;align-items:center;gap:10px;"
                f"font-size:13px;color:#5f4307;line-height:1.4'>"
                f"<span style='font-size:18px'>📌</span>"
                f"<div><b>Posto selecionado</b><br>"
                f"<span style='color:#333'>{_ban_nome}"
                f"{'  ·  ' + _ban_geo if _ban_geo else ''}</span></div>"
                f"</div>",
                unsafe_allow_html=True,
            )
        with c_maps:
            if _ban_maps:
                st.link_button("📍 Maps", _ban_maps, use_container_width=True)
        with c_limpar:
            if st.button(
                "✕ Limpar",
                key=f"_desel_{key}",
                use_container_width=True,
                help="Remover seleção e voltar ao mapa limpo",
            ):
                st.session_state.pop(_sel_key, None)
                st.session_state[_ver_key] = _ver + 1  # troca a key → re-render sem seleção
                st.rerun()

    # ── Renderiza o gráfico ─────────────────────────────────────────────────────
    # Config mobile-friendly: botões da toolbar reduzidos, sem mode bar em telas pequenas
    _plotly_cfg = {
        "scrollZoom":          True,
        "displaylogo":         False,
        "responsive":          True,
        "modeBarButtonsToRemove": [
            "select2d", "lasso2d", "autoScale2d",
            "hoverClosestCartesian", "hoverCompareCartesian",
        ],
        "toImageButtonOptions": {"format": "png", "filename": "mapa_rede"},
    }
    try:
        evt = st.plotly_chart(
            fig,
            use_container_width=True,
            config=_plotly_cfg,
            height=height,
            on_select="rerun",
            key=_chart_key,
        )
    except TypeError:
        st.plotly_chart(fig, use_container_width=True,
                        config={"scrollZoom": True}, height=height)
        return

    # ── Captura seleção ─────────────────────────────────────────────────────────
    pts = (evt or {}).get("selection", {}).get("points", [])
    if not pts:
        return

    pt = pts[0]
    cd = pt.get("customdata") or []

    # customdata = [cnpj, nome, distribuidora, geo, lat, lon, pf, cercado, rr]
    _cnpj = cd[0] if len(cd) > 0 else ""
    _nome = cd[1] if len(cd) > 1 else pt.get("text", "Posto")
    _dist = cd[2] if len(cd) > 2 else ""
    _geo  = cd[3] if len(cd) > 3 else ""
    _lat  = cd[4] if len(cd) > 4 else pt.get("lat")
    _lon  = cd[5] if len(cd) > 5 else pt.get("lon")
    _pf   = bool(cd[6]) if len(cd) > 6 else False
    _cer  = bool(cd[7]) if len(cd) > 7 else False
    _rr   = bool(cd[8]) if len(cd) > 8 else False

    if _lat is None or _lon is None:
        return

    # Persiste no session_state (mantém painel visível em reruns futuros)
    _existing_cnpj = (st.session_state.get(_sel_key) or {}).get("cnpj", "")
    st.session_state[_sel_key] = dict(
        nome=_nome, cnpj=_cnpj, dist=_dist, geo=_geo,
        lat=_lat, lon=_lon, pf=_pf, cer=_cer, rr=_rr,
    )

    # ── Força re-render imediato para o banner aparecer na 1ª seleção ──
    # Sem este rerun o banner só apareceria no 2º clique, porque o
    # st.plotly_chart() (que produz os dados) vem depois do bloco "if _sel:"
    if _cnpj != _existing_cnpj:
        st.rerun()

    _maps_url = f"https://maps.google.com/?q={_lat:.6f},{_lon:.6f}"

    # ── Badges de categoria ─────────────────────────────────────────────────────
    _badges_html = ""
    if _pf:
        _badges_html += (
            "<span style='background:#1565c0;color:#fff;border-radius:4px;"
            "padding:2px 8px;font-size:11px;margin-right:5px'>⭐ Gestão de Frotas</span>"
        )
    if _rr:
        _rr_img = _img_rodo_rede_b64()
        if _rr_img:
            _badges_html += (
                f"<span style='display:inline-flex;align-items:center;gap:4px;"
                f"background:#fff3e0;border:1px solid {COR_RR_FILL};"
                f"border-radius:4px;padding:2px 7px;font-size:11px;margin-right:5px'>"
                f"<img src='{_rr_img}' style='height:14px;width:auto;object-fit:contain;"
                f"vertical-align:middle;border-radius:2px'> ⭐ Ipiranga RodoRede</span>"
            )
        else:
            _badges_html += (
                f"<span style='background:{COR_RR_FILL};color:#fff;border-radius:4px;"
                f"padding:2px 8px;font-size:11px;margin-right:5px'>⭐ Ipiranga RodoRede</span>"
            )
    if _cer:
        _badges_html += (
            "<span style='background:#FF8F00;color:#fff;border-radius:4px;"
            "padding:2px 8px;font-size:11px;margin-right:5px'>⚠️ Cercado</span>"
        )

    _cnpj_html = f"<span style='color:#666;font-size:12px'>📋 {_cnpj}</span>&nbsp;&nbsp;" if _cnpj else ""
    _geo_html  = f"<span style='color:#666;font-size:12px'>📍 {_geo}</span>" if _geo else ""
    _dist_html = f"&nbsp;·&nbsp;<span style='color:#888'>{_dist}</span>" if _dist else ""

    # ── Card de detalhe destacado (abaixo do mapa) ──────────────────────────────
    col_info, col_btn = st.columns([5, 1])
    with col_info:
        st.markdown(
            f"<div style='"
            f"background:linear-gradient(135deg,#e3f2fd 0%,#bbdefb 100%);"
            f"border-left:5px solid #1565c0;"
            f"border-radius:0 10px 10px 0;"
            f"padding:12px 16px;"
            f"box-shadow:0 3px 12px rgba(21,101,192,0.18);"
            f"margin-top:4px'>"
            f"<div style='font-size:15px;font-weight:700;color:#0d47a1'>"
            f"🏪 {_nome}{_dist_html}</div>"
            f"<div style='margin:6px 0 4px'>{_badges_html}</div>"
            f"<div style='margin-top:4px'>{_cnpj_html}{_geo_html}</div>"
            f"</div>",
            unsafe_allow_html=True,
        )
    with col_btn:
        st.link_button("📍 Maps", _maps_url, use_container_width=True)


# ═══════════════════════════════════════════════════════════════════
#  PREÇOS ANP — Levantamento Semanal de Preços de Combustíveis
#
#  Estrutura real do arquivo (resumo_semanal_lpc_*.xlsx):
#    Aba CAPITAIS  — colunas: DATA INICIAL, DATA FINAL, ESTADO, MUNICÍPIO, PRODUTO, Nº POSTOS, UNIDADE, PREÇO MÉDIO
#    Aba MUNICIPIOS— mesmas colunas de CAPITAIS
#    Aba ESTADOS   — colunas: DATA INICIAL, DATA FINAL, REGIAO, ESTADOS, PRODUTO, Nº POSTOS, UNIDADE, PREÇO MÉDIO
#    Aba REGIOES   — colunas: DATA INICIAL, DATA FINAL, REGIAO, PRODUTO, Nº POSTOS, UNIDADE, PREÇO MÉDIO, DESVIO
#    Aba BRASIL    — colunas: DATA INICIAL, DATA FINAL, BRASIL, PRODUTO, Nº POSTOS, UNIDADE, PREÇO MÉDIO, DESVIO
#  Cabeçalho sempre na linha 9 (0-indexed); linhas 0-7 são título/obs.
# ═══════════════════════════════════════════════════════════════════

def _anp_norm(s):
    """Remove acentos, uppercase, strip — para comparações robustas."""
    return unicodedata.normalize("NFD", str(s)).encode("ascii", "ignore").decode("ascii").upper().strip()


# Mapeamento reverso: nome normalizado → código UF (construído uma única vez)
_NOME_PARA_UF: dict = {
    _anp_norm(nome): codigo for codigo, nome in UF_NOME.items()
}


def _normalizar_uf(valor: str) -> str:
    """Converte 'Ceará', 'CEARÁ', 'ceara', 'CE', 'ce', etc. para o código de 2 letras.
    Retorna o valor original em maiúscula se não conseguir identificar.
    """
    v = str(valor).strip().upper()
    if v in UFS:                          # já é código válido
        return v
    return _NOME_PARA_UF.get(_anp_norm(v), v)


def _brl(v, d=2):
    """Formata número no padrão monetário brasileiro: ponto para milhar, vírgula para decimal.
    Exemplo: _brl(1234.5) → '1.234,50'  |  _brl(1.329, 3) → '1,329'
    """
    s = f"{v:,.{d}f}"          # '1,234.50' (padrão US)
    return s.replace(",", "X").replace(".", ",").replace("X", ".")


def _fmt_int(v) -> str:
    """Formata inteiro com ponto como separador de milhar (padrão BR).
    Exemplo: _fmt_int(14045) → '14.045'  |  _fmt_int(2962) → '2.962'
    """
    try:
        return f"{int(round(v)):,}".replace(",", ".")
    except (TypeError, ValueError):
        return str(v)


def _anp_preco_brasil(sheets, pk):
    """Extrai o preço médio nacional (aba BRASIL) para um produto pk.
    Retorna float ou None.
    """
    df_b = sheets.get("brasil")
    if df_b is None or df_b.empty:
        return None
    c_prod = _anp_col(df_b, "produto")
    c_med  = _anp_col(df_b, "medio revenda", "media revenda", "preco medio", "medio")
    if not c_prod or not c_med:
        return None
    for prod_raw, grp in df_b.groupby(c_prod):
        if _anp_norm(str(prod_raw)) == pk:
            val = pd.to_numeric(grp[c_med], errors="coerce").mean()
            if not pd.isna(val):
                return float(round(val, 3))
    return None


def _anp_preco_uf(sheets, pk, uf):
    """Extrai o preço médio estadual (aba ESTADOS) para produto pk e UF.
    Retorna float ou None.
    """
    df_e = sheets.get("estados")
    if df_e is None or df_e.empty:
        return None
    c_prod = _anp_col(df_e, "produto")
    c_med  = _anp_col(df_e, "medio revenda", "media revenda", "preco medio", "medio")
    c_est  = _anp_col(df_e, "estado", "estados", "uf")
    if not c_prod or not c_med or not c_est:
        return None
    uf_norm = _anp_norm(UF_NOME.get(uf, uf))
    for (prod_raw, est_raw), grp in df_e.groupby([c_prod, c_est]):
        if (_anp_norm(str(prod_raw)) == pk
                and (_anp_norm(str(est_raw)) == uf_norm
                     or _anp_norm(str(est_raw)) == _anp_norm(uf))):
            val = pd.to_numeric(grp[c_med], errors="coerce").mean()
            if not pd.isna(val):
                return float(round(val, 3))
    return None


def _tendencia_badge(preco_atual, preco_anterior, inline=True):
    """Gera um badge HTML com seta de tendência comparando dois preços ANP.

    Parâmetros:
        preco_atual:    float — preço ANP da semana atual
        preco_anterior: float — preço ANP da semana anterior
        inline:         bool  — True = badge pequeno (para cards); False = bloco expandido

    Retorna string HTML ou "" se qualquer preço for None.

    Regras:
        |delta| ≤ 0.5% do anterior → ≈ estável (cinza)
        delta  > 0  → ↑ subiu  (vermelho)
        delta  < 0  → ↓ caiu   (verde)
    """
    if preco_atual is None or preco_anterior is None:
        return ""
    delta     = preco_atual - preco_anterior
    delta_pct = (delta / preco_anterior * 100) if preco_anterior else 0
    tol       = abs(preco_anterior) * 0.005   # 0.5% de tolerância

    if abs(delta) <= tol:
        seta, cor_bg, cor_txt, texto = "≈", "#f3f4f6", "#666", "estável"
    elif delta > 0:
        seta, cor_bg, cor_txt, texto = "↑", "#ffebee", "#c62828", f"+{_brl(delta, 3)}/L"
    else:
        seta, cor_bg, cor_txt, texto = "↓", "#e8f5e9", "#2e7d32", f"{_brl(delta, 3)}/L"

    if inline:
        return (
            f"<span title='Var. semana anterior: {_brl(delta,3)} R$/L ({delta_pct:+.1f}%)' "
            f"style='display:inline-flex;align-items:center;gap:2px;"
            f"background:{cor_bg};color:{cor_txt};"
            f"border-radius:4px;padding:1px 5px;font-size:11px;"
            f"font-weight:700;white-space:nowrap;cursor:help'>"
            f"{seta}&thinsp;{texto}</span>"
        )
    else:
        return (
            f"<div style='display:flex;align-items:center;gap:6px;"
            f"background:{cor_bg};border-radius:6px;padding:4px 10px;margin-top:4px'>"
            f"<span style='font-size:18px;color:{cor_txt};font-weight:800'>{seta}</span>"
            f"<span style='font-size:11px;color:{cor_txt}'>"
            f"vs sem. anterior: <b>{_brl(preco_anterior, 3)}</b> → "
            f"<b>{_brl(preco_atual, 3)}</b> "
            f"(<b>{delta_pct:+.1f}%</b>)</span>"
            f"</div>"
        )


def _anp_col(df, *termos):
    """Retorna a primeira coluna cujo nome normalizado contenha qualquer dos termos."""
    for col in df.columns:
        cn = _anp_norm(col)
        if any(_anp_norm(t) in cn for t in termos):
            return col
    return None


def _anp_ler_aba(xls, aba):
    """Lê aba da planilha ANP, pulando linhas de título/notas.

    O cabeçalho real fica na linha que contém 'PRODUTO' + 'DATA' +
    algum identificador geográfico. Requer 'DATA' para não confundir
    com a linha de OBS que cita o produto mas não é cabeçalho.
    """
    df_raw = pd.read_excel(xls, sheet_name=aba, header=None)
    header_row = 0
    geo_keys = ["ESTADO", "MUNICIPIO", "REGIAO", "BRASIL", "PAIS"]
    for i in range(min(20, len(df_raw))):
        row_str = _anp_norm(" ".join(str(v) for v in df_raw.iloc[i].values))
        if ("PRODUTO" in row_str and "DATA" in row_str
                and any(k in row_str for k in geo_keys)):
            header_row = i
            break
    df = pd.read_excel(xls, sheet_name=aba, header=header_row)
    df.columns = [str(c).strip() for c in df.columns]
    df = df.dropna(how="all").reset_index(drop=True)
    col_p = _anp_col(df, "produto")
    if col_p:
        df = df[df[col_p].notna() & (df[col_p].astype(str).str.strip() != "")].copy()
    return df


def _anp_processar_arquivo(buf):
    """Lê todas as abas do arquivo ANP e retorna dict {chave: DataFrame}.

    Abas reconhecidas: MUNICIPIOS, ESTADOS, CAPITAIS, REGIOES, BRASIL
    """
    xls = pd.ExcelFile(buf)
    sheets = {}
    # Nota: usar prefixos que são substrings reais dos nomes normalizados:
    #   "CAPITAL" ≠ substring de "CAPITAIS" (pos6: L vs I)
    #   "REGIAO"  ≠ substring de "REGIOES"  (pos4: A vs O, pois ã→a mas ões→oes)
    #   → usar "CAPITA" (em CAPITAIS) e "REGIO" (em REGIOES)
    mapa = {
        "municipios": ["munic"],
        "estados":    ["estado"],
        "capitais":   ["capita"],   # "CAPITA" ⊂ "CAPITAIS"
        "regioes":    ["regio"],    # "REGIO"  ⊂ "REGIOES"
        "brasil":     ["brasil", "pais", "nacional"],
    }
    for tipo, palavras in mapa.items():
        aba = next(
            (s for s in xls.sheet_names if any(_anp_norm(p) in _anp_norm(s) for p in palavras)),
            None,
        )
        if aba:
            try:
                df = _anp_ler_aba(xls, aba)
                if not df.empty:
                    sheets[tipo] = df
            except Exception:
                pass
    return sheets


def buscar_precos_anp():
    """Tenta baixar a planilha de preços ANP automaticamente.
    Retorna (bytes_xlsx | None, semana_str | None, erro_str | None).
    Também tenta baixar a semana anterior e salvar em session_state
    como '_precos_anp_cache_anterior' para o indicador de tendência.
    NÃO usa @st.cache_data para que falhas não fiquem em cache.
    O resultado bem-sucedido é guardado em st.session_state pelo chamador.
    """
    try:
        headers_page = {**HEADERS_ANP,
                        "Accept": "text/html,application/xhtml+xml,*/*",
                        "Referer": "https://www.gov.br/anp/"}
        resp = requests.get(ANP_PRECOS_URL, headers=headers_page, timeout=20)
        resp.raise_for_status()
        from bs4 import BeautifulSoup
        soup = BeautifulSoup(resp.text, "html.parser")
        links_xlsx = [
            a["href"] for a in soup.find_all("a", href=True)
            if a["href"].lower().endswith(".xlsx")
        ]
        if not links_xlsx:
            return None, None, "Nenhum link .xlsx encontrado na página da ANP."
        link = links_xlsx[0]
        if not link.startswith("http"):
            link = "https://www.gov.br" + link
        r2 = requests.get(link, headers=HEADERS_ANP, timeout=60)
        r2.raise_for_status()
        semana = link.split("/")[-1].replace(".xlsx", "")

        # ── Tenta buscar a semana anterior (segundo link) ─────────
        if len(links_xlsx) >= 2:
            try:
                link_ant = links_xlsx[1]
                if not link_ant.startswith("http"):
                    link_ant = "https://www.gov.br" + link_ant
                _ant_ja = st.session_state.get("_precos_anp_cache_anterior", {})
                _sem_ant = link_ant.split("/")[-1].replace(".xlsx", "")
                if _ant_ja.get("semana") != _sem_ant:
                    r_ant = requests.get(link_ant, headers=HEADERS_ANP, timeout=60)
                    r_ant.raise_for_status()
                    _sh_ant = _anp_processar_arquivo(io.BytesIO(r_ant.content))
                    if _sh_ant:
                        st.session_state["_precos_anp_cache_anterior"] = {
                            "sheets": _sh_ant, "semana": _sem_ant
                        }
            except Exception:
                pass  # semana anterior é opcional — falha silenciosa

        return r2.content, semana, None
    except Exception as ex:
        return None, None, str(ex)


def _anp_preco_medio(df, col_geo, val_geo, col_prod, col_med, col_uni, col_npos):
    """Extrai preços médios por produto de um DataFrame ANP, opcionalmente filtrado por geo."""
    ordem = {_anp_norm(k): i for i, k in enumerate(PRODUTOS_CHAVE)}
    linhas = []
    df_fil = df.copy()
    if val_geo is not None and col_geo:
        df_fil = df_fil[df_fil[col_geo].apply(_anp_norm) == _anp_norm(str(val_geo))]
    if df_fil.empty:
        return linhas
    for prod_raw, grp in df_fil.groupby(col_prod):
        pk  = _anp_norm(str(prod_raw))
        med = pd.to_numeric(grp[col_med], errors="coerce").mean()
        if pd.isna(med):
            continue
        uni = str(grp[col_uni].iloc[0]).strip() if col_uni else "R$/L"
        nps = int(pd.to_numeric(grp[col_npos], errors="coerce").sum()) if col_npos else None
        linhas.append({
            "_ordem":      ordem.get(pk, 99),
            "_pk":         pk,
            "Combustível": PRODUTO_CURTO.get(pk, str(prod_raw).title()),
            "Preço Médio": round(float(med), 3),
            "Unidade":     uni,
            "Postos":      nps,
        })
    linhas.sort(key=lambda x: x["_ordem"])
    return linhas


def _anp_extrair_precos(sheets, uf=None, municipio=None):
    """Extrai preços para UF/município com referências em múltiplos níveis.

    Hierarquia de dados: Município → Capital → Estado → Região → Brasil
    """
    nome_uf = _anp_norm(UF_NOME.get(uf or "", "")) if uf else None

    def _cols(df):
        return {
            "est":  _anp_col(df, "estado", "estados"),
            "mun":  _anp_col(df, "munic"),
            "reg":  _anp_col(df, "regiao"),
            "prod": _anp_col(df, "produto"),
            "med":  _anp_col(df, "medio revenda", "media revenda", "preco medio"),
            "uni":  _anp_col(df, "unidade"),
            "npos": _anp_col(df, "postos pesq", "numero de postos", "n postos", "numero postos"),
        }

    # ── Referência Brasil ──────────────────────────────────────────
    ref_brasil: dict = {}
    if "brasil" in sheets:
        df_b = sheets["brasil"]
        c = _cols(df_b)
        if c["prod"] and c["med"]:
            for r in _anp_preco_medio(df_b, None, None, c["prod"], c["med"], c["uni"], c["npos"]):
                ref_brasil[r["_pk"]] = r["Preço Médio"]

    # ── Referência Região (descobre a região do estado) ────────────
    ref_regiao: dict = {}
    nome_regiao = None
    if "regioes" in sheets and uf and "estados" in sheets:
        df_reg = sheets["regioes"]
        cr = _cols(df_reg)
        df_est = sheets["estados"]
        ce = _cols(df_est)
        if ce["est"] and ce["reg"] and nome_uf and cr["reg"] and cr["prod"] and cr["med"]:
            row_est = df_est[df_est[ce["est"]].apply(_anp_norm) == nome_uf]
            if not row_est.empty:
                nome_regiao = _anp_norm(str(row_est.iloc[0][ce["reg"]]))
                for r in _anp_preco_medio(df_reg, cr["reg"], nome_regiao,
                                          cr["prod"], cr["med"], cr["uni"], cr["npos"]):
                    ref_regiao[r["_pk"]] = r["Preço Médio"]

    # ── Referência Estado ──────────────────────────────────────────
    ref_estado: dict = {}
    if "estados" in sheets and uf:
        df_est = sheets["estados"]
        ce = _cols(df_est)
        if ce["est"] and ce["prod"] and ce["med"] and nome_uf:
            for r in _anp_preco_medio(df_est, ce["est"], nome_uf,
                                       ce["prod"], ce["med"], ce["uni"], ce["npos"]):
                ref_estado[r["_pk"]] = r["Preço Médio"]

    # ── Nível primário ─────────────────────────────────────────────
    linhas_base = []
    nivel_real  = "Estado"

    # 1. Município
    if municipio and "municipios" in sheets:
        df_mun = sheets["municipios"]
        cm = _cols(df_mun)
        if cm["est"] and cm["mun"] and cm["prod"] and cm["med"] and nome_uf:
            mun_n = _anp_norm(municipio)
            df_fil = df_mun[
                (df_mun[cm["est"]].apply(_anp_norm) == nome_uf) &
                (df_mun[cm["mun"]].apply(_anp_norm).str.contains(mun_n, na=False))
            ]
            if not df_fil.empty:
                linhas_base = _anp_preco_medio(df_fil, None, None,
                                               cm["prod"], cm["med"], cm["uni"], cm["npos"])
                nivel_real = "Município"

    # 2. Capital (fallback se município não encontrado em MUNICIPIOS)
    if not linhas_base and municipio and "capitais" in sheets:
        df_cap = sheets["capitais"]
        cc = _cols(df_cap)
        if cc["est"] and cc["mun"] and cc["prod"] and cc["med"] and nome_uf:
            mun_n = _anp_norm(municipio)
            df_fil = df_cap[
                (df_cap[cc["est"]].apply(_anp_norm) == nome_uf) &
                (df_cap[cc["mun"]].apply(_anp_norm).str.contains(mun_n, na=False))
            ]
            if not df_fil.empty:
                linhas_base = _anp_preco_medio(df_fil, None, None,
                                               cc["prod"], cc["med"], cc["uni"], cc["npos"])
                nivel_real = "Capital"

    # 3. Estado
    if not linhas_base and uf and "estados" in sheets:
        df_est = sheets["estados"]
        ce = _cols(df_est)
        if ce["est"] and ce["prod"] and ce["med"] and nome_uf:
            linhas_base = _anp_preco_medio(df_est, ce["est"], nome_uf,
                                           ce["prod"], ce["med"], ce["uni"], ce["npos"])
            nivel_real = "Estado"

    if not linhas_base:
        return []

    resultado = []
    for r in linhas_base:
        pk = r["_pk"]
        resultado.append({
            **r,
            "Nível":        nivel_real,
            "Ref. Estado":  ref_estado.get(pk),
            "Ref. Região":  ref_regiao.get(pk),
            "Ref. Brasil":  ref_brasil.get(pk),
            "Nome Região":  nome_regiao,
        })
    return resultado


def _parse_label_localizacao(label: str) -> tuple:
    """Extrai (municipio_str | None, uf_str | None) de um label de seleção de local.

    Formatos suportados:
      "CAMPINAS – SP"          → ("CAMPINAS", "SP")
      "Campinas – São Paulo"   → ("Campinas", "SP")
      "São Paulo"              → (None, "SP")   se for nome de estado
      "PALMAS – TO | Posto X"  → ("PALMAS", "TO")
    """
    if not label:
        return None, None

    # Remove sufixo de posto, se houver
    label = label.split(" | ")[0].strip()

    if " – " in label or " - " in label:
        sep = " – " if " – " in label else " - "
        parts = label.split(sep, 1)
        cidade     = parts[0].strip()
        uf_ou_est  = parts[1].strip()

        # 2-letter UF code
        if len(uf_ou_est) == 2 and uf_ou_est.upper() in UF_NOME:
            return cidade, uf_ou_est.upper()

        # Full state name → reverse lookup
        n = _anp_norm(uf_ou_est)
        for uf_k, nome_k in UF_NOME.items():
            if _anp_norm(nome_k) == n:
                return cidade, uf_k
        # Partial match
        for uf_k, nome_k in UF_NOME.items():
            if _anp_norm(nome_k) in n or n in _anp_norm(nome_k):
                return cidade, uf_k
        return cidade, None

    # Sem separador — pode ser nome de estado, sigla, ou "Estado XX"
    # Trata "Estado SP" / "Estado São Paulo" / "estado ES"
    label_strip = label.strip()
    prefixos = ("estado ", "state ", "uf ")
    label_sem_prefixo = label_strip
    for pfx in prefixos:
        if _anp_norm(label_strip).startswith(_anp_norm(pfx)):
            label_sem_prefixo = label_strip[len(pfx):].strip()
            break

    # Sigla de UF (2 letras)
    if len(label_sem_prefixo) == 2 and label_sem_prefixo.upper() in UF_NOME:
        return None, label_sem_prefixo.upper()

    # Nome completo de estado
    n = _anp_norm(label_sem_prefixo)
    for uf_k, nome_k in UF_NOME.items():
        if _anp_norm(nome_k) == n:
            return None, uf_k

    # Tentativa com label original (sem remover prefixo)
    n_orig = _anp_norm(label_strip)
    for uf_k, nome_k in UF_NOME.items():
        if _anp_norm(nome_k) == n_orig:
            return None, uf_k

    return label_strip, None


def _anp_preco_ponto(sheets: dict, label: str, combustivel_pk: str) -> tuple:
    """Resolve o preço médio de um combustível para um ponto (label de cidade/estado).

    Hierarquia: Município → Capital → Estado → Região → Brasil
    Retorna (preco_float | None, nivel_str, descricao_str).
    """
    municipio, uf = _parse_label_localizacao(label)
    if not uf and not municipio:
        return None, "—", "—"

    # Tenta extrair preço pelo nível mais detalhado disponível
    rows = _anp_extrair_precos(sheets, uf=uf, municipio=municipio)
    if rows:
        for r in rows:
            if r["_pk"] == combustivel_pk:
                desc = municipio or UF_NOME.get(uf or "", uf or "")
                return r["Preço Médio"], r["Nível"], desc

    # Fallback: apenas estado
    if uf:
        rows_uf = _anp_extrair_precos(sheets, uf=uf)
        for r in rows_uf:
            if r["_pk"] == combustivel_pk:
                return r["Preço Médio"], "Estado", UF_NOME.get(uf, uf)

    # Fallback: Brasil
    if "brasil" in sheets:
        df_b   = sheets["brasil"]
        b_prod = _anp_col(df_b, "produto")
        b_med  = _anp_col(df_b, "medio revenda", "media revenda", "preco medio")
        b_uni  = _anp_col(df_b, "unidade")
        b_npos = _anp_col(df_b, "postos pesq", "numero de postos", "n postos")
        if b_prod and b_med:
            for r in _anp_preco_medio(df_b, None, None, b_prod, b_med, b_uni, b_npos):
                if r["_pk"] == combustivel_pk:
                    return r["Preço Médio"], "Brasil", "Média Nacional"

    return None, "—", "—"


def _anp_precos_por_fuel_brasil(sheets):
    """Retorna dict {fuel_pk: preco_medio} do nível Brasil."""
    if "brasil" not in sheets:
        return {}
    df_b  = sheets["brasil"]
    c_prod = _anp_col(df_b, "produto")
    c_med  = _anp_col(df_b, "medio revenda", "media revenda", "preco medio")
    c_uni  = _anp_col(df_b, "unidade")
    c_npos = _anp_col(df_b, "postos pesq", "numero de postos")
    if not c_prod or not c_med:
        return {}
    return {r["_pk"]: r["Preço Médio"]
            for r in _anp_preco_medio(df_b, None, None, c_prod, c_med, c_uni, c_npos)}


def _anp_precos_por_fuel_ufs(sheets, ufs):
    """Retorna dict {fuel_pk: preco_medio} média dos estados listados."""
    acum: dict = {}
    for uf in ufs:
        for r in _anp_extrair_precos(sheets, uf=uf):
            acum.setdefault(r["_pk"], []).append(r["Preço Médio"])
    return {pk: sum(v) / len(v) for pk, v in acum.items()}


def _anp_precos_por_fuel_por_uf(sheets, ufs):
    """Retorna dict {fuel_pk: {uf: preco_medio}} para cada estado."""
    result: dict = {}
    for uf in ufs:
        for r in _anp_extrair_precos(sheets, uf=uf):
            result.setdefault(r["_pk"], {})[uf] = r["Preço Médio"]
    return result


def _calcular_comparativo_pf_anp(df_pp, cnpjs_pf, sheets_anp, ufs=None):
    """
    Calcula comparativo Gestão de Frotas (Preço Posto) vs ANP.

    Retorna lista de dicts com:
      combustivel_label, combustivel_pk,
      preco_pf_med, preco_pf_min, preco_pf_max, n_postos_pf,
      preco_anp, nivel_anp,
      delta_abs, delta_pct, economia_100l, data_atualizacao
    """
    if df_pp is None or df_pp.empty or not cnpjs_pf or sheets_anp is None:
        return []

    # Filtra só postos Gestão de Frotas
    df_pf = df_pp[df_pp["cnpj_norm"].isin(cnpjs_pf)].copy()
    if df_pf.empty:
        return []

    # Normaliza os PKs da planilha PP → PKs canônicos ANP
    # Isso consolida variantes como "DIESEL S-10 COMUM" + "DIESEL S-10 ADITIVADO"
    # em um único grupo "OLEO DIESEL S10" para gerar um card por tipo de combustível.
    df_pf["combustivel_pk"] = df_pf["combustivel_pk"].map(
        lambda pk: _PP_PARA_ANP_PK.get(pk, pk)
    )

    # Preços ANP de referência
    if ufs:
        precos_anp = _anp_precos_por_fuel_ufs(sheets_anp, ufs)
        nivel_anp  = UF_NOME.get(ufs[0], ufs[0]) if len(ufs) == 1 else f"{len(ufs)} estados"
    else:
        precos_anp = _anp_precos_por_fuel_brasil(sheets_anp)
        nivel_anp  = "Brasil"

    resultado = []
    for pk, grp in df_pf.groupby("combustivel_pk"):
        # ── Resolução de PK: PP → ANP ──────────────────────────────
        # 1) Tenta match direto no dicionário ANP
        preco_anp = precos_anp.get(pk)

        # 2) Usa tabela de mapeamento explícita (cobre "DIESEL COMUM" → "OLEO DIESEL" etc.)
        if preco_anp is None:
            anp_canonical = _PP_PARA_ANP_PK.get(pk)
            if anp_canonical:
                preco_anp = precos_anp.get(anp_canonical)

        # 3) Fallback: substring bidirecional (último recurso)
        if preco_anp is None:
            for anp_pk, anp_p in precos_anp.items():
                if pk in anp_pk or anp_pk in pk:
                    preco_anp = anp_p
                    break

        if preco_anp is None:
            continue

        preco_pf_med = grp["preco"].mean()
        preco_pf_min = grp["preco"].min()
        preco_pf_max = grp["preco"].max()
        n_postos     = int(grp["cnpj_norm"].nunique())
        # Label: usa PRODUTO_CURTO com o pk direto, depois tenta o canonical ANP, depois raw
        anp_can_lbl = _PP_PARA_ANP_PK.get(pk, pk)
        label = (PRODUTO_CURTO.get(pk)
                 or PRODUTO_CURTO.get(anp_can_lbl)
                 or grp["combustivel_label"].iloc[0].title())
        data_atz     = ""
        datas = grp["data_atualizacao"].dropna()
        if not datas.empty:
            data_atz = datas.iloc[0]

        delta_abs    = preco_pf_med - preco_anp
        delta_pct    = (delta_abs / preco_anp) * 100 if preco_anp else 0
        economia_100 = (preco_anp - preco_pf_med) * 100   # economia em 100 L

        # Detalhamento por estado (só quando há múltiplos UFs)
        por_uf = []
        if ufs and len(ufs) > 1:
            _pp_uf = _anp_precos_por_fuel_por_uf(sheets_anp, ufs)
            # Usa PK canonico ANP para buscar no dicionário por-UF
            _pk_anp = _PP_PARA_ANP_PK.get(pk, pk)
            for uf_i in ufs:
                p_uf = (_pp_uf.get(pk, {}).get(uf_i)
                        or _pp_uf.get(_pk_anp, {}).get(uf_i))
                if p_uf is not None:
                    d_abs_i = preco_pf_med - p_uf
                    d_pct_i = (d_abs_i / p_uf) * 100 if p_uf else 0
                    por_uf.append({
                        "uf":        uf_i,
                        "nome":      UF_NOME.get(uf_i, uf_i),
                        "preco_anp": round(p_uf, 3),
                        "delta_abs": round(d_abs_i, 3),
                        "delta_pct": round(d_pct_i, 1),
                    })

        resultado.append({
            "combustivel_label": label,
            "combustivel_pk":    pk,
            "preco_pf_med":      round(preco_pf_med, 3),
            "preco_pf_min":      round(preco_pf_min, 3),
            "preco_pf_max":      round(preco_pf_max, 3),
            "n_postos_pf":       n_postos,
            "preco_anp":         round(preco_anp, 3),
            "nivel_anp":         nivel_anp,
            "delta_abs":         round(delta_abs, 3),
            "delta_pct":         round(delta_pct, 1),
            "economia_100l":     round(economia_100, 2),
            "data_atualizacao":  data_atz,
            "por_uf":            por_uf,
        })

    resultado.sort(key=lambda x: x["combustivel_label"])
    return resultado


def _renderizar_comparativo_pf_anp(comparativo, subtitulo=""):
    """Renderiza cards de comparativo Gestão de Frotas vs ANP — visual aprimorado."""
    if not comparativo:
        st.info("ℹ️ Sem dados suficientes para o comparativo. "
                "Carregue a planilha **Preço Posto** e a planilha **ANP** em ⚙️ Configurações.")
        return

    # ── Cabeçalho ─────────────────────────────────────────────────────────
    st.markdown(
        f"<div style='background:linear-gradient(135deg,#0d1b4b 0%,#1565c0 100%);"
        f"border-radius:14px;padding:18px 24px 14px;margin-bottom:20px'>"
        f"<div style='color:#fff;font-size:18px;font-weight:800;letter-spacing:.3px'>"
        f"📊 Preços Gestão de Frotas vs ANP</div>"
        f"<div style='color:rgba(255,255,255,.8);font-size:13px;margin-top:4px'>"
        f"{subtitulo or 'Comparativo por combustível — preço médio dos postos credenciados'}"
        f"</div>"
        f"</div>",
        unsafe_allow_html=True,
    )

    # ── CSS + grid de cards em um único bloco HTML ────────────────────────
    # Emitido como bloco único para que CSS grid controle altura uniforme
    # (align-items:stretch + flex-direction:column por card)

    css = """
<style>
.cmp-grid{display:grid;grid-template-columns:repeat(3,1fr);
          gap:14px;margin-bottom:20px;align-items:stretch}
.cmp-card{border-radius:14px;overflow:hidden;
          box-shadow:0 3px 10px rgba(0,0,0,.1);
          display:flex;flex-direction:column}
.cmp-head{padding:10px 16px;display:flex;
          justify-content:space-between;align-items:center}
.cmp-head-nome{color:#fff;font-weight:800;font-size:15px}
.cmp-head-nivel{color:rgba(255,255,255,.85);font-size:11px}
.cmp-body{padding:14px 16px;display:flex;flex-direction:column;flex:1}
.cmp-prices{display:flex;justify-content:space-between;
            align-items:baseline;margin-bottom:6px}
.cmp-lbl{font-size:10px;color:#888;margin-bottom:2px}
.cmp-pf{font-size:24px;font-weight:900;color:#1565c0;line-height:1}
.cmp-anp{font-size:18px;font-weight:700;color:#555;line-height:1}
.cmp-delta{border-radius:10px;padding:10px 14px;margin-top:10px;
           display:flex;justify-content:space-between;align-items:center}
.cmp-pct{font-size:28px;font-weight:900;line-height:1}
.cmp-abs{font-size:16px;font-weight:800}
.cmp-eco{font-size:11px;margin-top:2px}
.cmp-minmax{display:flex;justify-content:space-between;
            font-size:10px;color:#999;margin-top:8px}
.cmp-uf-wrap{margin-top:12px;border-top:1px solid #e0e0e0;
             padding-top:10px;flex:1}
.cmp-uf-titulo{font-size:10px;font-weight:700;color:#888;
               letter-spacing:.8px;margin-bottom:6px}
.cmp-uf-table{width:100%;font-size:11px;border-collapse:collapse}
.cmp-uf-table th{text-align:left;padding:2px 6px;font-weight:600;color:#999}
.cmp-uf-table th:not(:first-child){text-align:right}
.cmp-uf-table td{padding:4px 6px;color:#333;border-top:1px solid #f0f0f0}
.cmp-uf-table td:not(:first-child){text-align:right}
.cmp-footer{font-size:10px;color:#aaa;text-align:right;margin-top:auto;
            padding-top:8px}
@media(max-width:720px){.cmp-grid{grid-template-columns:1fr}}
</style>
"""

    # ── Preços ANP semana anterior (para tendência nos cards) ────────
    _ant_cache_cmp  = st.session_state.get("_precos_anp_cache_anterior", {})
    _ant_sheets_cmp = _ant_cache_cmp.get("sheets", {})
    _ant_semana_cmp = _ant_cache_cmp.get("semana", "")

    cards_html = ""
    for item in comparativo:
        cheaper   = item["delta_abs"] < 0
        cor_brd   = "#2e7d32" if cheaper else "#c62828"
        cor_bg    = "#e8f5e9" if cheaper else "#ffebee"
        cor_delta = "#1b5e20" if cheaper else "#b71c1c"
        sinal     = "▼" if cheaper else "▲"
        icone_eco = "💚" if cheaper else "🔴"
        txt_eco   = (f"Economia de R$ {_brl(abs(item['economia_100l']))}/100 L"
                     if cheaper
                     else f"Custo adicional de R$ {_brl(abs(item['economia_100l']))}/100 L")

        # Tendência ANP semana anterior
        _pk_cmp     = _anp_norm(item.get("combustivel_pk", item.get("combustivel_label", "")))
        _anp_ant_v  = None
        if _ant_sheets_cmp:
            _ufs_item = list({pu["uf"] for pu in item.get("por_uf", []) if "uf" in pu})
            if _ufs_item:
                _vals_ant = [_anp_preco_uf(_ant_sheets_cmp, _pk_cmp, u) for u in _ufs_item]
                _vals_ant = [v for v in _vals_ant if v is not None]
                _anp_ant_v = sum(_vals_ant) / len(_vals_ant) if _vals_ant else None
            if _anp_ant_v is None:
                _anp_ant_v = _anp_preco_brasil(_ant_sheets_cmp, _pk_cmp)
        _tend_cmp = _tendencia_badge(item["preco_anp"], _anp_ant_v, inline=True) if _anp_ant_v else ""

        # Tabela por estado (só rota com múltiplos UFs)
        por_uf = item.get("por_uf", [])
        if por_uf:
            linhas_uf = "".join(
                f"<tr>"
                f"<td>{pu['nome']}</td>"
                f"<td>R$ {_brl(pu['preco_anp'], 3)}</td>"
                f"<td style='font-weight:700;color:{'#1b5e20' if pu['delta_abs']<0 else '#b71c1c'}'>"
                f"{'▼' if pu['delta_abs']<0 else '▲'} {abs(pu['delta_pct']):.1f}%</td>"
                f"</tr>"
                for pu in por_uf
            )
            tabela_uf_html = (
                "<div class='cmp-uf-wrap'>"
                "<div class='cmp-uf-titulo'>PREÇO ANP POR ESTADO</div>"
                "<table class='cmp-uf-table'>"
                "<thead><tr><th>Estado</th><th>ANP</th><th>vs GF</th></tr></thead>"
                f"<tbody>{linhas_uf}</tbody>"
                "</table></div>"
            )
        else:
            tabela_uf_html = ""

        n_p   = item['n_postos_pf']
        data_ = item['data_atualizacao']
        footer = (
            f"<div class='cmp-footer'>"
            f"{n_p} posto{'s' if n_p != 1 else ''} GF"
            f"{' · atualizado ' + data_ if data_ else ''}"
            f"</div>"
        )

        cards_html += (
            f"<div class='cmp-card' style='border:2px solid {cor_brd}'>"
            f"<div class='cmp-head' style='background:{cor_brd}'>"
            f"<span class='cmp-head-nome'>⛽ {item['combustivel_label']}</span>"
            f"<span class='cmp-head-nivel'>ref. ANP: {item['nivel_anp']}</span>"
            f"</div>"
            f"<div class='cmp-body'>"
            f"<div class='cmp-prices'>"
            f"<div><div class='cmp-lbl'>⭐ Gestão de Frotas médio</div>"
            f"<div class='cmp-pf'>R$ {_brl(item['preco_pf_med'], 3)}</div></div>"
            f"<div style='text-align:right'><div class='cmp-lbl'>📊 Ref. ANP</div>"
            f"<div style='display:flex;align-items:baseline;gap:6px;justify-content:flex-end'>"
            f"<span class='cmp-anp'>R$ {_brl(item['preco_anp'], 3)}</span>"
            f"{_tend_cmp}"
            f"</div></div>"
            f"</div>"
            f"<div class='cmp-delta' style='background:{cor_bg}'>"
            f"<div class='cmp-pct' style='color:{cor_delta}'>{sinal} {abs(item['delta_pct']):.1f}%</div>"
            f"<div style='text-align:right'>"
            f"<div class='cmp-abs' style='color:{cor_delta}'>{sinal} R$ {_brl(abs(item['delta_abs']),3)}/L</div>"
            f"<div class='cmp-eco' style='color:{cor_delta}'>{icone_eco} {txt_eco}</div>"
            f"</div></div>"
            f"<div class='cmp-minmax'>"
            f"<span>PF mín: R$ {_brl(item['preco_pf_min'], 3)}</span>"
            f"<span>PF máx: R$ {_brl(item['preco_pf_max'], 3)}</span>"
            f"</div>"
            f"{tabela_uf_html}"
            f"{footer}"
            f"</div>"  # end body
            f"</div>"  # end card
        )

    st.markdown(
        css + f"<div class='cmp-grid'>{cards_html}</div>",
        unsafe_allow_html=True,
    )


def _renderizar_precos_anp(uf, municipio=None, ufs_multiplas=None):
    """Renderiza aba de preços ANP.

    Modo 1 (sem rota): indicadores por Município/Capital/Estado + referências Região e Brasil
    Modo 2 (com rota): tabela pivot Estado × Combustível + referências regional e nacional

    Fluxo de carregamento:
      1. Se já há dados em session_state → exibe direto
      2. Se não há dados → tenta auto-fetch (uma vez por sessão)
      3. Se auto-fetch falha ou nunca tentado → exibe widget de upload limpo
    """
    _cache = st.session_state.get("_precos_anp_cache", {})
    sheets = _cache.get("sheets")
    semana = _cache.get("semana")

    # ── 1. Se não há dados → UI de upload amigável ─────────────────────
    if sheets is None:
        st.markdown(
            "<div style='font-size:15px;font-weight:700;color:#1565c0;margin-bottom:14px'>"
            "💰 Preços Médios ANP</div>",
            unsafe_allow_html=True,
        )

        # Card de instrução — link + passos visuais
        st.markdown(
            "<div style='background:#e3f2fd;border-radius:12px;padding:16px 20px;"
            "border:1px solid #90caf9;margin-bottom:16px'>"
            "<div style='font-size:14px;font-weight:700;color:#1565c0;margin-bottom:10px'>"
            "📥 Como carregar os preços de referência</div>"
            "<div style='display:flex;flex-direction:column;gap:8px'>"
            # Passo 1
            "<div style='display:flex;align-items:flex-start;gap:10px'>"
            "<div style='min-width:24px;height:24px;background:#1565c0;color:#fff;"
            "border-radius:50%;font-size:12px;font-weight:700;"
            "display:flex;align-items:center;justify-content:center'>1</div>"
            "<div style='font-size:13px;color:#1a1a1a;line-height:1.5'>"
            "Acesse o site da ANP: "
            "<a href='https://www.gov.br/anp/pt-br/assuntos/precos-e-defesa-da-concorrencia"
            "/precos/levantamento-de-precos-de-combustiveis-ultimas-semanas-pesquisadas' "
            "target='_blank' style='color:#1565c0;font-weight:600'>"
            "gov.br/anp → Levantamento de Preços</a>"
            "</div></div>"
            # Passo 2
            "<div style='display:flex;align-items:flex-start;gap:10px'>"
            "<div style='min-width:24px;height:24px;background:#1565c0;color:#fff;"
            "border-radius:50%;font-size:12px;font-weight:700;"
            "display:flex;align-items:center;justify-content:center'>2</div>"
            "<div style='font-size:13px;color:#1a1a1a;line-height:1.5'>"
            "Baixe o arquivo <b>resumo_semanal_lpc_*.xlsx</b> da última semana disponível."
            "</div></div>"
            # Passo 3
            "<div style='display:flex;align-items:flex-start;gap:10px'>"
            "<div style='min-width:24px;height:24px;background:#1565c0;color:#fff;"
            "border-radius:50%;font-size:12px;font-weight:700;"
            "display:flex;align-items:center;justify-content:center'>3</div>"
            "<div style='font-size:13px;color:#1a1a1a;line-height:1.5'>"
            "Faça o upload do arquivo aqui embaixo — ele fica disponível durante toda a sessão."
            "</div></div>"
            "</div></div>",
            unsafe_allow_html=True,
        )

        arq = st.file_uploader(
            "📎 Selecione o arquivo baixado (resumo_semanal_lpc_*.xlsx)",
            type=["xlsx", "xls"],
            key="upload_precos_anp",
            help="Planilha da ANP com preços médios semanais de combustíveis por município/estado",
        )
        if arq:
            with st.spinner("🔍 Processando planilha…"):
                try:
                    _sheets = _anp_processar_arquivo(io.BytesIO(arq.read()))
                    if not _sheets:
                        st.error("❌ Nenhuma aba reconhecida. Verifique se é a planilha correta da ANP.")
                    else:
                        _sem = arq.name.replace(".xlsx", "").replace(".xls", "")
                        # Rotaciona: cache atual → anterior (para tendência de preço)
                        _cache_atual = st.session_state.get("_precos_anp_cache", {})
                        if _cache_atual.get("sheets") and _cache_atual.get("semana") != _sem:
                            st.session_state["_precos_anp_cache_anterior"] = _cache_atual
                        st.session_state["_precos_anp_cache"] = {"sheets": _sheets, "semana": _sem}
                        st.rerun()
                except Exception as ex:
                    st.error(f"❌ Erro ao ler arquivo: {ex}")
        return

    # ── 2. Dados disponíveis ── cabeçalho compacto com opção de substituir ──
    with st.expander(f"✅ Planilha carregada: **{semana}** · Clique para trocar", expanded=False):
        st.markdown(
            "<div style='font-size:12px;color:#555;margin-bottom:8px'>"
            "Para atualizar, baixe a planilha mais recente em "
            "<a href='https://www.gov.br/anp/pt-br/assuntos/precos-e-defesa-da-concorrencia"
            "/precos/levantamento-de-precos-de-combustiveis-ultimas-semanas-pesquisadas' "
            "target='_blank'>gov.br/anp → Levantamento de Preços</a> "
            "e faça upload abaixo.</div>",
            unsafe_allow_html=True,
        )
        arq2 = st.file_uploader(
            "📎 Substituir por arquivo mais recente",
            type=["xlsx", "xls"],
            key="upload_precos_anp_sub",
        )
        if arq2:
            try:
                _sh2 = _anp_processar_arquivo(io.BytesIO(arq2.read()))
                if _sh2:
                    _sem2 = arq2.name.replace(".xlsx", "")
                    _cache_atual2 = st.session_state.get("_precos_anp_cache", {})
                    if _cache_atual2.get("sheets") and _cache_atual2.get("semana") != _sem2:
                        st.session_state["_precos_anp_cache_anterior"] = _cache_atual2
                    st.session_state["_precos_anp_cache"] = {"sheets": _sh2, "semana": _sem2}
                    st.rerun()
                else:
                    st.error("❌ Arquivo não reconhecido. Verifique se é a planilha correta da ANP.")
            except Exception as _ex2:
                st.error(f"❌ Erro ao ler arquivo: {_ex2}")

    st.divider()

    # ══════════════════════════════════════════════════════════════
    # MODO 2 — Rota: cards por combustível + referências
    # ══════════════════════════════════════════════════════════════
    if ufs_multiplas:

        if "estados" not in sheets:
            st.warning("Aba de estados não encontrada na planilha.")
            return

        linhas_rota = []
        for uf_r in ufs_multiplas:
            for r in _anp_extrair_precos(sheets, uf=uf_r):
                linhas_rota.append({"UF": uf_r, "Estado": UF_NOME.get(uf_r, uf_r), **r})

        if not linhas_rota:
            st.info("Preços não encontrados para os estados desta rota na planilha carregada.")
            return

        df_rota = pd.DataFrame(linhas_rota)

        # ══════════════════════════════════════════════════════════
        # CALCULADORA DE CUSTO DA ROTA
        # ══════════════════════════════════════════════════════════
        label_orig = st.session_state.get("label_orig", "")
        label_dest = st.session_state.get("label_dest", "")
        dist_km    = st.session_state.get("dist_km", 0)

        # Combustíveis disponíveis na planilha (ordenados)
        comb_disponiveis = df_rota.drop_duplicates("Combustível")[["Combustível","_pk"]].values.tolist()
        comb_nomes = [c[0] for c in comb_disponiveis]
        comb_pks   = {c[0]: c[1] for c in comb_disponiveis}

        # Filtra combustíveis relevantes para frota (não GLP)
        comb_frota = [c for c in comb_nomes if "GLP" not in c.upper()]

        st.markdown(
            "<div style='background:linear-gradient(135deg,#0d1b4b 0%,#1565c0 100%);"
            "border-radius:14px;padding:20px 24px 16px;margin-bottom:20px'>"
            "<div style='color:#fff;font-size:18px;font-weight:700;margin-bottom:4px'>"
            "⛽ Calculadora de Custo da Rota</div>"
            f"<div style='color:rgba(255,255,255,.75);font-size:13px'>"
            f"{label_orig or '—'}  →  {label_dest or '—'}"
            f"{'  ·  ' + str(round(dist_km)) + ' km' if dist_km else ''}</div>"
            "</div>",
            unsafe_allow_html=True,
        )

        col_comb, col_cons = st.columns([2, 1])
        with col_comb:
            comb_sel = st.selectbox(
                "Combustível", comb_frota or comb_nomes,
                key="calc_comb_sel",
                label_visibility="visible",
            )
        with col_cons:
            consumo = st.number_input(
                "Consumo (km/L)", min_value=1.0, max_value=40.0,
                value=12.0, step=0.5, key="calc_consumo",
                help="Consumo médio do veículo em km por litro",
            )

        if comb_sel and dist_km and consumo:
            pk_sel = comb_pks.get(comb_sel, _anp_norm(comb_sel))

            # Preço origem
            p_orig,  niv_orig,  desc_orig  = _anp_preco_ponto(sheets, label_orig, pk_sel)
            # Preço destino
            p_dest,  niv_dest,  desc_dest  = _anp_preco_ponto(sheets, label_dest, pk_sel)
            # Preços por estado da rota
            precos_rota = {}
            for uf_r in ufs_multiplas:
                rows_u = _anp_extrair_precos(sheets, uf=uf_r)
                for r in rows_u:
                    if r["_pk"] == pk_sel:
                        precos_rota[uf_r] = r["Preço Médio"]
            p_med_rota = round(sum(precos_rota.values()) / len(precos_rota), 3) if precos_rota else None
            p_min_rota = min(precos_rota.values()) if precos_rota else None
            uf_min_rota = min(precos_rota, key=precos_rota.get) if precos_rota else None

            litros = dist_km / consumo

            def _custo(p):
                return round(p * litros, 2) if p else None

            custo_orig    = _custo(p_orig)
            custo_dest    = _custo(p_dest)
            custo_med     = _custo(p_med_rota)
            custo_min     = _custo(p_min_rota)
            custo_med_val = custo_med   # garante que a variável existe mesmo quando o card de média não é renderizado

            # Montagem do HTML dos cards de custo
            def _card_custo(titulo, subtitulo, preco, custo, nivel, cor_header="#1565c0", destaque=False):
                if preco is None:
                    return (f"<div class='cc-card'>"
                            f"<div class='cc-head' style='background:{cor_header}'>"
                            f"<div class='cc-titulo'>{titulo}</div>"
                            f"<div class='cc-sub'>{subtitulo}</div></div>"
                            f"<div class='cc-body'><div class='cc-nd'>Preço não disponível</div></div></div>")
                econ = ""
                if custo_orig and custo and custo < custo_orig:
                    econ_val = custo_orig - custo
                    econ = (f"<div class='cc-econ'>💚 Economia de "
                            f"<b>R$ {_brl(econ_val)}</b> vs origem</div>")
                badge = "<span class='cc-best'>✦ MAIS BARATO</span>" if destaque else ""
                return (
                    f"<div class='cc-card{' cc-best-card' if destaque else ''}'>"
                    f"<div class='cc-head' style='background:{cor_header}'>"
                    f"<div class='cc-titulo'>{titulo}{badge}</div>"
                    f"<div class='cc-sub'>{subtitulo} · {nivel}</div></div>"
                    f"<div class='cc-body'>"
                    f"<div class='cc-preco-label'>Preço médio</div>"
                    f"<div class='cc-preco'>R$ {_brl(preco, 3)}<span class='cc-unidade'>/L</span></div>"
                    f"<div class='cc-litros'>{litros:.1f} L necessários</div>"
                    f"<div class='cc-custo-label'>Custo estimado</div>"
                    f"<div class='cc-custo'>R$ {_brl(custo)}</div>"
                    f"{econ}"
                    f"</div></div>"
                )

            # Determina qual é o mais barato
            opcoes_preco = {k: v for k, v in {
                "orig": p_orig, "dest": p_dest, "min": p_min_rota
            }.items() if v is not None}
            mais_barato = min(opcoes_preco, key=opcoes_preco.get) if opcoes_preco else None

            cards_html = (
                _card_custo(
                    f"📍 {label_orig or 'Origem'}",
                    desc_orig, p_orig, custo_orig, niv_orig,
                    cor_header="#1565c0",
                    destaque=(mais_barato == "orig"),
                ) +
                _card_custo(
                    f"🏁 {label_dest or 'Destino'}",
                    desc_dest, p_dest, custo_dest, niv_dest,
                    cor_header="#37474f",
                    destaque=(mais_barato == "dest"),
                ) +
                _card_custo(
                    f"🟢 Menor preço ({uf_min_rota or '—'})",
                    UF_NOME.get(uf_min_rota or "", uf_min_rota or "—"),
                    p_min_rota, custo_min, "Estado",
                    cor_header="#2e7d32",
                    destaque=(mais_barato == "min"),
                )
            )

            # Adiciona card de preço médio da rota se diferente do mínimo
            if p_med_rota and p_med_rota != p_min_rota:
                custo_med_val = _custo(p_med_rota)
                cards_html += (
                    f"<div class='cc-card'>"
                    f"<div class='cc-head' style='background:#4527a0'>"
                    f"<div class='cc-titulo'>📊 Média da Rota</div>"
                    f"<div class='cc-sub'>{len(precos_rota)} estados</div></div>"
                    f"<div class='cc-body'>"
                    f"<div class='cc-preco-label'>Preço médio</div>"
                    f"<div class='cc-preco'>R$ {_brl(p_med_rota, 3)}<span class='cc-unidade'>/L</span></div>"
                    f"<div class='cc-litros'>{litros:.1f} L necessários</div>"
                    f"<div class='cc-custo-label'>Custo estimado</div>"
                    f"<div class='cc-custo'>R$ {_brl(custo_med_val)}</div>"
                    f"</div></div>"
                )

            # ── Card Gestão de Frotas (Preço Posto real) ───────────────────
            _pp_df_calc    = st.session_state.get("_pp_df")
            _cnpjs_pf_calc = st.session_state.get("cnpjs_pro_frotas", set())
            if _pp_df_calc is not None and _cnpjs_pf_calc:
                _df_pf_calc = _pp_df_calc[_pp_df_calc["cnpj_norm"].isin(_cnpjs_pf_calc)]
                # Filtra pelo pk do combustível selecionado (match exato ou parcial)
                _pk_mask = _df_pf_calc["combustivel_pk"].apply(
                    lambda x: x == pk_sel or pk_sel in x or x in pk_sel)
                _df_pf_fuel = _df_pf_calc[_pk_mask]
                if not _df_pf_fuel.empty:
                    _p_pf_real   = _df_pf_fuel["preco"].mean()
                    _n_pf_real   = _df_pf_fuel["cnpj_norm"].nunique()
                    _custo_pf    = _custo(_p_pf_real)
                    _data_pf     = _df_pf_fuel["data_atualizacao"].dropna().iloc[0] \
                                   if not _df_pf_fuel["data_atualizacao"].dropna().empty else ""
                    # Delta vs ANP médio da rota
                    _delta_pf    = _p_pf_real - p_med_rota if p_med_rota else None
                    _delta_pct   = (_delta_pf / p_med_rota * 100) if p_med_rota and _delta_pf is not None else None
                    _delta_html  = ""
                    if _delta_pf is not None:
                        _d_cor   = "#2e7d32" if _delta_pf < 0 else "#c62828"
                        _d_sinal = "▼" if _delta_pf < 0 else "▲"
                        _d_txt   = "abaixo" if _delta_pf < 0 else "acima"
                        _delta_html = (
                            f"<div style='font-size:10px;color:{_d_cor};"
                            f"background:{'#e8f5e9' if _delta_pf<0 else '#ffebee'};"
                            f"border-radius:4px;padding:3px 6px;margin-top:6px;text-align:center'>"
                            f"{_d_sinal} R$ {_brl(abs(_delta_pf), 3)} "
                            f"({abs(_delta_pct):.1f}%) {_d_txt} da média ANP</div>"
                        )
                    _pf_dest = (_custo_pf is not None and
                                all(v is None or _custo_pf <= v
                                    for v in [custo_orig, custo_dest, custo_min, custo_med_val]))
                    _pf_best_cls   = " cc-best-card" if _pf_dest else ""
                    _pf_best_badge = "<span class='cc-best'>✦ MAIS BARATO</span>" if _pf_dest else ""
                    _pf_data_html  = f" · {_data_pf}" if _data_pf else ""
                    cards_html += (
                        f"<div class='cc-card{_pf_best_cls}'>"
                        f"<div class='cc-head' style='background:#0d47a1'>"
                        f"<div class='cc-titulo'>⭐ Gestão de Frotas{_pf_best_badge}</div>"
                        f"<div class='cc-sub'>{_n_pf_real} postos{_pf_data_html}</div></div>"
                        f"<div class='cc-body'>"
                        f"<div class='cc-preco-label'>Preço médio real</div>"
                        f"<div class='cc-preco'>R$ {_brl(_p_pf_real, 3)}"
                        f"<span class='cc-unidade'>/L</span></div>"
                        f"<div class='cc-litros'>{litros:.1f} L necessários</div>"
                        f"<div class='cc-custo-label'>Custo estimado</div>"
                        f"<div class='cc-custo'>R$ {_brl(_custo_pf)}</div>"
                        f"{_delta_html}"
                        f"</div></div>"
                    )

            st.markdown(f"""
<style>
.cc-grid{{display:grid;grid-template-columns:repeat(auto-fill,minmax(200px,1fr));
          gap:12px;margin-bottom:24px}}
.cc-card{{border-radius:12px;overflow:hidden;
          box-shadow:0 2px 10px rgba(0,0,0,.12);background:#fff}}
.cc-best-card{{box-shadow:0 4px 16px rgba(46,125,50,.35);
               outline:2px solid #2e7d32}}
.cc-head{{padding:12px 16px 10px;color:#fff}}
.cc-titulo{{font-size:13px;font-weight:700;line-height:1.3;
            display:flex;align-items:center;gap:6px;flex-wrap:wrap}}
.cc-sub{{font-size:11px;opacity:.8;margin-top:3px}}
.cc-best{{background:rgba(255,255,255,.25);font-size:9px;font-weight:800;
          padding:2px 6px;border-radius:8px;letter-spacing:.4px}}
.cc-body{{padding:14px 16px 12px}}
.cc-preco-label,.cc-custo-label{{font-size:10px;color:#888;
                                  text-transform:uppercase;letter-spacing:.5px}}
.cc-preco{{font-size:22px;font-weight:800;color:#0d1b4b;margin:2px 0 4px}}
.cc-unidade{{font-size:12px;font-weight:400;color:#666}}
.cc-litros{{font-size:11px;color:#888;margin-bottom:10px}}
.cc-custo{{font-size:20px;font-weight:800;color:#1565c0}}
.cc-nd{{font-size:13px;color:#999;padding:20px 0;text-align:center}}
.cc-econ{{font-size:11px;color:#2e7d32;margin-top:6px;
          background:#e8f5e9;padding:4px 8px;border-radius:6px}}
</style>
<div class='cc-grid'>{cards_html}</div>
""", unsafe_allow_html=True)

            # Nota de rodapé
            st.caption(
                f"*Cálculo baseado em {dist_km:.0f} km ÷ {consumo:.1f} km/L = {litros:.1f} litros."
                f" Preços: ANP semana {semana or '—'}. Valores estimados.*"
            )

        st.divider()

        # ── Cabeçalho visual ──────────────────────────────────────
        ufs_ord = list(dict.fromkeys(df_rota["UF"].tolist()))   # ordem original da rota
        st.markdown(
            "<h3 style='margin:0 0 4px 0;font-size:20px;color:#0d1b4b'>💰 Preços Médios por Estado</h3>"
            f"<p style='margin:0 0 18px 0;font-size:13px;color:#555'>Rota: "
            f"{'  →  '.join(f'<b>{u}</b>' for u in ufs_ord)}"
            f" &nbsp;·&nbsp; Semana: <b>{semana or '—'}</b></p>",
            unsafe_allow_html=True,
        )

        # ── Cards de combustível × estado ─────────────────────────
        combustiveis = df_rota["Combustível"].unique().tolist()
        ordem_comb = {_anp_norm(k): i for i, k in enumerate(PRODUTOS_CHAVE)}
        combustiveis.sort(key=lambda c: ordem_comb.get(_anp_norm(c), 99))

        # Preços semana anterior (por UF, para tendência)
        _ant_cache_pc  = st.session_state.get("_precos_anp_cache_anterior", {})
        _ant_sheets_pc = _ant_cache_pc.get("sheets", {})

        for comb in combustiveis:
            df_c = df_rota[df_rota["Combustível"] == comb]
            precos = {row["UF"]: row["Preço Médio"] for _, row in df_c.iterrows()}
            if not precos: continue

            _pk_pc   = _anp_norm(comb)
            p_vals   = list(precos.values())
            p_min    = min(p_vals)
            p_max    = max(p_vals)
            unidade  = df_c["Unidade"].iloc[0] if not df_c.empty else "R$/L"

            # Linha HTML por combustível
            cells = ""
            for uf_r in ufs_ord:
                preco = precos.get(uf_r)
                if preco is None:
                    cells += f"<div class='pc-cell pc-na'><span class='pc-uf'>{uf_r}</span><span class='pc-val'>—</span></div>"
                    continue
                diff = preco - p_min
                pct  = (diff / (p_max - p_min) * 100) if p_max > p_min else 0
                # Cor: verde (min) → amarelo → vermelho (max)
                if pct < 1:
                    bg, txt, badge = "#e8f5e9", "#1b5e20", "MIN"
                elif pct > 98:
                    bg, txt, badge = "#ffebee", "#b71c1c", "MAX"
                else:
                    bg, txt, badge = "#fff8e1", "#4e342e", ""
                badge_html = f"<span class='pc-badge' style='background:{txt};color:#fff'>{badge}</span>" if badge else ""
                # Tendência: preço ANP desta semana vs semana anterior para o mesmo UF
                _ant_p_pc = None
                if _ant_sheets_pc:
                    _ant_p_pc = _anp_preco_uf(_ant_sheets_pc, _pk_pc, uf_r)
                    if _ant_p_pc is None:
                        _ant_p_pc = _anp_preco_brasil(_ant_sheets_pc, _pk_pc)
                _tend_pc = _tendencia_badge(preco, _ant_p_pc, inline=True) if _ant_p_pc else ""
                cells += (
                    f"<div class='pc-cell' style='background:{bg}'>"
                    f"<span class='pc-uf'>{uf_r}</span>"
                    f"<span class='pc-val' style='color:{txt}'>R$ {_brl(preco, 3)}</span>"
                    f"<span class='pc-uni'>{unidade}</span>"
                    f"{_tend_pc}"
                    f"{badge_html}</div>"
                )

            st.markdown(f"""
<div class='pc-row'>
  <div class='pc-label'>{comb}</div>
  <div class='pc-cells'>{cells}</div>
</div>""", unsafe_allow_html=True)

        # CSS injetado uma vez
        st.markdown("""
<style>
.pc-row{display:flex;align-items:stretch;margin-bottom:8px;border-radius:10px;
        overflow:hidden;box-shadow:0 1px 4px rgba(0,0,0,.08)}
.pc-label{min-width:160px;max-width:160px;background:#0d1b4b;color:#fff;
          display:flex;align-items:center;padding:10px 14px;
          font-size:13px;font-weight:600;line-height:1.3}
.pc-cells{display:flex;flex:1;flex-wrap:wrap}
.pc-cell{flex:1;min-width:110px;display:flex;flex-direction:column;
         align-items:center;justify-content:center;padding:10px 8px;
         border-right:1px solid rgba(0,0,0,.06);gap:2px}
.pc-cell:last-child{border-right:none}
.pc-na{background:#f5f5f5}
.pc-uf{font-size:11px;font-weight:700;color:#555;letter-spacing:.5px;text-transform:uppercase}
.pc-val{font-size:15px;font-weight:700}
.pc-uni{font-size:10px;color:#888}
.pc-badge{font-size:9px;font-weight:700;padding:1px 5px;border-radius:10px;
          margin-top:2px;letter-spacing:.3px}
</style>""", unsafe_allow_html=True)

        # ── Referências (Região + Brasil) ─────────────────────────
        regioes_rota = [r for r in df_rota["Nome Região"].dropna().unique() if r]
        ref_rows: dict = {}   # pk → {brasil, regioes...}

        if "brasil" in sheets:
            df_b  = sheets["brasil"]
            b_p   = _anp_col(df_b, "produto")
            b_m   = _anp_col(df_b, "medio revenda", "media revenda", "preco medio")
            b_u   = _anp_col(df_b, "unidade")
            b_n   = _anp_col(df_b, "postos pesq", "numero de postos", "n postos")
            if b_p and b_m:
                for r in _anp_preco_medio(df_b, None, None, b_p, b_m, b_u, b_n):
                    ref_rows.setdefault(r["Combustível"], {})["🌎 Brasil"] = r["Preço Médio"]

        if regioes_rota and "regioes" in sheets:
            df_reg = sheets["regioes"]
            rg_r   = _anp_col(df_reg, "regiao")
            rg_p   = _anp_col(df_reg, "produto")
            rg_m   = _anp_col(df_reg, "medio revenda", "media revenda", "preco medio")
            rg_u   = _anp_col(df_reg, "unidade")
            rg_n   = _anp_col(df_reg, "postos pesq", "numero de postos", "n postos")
            if rg_r and rg_p and rg_m:
                for reg in regioes_rota:
                    for r in _anp_preco_medio(df_reg, rg_r, reg, rg_p, rg_m, rg_u, rg_n):
                        ref_rows.setdefault(r["Combustível"], {})[f"📍 {reg.title()}"] = r["Preço Médio"]

        if ref_rows:
            st.markdown(
                "<div style='margin:24px 0 10px;font-size:14px;font-weight:700;"
                "color:#0d1b4b;border-left:4px solid #1565c0;padding-left:10px'>"
                "Referências: Região e Brasil</div>",
                unsafe_allow_html=True,
            )
            # Constrói tabela HTML de referência
            combust_keys = [c for c in [row["Combustível"] for row in linhas_rota] if c in ref_rows]
            seen = set(); combust_keys = [x for x in combust_keys if not (x in seen or seen.add(x))]
            ref_cols = []
            for v in ref_rows.values():
                for k in v: 
                    if k not in ref_cols: ref_cols.append(k)

            header_cells = "".join(f"<th>{c}</th>" for c in ref_cols)
            rows_html = ""
            for comb in combust_keys:
                vals = ref_rows.get(comb, {})
                row_cells = ""
                for col in ref_cols:
                    v = vals.get(col)
                    row_cells += f"<td>{'R$ '+_brl(v,3) if v else '—'}</td>"
                rows_html += f"<tr><td class='rt-comb'>{comb}</td>{row_cells}</tr>"

            st.markdown(f"""
<style>
.ref-table{{width:100%;border-collapse:collapse;font-size:13px;margin-bottom:16px}}
.ref-table th{{background:#1565c0;color:#fff;padding:8px 12px;text-align:right;font-weight:600}}
.ref-table th:first-child{{text-align:left}}
.ref-table td{{padding:7px 12px;border-bottom:1px solid #e8eaf6;text-align:right;color:#333}}
.ref-table tr:hover td{{background:#f3f4ff}}
.ref-table .rt-comb{{text-align:left;font-weight:600;color:#0d1b4b}}
.ref-table tr:last-child td{{border-bottom:none}}
</style>
<table class='ref-table'>
<thead><tr><th>Combustível</th>{header_cells}</tr></thead>
<tbody>{rows_html}</tbody>
</table>""", unsafe_allow_html=True)
        return

    # ══════════════════════════════════════════════════════════════
    # MODO 1 — UF / Município selecionado
    # ══════════════════════════════════════════════════════════════
    rows = _anp_extrair_precos(sheets, uf=uf, municipio=municipio or None)

    if not rows:
        st.warning(
            f"Preços não encontrados para "
            f"**{municipio or UF_NOME.get(uf or '', uf or '')}** na planilha carregada."
        )
        return

    nivel       = rows[0]["Nível"]
    nome_regiao = rows[0].get("Nome Região")
    scope_label = (
        municipio if nivel in ("Município", "Capital")
        else UF_NOME.get(uf or "", uf or "")
    )
    uf_label    = f" ({uf})" if uf else ""
    col_est     = UF_NOME.get(uf or "", uf or "")
    col_reg     = nome_regiao.title() if nome_regiao else "Região"
    mostrar_est = nivel in ("Município", "Capital")
    semana_str  = semana or "—"

    # ── Cabeçalho compacto ────────────────────────────────────────
    _b_branco  = "<b style='color:#fff'>"
    _reg_trecho = f" &nbsp;·&nbsp; Região: {_b_branco}{col_reg}</b>" if nome_regiao else ""
    _cab_html = (
        "<div style='background:linear-gradient(90deg,#0d1b4b 0%,#1565c0 100%);"
        "border-radius:10px;padding:12px 18px;margin-bottom:16px;display:flex;"
        "align-items:center;justify-content:space-between;flex-wrap:wrap;gap:4px'>"
        f"<span style='color:#fff;font-size:16px;font-weight:700'>"
        f"💰 Preços ANP — {scope_label}{uf_label}</span>"
        f"<span style='color:rgba(255,255,255,.75);font-size:12px'>"
        f"Semana: {_b_branco}{semana_str}</b>"
        f"{_reg_trecho}"
        f" &nbsp;·&nbsp; Nível: {nivel}</span>"
        "</div>"
    )
    st.markdown(_cab_html, unsafe_allow_html=True)

    # ── CSS dos cards (injetado uma vez) ─────────────────────────
    st.markdown("""
<style>
.fc{background:#fff;border-radius:12px;padding:16px 18px 12px;
    box-shadow:0 2px 10px rgba(0,0,0,.08);
    border-left:5px solid #ddd;margin-bottom:4px;
    transition:box-shadow .15s}
.fc:hover{box-shadow:0 4px 18px rgba(0,0,0,.13)}
.fc-nome{font-size:12px;font-weight:700;color:#555;
         letter-spacing:.4px;text-transform:uppercase;margin-bottom:6px}
.fc-preco{font-size:26px;font-weight:800;letter-spacing:-.5px;line-height:1}
.fc-uni{font-size:11px;font-weight:400;color:#999;margin-left:3px}
.fc-refs{margin-top:10px;border-top:1px solid #f0f0f0;padding-top:8px}
.fc-ref{display:flex;justify-content:space-between;align-items:center;
        padding:3px 0;font-size:12px}
.fc-ref-label{color:#888}
.fc-ref-val{font-weight:600;color:#333;display:flex;align-items:center;gap:5px}
.fc-delta-up{background:#ffebee;color:#c62828;font-size:10px;font-weight:700;
             padding:1px 6px;border-radius:20px;white-space:nowrap}
.fc-delta-dn{background:#e8f5e9;color:#2e7d32;font-size:10px;font-weight:700;
             padding:1px 6px;border-radius:20px;white-space:nowrap}
.fc-delta-eq{background:#f3f4f6;color:#666;font-size:10px;font-weight:700;
             padding:1px 6px;border-radius:20px;white-space:nowrap}
.fc-postos{font-size:10px;color:#bbb;margin-top:6px;text-align:right}
</style>""", unsafe_allow_html=True)

    # ── Helper: badge de variação ─────────────────────────────────
    def _ref_row_html(label, ref, pm):
        if ref is None:
            return (
                f"<div class='fc-ref'>"
                f"<span class='fc-ref-label'>{label}</span>"
                f"<span class='fc-ref-val'><span style='color:#ccc'>—</span></span>"
                f"</div>"
            )
        diff = round(pm - ref, 3)
        val_str = f"R$ {_brl(ref, 3)}"
        if abs(diff) < 0.001:
            delta_html = f"<span class='fc-delta-eq'>= igual</span>"
        elif diff > 0:
            delta_html = f"<span class='fc-delta-up'>▲ {_brl(diff, 3)}</span>"
        else:
            delta_html = f"<span class='fc-delta-dn'>▼ {_brl(abs(diff), 3)}</span>"
        return (
            f"<div class='fc-ref'>"
            f"<span class='fc-ref-label'>{label}</span>"
            f"<span class='fc-ref-val'>{val_str} {delta_html}</span>"
            f"</div>"
        )

    # ── Pré-carrega preços da semana anterior (para tendência) ───
    _cache_ant  = st.session_state.get("_precos_anp_cache_anterior", {})
    _sheets_ant = _cache_ant.get("sheets", {})
    _semana_ant = _cache_ant.get("semana", "")
    # Monta dict {pk: preco_anterior} usando o mesmo escopo (uf/brasil)
    _ant_precos: dict = {}
    if _sheets_ant:
        _rows_ant = _anp_extrair_precos(_sheets_ant, uf=uf, municipio=municipio or None)
        if not _rows_ant and uf:
            _rows_ant = _anp_extrair_precos(_sheets_ant, uf=uf)
        for _r_ant in _rows_ant:
            _ant_precos[_anp_norm(_r_ant["Combustível"])] = _r_ant["Preço Médio"]

    # ── Grid 3 colunas ────────────────────────────────────────────
    n_cols = 3
    for chunk_start in range(0, len(rows), n_cols):
        chunk = rows[chunk_start: chunk_start + n_cols]
        cols  = st.columns(n_cols)
        for j, r in enumerate(chunk):
            pm     = r["Preço Médio"]
            r_est  = r.get("Ref. Estado")
            r_reg  = r.get("Ref. Região")
            r_br   = r.get("Ref. Brasil")
            uni    = r.get("Unidade", "R$/L")
            postos = r.get("Postos") or "?"

            # Cor da borda e do preço: verde=abaixo do nacional, vermelho=acima
            ref_main = r_br if r_br is not None else r_reg
            if ref_main is not None:
                d = pm - ref_main
                borda = "#2e7d32" if d < -0.005 else ("#e53935" if d > 0.005 else "#1565c0")
                preco_cor = "#1b5e20" if d < -0.005 else ("#b71c1c" if d > 0.005 else "#0d1b4b")
            else:
                borda, preco_cor = "#1565c0", "#0d1b4b"

            # ── Tendência vs semana anterior ─────────────────────
            _pk_r         = _anp_norm(r["Combustível"])
            _preco_ant_r  = _ant_precos.get(_pk_r)
            _tend_badge   = _tendencia_badge(pm, _preco_ant_r, inline=True)
            _tend_bloco   = _tendencia_badge(pm, _preco_ant_r, inline=False) if _preco_ant_r else ""
            _ant_caption  = (
                f"<div style='font-size:10px;color:#aaa;margin-top:2px'>"
                f"Sem. ant.: R$ {_brl(_preco_ant_r, 3)}</div>"
            ) if _preco_ant_r else ""

            # Monta referências
            refs = ""
            if mostrar_est:
                refs += _ref_row_html(f"vs {col_est}", r_est, pm)
            refs += _ref_row_html(f"vs {col_reg}", r_reg, pm)
            refs += _ref_row_html("vs Brasil", r_br, pm)

            card = (
                f"<div class='fc' style='border-left-color:{borda}'>"
                f"<div class='fc-nome'>{r['Combustível']}</div>"
                f"<div style='display:flex;align-items:baseline;gap:8px;flex-wrap:wrap'>"
                f"<span class='fc-preco' style='color:{preco_cor}'>"
                f"R$ {_brl(pm, 3)}</span>"
                f"<span class='fc-uni'>{uni}</span>"
                f"{_tend_badge}"
                f"</div>"
                f"{_ant_caption}"
                f"<div class='fc-refs'>{refs}</div>"
                f"<div class='fc-postos'>{postos} postos pesquisados</div>"
                f"</div>"
            )
            with cols[j]:
                st.markdown(card, unsafe_allow_html=True)

    _tend_nota = (
        f" · Tendência vs semana <b>{_semana_ant}</b>"
        if _semana_ant else
        " · Sem dados de semana anterior para tendência"
    )
    st.caption(
        f"Legenda de borda: 🟢 abaixo da média nacional · "
        f"🔴 acima · 🔵 igual  |  Semana ANP: {semana_str}"
        + (_tend_nota if _sheets_ant else "")
    )


# ═══════════════════════════════════════════════════════════════════
#  EXPORTAÇÃO — Base Nacional de Postos (Excel)
# ═══════════════════════════════════════════════════════════════════

def _gerar_excel_base_brasil() -> tuple:
    """Consolida todos os estados em cache, marca Gestão de Frotas e gera um .xlsx.

    Retorna (bytes_do_arquivo | None, mensagem_str).

    Estrutura do arquivo:
      • Aba "Postos ANP"  — todos os postos, ordenados por UF > Município > Razão Social
      • Cabeçalho azul escuro (#0D47A1) com texto branco
      • Linhas Gestão de Frotas destacadas em azul claro via formatação condicional
        (avaliado pelo Excel, sem loop Python linha-a-linha — suporta 60 000+ linhas)
      • Coluna "Gestão de Frotas" com valor "SIM" nos postos credenciados
    """
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.formatting.rule import FormulaRule
    from openpyxl.utils import get_column_letter

    estados = st.session_state.get("_estados_precarregados", [])
    if not estados:
        return None, "⚠️ Base não carregada. Use 'Pré-carregar Base Nacional' primeiro."

    # ── Consolida todos os estados ────────────────────────────────
    frames = []
    for uf in estados:
        try:
            df_uf = buscar_postos(uf=uf)
            if not df_uf.empty:
                frames.append(df_uf)
        except Exception:
            continue

    if not frames:
        return None, "❌ Nenhum dado encontrado nos estados carregados."

    df_all = pd.concat(frames, ignore_index=True)

    # ── Marca Gestão de Frotas ──────────────────────────────────────────
    cnpjs_pf = st.session_state.get("cnpjs_pro_frotas", set())
    df_all   = marcar_pro_frotas(df_all, cnpjs_pf)

    # ── Formata CNPJ ──────────────────────────────────────────────
    if "cnpj" in df_all.columns:
        df_all["cnpj"] = df_all["cnpj"].fillna("").apply(
            lambda x: _formatar_cnpj(str(x)) if x else "")

    # Coluna legível para Gestão de Frotas
    df_all["_pf_txt"] = df_all["_pro_frotas"].map({True: "SIM", False: ""})

    # ── Seleciona e renomeia colunas ──────────────────────────────
    _COL_MAP = [
        ("uf",                "UF"),
        ("municipio",         "Município"),
        ("razaoSocial",       "Razão Social"),
        ("cnpj",              "CNPJ"),
        ("distribuidora",     "Distribuidora / Bandeira"),
        ("_pf_txt",           "Gestão de Frotas"),
        ("endereco",          "Endereço"),
        ("bairro",            "Bairro"),
        ("cep",               "CEP"),
        ("autorizacao",       "Autorização ANP"),
        ("situacaoConstatada","Situação"),
        ("statusSIGAF",       "Status SIGAF"),
        ("_lat",              "Latitude"),
        ("_lon",              "Longitude"),
    ]
    cols_src = [c for c, _ in _COL_MAP if c in df_all.columns]
    cols_dst = [d for c, d in _COL_MAP if c in df_all.columns]

    df_exp = (df_all[cols_src]
              .rename(columns=dict(zip(cols_src, cols_dst)))
              .sort_values(["UF", "Município", "Razão Social"])
              .reset_index(drop=True))

    n_total = len(df_exp)
    n_pf_exp = int(df_exp["Gestão de Frotas"].eq("SIM").sum()) if "Gestão de Frotas" in df_exp.columns else 0

    # ── Gera Excel em memória ─────────────────────────────────────
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        df_exp.to_excel(writer, index=False, sheet_name="Postos ANP")
        ws = writer.sheets["Postos ANP"]

        n_rows = n_total + 1   # +1 cabeçalho
        n_cols = len(df_exp.columns)

        # Estilos base
        hdr_fill  = PatternFill("solid", fgColor="0D47A1")
        hdr_font  = Font(color="FFFFFF", bold=True, size=10)
        hdr_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        thin_side = Side(style="thin", color="C5CAE9")
        thin_brd  = Border(left=thin_side, right=thin_side,
                           top=thin_side, bottom=thin_side)

        # Cabeçalho formatado
        ws.row_dimensions[1].height = 28
        for col_idx, cell in enumerate(ws[1], start=1):
            cell.fill  = hdr_fill
            cell.font  = hdr_font
            cell.alignment = hdr_align
            cell.border    = thin_brd

        # Formatação condicional — linhas Gestão de Frotas em azul claro
        # Usa fórmula Excel avaliada pelo próprio app (rápido para 60 000+ linhas)
        if "Gestão de Frotas" in df_exp.columns:
            pf_col_idx    = df_exp.columns.get_loc("Gestão de Frotas") + 1   # 1-based
            pf_col_letter = get_column_letter(pf_col_idx)
            last_cell     = f"{get_column_letter(n_cols)}{n_rows}"
            pf_fill       = PatternFill("solid", fgColor="DBEAFE")
            pf_font       = Font(bold=True, size=10)
            # A fórmula usa referência absoluta na coluna GF e relativa na linha
            formula = [f'${pf_col_letter}2="SIM"']
            ws.conditional_formatting.add(
                f"A2:{last_cell}",
                FormulaRule(formula=formula, fill=pf_fill, font=pf_font),
            )

        # Painel informativo — linha abaixo dos dados
        info_row = n_rows + 2
        ws.cell(info_row, 1,
                f"Gerado em: {_agora()}  |  "
                f"{_n(n_total)} postos  |  "
                f"{_n(len(estados))} estados  |  "
                f"{_n(n_pf_exp)} Gestão de Frotas  |  "
                f"Fonte: API ANP — revendedoresapi.anp.gov.br"
                ).font = Font(italic=True, color="757575", size=9)

        # Largura automática das colunas (amostragem das primeiras 500 linhas)
        for col_idx, col_cells in enumerate(ws.iter_cols(
                min_row=1, max_row=min(n_rows, 501), max_col=n_cols), start=1):
            max_len = max((len(str(c.value or "")) for c in col_cells), default=8)
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 3, 42)

        # Congela linha do cabeçalho
        ws.freeze_panes = "A2"

        # Filtro automático
        ws.auto_filter.ref = f"A1:{get_column_letter(n_cols)}1"

    buf.seek(0)
    data = buf.read()
    msg  = (f"✅ {_n(n_total)} postos exportados "
            f"({_n(len(estados))} estados)  |  "
            f"⭐ {_n(n_pf_exp)} Gestão de Frotas identificados")
    return data, msg


# ═══════════════════════════════════════════════════════════════════
#  HELPER
# ═══════════════════════════════════════════════════════════════════

def _normalizar_distribuidora(serie: "pd.Series") -> "pd.Series":
    """Normaliza a coluna distribuidora para Title Case sem acentos alterados.

    Garante que 'IPIRANGA', 'ipiranga' e 'Ipiranga' virem todos 'Ipiranga',
    eliminando duplicatas em filtros, gráficos e tabelas.

    Regras:
    • strip de espaços externos
    • Title Case (cada palavra com inicial maiúscula)
    • Palavras-chave que devem ficar em maiúsculas preservadas: SA, LTDA, ME, EPP, BR, ANP, SBP
    """
    _upper_always = {"Sa", "Ltda", "Me", "Epp", "Br", "Anp", "Sbp", "S/A", "S.A"}

    def _fmt(val: str) -> str:
        v = str(val).strip()
        if not v or v.lower() in ("nan", "none", ""):
            return ""
        words = v.title().split()
        return " ".join(w.upper() if w in _upper_always else w for w in words)

    return serie.fillna("").apply(_fmt)


def _marcar_df_completo(df_raw: pd.DataFrame) -> pd.DataFrame:
    """
    Aplica todas as marcações (pro_frotas, cercados, perfil_venda) em UMA ÚNICA cópia
    do DataFrame, usando operações vetorizadas (str.replace em vez de apply).
    Resultado é guardado em session_state para evitar reprocessamento a cada rerun.
    """
    cnpjs_pf   = st.session_state.get("cnpjs_pro_frotas", set())
    cnpjs_cer  = st.session_state.get("cnpjs_cercados",   set())
    perfil_map = st.session_state.get("perfil_venda_map", {})

    df = df_raw.copy()   # ← única cópia

    # ── Normaliza distribuidora — elimina duplicatas por diferença de case ──
    if "distribuidora" in df.columns:
        df["distribuidora"] = _normalizar_distribuidora(df["distribuidora"])

    # ── CNPJ normalizado (vetorizado, C-level) ───────────────────────
    if "cnpj" in df.columns:
        df["_cnpj_norm"] = df["cnpj"].fillna("").str.replace(r'\D', '', regex=True)
    else:
        df["_cnpj_norm"] = ""

    # ── Gestão de Frotas ───────────────────────────────────────────────────
    df["_pro_frotas"] = (
        df["_cnpj_norm"].isin(cnpjs_pf) if cnpjs_pf else False
    )

    # ── Cercados ─────────────────────────────────────────────────────
    df["_cercado"] = (
        df["_cnpj_norm"].isin(cnpjs_cer) if cnpjs_cer else False
    )

    # ── Perfil de Venda / Rodo Rede ──────────────────────────────────
    if perfil_map:
        df["_perfil_venda"] = df["_cnpj_norm"].map(perfil_map).fillna("")
        df["_rodo_rede"] = df["_perfil_venda"].str.upper().str.strip() == PERFIL_RODO_REDE
    else:
        df["_perfil_venda"] = ""
        df["_rodo_rede"] = False

    return df


def preparar_df(df_raw, distribuidoras_filtro, perfis_filtro=None,
                filtro_servicos=None, filtro_24h=False):
    """
    Retorna df filtrado e marcado.
    A etapa de marcação (cara) é cacheada em session_state:
    só reprocessa quando df_raw ou os conjuntos GF/cercados/perfis mudam.

    Parâmetros extras:
      filtro_servicos: list com zero ou mais de ['pista_caminhao','arla','conveniencia']
      filtro_24h: bool — se True, exibe apenas postos com funciona_24h == True
    """
    cnpjs_pf   = st.session_state.get("cnpjs_pro_frotas", set())
    cnpjs_cer  = st.session_state.get("cnpjs_cercados",   set())
    perfil_map = st.session_state.get("perfil_venda_map", {})

    # Chave de cache: muda só quando os dados ou marcadores mudam.
    _first_cnpj = df_raw["cnpj"].iloc[0]  if (not df_raw.empty and "cnpj" in df_raw.columns) else ""
    _last_cnpj  = df_raw["cnpj"].iloc[-1] if (not df_raw.empty and "cnpj" in df_raw.columns) else ""
    _mark_key = (
        id(df_raw),
        len(df_raw),
        _first_cnpj,
        _last_cnpj,
        len(cnpjs_pf),
        len(cnpjs_cer),
        len(perfil_map),
    )

    if st.session_state.get("_df_marcado_key") == _mark_key:
        df = st.session_state["_df_marcado"]
    else:
        df = _marcar_df_completo(df_raw)
        st.session_state["_df_marcado"]     = df
        st.session_state["_df_marcado_key"] = _mark_key

    # Filtros leves (indexação booleana — sem cópia extra desnecessária)
    if distribuidoras_filtro:
        df = df[df["distribuidora"].isin(distribuidoras_filtro)]
    if perfis_filtro and "_perfil_venda" in df.columns:
        df = df[df["_perfil_venda"].isin(perfis_filtro)]

    # ── Filtro 24h ────────────────────────────────────────────
    if filtro_24h and "funciona_24h" in df.columns:
        _mask_24h = df["funciona_24h"].fillna(False).astype(bool)
        # Aplica apenas a postos que têm dado (não None); sem dado → não remove
        _tem_dado = df["funciona_24h"].notna()
        df = df[~_tem_dado | _mask_24h]

    # ── Filtro de Serviços ────────────────────────────────────
    if filtro_servicos:
        for _svc in filtro_servicos:
            if _svc in df.columns:
                _mask_svc = df[_svc].fillna(False).astype(bool)
                _tem_dado_svc = df[_svc].notna()
                df = df[~_tem_dado_svc | _mask_svc]

    return df


def n_pf(df):
    return int(df["_pro_frotas"].sum()) if "_pro_frotas" in df.columns else 0


# Emojis de medalha para rank 1-5
_RANK_EMOJI = {1: "🥇", 2: "🥈", 3: "🥉", 4: "4️⃣", 5: "5️⃣"}


def _calcular_top5_baratos(df: "pd.DataFrame", fuel_label: str = None) -> dict:
    """Calcula os 5 postos mais baratos da consulta atual.

    Parâmetros:
        df:          DataFrame de postos (df_show) — deve ter coluna '_cnpj_norm'
        fuel_label:  rótulo do combustível (ex: 'GASOLINA COMUM'). Se None, usa
                     o menor preço de qualquer combustível por posto.

    Retorna:
        dict {cnpj_norm: (rank, preco, combustivel_label)}  — só top-5 presentes em df.
        Rank 1 = mais barato.  Vazio se _pp_df não carregado ou df sem CNPJs.
    """
    _pp = st.session_state.get("_pp_df")
    if _pp is None or df.empty or "_cnpj_norm" not in df.columns:
        return {}

    cnpjs_visíveis = set(df["_cnpj_norm"].dropna().tolist())
    if not cnpjs_visíveis:
        return {}

    # Filtra _pp_df pelos postos visíveis
    _pp_vis = _pp[_pp["cnpj_norm"].isin(cnpjs_visíveis)].copy()
    if _pp_vis.empty:
        return {}

    # Filtra combustível se especificado
    if fuel_label and "combustivel_label" in _pp_vis.columns:
        _pp_vis = _pp_vis[_pp_vis["combustivel_label"].str.strip() == fuel_label]
    if _pp_vis.empty:
        return {}

    # Preço mínimo por posto (caso haja múltiplos combustíveis)
    _pp_min = (
        _pp_vis.sort_values("preco")
        .groupby("cnpj_norm", sort=False)
        .first()
        .reset_index()
    )[["cnpj_norm", "preco", "combustivel_label"]]

    # Ordena e pega top 5
    _top = _pp_min.nsmallest(5, "preco").reset_index(drop=True)
    result = {}
    for i, row in _top.iterrows():
        rank = i + 1
        result[row["cnpj_norm"]] = (rank, float(row["preco"]), str(row.get("combustivel_label", "")))
    return result


def _aplicar_rank_barato(df: "pd.DataFrame", top5: dict) -> "pd.DataFrame":
    """Adiciona coluna '_rank_barato' (0 = não ranqueado, 1-5 = posição)."""
    if "_cnpj_norm" not in df.columns:
        return df
    df = df.copy()
    df["_rank_barato"]  = df["_cnpj_norm"].map(lambda x: top5.get(x, (0,))[0]).fillna(0).astype(int)
    df["_preco_barato"] = df["_cnpj_norm"].map(lambda x: top5.get(x, (0, None))[1])
    df["_comb_barato"]  = df["_cnpj_norm"].map(lambda x: top5.get(x, (0, None, ""))[2] if x in top5 else "")
    return df


def _agora() -> str:
    """Retorna data e hora atual formatada: 06/05/2026 às 20:53."""
    return datetime.now().strftime("%d/%m/%Y às %H:%M")


def _n(valor, dec: int = 0) -> str:
    """Formata número com ponto como separador de milhar (padrão BR).
    Exemplos: _n(1234) → '1.234'  |  _n(1234.5, 1) → '1.234,5'
    """
    if dec == 0:
        return f"{int(round(valor)):,}".replace(",", ".")
    s = f"{valor:,.{dec}f}"
    return s.replace(",", "X").replace(".", ",").replace("X", ".")


def _sem_acento(texto: str) -> str:
    """Remove acentos e retorna texto normalizado em maiúsculas.
    Permite comparar 'Ribeirao Preto' com 'RIBEIRÃO PRETO'.
    """
    return "".join(
        c for c in unicodedata.normalize("NFD", texto.upper())
        if unicodedata.category(c) != "Mn"
    )


def _precarregar_estados_paralelo(max_workers: int = 5):
    """Carrega todos os 27 estados em paralelo (até max_workers simultâneos).
    - Hits de cache são instantâneos (< 1s total).
    - Chamadas reais à API rodam em paralelo: ~5-10s no lugar de ~50s sequencial.
    Retorna (lista_ok, lista_err).
    """
    def _load(uf):
        try:
            df = buscar_postos(uf=uf)
            return uf, not df.empty
        except Exception:
            return uf, False

    ok, err = [], []
    with ThreadPoolExecutor(max_workers=max_workers) as ex:
        futures = {ex.submit(_load, uf): uf for uf in UFS}
        for f in as_completed(futures):
            uf, success = f.result()
            (ok if success else err).append(uf)
    return sorted(ok), sorted(err)


def _buscar_cidades_cache(texto: str, max_results: int = 6) -> list:
    """Busca municípios em duas fontes (Gestão de Frotas e ANP) — sem depender do Nominatim.
    Normaliza acentos: 'Ribeirao Preto' encontra 'RIBEIRÃO PRETO', 'Vitoria' → 'VITÓRIA'.

    Ordem de prioridade:
      1. pf_coords_df (planilha Gestão de Frotas local) — sempre disponível e sem limite de API.
      2. Cache ANP por UF (buscar_postos) — complementa cidades não cobertas pela planilha.

    Retorna lista de dicts {label, lat, lon, tipo='cidade'}.
    """
    texto_norm = _sem_acento(texto.strip())
    if len(texto_norm) < 2:
        return []

    vistos: set   = set()
    resultados: list = []

    # ── 1. Planilha Gestão de Frotas (pf_coords_df) ─────────────────────────────────
    _pf_df = st.session_state.get("pf_coords_df", pd.DataFrame())
    if not _pf_df.empty and "municipio" in _pf_df.columns and "uf" in _pf_df.columns:
        _mask_pf = _pf_df["municipio"].fillna("").apply(
            lambda x: texto_norm in _sem_acento(x)
        )
        for (mun, uf_), grupo in _pf_df[_mask_pf].groupby(["municipio", "uf"]):
            chave = f"{mun}|{uf_}"
            if chave in vistos:
                continue
            vistos.add(chave)
            resultados.append({
                "label": f"{mun} – {uf_}",
                "lat":   float(grupo["_lat"].mean()),
                "lon":   float(grupo["_lon"].mean()),
                "tipo":  "cidade",
            })
            if len(resultados) >= max_results:
                return resultados

    if len(resultados) >= max_results:
        return resultados

    # ── 2. Cache ANP (complemento quando planilha não cobre) ──────────────────
    estados = st.session_state.get("_estados_precarregados", [])
    for uf in estados:
        try:
            df = buscar_postos(uf=uf)
            if df.empty or "municipio" not in df.columns:
                continue
            mask = df["municipio"].fillna("").apply(
                lambda x: texto_norm in _sem_acento(x)
            )
            for mun, grupo in df[mask].groupby("municipio"):
                chave = f"{mun}|{uf}"
                if chave in vistos:
                    continue
                vistos.add(chave)
                resultados.append({
                    "label": f"{mun} – {uf}",
                    "lat":   float(grupo["_lat"].mean()),
                    "lon":   float(grupo["_lon"].mean()),
                    "tipo":  "cidade",
                })
                if len(resultados) >= max_results:
                    return resultados
        except Exception:
            continue

    return resultados


# ═══════════════════════════════════════════════════════════════════
#  ROTAS SALVAS — persistência em arquivo JSON local
# ═══════════════════════════════════════════════════════════════════

_ROTAS_FILE = os.path.join(_DIR, "rotas_salvas.json")
_TOUR_FILE  = os.path.join(_DIR, "tour_done.flag")


def _tour_primeira_visita() -> bool:
    """Retorna True se o usuário nunca completou o tour (arquivo de flag ausente)."""
    return not os.path.exists(_TOUR_FILE)


def _marcar_tour_concluido():
    """Grava o arquivo de flag para não mostrar o tour automaticamente novamente."""
    try:
        with open(_TOUR_FILE, "w", encoding="utf-8") as _f:
            _f.write(_agora())
    except Exception:
        pass


def _carregar_rotas_salvas_local() -> list:
    """Lê rotas_salvas.json local (fallback sem banco)."""
    try:
        if os.path.exists(_ROTAS_FILE):
            with open(_ROTAS_FILE, "r", encoding="utf-8") as _f:
                _data = _json_mod.load(_f)
                return _data if isinstance(_data, list) else []
    except Exception:
        pass
    return []


def _gravar_rotas_salvas_local(rotas: list) -> bool:
    """Persiste a lista no arquivo JSON local (fallback sem banco)."""
    try:
        with open(_ROTAS_FILE, "w", encoding="utf-8") as _f:
            _json_mod.dump(rotas, _f, ensure_ascii=False, indent=2, default=str)
        return True
    except Exception:
        return False


def _salvar_rota_nova_local(nome: str, tipo: str, dados: dict) -> bool:
    """Salva rota no JSON local (fallback sem banco)."""
    rotas = _carregar_rotas_salvas_local()
    _id   = f"{int(time.time())}_{len(rotas)}"
    rotas.append({"id": _id, "nome": nome.strip() or f"Rota {len(rotas)+1}",
                  "tipo": tipo, "criado_em": _agora(), **dados})
    return _gravar_rotas_salvas_local(rotas)


def _deletar_rota_local(rota_id: str) -> bool:
    """Remove rota do JSON local (fallback sem banco)."""
    rotas = _carregar_rotas_salvas_local()
    antes = len(rotas)
    rotas = [r for r in rotas if r.get("id") != rota_id]
    if len(rotas) < antes:
        return _gravar_rotas_salvas_local(rotas)
    return False


# Aliases públicos — usam Supabase se disponível, JSON local como fallback
def _carregar_rotas_salvas() -> list:   return _db_carregar_rotas()
def _salvar_rota_nova(nome, tipo, dados): return _db_salvar_rota(nome, tipo, dados)
def _deletar_rota(rota_id):             return _db_deletar_rota(rota_id)


def _icone_tipo(tipo: str) -> str:
    return {"estado": "📍", "rota": "🗺️", "busca": "🔍", "roteirizacao": "🛣️"}.get(tipo, "📌")


# ═══════════════════════════════════════════════════════════════════
#  INTERFACE — BARRA SUPERIOR
# ═══════════════════════════════════════════════════════════════════

cnpjs_pf_ativos = st.session_state.get("cnpjs_pro_frotas", set())
pf_badge_html = (
    f'<span class="topbar-badge">⭐ Gestão de Frotas: {len(cnpjs_pf_ativos)} CNPJs ativos</span>'
    if cnpjs_pf_ativos else
    '<span class="topbar-badge">⭐ Gestão de Frotas não carregado</span>'
)

st.markdown(f"""
<div class="topbar">
  <div style="min-width:0;">
    <div class="topbar-title">Estudo de Rede</div>
    <div class="topbar-sub">Gestão de Frotas</div>
  </div>
  {pf_badge_html}
</div>
""", unsafe_allow_html=True)


# ═══════════════════════════════════════════════════════════════════
#  AUTO-CARGA DA BASE NACIONAL (uma vez por sessão)
#  Carrega os 27 estados em paralelo ao abrir o app.
#  - Hits de cache (24h): quase instantâneo.
#  - Cache frio (1ª abertura ou restart): ~5-15s com 5 workers paralelos.
#  Garante que buscas por cidade, nome e CNPJ funcionem imediatamente.
# ═══════════════════════════════════════════════════════════════════

if not st.session_state.get("_base_auto_ok"):
    with st.spinner("⏳ Carregando base nacional de postos… (apenas na primeira abertura)"):
        _auto_ok, _auto_err = _precarregar_estados_paralelo(max_workers=5)
    st.session_state.update({
        "_estados_precarregados": _auto_ok,
        "_preload_brasil_em":     _agora(),
        "_preload_brasil_ok":     len(_auto_ok),
        "_preload_brasil_err":    _auto_err,
        "_base_auto_ok":          True,
    })
    st.rerun()


# ═══════════════════════════════════════════════════════════════════
#  TOUR — inicialização de estado
# ═══════════════════════════════════════════════════════════════════

# Inicializa na primeira execução da sessão
if "_tour_ativo" not in st.session_state:
    st.session_state["_tour_ativo"] = _tour_primeira_visita()

# Botão oculto que JS vai clicar ao concluir/pular o tour
if st.button("​", key="btn_tour_done_hidden"):   # zero-width space
    _marcar_tour_concluido()
    st.session_state["_tour_ativo"] = False
    st.rerun()

if st.button("​​", key="btn_tour_open_hidden"):   # dois zero-width spaces
    st.session_state["_tour_ativo"] = True
    st.rerun()

# ═══════════════════════════════════════════════════════════════════
#  SIDEBAR
# ═══════════════════════════════════════════════════════════════════

with st.sidebar:

    # ── Header do Sidebar — imagem banner ─────────────────────────
    if _MENU_B64:
        st.markdown(
            # Imagem full-width sem margens, bordas arredondadas suaves
            f"<div style='margin:-1rem -1rem 0 -1rem;overflow:hidden'>"
            f"<img src='data:{_menu_mime};base64,{_MENU_B64}' "
            f"style='width:100%;display:block;object-fit:cover;"
            f"border-radius:0 0 8px 8px'>"
            f"</div>"
            # Barra separadora com gradiente alinhado ao tema azul/laranja dos botões
            f"<div style='height:4px;margin:0 -1rem 12px -1rem;"
            f"background:linear-gradient(90deg,#0D47A1 0%,#1565C0 40%,#1976D2 70%,#E65100 100%)'>"
            f"</div>",
            unsafe_allow_html=True,
        )
    else:
        # ── Fallback: logo Gestão de Frotas + subtítulo ─────────────────────
        st.markdown(f"""
        <div style='
            margin: -1rem -1rem 0 -1rem;
            background: #ffffff;
            padding: 22px 16px 14px;
            text-align: center;
            border-bottom: 4px solid transparent;
            border-image: linear-gradient(90deg, #0d1b4b 0%, #1565c0 55%, #0288d1 100%) 1;
            box-shadow: 0 4px 14px rgba(13,27,75,0.10);
            margin-bottom: 14px;
        '>
          {_LOGO_SIDEBAR}
          <div style='
              font-size: 10px;
              color: #1565c0;
              margin-top: 9px;
              letter-spacing: 1px;
              text-transform: uppercase;
              font-weight: 700;
              opacity: 0.75;
          '>Estudo de Rede · ANP</div>
        </div>
        """, unsafe_allow_html=True)

    # ── Card do usuário autenticado ────────────────────────────────
    _auth_u = st.session_state.get("_auth_user")
    if _auth_u:
        _nome_u   = _auth_u.get("name",     "Usuário")
        _email_u  = _auth_u.get("email",    "")
        _pic_u    = _auth_u.get("picture",  "")
        _prov_u   = _auth_u.get("provider", "")
        _prov_ico = "🪟" if _prov_u == "microsoft" else "🔴"

        # Avatar: foto de perfil ou iniciais
        if _pic_u:
            _avatar_html = (
                f"<img src='{_pic_u}' style='"
                f"width:36px;height:36px;border-radius:50%;"
                f"object-fit:cover;border:2px solid #e3e8f0'>"
            )
        else:
            _ini = "".join(w[0].upper() for w in _nome_u.split()[:2]) if _nome_u else "?"
            _avatar_html = (
                f"<div style='width:36px;height:36px;border-radius:50%;"
                f"background:#0D47A1;color:#fff;display:flex;align-items:center;"
                f"justify-content:center;font-weight:700;font-size:13px;"
                f"flex-shrink:0'>{_ini}</div>"
            )

        st.markdown(f"""
        <div style='background:#f5f7fd;border:1px solid #dde3ee;border-radius:11px;
                    padding:9px 11px;margin:4px 0 6px;
                    display:flex;align-items:center;gap:9px'>
          {_avatar_html}
          <div style='min-width:0;flex:1;overflow:hidden'>
            <div style='font-weight:600;font-size:12.5px;color:#1a237e;
                        white-space:nowrap;overflow:hidden;text-overflow:ellipsis'
            >{_nome_u}</div>
            <div style='font-size:10.5px;color:#607d8b;
                        white-space:nowrap;overflow:hidden;text-overflow:ellipsis'
            >{_email_u}</div>
          </div>
          <span title='{_prov_u}' style='font-size:15px;flex-shrink:0'>{_prov_ico}</span>
        </div>
        """, unsafe_allow_html=True)

        if st.button(
            "🚪 Sair",
            use_container_width=True,
            type="secondary",
            key="btn_logout",
            help="Encerrar sessão e voltar ao login",
        ):
            st.session_state["_auth_user"] = None
            st.rerun()

    # ── Auto-carregamento do repositório ─────────────────────
    # Tenta UMA VEZ por sessão — usa flag para não repetir
    if not st.session_state.get("cnpjs_pro_frotas") and not st.session_state.get("_repo_tentado"):
        st.session_state["_repo_tentado"] = True   # evita loop
        _cnpjs_repo, _msg_repo, _prev_repo, _perfil_repo, _coords_repo = _auto_carregar_pro_frotas_repo()
        if _cnpjs_repo:
            st.session_state["cnpjs_pro_frotas"]  = _cnpjs_repo
            st.session_state["_pf_fonte"]         = "repo"
            st.session_state["_pf_carregado_em"]  = _agora()
        if _perfil_repo:
            st.session_state["perfil_venda_map"]    = _perfil_repo
            st.session_state["perfis_pf_lista"]     = sorted(set(_perfil_repo.values()))
        if _coords_repo is not None and not _coords_repo.empty:
            st.session_state["pf_coords_df"] = _coords_repo
            _atualizar_servicos_pf(_coords_repo)

    # Reconstrói labels de serviços se pf_coords_df já existe mas labels ainda não foram gerados
    # (ex: primeira rerun após cache hit)
    if ("pf_coords_df" in st.session_state
            and not st.session_state["pf_coords_df"].empty
            and not st.session_state.get("_servicos_pf_labels")):
        _atualizar_servicos_pf(st.session_state["pf_coords_df"])

    # Fallback: CNPJs já carregados mas perfil_venda_map ou pf_coords_df ainda ausentes/vazio
    # (ocorre quando o cache antigo não incluía lat/lon ou pf_coords_df foi armazenado vazio)
    _pf_coords_ok = (
        "pf_coords_df" in st.session_state
        and not st.session_state["pf_coords_df"].empty
    )
    _needs_reload = (
        st.session_state.get("cnpjs_pro_frotas") and (
            not st.session_state.get("perfil_venda_map") or
            not _pf_coords_ok
        )
    )
    if _needs_reload and not st.session_state.get("_pf_coords_reload_feito"):
        st.session_state["_pf_coords_reload_feito"] = True
        _auto_carregar_pro_frotas_repo.clear()   # garante releitura com código atual
        _cnpjs_r2, _, _, _perfil_r2, _coords_r2 = _auto_carregar_pro_frotas_repo()
        if _perfil_r2:
            st.session_state["perfil_venda_map"]  = _perfil_r2
            st.session_state["perfis_pf_lista"]   = sorted(set(_perfil_r2.values()))
        if _coords_r2 is not None and not _coords_r2.empty:
            st.session_state["pf_coords_df"] = _coords_r2
            _atualizar_servicos_pf(_coords_r2)

    # Auto-load Postos Cercados (uma vez por sessão)
    if not st.session_state.get("cnpjs_cercados") and not st.session_state.get("_cercados_tentado"):
        st.session_state["_cercados_tentado"] = True
        _cnpjs_cer, _msg_cer, _ = _auto_carregar_cercados_repo()
        if _cnpjs_cer:
            st.session_state["cnpjs_cercados"]          = _cnpjs_cer
            st.session_state["_cercados_fonte"]         = "repo"
            st.session_state["_cercados_carregado_em"]  = _agora()

    # Auto-load Preço Posto — re-parseia se a versão do parser mudou
    _pp_ver_atual = st.session_state.get("_pp_parser_ver")
    if _pp_ver_atual != _PP_PARSER_VERSION:
        # Parser foi atualizado: descarta dado antigo e re-parseia
        st.session_state.pop("_pp_df", None)
        st.session_state.pop("_pp_tentado", None)
        _auto_carregar_precos_postos_repo.clear()
        st.session_state["_pp_parser_ver"] = _PP_PARSER_VERSION

    if st.session_state.get("_pp_df") is None and not st.session_state.get("_pp_tentado"):
        st.session_state["_pp_tentado"] = True
        _pp_df_tmp, _pp_msg_tmp, _ = _auto_carregar_precos_postos_repo()
        if _pp_df_tmp is not None:
            st.session_state["_pp_df"]         = _pp_df_tmp
            st.session_state["_pp_fonte"]       = "repo"
            st.session_state["_pp_carregado_em"] = _agora()

    # ── Modo de consulta — toggle buttons ─────────────────────
    # Banner: sutil quando Menu.jpg existe, com gradiente quando não existe
    if _MENU_B64:
        st.markdown(
            "<div style='display:flex;align-items:center;gap:8px;margin-bottom:8px'>"
            "<div style='flex:1;height:1px;background:linear-gradient(90deg,"
            "#0D47A1,#E65100)'></div>"
            "<div style='font-size:9px;font-weight:700;color:#1565c0;"
            "letter-spacing:1.5px;text-transform:uppercase;white-space:nowrap'>"
            "Modo de Consulta</div>"
            "<div style='flex:1;height:1px;background:linear-gradient(90deg,"
            "#E65100,#0D47A1)'></div>"
            "</div>",
            unsafe_allow_html=True,
        )
    else:
        st.markdown(
            "<div style='"
            "background:linear-gradient(135deg,#0D47A1 0%,#1565C0 45%,#E65100 100%);"
            "border-radius:12px;padding:10px 14px 8px;margin-bottom:10px'>"
            "<div style='color:rgba(255,255,255,.65);font-size:9px;font-weight:600;"
            "letter-spacing:1.2px;text-transform:uppercase;margin-bottom:2px'>Modo de Consulta</div>"
            "<div style='color:#fff;font-size:13px;font-weight:700;line-height:1.2'>"
            "Selecione como deseja buscar</div>"
            "</div>",
            unsafe_allow_html=True,
        )

    # CSS exclusivo para os 3 botões de modo (seletores st-key-* + data-testid)
    st.markdown("""
<style>
/* ── base: altura e texto para os 3 botões de modo ── */
.st-key-btn_modo_estado button,
.st-key-btn_modo_rota    button,
.st-key-btn_modo_consulta button {
    height: 48px !important;
    min-height: 48px !important;
    border-radius: 10px !important;
    font-weight: 700 !important;
    letter-spacing: 0.2px !important;
    transition: all .2s ease !important;
    white-space: nowrap !important;
    overflow: hidden !important;
    text-overflow: ellipsis !important;
    padding: 0 6px !important;
}
.st-key-btn_modo_estado button p,
.st-key-btn_modo_rota    button p,
.st-key-btn_modo_consulta button p {
    font-size: 12px !important;
    margin: 0 !important;
    font-weight: 700 !important;
}

/* ── ATIVO: gradiente azul → laranja ── */
.st-key-btn_modo_estado    [data-testid="stBaseButton-primary"],
.st-key-btn_modo_rota      [data-testid="stBaseButton-primary"],
.st-key-btn_modo_consulta  [data-testid="stBaseButton-primary"] {
    background: linear-gradient(135deg, #0D47A1 0%, #1565C0 50%, #E65100 100%) !important;
    border: none !important;
    color: #fff !important;
    box-shadow: 0 3px 10px rgba(13,71,161,.40), 0 1px 3px rgba(230,81,0,.30) !important;
}
.st-key-btn_modo_estado    [data-testid="stBaseButton-primary"]:hover,
.st-key-btn_modo_rota      [data-testid="stBaseButton-primary"]:hover,
.st-key-btn_modo_consulta  [data-testid="stBaseButton-primary"]:hover {
    background: linear-gradient(135deg, #1565C0 0%, #1976D2 50%, #F57C00 100%) !important;
    box-shadow: 0 5px 14px rgba(13,71,161,.50) !important;
    transform: translateY(-1px) !important;
}
.st-key-btn_modo_estado    [data-testid="stBaseButton-primary"] p,
.st-key-btn_modo_rota      [data-testid="stBaseButton-primary"] p,
.st-key-btn_modo_consulta  [data-testid="stBaseButton-primary"] p {
    color: #fff !important;
}

/* ── INATIVO: outline azul, hover com leve gradiente ── */
.st-key-btn_modo_estado    [data-testid="stBaseButton-secondary"],
.st-key-btn_modo_rota      [data-testid="stBaseButton-secondary"],
.st-key-btn_modo_consulta  [data-testid="stBaseButton-secondary"] {
    background: rgba(255,255,255,.92) !important;
    border: 2px solid #1565C0 !important;
    color: #1565C0 !important;
    box-shadow: none !important;
}
.st-key-btn_modo_estado    [data-testid="stBaseButton-secondary"]:hover,
.st-key-btn_modo_rota      [data-testid="stBaseButton-secondary"]:hover,
.st-key-btn_modo_consulta  [data-testid="stBaseButton-secondary"]:hover {
    background: linear-gradient(135deg,rgba(13,71,161,.07) 0%,rgba(230,81,0,.07) 100%) !important;
    border-color: #E65100 !important;
    color: #E65100 !important;
    transform: translateY(-1px) !important;
}
.st-key-btn_modo_estado    [data-testid="stBaseButton-secondary"] p,
.st-key-btn_modo_rota      [data-testid="stBaseButton-secondary"] p,
.st-key-btn_modo_consulta  [data-testid="stBaseButton-secondary"] p {
    color: inherit !important;
}

/* ── Botão Roteirização ── */
.st-key-btn_modo_roteirizacao button {
    height: 40px !important; min-height: 40px !important;
    border-radius: 10px !important; font-weight: 700 !important;
    letter-spacing: 0.2px !important; transition: all .2s ease !important;
}
.st-key-btn_modo_roteirizacao button p { font-size: 12px !important; margin: 0 !important; font-weight: 700 !important; }
.st-key-btn_modo_roteirizacao [data-testid="stBaseButton-primary"] {
    background: linear-gradient(135deg, #004D40 0%, #00796B 55%, #E65100 100%) !important;
    border: none !important; color: #fff !important;
    box-shadow: 0 3px 10px rgba(0,77,64,.45), 0 1px 3px rgba(230,81,0,.25) !important;
}
.st-key-btn_modo_roteirizacao [data-testid="stBaseButton-primary"]:hover {
    background: linear-gradient(135deg, #00695C 0%, #00897B 55%, #F57C00 100%) !important;
    transform: translateY(-1px) !important;
    box-shadow: 0 5px 14px rgba(0,77,64,.55) !important;
}
.st-key-btn_modo_roteirizacao [data-testid="stBaseButton-primary"] p { color: #fff !important; }
.st-key-btn_modo_roteirizacao [data-testid="stBaseButton-secondary"] {
    background: rgba(255,255,255,.92) !important;
    border: 2px solid #00796B !important; color: #00796B !important;
    box-shadow: none !important;
}
.st-key-btn_modo_roteirizacao [data-testid="stBaseButton-secondary"]:hover {
    border-color: #E65100 !important; color: #E65100 !important;
    transform: translateY(-1px) !important;
}
.st-key-btn_modo_roteirizacao [data-testid="stBaseButton-secondary"] p { color: inherit !important; }

/* ── Botão Rotas Salvas ── */
.st-key-btn_rotas_salvas button {
    height: 40px !important;
    min-height: 40px !important;
    border-radius: 10px !important;
    font-weight: 700 !important;
    letter-spacing: 0.2px !important;
    transition: all .2s ease !important;
}
.st-key-btn_rotas_salvas button p { font-size: 12px !important; margin: 0 !important; font-weight: 700 !important; }
.st-key-btn_rotas_salvas [data-testid="stBaseButton-primary"] {
    background: linear-gradient(135deg, #1B5E20 0%, #2E7D32 50%, #F57F17 100%) !important;
    border: none !important; color: #fff !important;
    box-shadow: 0 3px 10px rgba(27,94,32,.40) !important;
}
.st-key-btn_rotas_salvas [data-testid="stBaseButton-primary"]:hover {
    background: linear-gradient(135deg, #2E7D32 0%, #388E3C 50%, #F9A825 100%) !important;
    transform: translateY(-1px) !important;
}
.st-key-btn_rotas_salvas [data-testid="stBaseButton-primary"] p { color: #fff !important; }
.st-key-btn_rotas_salvas [data-testid="stBaseButton-secondary"] {
    background: rgba(255,255,255,.92) !important;
    border: 2px solid #2E7D32 !important;
    color: #2E7D32 !important;
    box-shadow: none !important;
}
.st-key-btn_rotas_salvas [data-testid="stBaseButton-secondary"]:hover {
    border-color: #F57F17 !important; color: #F57F17 !important;
    transform: translateY(-1px) !important;
}
.st-key-btn_rotas_salvas [data-testid="stBaseButton-secondary"] p { color: inherit !important; }

/* ── Botão Dashboard ── */
.st-key-btn_dashboard button {
    height: 40px !important;
    min-height: 40px !important;
    border-radius: 10px !important;
    font-weight: 700 !important;
    letter-spacing: 0.2px !important;
    transition: all .2s ease !important;
}
.st-key-btn_dashboard button p { font-size: 12px !important; margin: 0 !important; font-weight: 700 !important; }
.st-key-btn_dashboard [data-testid="stBaseButton-primary"] {
    background: linear-gradient(135deg, #0D47A1 0%, #1565C0 50%, #E65100 100%) !important;
    border: none !important; color: #fff !important;
    box-shadow: 0 3px 10px rgba(13,71,161,.40) !important;
}
.st-key-btn_dashboard [data-testid="stBaseButton-primary"]:hover {
    background: linear-gradient(135deg, #1565C0 0%, #1976D2 50%, #F57C00 100%) !important;
    transform: translateY(-1px) !important;
}
.st-key-btn_dashboard [data-testid="stBaseButton-primary"] p { color: #fff !important; }
.st-key-btn_dashboard [data-testid="stBaseButton-secondary"] {
    background: rgba(255,255,255,.92) !important;
    border: 2px solid #1565C0 !important;
    color: #1565C0 !important;
    box-shadow: none !important;
}
.st-key-btn_dashboard [data-testid="stBaseButton-secondary"]:hover {
    border-color: #E65100 !important; color: #E65100 !important;
    transform: translateY(-1px) !important;
}
.st-key-btn_dashboard [data-testid="stBaseButton-secondary"] p { color: inherit !important; }
/* ── Botão Inteligência ── */
.st-key-btn_inteligencia button {
    height: 40px !important;
    min-height: 40px !important;
    border-radius: 10px !important;
    font-weight: 700 !important;
    letter-spacing: 0.2px !important;
    transition: all .2s ease !important;
}
.st-key-btn_inteligencia button p { font-size: 12px !important; margin: 0 !important; font-weight: 700 !important; }
.st-key-btn_inteligencia [data-testid="stBaseButton-primary"] {
    background: linear-gradient(135deg, #4A148C 0%, #6A1B9A 50%, #0D47A1 100%) !important;
    border: none !important; color: #fff !important;
    box-shadow: 0 3px 10px rgba(74,20,140,.40) !important;
}
.st-key-btn_inteligencia [data-testid="stBaseButton-primary"]:hover {
    background: linear-gradient(135deg, #6A1B9A 0%, #7B1FA2 50%, #1565C0 100%) !important;
    transform: translateY(-1px) !important;
}
.st-key-btn_inteligencia [data-testid="stBaseButton-primary"] p { color: #fff !important; }
.st-key-btn_inteligencia [data-testid="stBaseButton-secondary"] {
    background: rgba(255,255,255,.92) !important;
    border: 2px solid #6A1B9A !important;
    color: #6A1B9A !important;
    box-shadow: none !important;
}
.st-key-btn_inteligencia [data-testid="stBaseButton-secondary"]:hover {
    border-color: #0D47A1 !important; color: #0D47A1 !important;
    transform: translateY(-1px) !important;
}
.st-key-btn_inteligencia [data-testid="stBaseButton-secondary"] p { color: inherit !important; }
</style>""", unsafe_allow_html=True)

    if "modo_selecionado" not in st.session_state:
        st.session_state["modo_selecionado"] = "📍 Por UF/Município"
    _modo_atual = st.session_state["modo_selecionado"]
    _col_m1, _col_m2, _col_m3 = st.columns(3)
    with _col_m1:
        if st.button(
            "📍 Por UF",
            use_container_width=True,
            type="primary" if _modo_atual == "📍 Por UF/Município" else "secondary",
            key="btn_modo_estado",
        ):
            st.session_state["modo_selecionado"] = "📍 Por UF/Município"
            _log_acesso("MODO_SELECIONADO", "📍 Por UF/Município", modo_override="📍 Por UF/Município")
            st.rerun()
    with _col_m2:
        if st.button(
            "🗺️ Rota",
            use_container_width=True,
            type="primary" if _modo_atual == "🗺️ Por Rota" else "secondary",
            key="btn_modo_rota",
        ):
            st.session_state["modo_selecionado"] = "🗺️ Por Rota"
            _log_acesso("MODO_SELECIONADO", "🗺️ Por Rota", modo_override="🗺️ Por Rota")
            st.rerun()
    with _col_m3:
        if st.button(
            "🔍 Busca",
            use_container_width=True,
            type="primary" if _modo_atual == "🔍 Consulta por Posto" else "secondary",
            key="btn_modo_consulta",
        ):
            st.session_state["modo_selecionado"] = "🔍 Consulta por Posto"
            _log_acesso("MODO_SELECIONADO", "🔍 Consulta por Posto", modo_override="🔍 Consulta por Posto")
            st.rerun()

    # ── Botão Roteirização (largura total) ──────────────────────
    if st.button(
        "🛣️ Roteirização",
        use_container_width=True,
        type="primary" if _modo_atual == "🛣️ Roteirização" else "secondary",
        key="btn_modo_roteirizacao",
        help="Planejar rota com otimização de abastecimento",
    ):
        st.session_state["modo_selecionado"] = "🛣️ Roteirização"
        _log_acesso("MODO_SELECIONADO", "🛣️ Roteirização", modo_override="🛣️ Roteirização")
        st.rerun()

    # ── Botão Rotas Salvas (largura total, abaixo dos modos) ──────
    _n_rotas_sb = len(_carregar_rotas_salvas())
    _label_rotas = f"📋 Rotas Salvas{f'  ({_n_rotas_sb})' if _n_rotas_sb else ''}"
    if st.button(
        _label_rotas,
        use_container_width=True,
        type="primary" if _modo_atual == "📋 Rotas Salvas" else "secondary",
        key="btn_rotas_salvas",
        help="Ver e restaurar consultas salvas anteriormente",
    ):
        st.session_state["modo_selecionado"] = "📋 Rotas Salvas"
        _log_acesso("MODO_SELECIONADO", "📋 Rotas Salvas", modo_override="📋 Rotas Salvas")
        st.rerun()

    # ── Botão Dashboard (largura total) ─────────────────────────
    if st.button(
        "📊 Dashboard",
        use_container_width=True,
        type="primary" if _modo_atual == "📊 Dashboard" else "secondary",
        key="btn_dashboard",
        help="KPIs de cobertura e penetração GF por estado",
    ):
        st.session_state["modo_selecionado"] = "📊 Dashboard"
        _log_acesso("MODO_SELECIONADO", "📊 Dashboard", modo_override="📊 Dashboard")
        st.rerun()

    # ── Botão Inteligência de Dados (largura total) ───────────────
    if st.button(
        "🧠 Inteligência",
        use_container_width=True,
        type="primary" if _modo_atual == "🧠 Inteligência" else "secondary",
        key="btn_inteligencia",
        help="Histórico de preços, score de postos e relatório de alertas",
    ):
        st.session_state["modo_selecionado"] = "🧠 Inteligência"
        _log_acesso("MODO_SELECIONADO", "🧠 Inteligência", modo_override="🧠 Inteligência")
        st.rerun()

    # ── Botão Admin (visível só para o administrador) ─────────────
    _email_atual = (st.session_state.get("_auth_user") or {}).get("email", "")
    if _email_atual.lower() == _ADMIN_EMAIL.lower():
        st.divider()
        if st.button(
            "🔐 Admin",
            use_container_width=True,
            type="primary" if _modo_atual == "🔐 Admin" else "secondary",
            key="btn_admin",
            help="Painel de controle de acesso de usuários",
        ):
            st.session_state["modo_selecionado"] = "🔐 Admin"
            st.rerun()

    modo = _modo_atual
    st.divider()

    # ── Modo 1 ────────────────────────────────────────────────
    if modo == "📍 Por UF/Município":  # noqa: E501
        _fk_m1 = st.session_state.get("_form_key_m1", 0)
        # Pré-preenche com valores restaurados (de Rotas Salvas), se presentes
        _restore_uf  = st.session_state.pop("_restore_uf",  None)
        _restore_mun = st.session_state.pop("_restore_mun", None)
        _uf_default_idx = 0
        if _restore_uf and _restore_uf in UFS:
            _uf_default_idx = UFS.index(_restore_uf) + 1  # +1 por "— Selecione —"
        st.markdown("<div class='sb-label'>Localização</div>", unsafe_allow_html=True)
        uf = st.selectbox("Estado (UF)", ["— Selecione —"] + UFS, index=_uf_default_idx,
                          key=f"sel_uf_{_fk_m1}",
                          help="Selecione o estado para carregar os postos")
        uf = "" if uf == "— Selecione —" else uf

        municipio_input = st.text_input("🏙️ Município (opcional)",
                                         value=_restore_mun or "",
                                         placeholder="Ex: Teresina",
                                         key=f"txt_mun_{_fk_m1}",
                                         help="Filtra os postos por município dentro do estado")

        if st.button("🗑️ Limpar Consulta", use_container_width=True,
                     help="Limpa estado, município, filtros e seleção de rota"):
            for _k in ["_map_orig", "_map_dest", "_map_rota_result", "_map_posto_sel",
                       "_uf_carregada", "df_raw_full", "distribuidoras_disponiveis"]:
                st.session_state.pop(_k, None)
            st.session_state["_form_key_m1"] = _fk_m1 + 1
            st.rerun()

        distribuidoras_filtro = []
        if st.session_state.get("distribuidoras_disponiveis"):
            st.markdown("<div class='sb-label'>Filtrar por Bandeira</div>", unsafe_allow_html=True)
            distribuidoras_filtro = st.multiselect(
                "Bandeiras", st.session_state["distribuidoras_disponiveis"],
                placeholder="Todas as bandeiras", label_visibility="collapsed",
                key=f"mult_dist_{_fk_m1}")

        # Filtro de Perfil de Venda (Gestão de Frotas)
        perfis_filtro_m1 = []
        _perfis_lista_m1 = st.session_state.get("perfis_pf_lista", [])
        if _perfis_lista_m1:
            st.markdown("<div class='sb-label'>Perfil de Venda ⭐</div>", unsafe_allow_html=True)
            perfis_filtro_m1 = st.multiselect(
                "Perfil de Venda", _perfis_lista_m1,
                placeholder="Todos os perfis", label_visibility="collapsed",
                key=f"mult_perfil_{_fk_m1}",
                help="Filtra os postos Gestão de Frotas pelo perfil de venda. Postos não-GF sempre exibidos.",
            )

        # ── Favoritos ─────────────────────────────────────────────────────
        if "fav_cnpjs" not in st.session_state:
            _favs_raw = _db_favoritos()
            st.session_state["fav_cnpjs"] = {r["cnpj"] for r in _favs_raw}

        _filtro_apenas_favoritos_m1 = st.checkbox(
            "⭐ Apenas Favoritos",
            key="chk_apenas_fav_m1",
            help="Exibe somente os postos que você marcou como favoritos",
        )

        # Painel de gerenciamento de favoritos
        _fav_list = _db_favoritos() if st.session_state.get("fav_cnpjs") else []
        if _fav_list:
            with st.expander(f"⭐ Meus Favoritos ({len(_fav_list)})", expanded=False):
                for _fv in _fav_list:
                    _fv_col1, _fv_col2 = st.columns([4, 1])
                    with _fv_col1:
                        st.markdown(
                            f"<div style='font-size:11px;font-weight:600;line-height:1.3'>"
                            f"{_fv.get('razao_social','—')[:35]}</div>"
                            f"<div style='font-size:10px;color:#888'>"
                            f"{_fv.get('municipio','')} / {_fv.get('uf','')}</div>",
                            unsafe_allow_html=True,
                        )
                    with _fv_col2:
                        if st.button("✕", key=f"rm_fav_{_fv['cnpj']}",
                                     help="Remover dos favoritos"):
                            _db_remove_favorito(_fv["cnpj"])
                            st.session_state["fav_cnpjs"].discard(_fv["cnpj"])
                            st.rerun()

        # ── Filtros Avançados ─────────────────────────────────────────────
        _pp_df_adv   = st.session_state.get("_pp_df")
        _svc_cols    = st.session_state.get("_servicos_cols_disponiveis", [])
        _tem_preco   = _pp_df_adv is not None and "preco" in _pp_df_adv.columns
        _tem_servico = bool(_svc_cols)

        if _tem_preco or _tem_servico:
            with st.expander("🔍 Filtros Avançados", expanded=False):
                # ── Faixa de Preço (R$/L) ─────────────────────────────
                _preco_min_m1 = _preco_max_m1 = None
                _preco_faixa_m1 = None
                _fuel_sel_m1 = None
                if _tem_preco:
                    _fuels_m1 = sorted(
                        _pp_df_adv["combustivel_label"].dropna().str.strip().unique().tolist()
                    )
                    _fuel_sel_m1 = st.selectbox(
                        "⛽ Combustível (preço)",
                        ["— Todos —"] + _fuels_m1,
                        key=f"adv_fuel_m1_{_fk_m1}",
                        label_visibility="visible",
                    )
                    if _fuel_sel_m1 and _fuel_sel_m1 != "— Todos —":
                        _df_fuel_m1 = _pp_df_adv[
                            _pp_df_adv["combustivel_label"].str.strip() == _fuel_sel_m1
                        ]
                        if not _df_fuel_m1.empty:
                            _pmin = float(_df_fuel_m1["preco"].min())
                            _pmax = float(_df_fuel_m1["preco"].max())
                            if _pmin < _pmax:
                                _preco_min_m1, _preco_max_m1 = _pmin, _pmax
                                _preco_faixa_m1 = st.slider(
                                    "💰 Faixa de Preço (R$/L)",
                                    min_value=round(_pmin, 2),
                                    max_value=round(_pmax, 2),
                                    value=(round(_pmin, 2), round(_pmax, 2)),
                                    step=0.01,
                                    format="R$ %.2f",
                                    key=f"adv_preco_m1_{_fk_m1}",
                                    help="Exibe apenas postos com preço registrado dentro da faixa selecionada",
                                )
                            else:
                                st.caption(f"Preço único: R$ {_pmin:.3f}/L")

                # ── Horário de Funcionamento ───────────────────────────
                _filtro_24h_m1 = False
                if "funciona_24h" in _svc_cols:
                    _filtro_24h_m1 = st.checkbox(
                        "🕐 Somente postos 24h",
                        key=f"adv_24h_m1_{_fk_m1}",
                        help="Filtra postos que funcionam 24 horas (coluna 'FUNCIONA_24H' da planilha)",
                    )

                # ── Serviços Disponíveis ───────────────────────────────
                _filtro_servicos_m1 = []
                _svc_pf_lbl_m1 = st.session_state.get("_servicos_pf_labels", {})
                # Labels legadas (colunas genéricas) — só aparecem se não substituídas
                _SVC_LEG_LBL = {"pista_caminhao": "🚛 Pista Caminhão",
                                 "arla":           "🧪 ARLA 32",
                                 "conveniencia":   "🛒 Conveniência"}
                _conceitos_cobertos_m1 = set(_svc_pf_lbl_m1.keys())
                _svc_map_m1: dict = {}
                # 1) Colunas dinâmicas da planilha ("Possui X?")
                for _ck, _lbl in _svc_pf_lbl_m1.items():
                    if _ck in _svc_cols:
                        _svc_map_m1[_lbl] = _ck
                # 2) Legadas (se o conceito não foi substituído)
                for _ck, _lbl in _SVC_LEG_LBL.items():
                    if _ck in _svc_cols:
                        _sup = _SVC_LEGADO_SUPERSEDE.get(_ck, set())
                        if not (_sup & _conceitos_cobertos_m1):
                            _svc_map_m1[_lbl] = _ck
                _svc_opts = list(_svc_map_m1.keys())
                if _svc_opts:
                    _svc_sel_m1 = st.multiselect(
                        "🔧 Serviços disponíveis",
                        _svc_opts,
                        placeholder="Qualquer serviço",
                        key=f"adv_svc_m1_{_fk_m1}",
                        help="Exibe apenas postos que oferecem os serviços selecionados",
                    )
                    _filtro_servicos_m1 = [_svc_map_m1[s] for s in _svc_sel_m1]

                if not _tem_servico:
                    st.caption(
                        "ℹ️ Colunas de serviço não encontradas na planilha. "
                        "Adicione colunas como 'Possui restaurante?', 'Possui estacionamento?', etc."
                    )
        else:
            _preco_faixa_m1 = None
            _fuel_sel_m1    = None
            _filtro_24h_m1  = False
            _filtro_servicos_m1 = []

    # ── Modo 3 ────────────────────────────────────────────────
    elif modo == "🔍 Consulta por Posto":

        # ── Status Origem / Destino ─────────────────────────────
        _m3sb_map_o = st.session_state.get("_map_orig")
        _m3sb_map_d = st.session_state.get("_map_dest")
        _m3sb_o_ok  = bool(_m3sb_map_o)
        _m3sb_d_ok  = bool(_m3sb_map_d)

        # Pílulas de progresso
        def _prog_pill_m3(label, ok, ativo):
            if ok:
                _bg, _brd, _c, _ic = "#e8f5e9","#a5d6a7","#2e7d32","✔"
            elif ativo:
                _bg, _brd, _c, _ic = "#e3f2fd","#90caf9","#1565c0","●"
            else:
                _bg, _brd, _c, _ic = "#f5f5f5","#e0e0e0","#bdbdbd","○"
            return (
                f"<div style='flex:1;background:{_bg};border:1px solid {_brd};"
                f"border-radius:8px;padding:5px 6px;text-align:center'>"
                f"<div style='font-size:10px;color:{_c};font-weight:700'>{_ic}</div>"
                f"<div style='font-size:9px;color:{_c};line-height:1.2'>{label}</div>"
                f"</div>"
            )
        st.markdown(
            f"<div style='display:flex;gap:4px;margin-bottom:10px'>"
            f"{_prog_pill_m3('Origem', _m3sb_o_ok, not _m3sb_o_ok)}"
            f"{_prog_pill_m3('Destino', _m3sb_d_ok, _m3sb_o_ok and not _m3sb_d_ok)}"
            f"{_prog_pill_m3('Traçar', False, _m3sb_o_ok and _m3sb_d_ok)}"
            f"</div>",
            unsafe_allow_html=True,
        )

        # Mini-cards mostrando seleções atuais
        if _m3sb_o_ok or _m3sb_d_ok:
            def _mini_card_m3sb(label, cor, sel):
                if sel:
                    nm = sel.get("label", "?")[:34]
                    lc = f"{sel.get('municipio','')}/{sel.get('uf','')}"
                    return (
                        f"<div style='border-left:3px solid {cor};background:#fafafa;"
                        f"border-radius:0 8px 8px 0;padding:6px 10px;margin-bottom:4px'>"
                        f"<div style='font-size:9px;font-weight:700;color:{cor};"
                        f"text-transform:uppercase;letter-spacing:0.6px'>{label} ✔</div>"
                        f"<div style='font-size:11px;font-weight:600;color:#1a1a1a;"
                        f"line-height:1.3;margin-top:2px'>{nm}</div>"
                        f"<div style='font-size:10px;color:#666'>📍 {lc}</div>"
                        f"</div>"
                    )
                else:
                    return (
                        f"<div style='border-left:3px dashed #d0d0d0;background:#f9f9f9;"
                        f"border-radius:0 8px 8px 0;padding:6px 10px;margin-bottom:4px'>"
                        f"<div style='font-size:9px;font-weight:700;color:#bbb;"
                        f"text-transform:uppercase;letter-spacing:0.6px'>{label}</div>"
                        f"<div style='font-size:10px;color:#aaa;font-style:italic'>Não definido</div>"
                        f"</div>"
                    )
            st.markdown(
                _mini_card_m3sb("🟢 Origem", "#2E7D32", _m3sb_map_o)
                + _mini_card_m3sb("🔴 Destino", "#C62828", _m3sb_map_d),
                unsafe_allow_html=True,
            )
            if st.button(
                "🗑️ Limpar Origem / Destino",
                use_container_width=True,
                key="btn_limpar_od_m3",
                help="Limpar pontos de origem e destino selecionados",
            ):
                for _k in ["_map_orig", "_map_dest", "_map_rota_result"]:
                    st.session_state.pop(_k, None)
                st.rerun()
            st.divider()

        # ── Dica de busca ─────────────────────────────────────────
        st.markdown(
            "<div style='background:#e8f5e9;border-radius:8px;padding:10px 12px;"
            "font-size:12px;color:#2e7d32;margin-bottom:12px'>"
            "🔍 Busque por <b>nome do posto</b>, <b>razão social</b> ou <b>CNPJ</b> "
            "e selecione como <b>Origem</b> ou <b>Destino</b>."
            "</div>",
            unsafe_allow_html=True,
        )
        _fk_m3 = st.session_state.get("_form_key_m3", 0)
        _termo_m3 = st.text_input(
            "Nome, razão social ou CNPJ",
            placeholder="Ex: Auto Posto Silva  ·  12.345.678/0001-99",
            key=f"txt_consulta_{_fk_m3}",
            help="Digite ao menos 3 caracteres do nome ou o CNPJ completo",
        )
        _uf_m3_sel = st.selectbox(
            "Estado (opcional)",
            ["Todos os estados"] + UFS,
            index=0,
            key=f"sel_uf_m3_{_fk_m3}",
            help="Filtre por estado para resultados mais rápidos",
        )
        _uf_m3 = "" if _uf_m3_sel == "Todos os estados" else _uf_m3_sel

        _buscar_m3 = st.button(
            "🔍 Buscar Posto",
            use_container_width=True,
            type="primary",
            key=f"btn_buscar_m3_{_fk_m3}",
            disabled=len(_termo_m3.strip()) < 3,
        )
        if _buscar_m3:
            _log_acesso("CONSULTA_POSTO", f"termo={_termo_m3.strip()} | uf={_uf_m3}")
            st.session_state["_m3_termo"]     = _termo_m3.strip()
            st.session_state["_m3_uf"]        = _uf_m3
            st.session_state["_m3_resultado"] = None
            st.rerun()

        if st.button("🗑️ Limpar Busca", use_container_width=True, key="btn_limpar_m3"):
            for _k in ["_m3_termo", "_m3_uf", "_m3_resultado"]:
                st.session_state.pop(_k, None)
            st.session_state["_form_key_m3"] = _fk_m3 + 1
            st.rerun()

    # ── Modo 2 ────────────────────────────────────────────────
    elif modo == "🗺️ Por Rota":

        _n_paradas  = st.session_state.get("_paradas_count", 0)
        _orig_pronto = bool(st.session_state.get("orig_sel"))
        _dest_pronto = bool(st.session_state.get("dest_sel"))

        # ── Mini barra de progresso ─────────────────────────────────
        def _prog_pill(label, ok, ativo):
            if ok:
                _bg, _brd, _c = "#e8f5e9","#a5d6a7","#2e7d32"
                _ic = "✔"
            elif ativo:
                _bg, _brd, _c = "#e3f2fd","#90caf9","#1565c0"
                _ic = "●"
            else:
                _bg, _brd, _c = "#f5f5f5","#e0e0e0","#bdbdbd"
                _ic = "○"
            return (
                f"<div style='flex:1;background:{_bg};border:1px solid {_brd};"
                f"border-radius:8px;padding:5px 6px;text-align:center'>"
                f"<div style='font-size:10px;color:{_c};font-weight:700'>{_ic}</div>"
                f"<div style='font-size:9px;color:{_c};line-height:1.2'>{label}</div>"
                f"</div>"
            )
        st.markdown(
            f"<div style='display:flex;gap:4px;margin-bottom:10px'>"
            f"{_prog_pill('Origem', _orig_pronto, not _orig_pronto)}"
            f"{_prog_pill('Destino', _dest_pronto, _orig_pronto and not _dest_pronto)}"
            f"{_prog_pill('Traçar', False, _orig_pronto and _dest_pronto)}"
            f"</div>",
            unsafe_allow_html=True,
        )

        # ── Dica de preenchimento ───────────────────────────────────
        st.markdown(
            "<div style='background:#e3f2fd;border-radius:8px;padding:7px 10px;"
            "font-size:10px;color:#1565c0;margin-bottom:8px;line-height:1.5'>"
            "💡 Digite <b>UF</b> (ex: SP), <b>cidade</b>, <b>nome do posto</b> ou "
            "<b>CNPJ</b> e selecione a sugestão.</div>",
            unsafe_allow_html=True,
        )

        # ── rail pontilhado vertical ───────────────────────────────
        def _rail(color: str = "#90CAF9", height: int = 10):
            st.markdown(
                f"<div style='margin:0 0 0 6px;border-left:2px dashed {color};"
                f"height:{height}px'></div>",
                unsafe_allow_html=True,
            )

        # ══ ORIGEM ══════════════════════════════════════════════════
        st.markdown(
            "<div style='display:flex;align-items:center;gap:6px;margin-bottom:4px'>"
            "<span style='width:10px;height:10px;border-radius:50%;"
            "background:#2E7D32;display:inline-block;flex-shrink:0'></span>"
            "<span style='font-size:10px;font-weight:700;color:#2E7D32;"
            "text-transform:uppercase;letter-spacing:0.8px'>Ponto de Origem</span>"
            + (" <span style='font-size:10px;color:#2E7D32'>✔</span>" if _orig_pronto else
               " <span style='font-size:9px;color:#aaa;font-style:italic'>— informe abaixo</span>")
            + "</div>",
            unsafe_allow_html=True,
        )
        _campo_rota_compacto(
            "UF · Cidade · Nome do posto · CNPJ",
            "txt_origem", "orig_sel",
            icon_bg="#2E7D32",
            action_help="Limpar origem",
        )
        orig_sel = st.session_state.get("orig_sel")

        # ══ PARADAS ═════════════════════════════════════════════════
        for _p_idx in range(1, _n_paradas + 1):
            _rail(color="#FF8F00", height=8)
            _parada_ok = bool(st.session_state.get(f"parada_sel_{_p_idx}"))
            st.markdown(
                f"<div style='display:flex;align-items:center;gap:6px;margin-bottom:4px'>"
                f"<span style='width:10px;height:10px;border-radius:50%;"
                f"background:#FF8F00;display:inline-block;flex-shrink:0'></span>"
                f"<span style='font-size:10px;font-weight:700;color:#E65100;"
                f"text-transform:uppercase;letter-spacing:0.8px'>Parada {_p_idx}</span>"
                + (f" <span style='font-size:10px;color:#E65100'>✔</span>" if _parada_ok else "")
                + "</div>",
                unsafe_allow_html=True,
            )
            _deleted = _campo_rota_compacto(
                "UF · Cidade · Nome do posto · CNPJ",
                f"txt_parada_{_p_idx}",
                f"parada_sel_{_p_idx}",
                icon_bg="#FF8F00",
                icon_number=str(_p_idx),
                always_show_action=True,
                action_key=f"btn_del_parada_{_p_idx}",
                action_help=f"Remover parada {_p_idx}",
            )
            if _deleted:
                for _j in range(_p_idx, _n_paradas):
                    st.session_state[f"parada_sel_{_j}"] = st.session_state.get(
                        f"parada_sel_{_j+1}")
                    st.session_state[f"txt_parada_{_j}"] = st.session_state.get(
                        f"txt_parada_{_j+1}", "")
                st.session_state.pop(f"parada_sel_{_n_paradas}", None)
                st.session_state.pop(f"txt_parada_{_n_paradas}", None)
                st.session_state["_paradas_count"] = _n_paradas - 1
                st.rerun()

        # ══ + PARADA ════════════════════════════════════════════════
        _rail(color="#BBDEFB", height=6)
        if _n_paradas < 10:
            _col_btn, _col_cnt2 = st.columns([3, 1])
            with _col_btn:
                if st.button(
                    "＋ Adicionar Parada",
                    key="btn_add_parada",
                    use_container_width=True,
                    help=f"Adicionar ponto intermediário ({_n_paradas}/10 usados)",
                ):
                    st.session_state["_paradas_count"] = _n_paradas + 1
                    st.rerun()
            with _col_cnt2:
                if _n_paradas > 0:
                    st.markdown(
                        f"<div style='padding-top:9px;font-size:10px;"
                        f"color:#90a4ae;text-align:center'>{_n_paradas}/10</div>",
                        unsafe_allow_html=True,
                    )
        _rail(color="#BBDEFB", height=6)

        # ══ DESTINO ═════════════════════════════════════════════════
        st.markdown(
            "<div style='display:flex;align-items:center;gap:6px;margin-bottom:4px'>"
            "<span style='width:10px;height:10px;border-radius:50%;"
            "background:#C62828;display:inline-block;flex-shrink:0'></span>"
            "<span style='font-size:10px;font-weight:700;color:#C62828;"
            "text-transform:uppercase;letter-spacing:0.8px'>Ponto de Destino</span>"
            + (" <span style='font-size:10px;color:#C62828'>✔</span>" if _dest_pronto else
               " <span style='font-size:9px;color:#aaa;font-style:italic'>— informe abaixo</span>")
            + "</div>",
            unsafe_allow_html=True,
        )
        _campo_rota_compacto(
            "UF · Cidade · Nome do posto · CNPJ",
            "txt_destino", "dest_sel",
            icon_bg="#C62828",
            action_help="Limpar destino",
        )
        dest_sel = st.session_state.get("dest_sel")
        st.divider()

        # ── Raio ────────────────────────────────────────────────────
        st.markdown(
            "<div style='font-size:10px;font-weight:700;color:#555;"
            "text-transform:uppercase;letter-spacing:0.8px;margin-bottom:4px'>"
            "📏 Raio de busca ao longo da rota</div>",
            unsafe_allow_html=True,
        )
        raio = st.slider("Raio (m)", min_value=200, max_value=2000, value=500, step=100,
                         label_visibility="collapsed",
                         help="Postos dentro deste raio ao redor da rota serão exibidos")
        st.caption(f"Buscando postos a até **{raio} m** da rota")

        # ── Botão Traçar Rota (com contexto) ───────────────────────
        if _orig_pronto and _dest_pronto:
            _o_lbl = str(st.session_state["orig_sel"].get("label",""))[:18]
            _d_lbl = str(st.session_state["dest_sel"].get("label",""))[:18]
            buscar_rota_btn = st.button(
                f"🗺️  {_o_lbl}  →  {_d_lbl}",
                use_container_width=True, type="primary",
                help="Calcular rota e exibir postos próximos",
            )
        elif not _orig_pronto:
            st.markdown(
                "<div style='background:#fff8e1;border:1px solid #ffe082;"
                "border-radius:8px;padding:8px 10px;font-size:11px;color:#f57f17;"
                "text-align:center'>⚠️ Informe o <b>Ponto de Origem</b> acima</div>",
                unsafe_allow_html=True,
            )
            buscar_rota_btn = False
        else:
            st.markdown(
                "<div style='background:#fff8e1;border:1px solid #ffe082;"
                "border-radius:8px;padding:8px 10px;font-size:11px;color:#f57f17;"
                "text-align:center'>⚠️ Informe o <b>Ponto de Destino</b> acima</div>",
                unsafe_allow_html=True,
            )
            buscar_rota_btn = False

        st.markdown("<div style='height:4px'></div>", unsafe_allow_html=True)
        if st.button("🗑️ Limpar tudo e recomeçar", use_container_width=True,
                     help="Remove os resultados e limpa todos os campos"):
            for _k in [
                "df_rota", "coords_rota",
                "lat_orig", "lon_orig", "label_orig",
                "lat_dest", "lon_dest", "label_dest",
                "dist_km", "dur_min", "raio_usado", "linha_reta",
                "distribuidoras_rota",
                "orig_sel", "dest_sel",
                "_orig_sel_txt_ant", "_dest_sel_txt_ant",
                "_paradas_data", "_ufs_rota_atual",
            ]:
                st.session_state.pop(_k, None)
            _n_p_clr = st.session_state.get("_paradas_count", 0)
            for _pi in range(1, _n_p_clr + 1):
                st.session_state.pop(f"parada_sel_{_pi}", None)
                st.session_state.pop(f"txt_parada_{_pi}", None)
            st.session_state["_paradas_count"] = 0
            st.session_state["_form_key"] = st.session_state.get("_form_key", 0) + 1
            st.rerun()

        distribuidoras_filtro = []
        if st.session_state.get("distribuidoras_rota"):
            st.divider()
            st.markdown("<div class='sb-label'>Filtrar por Bandeira</div>", unsafe_allow_html=True)
            distribuidoras_filtro = st.multiselect(
                "Bandeiras", st.session_state["distribuidoras_rota"],
                placeholder="Todas as bandeiras", label_visibility="collapsed")

        # Filtro de Perfil de Venda — Modo Rota
        perfis_filtro_m2 = []
        _perfis_lista_m2 = st.session_state.get("perfis_pf_lista", [])
        if _perfis_lista_m2:
            st.markdown("<div class='sb-label'>Perfil de Venda ⭐</div>", unsafe_allow_html=True)
            perfis_filtro_m2 = st.multiselect(
                "Perfil de Venda", _perfis_lista_m2,
                placeholder="Todos os perfis", label_visibility="collapsed",
                key="mult_perfil_m2",
                help="Filtra os postos Gestão de Frotas pelo perfil de venda.",
            )

        # ── Filtros Avançados — Modo Rota ─────────────────────────────────
        _pp_df_adv_m2 = st.session_state.get("_pp_df")
        _svc_cols_m2  = st.session_state.get("_servicos_cols_disponiveis", [])
        _tem_preco_m2   = _pp_df_adv_m2 is not None and "preco" in _pp_df_adv_m2.columns
        _tem_servico_m2 = bool(_svc_cols_m2)

        if _tem_preco_m2 or _tem_servico_m2:
            with st.expander("🔍 Filtros Avançados", expanded=False):
                _preco_faixa_m2 = None
                _fuel_sel_m2    = None
                if _tem_preco_m2:
                    _fuels_m2 = sorted(
                        _pp_df_adv_m2["combustivel_label"].dropna().str.strip().unique().tolist()
                    )
                    _fuel_sel_m2 = st.selectbox(
                        "⛽ Combustível (preço)",
                        ["— Todos —"] + _fuels_m2,
                        key="adv_fuel_m2",
                        label_visibility="visible",
                    )
                    if _fuel_sel_m2 and _fuel_sel_m2 != "— Todos —":
                        _df_fuel_m2 = _pp_df_adv_m2[
                            _pp_df_adv_m2["combustivel_label"].str.strip() == _fuel_sel_m2
                        ]
                        if not _df_fuel_m2.empty:
                            _pmin2 = float(_df_fuel_m2["preco"].min())
                            _pmax2 = float(_df_fuel_m2["preco"].max())
                            if _pmin2 < _pmax2:
                                _preco_faixa_m2 = st.slider(
                                    "💰 Faixa de Preço (R$/L)",
                                    min_value=round(_pmin2, 2),
                                    max_value=round(_pmax2, 2),
                                    value=(round(_pmin2, 2), round(_pmax2, 2)),
                                    step=0.01,
                                    format="R$ %.2f",
                                    key="adv_preco_m2",
                                    help="Exibe apenas postos com preço dentro da faixa selecionada",
                                )
                            else:
                                st.caption(f"Preço único: R$ {_pmin2:.3f}/L")

                _filtro_24h_m2 = False
                if "funciona_24h" in _svc_cols_m2:
                    _filtro_24h_m2 = st.checkbox(
                        "🕐 Somente postos 24h",
                        key="adv_24h_m2",
                        help="Filtra postos que funcionam 24 horas",
                    )

                _filtro_servicos_m2 = []
                _svc_pf_lbl_m2 = st.session_state.get("_servicos_pf_labels", {})
                _SVC_LEG_LBL_M2 = {"pista_caminhao": "🚛 Pista Caminhão",
                                    "arla":           "🧪 ARLA 32",
                                    "conveniencia":   "🛒 Conveniência"}
                _conceitos_cob_m2 = set(_svc_pf_lbl_m2.keys())
                _svc_map_m2: dict = {}
                # 1) Colunas dinâmicas da planilha ("Possui X?")
                for _ck, _lbl in _svc_pf_lbl_m2.items():
                    if _ck in _svc_cols_m2:
                        _svc_map_m2[_lbl] = _ck
                # 2) Legadas (se o conceito não foi substituído)
                for _ck, _lbl in _SVC_LEG_LBL_M2.items():
                    if _ck in _svc_cols_m2:
                        _sup = _SVC_LEGADO_SUPERSEDE.get(_ck, set())
                        if not (_sup & _conceitos_cob_m2):
                            _svc_map_m2[_lbl] = _ck
                _svc_opts_m2 = list(_svc_map_m2.keys())
                if _svc_opts_m2:
                    _svc_sel_m2 = st.multiselect(
                        "🔧 Serviços disponíveis",
                        _svc_opts_m2,
                        placeholder="Qualquer serviço",
                        key="adv_svc_m2",
                        help="Exibe apenas postos que oferecem os serviços selecionados",
                    )
                    _filtro_servicos_m2 = [_svc_map_m2[s] for s in _svc_sel_m2]
        else:
            _preco_faixa_m2     = None
            _fuel_sel_m2        = None
            _filtro_24h_m2      = False
            _filtro_servicos_m2 = []

    # ── Modo Roteirização — campos do veículo ─────────────────────────────────
    elif modo == "🛣️ Roteirização":

        st.markdown(
            "<div style='background:linear-gradient(135deg,#004D40,#00796B);"
            "border-radius:8px;padding:8px 12px;margin-bottom:10px'>"
            "<span style='color:#fff;font-weight:700;font-size:12px'>🛣️ Roteirização</span><br>"
            "<span style='color:#b2dfdb;font-size:10px'>Configure o veículo e trace a rota</span>"
            "</div>",
            unsafe_allow_html=True,
        )

        # ── Seletor de Perfis de Veículo ─────────────────────────────────
        _perfis_db = _db_perfis_veiculo()
        if _perfis_db:
            st.markdown("<div class='sb-label'>🚗 Perfis Salvos</div>", unsafe_allow_html=True)
            _perfis_opcoes = ["— Selecione um perfil —"] + [
                f"{p['nome']} · {p.get('placa','').upper() or '—'}" for p in _perfis_db
            ]
            _perfil_sel_idx = st.selectbox(
                "Perfil de veículo",
                range(len(_perfis_opcoes)),
                format_func=lambda i: _perfis_opcoes[i],
                key="rot_perfil_sel",
                label_visibility="collapsed",
            )
            if _perfil_sel_idx and _perfil_sel_idx > 0:
                _psel = _perfis_db[_perfil_sel_idx - 1]
                _c_load, _c_del = st.columns([3, 1])
                with _c_load:
                    if st.button("⬇️ Carregar perfil", use_container_width=True,
                                 key="btn_carregar_perfil"):
                        st.session_state["rot_placa"]      = _psel.get("placa", "")
                        st.session_state["rot_combustivel"] = _psel.get("combustivel", "")
                        st.session_state["rot_capacidade"] = float(_psel.get("tanque") or 80.0)
                        st.session_state["rot_autonomia"]  = float(_psel.get("autonomia") or 10.0)
                        st.toast(f"✅ Perfil **{_psel['nome']}** carregado!", icon="🚛")
                        st.rerun()
                with _c_del:
                    if st.button("🗑️", key="btn_del_perfil",
                                 help="Excluir este perfil"):
                        _db_deletar_perfil_veiculo(_psel["id"])
                        st.toast("Perfil excluído", icon="🗑️")
                        st.rerun()

        st.markdown("<div class='sb-label'>🚛 Dados do Veículo</div>", unsafe_allow_html=True)

        st.text_input(
            "Placa do Veículo",
            placeholder="Ex: ABC-1D23",
            key="rot_placa",
            help="Placa para identificar a roteirização salva",
        )

        # Opções de combustível — prioriza tipos do _pp_df
        _pp_combs_rot: list = []
        _pp_df_sidebar = st.session_state.get("_pp_df")
        if _pp_df_sidebar is not None and "combustivel_label" in _pp_df_sidebar.columns:
            _pp_combs_rot = sorted(
                _pp_df_sidebar["combustivel_label"].dropna().str.strip().unique().tolist()
            )
        if not _pp_combs_rot:
            _pp_combs_rot = [
                "GASOLINA COMUM", "GASOLINA ADITIVADA",
                "ÓLEO DIESEL", "ÓLEO DIESEL S10", "ETANOL",
            ]
        st.selectbox(
            "Combustível",
            _pp_combs_rot,
            key="rot_combustivel",
            help="Tipo de combustível a ser abastecido nos postos GF",
        )

        _c_cap, _c_aut = st.columns(2)
        with _c_cap:
            st.number_input(
                "Tanque (L)",
                min_value=10.0, max_value=1200.0,
                value=float(st.session_state.get("rot_capacidade") or 80.0),
                step=10.0, key="rot_capacidade",
                help="Capacidade total do tanque em litros (até 1.200 L)",
            )
        with _c_aut:
            st.number_input(
                "Autonomia (km/L)",
                min_value=1.0, max_value=40.0,
                value=float(st.session_state.get("rot_autonomia") or 10.0),
                step=0.5, key="rot_autonomia",
                help="Consumo médio do veículo em km por litro",
            )

        _cap_sb = float(st.session_state.get("rot_capacidade") or 80.0)
        _aut_sb = float(st.session_state.get("rot_autonomia") or 10.0)
        _min_sb = _cap_sb * 0.25
        _range_sb = (_cap_sb - _min_sb) * _aut_sb

        st.markdown(
            f"<div style='background:#e0f7fa;border:1px solid #80deea;border-radius:6px;"
            f"padding:7px 10px;font-size:11px;color:#004D40;margin-top:4px;line-height:1.6'>"
            f"⚠️ <b>Nível mínimo:</b> {_min_sb:.0f} L (25%)<br>"
            f"📏 <b>Alcance efetivo:</b> {_range_sb:.0f} km</div>",
            unsafe_allow_html=True,
        )

        # ── Salvar como perfil de veículo ────────────────────────────────
        st.markdown("<div style='height:6px'></div>", unsafe_allow_html=True)
        with st.expander("💾 Salvar como perfil", expanded=False):
            _nome_perfil = st.text_input(
                "Nome do perfil",
                placeholder="Ex: Scania R450 / Bitrem / Carreta",
                key="rot_nome_perfil",
                label_visibility="visible",
            )
            if st.button("💾 Salvar perfil", use_container_width=True,
                         key="btn_salvar_perfil_veiculo"):
                _placa_atual  = str(st.session_state.get("rot_placa") or "")
                _comb_atual   = str(st.session_state.get("rot_combustivel") or "")
                _tank_atual   = float(st.session_state.get("rot_capacidade") or 80.0)
                _aut_atual    = float(st.session_state.get("rot_autonomia") or 10.0)
                _nome_final   = _nome_perfil.strip() or _placa_atual or "Veículo"
                if _db_salvar_perfil_veiculo(_nome_final, _placa_atual, _comb_atual,
                                             _tank_atual, _aut_atual):
                    st.toast(f"✅ Perfil **{_nome_final}** salvo!", icon="🚛")
                    st.rerun()
                else:
                    st.error("❌ Erro ao salvar. Verifique a conexão com o banco.")

        st.caption("💡 Configure a origem, destino e paradas na área principal →")

    # ── Defaults para variáveis do Modo 2 quando outro modo está ativo ────────
    # Evita NameError quando o bloco elif "🗺️ Por Rota" não executou
    if "buscar_rota_btn" not in dir():
        buscar_rota_btn    = False
    if "raio" not in dir():
        raio               = 500
    if "distribuidoras_filtro" not in dir():
        distribuidoras_filtro = []
    if "perfis_filtro_m2" not in dir():
        perfis_filtro_m2   = []
    if "perfis_filtro_m1" not in dir():
        perfis_filtro_m1   = []
    if "filtro_24h_m1" not in dir():
        _filtro_24h_m1      = False
    if "_filtro_servicos_m1" not in dir():
        _filtro_servicos_m1 = []
    if "_preco_faixa_m1" not in dir():
        _preco_faixa_m1     = None
    if "_fuel_sel_m1" not in dir():
        _fuel_sel_m1        = None
    if "_filtro_24h_m2" not in dir():
        _filtro_24h_m2      = False
    if "_filtro_servicos_m2" not in dir():
        _filtro_servicos_m2 = []
    if "_preco_faixa_m2" not in dir():
        _preco_faixa_m2     = None
    if "_fuel_sel_m2" not in dir():
        _fuel_sel_m2        = None

    # ── Configurações (Gestão de Frotas · Cercados · Preços PP · Base · Exportar) ──
    st.markdown("---")
    _pf_fonte  = st.session_state.get("_pf_fonte",  "manual")
    _pf_set    = st.session_state.get("cnpjs_pro_frotas", set())
    _pf_ts     = st.session_state.get("_pf_carregado_em", "")
    _cer_set   = st.session_state.get("cnpjs_cercados",   set())
    _cer_fonte = st.session_state.get("_cercados_fonte",  "manual")
    _cer_ts    = st.session_state.get("_cercados_carregado_em", "")
    _pp_df_sb  = st.session_state.get("_pp_df")
    _pp_fonte  = st.session_state.get("_pp_fonte",  "manual")
    _pp_ts     = st.session_state.get("_pp_carregado_em", "")

    with st.expander("⚙️  Configurações", expanded=False):
        # ── Proteção por senha ────────────────────────────────────────────────
        if not st.session_state.get("_cfg_autenticado", False):
            _c1_lk, _c2_lk, _c3_lk = st.columns([1, 6, 1])
            with _c2_lk:
                st.markdown(
                    "<div style='text-align:center;padding:10px 0 6px'>"
                    "🔐 <b>Acesso restrito</b><br>"
                    "<span style='font-size:11px;color:#666'>"
                    "Informe a senha para acessar as configurações</span></div>",
                    unsafe_allow_html=True,
                )
                _senha_cfg = st.text_input(
                    "Senha de acesso",
                    type="password",
                    key="cfg_senha_input",
                    placeholder="Digite a senha…",
                    label_visibility="collapsed",
                )
                if st.button("🔓 Confirmar acesso", key="btn_cfg_entrar",
                             use_container_width=True):
                    if _senha_cfg == "Prototipo@2026":
                        st.session_state["_cfg_autenticado"] = True
                        st.session_state.pop("_cfg_senha_errada", None)
                        st.rerun()
                    else:
                        st.session_state["_cfg_senha_errada"] = True
                if st.session_state.get("_cfg_senha_errada", False):
                    st.error("❌ Senha incorreta. Tente novamente.")
            tab_pf = tab_cer = tab_pp = tab_base = tab_anp = tab_logs = tab_intel = None
        else:
            _col_cfg_lock, _ = st.columns([1, 5])
            with _col_cfg_lock:
                if st.button("🔒", key="btn_cfg_bloquear",
                             help="Bloquear Configurações", use_container_width=True):
                    st.session_state["_cfg_autenticado"] = False
                    st.session_state.pop("_cfg_senha_errada", None)
                    st.rerun()
            tab_pf, tab_cer, tab_pp, tab_base, tab_anp, tab_logs, tab_intel = st.tabs(
                ["⭐ Gestão de Frotas", "⚠️ Cercados", "💲 Preços PP",
                 "🗃️ Base", "🔵 Postos ANP", "📊 Logs de Uso", "🧠 Inteligência"]
            )

        # ── Tab Gestão de Frotas ────────────────────────────────────
        if tab_pf is not None:
         with tab_pf:
            _pf_ts_html = (f"<br><span style='font-size:10px;opacity:.8'>🕐 {_pf_ts}</span>"
                           if _pf_ts else "")
            if _pf_set:
                _c = ("#e8f5e9","#a5d6a7","#2e7d32","✅") if _pf_fonte == "repo" \
                     else ("#fff8e1","#ffe082","#f57f17","⭐")
                _src = (f"<span style='font-size:10px;opacity:.8'>"
                        f"Fonte: <code>{ARQUIVO_PF_REPO}</code></span>"
                        if _pf_fonte == "repo" else "")
                st.markdown(
                    f"<div style='background:{_c[0]};border:1px solid {_c[1]};"
                    f"border-radius:8px;padding:8px 11px;font-size:11px;color:{_c[2]}'>"
                    f"{_c[3]} <b>{_fmt_int(len(_pf_set))} CNPJs</b> carregados"
                    f"{_pf_ts_html}<br>{_src}</div>",
                    unsafe_allow_html=True,
                )
            else:
                st.markdown(
                    f"<div style='background:#fff3e0;border:1px solid #ffcc80;"
                    f"border-radius:8px;padding:8px 11px;font-size:11px;color:#e65100'>"
                    f"⚠️ <b>Não carregado</b><br>"
                    f"<span style='font-size:10px'>Adicione <code>{ARQUIVO_PF_REPO}</code> "
                    f"ou faça upload.</span></div>",
                    unsafe_allow_html=True,
                )
            st.markdown("")
            if st.button("🔄 Recarregar do repositório", use_container_width=True,
                         help="Força nova leitura do pro_frotas.xlsx no GitHub",
                         key="btn_reload_pf_cfg"):
                _auto_carregar_pro_frotas_repo.clear()
                with st.spinner(f"Lendo `{ARQUIVO_PF_REPO}`…"):
                    _cnpjs_r, _msg_r, _prev_r, _perfil_r, _coords_r = _auto_carregar_pro_frotas_repo()
                if _cnpjs_r:
                    st.session_state["cnpjs_pro_frotas"] = _cnpjs_r
                    st.session_state["_pf_fonte"]        = "repo"
                    st.session_state["_pf_carregado_em"] = _agora()
                    if _perfil_r:
                        st.session_state["perfil_venda_map"]  = _perfil_r
                        st.session_state["perfis_pf_lista"]   = sorted(set(_perfil_r.values()))
                    if _coords_r is not None and not _coords_r.empty:
                        st.session_state["pf_coords_df"] = _coords_r
                        _atualizar_servicos_pf(_coords_r)
                    st.success(f"✅ {_msg_r}")
                    time.sleep(1)
                    st.rerun()
                else:
                    st.error(_msg_r or f"❌ `{ARQUIVO_PF_REPO}` não encontrado.")
            st.markdown("<small><b>Upload manual</b></small>", unsafe_allow_html=True)
            arquivo_pf = st.file_uploader(
                "Planilha Gestão de Frotas", type=["xlsx","xls","csv"],
                key="upload_pf", label_visibility="collapsed",
            )
            if arquivo_pf is not None:
                _pf_file_id = f"{arquivo_pf.name}_{arquivo_pf.size}"
                if st.session_state.get("_pf_last_upload_id") != _pf_file_id:
                    with st.spinner("Lendo planilha…"):
                        cnpjs_pf, msg_pf, preview_pf, perfil_pf, coords_pf = ler_planilha_pro_frotas(arquivo_pf)
                    if cnpjs_pf is not None:
                        st.session_state["cnpjs_pro_frotas"]    = cnpjs_pf
                        st.session_state["_pf_fonte"]            = "manual"
                        st.session_state["_pf_carregado_em"]     = _agora()
                        st.session_state["_pf_last_upload_id"]   = _pf_file_id
                        if perfil_pf:
                            st.session_state["perfil_venda_map"] = perfil_pf
                            st.session_state["perfis_pf_lista"]  = sorted(set(perfil_pf.values()))
                        if coords_pf is not None and not coords_pf.empty:
                            st.session_state["pf_coords_df"] = coords_pf
                            _atualizar_servicos_pf(coords_pf)
                        st.success(msg_pf)
                        if preview_pf is not None:
                            with st.expander("Ver amostra dos CNPJs"):
                                st.dataframe(preview_pf, use_container_width=True)
                        st.rerun()
                    else:
                        st.error(msg_pf)
            if _pf_set:
                if st.button("🗑️ Remover Gestão de Frotas", use_container_width=True,
                             key="btn_rm_pf_cfg"):
                    st.session_state.pop("cnpjs_pro_frotas", None)
                    st.session_state.pop("_pf_fonte", None)
                    st.rerun()

        # ── Tab Postos Cercados ───────────────────────────────
        if tab_cer is not None:
         with tab_cer:
            _cer_ts_html = (f"<br><span style='font-size:10px;opacity:.8'>🕐 {_cer_ts}</span>"
                            if _cer_ts else "")
            if _cer_set:
                _cc = ("#fff8e1","#ffe082","#e65100") if _cer_fonte == "manual" \
                      else ("#fff3e0","#ffcc80","#bf360c")
                _src_cer = (
                    f"<span style='font-size:10px;opacity:.8'>"
                    f"Fonte: <code>{ARQUIVO_CERCADOS_REPO}</code></span>"
                    if _cer_fonte == "repo" else
                    "<span style='font-size:10px;opacity:.8'>Carregado manualmente</span>"
                )
                st.markdown(
                    f"<div style='background:{_cc[0]};border:1px solid {_cc[1]};"
                    f"border-radius:8px;padding:8px 11px;font-size:11px;color:{_cc[2]}'>"
                    f"⚠️ <b>{_fmt_int(len(_cer_set))} postos cercados</b> identificados"
                    f"{_cer_ts_html}<br>{_src_cer}</div>",
                    unsafe_allow_html=True,
                )
            else:
                st.markdown(
                    f"<div style='background:#f5f5f5;border:1px solid #ddd;"
                    f"border-radius:8px;padding:8px 11px;font-size:11px;color:#666'>"
                    f"Planilha não carregada.<br>"
                    f"<span style='font-size:10px'>Adicione <code>{ARQUIVO_CERCADOS_REPO}</code> "
                    f"ao repositório ou faça upload manual.</span></div>",
                    unsafe_allow_html=True,
                )
            st.markdown("")
            if st.button("🔄 Recarregar do repositório", use_container_width=True,
                         help=f"Força nova leitura de '{ARQUIVO_CERCADOS_REPO}' no GitHub",
                         key="btn_reload_cercados"):
                _auto_carregar_cercados_repo.clear()
                with st.spinner(f"Lendo `{ARQUIVO_CERCADOS_REPO}`…"):
                    _cnpjs_cer2, _msg_cer2, _ = _auto_carregar_cercados_repo()
                if _cnpjs_cer2:
                    st.session_state["cnpjs_cercados"]          = _cnpjs_cer2
                    st.session_state["_cercados_fonte"]         = "repo"
                    st.session_state["_cercados_carregado_em"]  = _agora()
                    st.success(f"✅ {_msg_cer2}")
                    time.sleep(1)
                    st.rerun()
                else:
                    st.error(_msg_cer2 or f"❌ `{ARQUIVO_CERCADOS_REPO}` não encontrado.")

            st.markdown("<small><b>Upload manual</b> — substitui nesta sessão:</small>",
                        unsafe_allow_html=True)
            arquivo_cer = st.file_uploader(
                "Planilha Postos Cercados", type=["xlsx","xls","csv"],
                key="upload_cercados", label_visibility="collapsed",
            )
            if arquivo_cer is not None:
                _cer_file_id = f"{arquivo_cer.name}_{arquivo_cer.size}"
                if st.session_state.get("_cer_last_upload_id") != _cer_file_id:
                    with st.spinner("Lendo planilha…"):
                        cnpjs_cer_up, msg_cer_up, prev_cer = ler_planilha_cercados(arquivo_cer)
                    if cnpjs_cer_up is not None:
                        st.session_state["cnpjs_cercados"]          = cnpjs_cer_up
                        st.session_state["_cercados_fonte"]         = "manual"
                        st.session_state["_cercados_carregado_em"]  = _agora()
                        st.session_state["_cer_last_upload_id"]     = _cer_file_id
                        st.success(msg_cer_up)
                        if prev_cer is not None:
                            with st.expander("Ver amostra"):
                                st.dataframe(prev_cer, use_container_width=True)
                        st.rerun()
                    else:
                        st.error(msg_cer_up)

            if _cer_set:
                if st.button("🗑️ Remover Cercados", use_container_width=True,
                             key="btn_rm_cercados"):
                    st.session_state.pop("cnpjs_cercados", None)
                    st.session_state.pop("_cercados_fonte", None)
                    st.session_state.pop("_cercados_carregado_em", None)
                    st.rerun()

        # ── Tab Preços PP ─────────────────────────────────────
        if tab_pp is not None:
         with tab_pp:
            _pp_ts_html = (f"<br><span style='font-size:10px;opacity:.8'>🕐 {_pp_ts}</span>"
                           if _pp_ts else "")
            if _pp_df_sb is not None:
                _pp_n2 = _pp_df_sb["cnpj_norm"].nunique() if "cnpj_norm" in _pp_df_sb.columns else 0
                _pp_c  = _pp_df_sb["combustivel_pk"].nunique() if "combustivel_pk" in _pp_df_sb.columns else 0
                _pp_src = (f"<span style='font-size:10px;opacity:.8'>Fonte: <code>{ARQUIVO_PP_REPO}</code></span>"
                           if _pp_fonte == "repo" else
                           "<span style='font-size:10px;opacity:.8'>Carregado manualmente</span>")
                st.markdown(
                    f"<div style='background:#e3f2fd;border:1px solid #90caf9;"
                    f"border-radius:8px;padding:8px 11px;font-size:11px;color:#1565c0'>"
                    f"💲 <b>{_fmt_int(_pp_n2)} postos</b> · {_pp_c} combustíveis"
                    f"{_pp_ts_html}<br>{_pp_src}</div>",
                    unsafe_allow_html=True,
                )
                # ── Diagnóstico de combustíveis detectados ──────────
                if "combustivel_pk" in _pp_df_sb.columns:
                    _pks_pp = sorted(_pp_df_sb["combustivel_pk"].dropna().unique().tolist())
                    _cnpjs_pf_diag = st.session_state.get("cnpjs_pro_frotas", set())
                    # Abre automaticamente se detectou menos de 4 combustíveis
                    _diag_auto = len(_pks_pp) < 4
                    with st.expander(
                        f"🔍 Combustíveis detectados: {len(_pks_pp)}",
                        expanded=_diag_auto,
                    ):
                        if _diag_auto:
                            st.warning(
                                "⚠️ Menos de 4 combustíveis detectados. "
                                "Verifique os nomes abaixo e os nomes das colunas da planilha."
                            )
                        st.caption("PKs lidos (após normalização) → mapeamento ANP:")
                        for _pk_d in _pks_pp:
                            _lbl_d = PRODUTO_CURTO.get(_pk_d) or PRODUTO_CURTO.get(
                                _PP_PARA_ANP_PK.get(_pk_d, ""), "—")
                            _n_rows = int((_pp_df_sb["combustivel_pk"] == _pk_d).sum())
                            _n_pf_d = 0
                            if _cnpjs_pf_diag:
                                _n_pf_d = int(
                                    _pp_df_sb[
                                        (_pp_df_sb["combustivel_pk"] == _pk_d) &
                                        (_pp_df_sb["cnpj_norm"].isin(_cnpjs_pf_diag))
                                    ]["cnpj_norm"].nunique()
                                )
                            _match_anp = _PP_PARA_ANP_PK.get(_pk_d, "❌ SEM MATCH")
                            _cor_match = "#2e7d32" if "❌" not in _match_anp else "#c62828"
                            st.markdown(
                                f"<div style='font-size:11px;padding:4px 0;"
                                f"border-bottom:1px solid #eee'>"
                                f"<b><code>{_pk_d}</code></b> → {_lbl_d} "
                                f"<span style='color:#888'>({_n_rows} linhas · {_n_pf_d} postos GF)</span>"
                                f"<br><span style='font-size:10px;color:{_cor_match}'>"
                                f"ANP: <code>{_match_anp}</code></span>"
                                f"</div>",
                                unsafe_allow_html=True,
                            )
                        # Nomes brutos dos labels (para identificar variantes)
                        if "combustivel_label" in _pp_df_sb.columns:
                            _labels_raw = sorted(
                                _pp_df_sb["combustivel_label"].dropna().unique().tolist()
                            )
                            st.caption(f"Nomes originais na planilha ({len(_labels_raw)}):")
                            st.code(", ".join(_labels_raw), language=None)
            else:
                st.markdown(
                    f"<div style='background:#f5f5f5;border:1px solid #ddd;"
                    f"border-radius:8px;padding:8px 11px;font-size:11px;color:#666'>"
                    f"Planilha não carregada.<br>"
                    f"<span style='font-size:10px'>Adicione <code>{ARQUIVO_PP_REPO}</code> "
                    f"ao repositório ou faça upload manual.</span></div>",
                    unsafe_allow_html=True,
                )
            st.markdown("")
            if st.button("🔄 Recarregar do repositório", use_container_width=True,
                         help=f"Força nova leitura de '{ARQUIVO_PP_REPO}' (cache 1 h)",
                         key="btn_reload_pp"):
                _auto_carregar_precos_postos_repo.clear()
                st.session_state["_pp_tentado"] = False
                with st.spinner(f"Lendo `{ARQUIVO_PP_REPO}`…"):
                    _pp_tmp, _pp_msg_tmp, _ = _auto_carregar_precos_postos_repo()
                if _pp_tmp is not None:
                    st.session_state["_pp_df"]          = _pp_tmp
                    st.session_state["_pp_fonte"]        = "repo"
                    st.session_state["_pp_carregado_em"] = _agora()
                    # ── Auto-registro no histórico de inteligência ──
                    try:
                        _hist_record_pp_df(_pp_tmp)
                    except Exception:
                        pass
                    st.success(f"✅ {_pp_msg_tmp}")
                    time.sleep(1)
                    st.rerun()
                else:
                    st.error(_pp_msg_tmp or f"❌ `{ARQUIVO_PP_REPO}` não encontrado.")
            st.markdown("<small><b>Upload manual</b> — substitui nesta sessão:</small>",
                        unsafe_allow_html=True)
            arquivo_pp = st.file_uploader(
                "Planilha Preço Posto", type=["xlsx","xls","csv"],
                key="upload_pp", label_visibility="collapsed",
            )
            if arquivo_pp is not None:
                # Evita loop infinito: só processa se for um arquivo novo
                _pp_file_id = f"{arquivo_pp.name}_{arquivo_pp.size}"
                if st.session_state.get("_pp_last_upload_id") != _pp_file_id:
                    with st.spinner("Lendo planilha…"):
                        _pp_up, _pp_msg_up, _ = ler_planilha_precos_postos(arquivo_pp)
                    if _pp_up is not None:
                        st.session_state["_pp_df"]             = _pp_up
                        st.session_state["_pp_fonte"]           = "manual"
                        st.session_state["_pp_carregado_em"]    = _agora()
                        st.session_state["_pp_last_upload_id"]  = _pp_file_id
                        # ── Auto-registro no histórico de inteligência ──
                        try:
                            _hist_record_pp_df(_pp_up)
                        except Exception:
                            pass
                        st.success(_pp_msg_up)
                        st.rerun()
                    else:
                        st.error(_pp_msg_up)
            if _pp_df_sb is not None:
                if st.button("🗑️ Remover Preços PP", use_container_width=True,
                             key="btn_rm_pp"):
                    st.session_state.pop("_pp_df", None)
                    st.session_state.pop("_pp_fonte", None)
                    st.session_state.pop("_pp_carregado_em", None)
                    st.session_state["_pp_tentado"] = False
                    st.rerun()

        # ── Tab Base Nacional ─────────────────────────────────
        if tab_base is not None:
         with tab_base:
            st.markdown(
                "<small>Carrega postos de <b>todos os 27 estados</b> antecipadamente "
                "e mantém em cache por <b>24 h</b>. "
                "Após isso, buscas por rota ficam <b>instantâneas</b>.</small>",
                unsafe_allow_html=True,
            )
            _preload_ts  = st.session_state.get("_preload_brasil_em", "")
            _preload_ok  = st.session_state.get("_preload_brasil_ok", 0)
            _preload_err = st.session_state.get("_preload_brasil_err", [])
            if _preload_ts:
                _cor_b = "#e8f5e9" if not _preload_err else "#fff8e1"
                _brd_b = "#a5d6a7" if not _preload_err else "#ffe082"
                _txt_b = "#2e7d32" if not _preload_err else "#f57f17"
                _ic_b  = "✅" if not _preload_err else "⚠️"
                _err_b = (f"<br><span style='font-size:10px'>Falha: {', '.join(_preload_err)}</span>"
                          if _preload_err else "")
                st.markdown(
                    f"<div style='background:{_cor_b};border:1px solid {_brd_b};"
                    f"border-radius:8px;padding:8px 11px;font-size:11px;color:{_txt_b};margin:6px 0'>"
                    f"{_ic_b} <b>{_preload_ok} estado(s)</b>{_err_b}<br>"
                    f"<span style='font-size:10px;opacity:.8'>🕐 {_preload_ts}</span>"
                    f"</div>",
                    unsafe_allow_html=True,
                )
            st.markdown("")
            if st.button("⚡ Carregar todos os estados",
                         use_container_width=True, key="btn_preload"):
                buscar_postos.clear()
                with st.spinner("📡 Recarregando base em paralelo…"):
                    carregados_pl, erros_pl = _precarregar_estados_paralelo(max_workers=5)
                st.session_state["_estados_precarregados"] = carregados_pl
                st.session_state["_preload_brasil_em"]     = _agora()
                st.session_state["_preload_brasil_ok"]     = len(carregados_pl)
                st.session_state["_preload_brasil_err"]    = erros_pl
                st.session_state["_base_auto_ok"]          = True
                if erros_pl:
                    st.warning(f"✅ {len(carregados_pl)} estados. ⚠️ Falha: {', '.join(erros_pl)}")
                else:
                    st.success(f"✅ Todos os {len(UFS)} estados carregados!")
                st.rerun()

            # ── Exportar (abaixo do botão Carregar) ──────────────
            st.markdown("---")
            st.markdown(
                "<small>📥 <b>Exportar:</b> gera <b>Excel (.xlsx)</b> com todos os postos "
                "dos estados carregados, com destaque para <b>Gestão de Frotas</b>.</small>",
                unsafe_allow_html=True,
            )
            st.markdown("")
            _n_est_exp = len(st.session_state.get("_estados_precarregados", []))
            if _n_est_exp == 0:
                st.warning("⚠️ Carregue a base primeiro (botão acima).")
            else:
                st.markdown(
                    f"<div style='font-size:11px;color:#555;margin-bottom:6px'>"
                    f"📦 <b>{_n_est_exp} estado(s)</b> disponíveis</div>",
                    unsafe_allow_html=True,
                )
                if st.button("📊 Gerar arquivo Excel",
                             use_container_width=True, key="btn_export_base"):
                    with st.spinner(f"⏳ Consolidando {_n_est_exp} estado(s)…"):
                        _exp_bytes, _exp_msg = _gerar_excel_base_brasil()
                    if _exp_bytes:
                        st.session_state["_base_export_bytes"] = _exp_bytes
                        st.session_state["_base_export_msg"]   = _exp_msg
                    else:
                        st.error(_exp_msg)
            if st.session_state.get("_base_export_bytes"):
                st.download_button(
                    label="⬇️  Baixar Excel",
                    data=st.session_state["_base_export_bytes"],
                    file_name=f"postos_anp_{datetime.now().strftime('%Y%m%d')}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True,
                    key="btn_download_base",
                )
                st.success(st.session_state.get("_base_export_msg", ""))
                if st.button("🗑️ Limpar exportação", use_container_width=True,
                             key="btn_clear_export"):
                    st.session_state.pop("_base_export_bytes", None)
                    st.session_state.pop("_base_export_msg", None)
                    st.rerun()


        # ── Tab Postos ANP ──────────────────────────────────────────
        if tab_anp is not None:
         with tab_anp:
            _anp_df_cfg = st.session_state.get("_anp_df_raw")
            _anp_ativo_cfg = _anp_df_cfg is not None
            if _anp_ativo_cfg:
                _n_anp_cfg = len(_anp_df_cfg)
                st.markdown(
                    f"<div style='background:#e3f2fd;border:1px solid #90caf9;"
                    f"border-radius:8px;padding:8px 11px;font-size:11px;"
                    f"color:#1565c0;margin-bottom:8px'>"
                    f"🔵 <b>{_fmt_int(_n_anp_cfg)} postos ANP</b> ativos como overlay em todos os modos"
                    f"</div>",
                    unsafe_allow_html=True,
                )
                if st.button("✕ Remover postos ANP", use_container_width=True,
                             key="btn_anp_rm_cfg"):
                    st.session_state.pop("_anp_df_raw",      None)
                    st.session_state.pop("_anp_upload_open", None)
                    st.session_state.pop("_uf_carregada",    None)
                    st.rerun()
            else:
                st.markdown(
                    "<div style='background:#fff3e0;border:1px solid #ffcc80;"
                    "border-radius:8px;padding:8px 11px;font-size:11px;color:#e65100;"
                    "margin-bottom:8px'>"
                    "⚠️ <b>Nenhum posto ANP carregado.</b><br>"
                    "<span style='font-size:10px'>Faça o upload do arquivo XLSX da ANP "
                    "para exibi-los como overlay no mapa.</span></div>",
                    unsafe_allow_html=True,
                )
                st.markdown(
                    "<div style='background:#e3f2fd;border:1px solid #90caf9;"
                    "border-radius:8px;padding:10px 12px;font-size:11px;color:#0d47a1;"
                    "margin-bottom:10px'>"
                    "📥 <b>Como carregar os postos ANP:</b><br>"
                    "① Acesse <a href='https://www.gov.br/anp/pt-br/assuntos/precos-e-defesa-da-concorrencia/"
                    "precos/precos-revenda-e-de-distribuicao-combustiveis/serie-historica-do-levantamento-de-precos' "
                    "target='_blank' style='color:#1565c0'>gov.br/anp</a> e baixe o arquivo XLSX<br>"
                    "② Extraia o arquivo da pasta ZIP, se necessário<br>"
                    "③ Faça o upload abaixo</div>",
                    unsafe_allow_html=True,
                )
                _anp_up_cfg = st.file_uploader(
                    "📎 Selecione o arquivo XLSX baixado da ANP",
                    type=["xlsx", "xls"],
                    key="anp_uploader_cfg",
                    help="Arquivo de postos revendedores da ANP (.xlsx)",
                )
                if _anp_up_cfg is not None:
                    with st.spinner("📂 Processando arquivo ANP…"):
                        _anp_r_cfg, _anp_msg_cfg = _processar_bytes_anp_postos(
                            _anp_up_cfg.name, _anp_up_cfg.read()
                        )
                    if _anp_r_cfg is not None:
                        st.session_state["_anp_df_raw"]   = _anp_r_cfg
                        st.session_state.pop("_uf_carregada", None)
                        st.success(f"✅ {_anp_msg_cfg}")
                        st.rerun()
                    else:
                        st.error(f"❌ {_anp_msg_cfg}")

        # ── Tab Logs de Uso ─────────────────────────────────────────────
        if tab_logs is not None:
         with tab_logs:
            _todos_logs = _log_ler_arquivo()

            if not _todos_logs:
                st.info("ℹ️ Nenhum evento registrado ainda. Os logs são gerados automaticamente "
                        "a cada acesso e interação com o app.")
            else:
                _log_df = pd.DataFrame(_todos_logs)

                # Garante colunas mínimas
                for _lc in _LOG_FIELDS:
                    if _lc not in _log_df.columns:
                        _log_df[_lc] = "—"

                # ── KPIs ────────────────────────────────────────────────
                _n_eventos    = len(_log_df)
                _n_sessoes    = _log_df["session_id"].nunique()
                _n_ips        = _log_df["ip"].replace("—", pd.NA).dropna().nunique()
                _log_df_hoje  = _log_df[_log_df["data"] == datetime.now().strftime("%d/%m/%Y")] \
                    if "data" in _log_df.columns else pd.DataFrame()
                _n_hoje       = len(_log_df_hoje)
                _emails_auth  = (_log_df["user_email"].replace("—", pd.NA).dropna()
                                 if "user_email" in _log_df.columns else pd.Series())
                _n_usuarios   = _emails_auth.nunique()

                _lk1, _lk2, _lk3, _lk4, _lk5 = st.columns(5)
                _lk1.metric("📋 Total de Eventos",  _fmt_int(_n_eventos))
                _lk2.metric("👤 Sessões Únicas",     _fmt_int(_n_sessoes))
                _lk3.metric("🔐 Usuários Google",    _fmt_int(_n_usuarios))
                _lk4.metric("🌐 IPs Únicos",         _fmt_int(_n_ips))
                _lk5.metric("📅 Eventos Hoje",       _fmt_int(_n_hoje))

                st.markdown("---")

                # ── Usuários autenticados ────────────────────────────────
                if "user_email" in _log_df.columns and _n_usuarios > 0:
                    st.markdown("##### 🔐 Usuários autenticados")
                    _df_logins = (
                        _log_df[_log_df["user_email"].replace("—", pd.NA).notna()]
                        .groupby(["user_email", "user_name", "auth_provider"])
                        .agg(
                            Acessos=("timestamp", "count"),
                            Último_acesso=("timestamp", "max"),
                        )
                        .reset_index()
                        .sort_values("Acessos", ascending=False)
                    )
                    _df_logins.columns = ["E-mail", "Nome", "Provider", "Acessos", "Último acesso"]
                    st.dataframe(
                        _df_logins.reset_index(drop=True),
                        use_container_width=True,
                        hide_index=True,
                        column_config={
                            "E-mail":        st.column_config.TextColumn("📧 E-mail"),
                            "Nome":          st.column_config.TextColumn("👤 Nome"),
                            "Provider":      st.column_config.TextColumn("🔑 Provider"),
                            "Acessos":       st.column_config.NumberColumn("Acessos", format="%d"),
                            "Último acesso": st.column_config.TextColumn("⏱️ Último acesso"),
                        },
                    )
                    st.markdown("---")

                # ── Top Modos ────────────────────────────────────────────
                _lt1, _lt2 = st.columns(2)

                with _lt1:
                    st.markdown("##### 🗂️ Ações mais frequentes")
                    _top_acoes = (
                        _log_df["acao"].value_counts().head(10)
                        .reset_index()
                    )
                    _top_acoes.columns = ["Ação", "Qtd"]
                    _fig_acao = go.Figure(go.Bar(
                        y=_top_acoes["Ação"],
                        x=_top_acoes["Qtd"],
                        orientation="h",
                        marker_color="#1565C0",
                        text=_top_acoes["Qtd"].astype(str),
                        textposition="outside",
                    ))
                    _fig_acao.update_layout(
                        height=max(250, len(_top_acoes) * 28 + 60),
                        margin=dict(l=10, r=40, t=20, b=10),
                        plot_bgcolor="rgba(0,0,0,0)",
                        paper_bgcolor="rgba(0,0,0,0)",
                        yaxis=dict(autorange="reversed"),
                        font=dict(size=11),
                    )
                    st.plotly_chart(_fig_acao, use_container_width=True)

                with _lt2:
                    st.markdown("##### 🗺️ Modos acessados")
                    _top_modos = (
                        _log_df[_log_df["modo"] != "—"]["modo"]
                        .value_counts().head(10)
                        .reset_index()
                    )
                    _top_modos.columns = ["Modo", "Qtd"]
                    if not _top_modos.empty:
                        _fig_modo = go.Figure(go.Bar(
                            y=_top_modos["Modo"],
                            x=_top_modos["Qtd"],
                            orientation="h",
                            marker_color="#00796B",
                            text=_top_modos["Qtd"].astype(str),
                            textposition="outside",
                        ))
                        _fig_modo.update_layout(
                            height=max(250, len(_top_modos) * 28 + 60),
                            margin=dict(l=10, r=40, t=20, b=10),
                            plot_bgcolor="rgba(0,0,0,0)",
                            paper_bgcolor="rgba(0,0,0,0)",
                            yaxis=dict(autorange="reversed"),
                            font=dict(size=11),
                        )
                        st.plotly_chart(_fig_modo, use_container_width=True)
                    else:
                        st.info("Sem dados de modo ainda.")

                # ── Top UFs ──────────────────────────────────────────────
                st.markdown("##### 📍 UFs mais consultadas")
                _top_ufs = (
                    _log_df[_log_df["uf"].replace("—", pd.NA).notna()]["uf"]
                    .value_counts().head(15)
                    .reset_index()
                )
                _top_ufs.columns = ["UF", "Acessos"]
                if not _top_ufs.empty:
                    _fig_uf = go.Figure(go.Bar(
                        x=_top_ufs["UF"],
                        y=_top_ufs["Acessos"],
                        marker_color="#E65100",
                        text=_top_ufs["Acessos"].astype(str),
                        textposition="outside",
                    ))
                    _fig_uf.update_layout(
                        height=280,
                        margin=dict(l=10, r=10, t=20, b=30),
                        plot_bgcolor="rgba(0,0,0,0)",
                        paper_bgcolor="rgba(0,0,0,0)",
                        font=dict(size=11),
                    )
                    st.plotly_chart(_fig_uf, use_container_width=True)

                # ── Tabela de eventos recentes ───────────────────────────
                st.markdown("##### 🕐 Últimos 50 eventos")
                _log_display = _log_df.tail(50).iloc[::-1].copy()
                # Garante colunas na ordem correta para exibição
                _cols_show = [c for c in [
                    "timestamp", "user_name", "user_email", "auth_provider",
                    "modo", "uf", "municipio", "acao", "detalhe", "ip",
                ] if c in _log_display.columns]
                _rename_log = {
                    "timestamp":     "Data/Hora",
                    "user_name":     "Usuário",
                    "user_email":    "E-mail",
                    "auth_provider": "Provider",
                    "modo":          "Modo",
                    "uf":            "UF",
                    "municipio":     "Município",
                    "acao":          "Ação",
                    "detalhe":       "Detalhe",
                    "ip":            "IP",
                }
                st.dataframe(
                    _log_display[_cols_show].rename(columns=_rename_log).reset_index(drop=True),
                    use_container_width=True,
                    hide_index=True,
                    column_config={
                        "Data/Hora": st.column_config.TextColumn("🕐 Data/Hora", width="medium"),
                        "Usuário":   st.column_config.TextColumn("👤 Usuário",   width="medium"),
                        "E-mail":    st.column_config.TextColumn("📧 E-mail",    width="medium"),
                        "Provider":  st.column_config.TextColumn("🔑",           width="small"),
                        "Modo":      st.column_config.TextColumn("Modo",         width="medium"),
                        "Ação":      st.column_config.TextColumn("Ação",         width="small"),
                        "Detalhe":   st.column_config.TextColumn("Detalhe",      width="large"),
                    },
                )

                # ── Exportar ──────────────────────────────────────────────
                st.markdown("---")
                _exp_l1, _exp_l2, _exp_l3 = st.columns([3, 1, 1])

                with _exp_l2:
                    # Exporta como XLSX (evita problema de vírgulas no user_agent)
                    _cols_export = [c for c in _LOG_FIELDS if c in _log_df.columns]
                    _rename_export = {
                        "timestamp": "Data/Hora", "data": "Data", "hora": "Hora",
                        "ip": "IP", "session_id": "Sessão",
                        "modo": "Modo", "uf": "UF", "municipio": "Município",
                        "acao": "Ação", "detalhe": "Detalhe", "user_agent": "Navegador",
                        "user_email": "E-mail", "user_name": "Usuário",
                        "auth_provider": "Provider",
                    }
                    import io as _io_mod
                    _xlsx_buf = _io_mod.BytesIO()
                    (_log_df[_cols_export]
                     .rename(columns=_rename_export)
                     .to_excel(_xlsx_buf, index=False, engine="openpyxl"))
                    _xlsx_buf.seek(0)
                    st.download_button(
                        label="📥 Exportar Excel",
                        data=_xlsx_buf.getvalue(),
                        file_name=f"logs_uso_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        use_container_width=True,
                        help="Baixar todos os eventos em Excel (.xlsx)",
                    )

                with _exp_l3:
                    if st.button("🗑️ Limpar Logs", use_container_width=True,
                                 key="btn_limpar_logs",
                                 help="Remove o arquivo de log do servidor"):
                        try:
                            if _os_mod.path.exists(_LOG_PATH):
                                _os_mod.remove(_LOG_PATH)
                            st.session_state.pop("_uso_logs", None)
                            st.success("✅ Logs limpos.")
                            st.rerun()
                        except Exception as _e_log:
                            st.error(f"❌ Não foi possível limpar: {_e_log}")

                with _exp_l1:
                    _arquivo_info = ""
                    try:
                        if _os_mod.path.exists(_LOG_PATH):
                            _sz = _os_mod.path.getsize(_LOG_PATH)
                            _arquivo_info = (f"📁 Arquivo: `_usage_logs.csv` · "
                                             f"{_sz/1024:.1f} KB · {_n_eventos} eventos")
                        else:
                            _arquivo_info = "⚠️ Arquivo não encontrado (apenas sessão em memória)"
                    except Exception:
                        _arquivo_info = "⚠️ Sem acesso ao arquivo"
                    st.caption(_arquivo_info)

        # ── Tab Inteligência de Dados ─────────────────────────────────────────
        if tab_intel is not None:
         with tab_intel:
            st.markdown("#### 🧠 Inteligência de Dados")

            # ── Seção 1: Histórico de Preços ──────────────────────────────────
            st.markdown("##### 📈 Histórico de Preços por Posto")
            _intel_d = _intel_load()
            _hist_all = _intel_d.get("historico", {})
            _n_cnpjs  = len(_hist_all)
            _n_obs    = sum(len(v) for v in _hist_all.values())

            if _n_obs == 0:
                st.info(
                    "Nenhum histórico registrado ainda.\n\n"
                    "Os preços são registrados automaticamente cada vez que você carrega "
                    "a **planilha de Preços PP** na aba *💲 Preços PP*. "
                    "Após algumas semanas, o histórico começa a mostrar a evolução dos preços."
                )
            else:
                _c_h1, _c_h2, _c_h3 = st.columns(3)
                _c_h1.metric("📍 Postos rastreados", _fmt_int(_n_cnpjs))
                _c_h2.metric("📊 Observações totais", _fmt_int(_n_obs))
                _datas_all = [e["data"] for v in _hist_all.values() for e in v]
                _dt_min = min(_datas_all) if _datas_all else "—"
                _dt_max = max(_datas_all) if _datas_all else "—"
                _c_h3.metric("📅 Período", f"{_dt_min} → {_dt_max}")

                # Busca por CNPJ
                st.markdown("**Visualizar histórico de um posto**")
                _cnpj_hist_inp = st.text_input(
                    "CNPJ do posto (somente números)",
                    key="intel_cnpj_hist_inp",
                    placeholder="Ex: 12345678000199",
                    max_chars=18,
                )
                if _cnpj_hist_inp:
                    _cnpj_h_n = re.sub(r"\D", "", _cnpj_hist_inp)
                    _hist_posto = _hist_all.get(_cnpj_h_n, [])
                    if not _hist_posto:
                        st.warning("Nenhum histórico encontrado para este CNPJ.")
                    else:
                        _nome_h = _hist_posto[0].get("nome", f"Posto {_cnpj_h_n}")
                        _fig_h  = _hist_chart_posto(_cnpj_h_n, _nome_h)
                        if _fig_h:
                            st.markdown(
                                f"<p style='font-weight:600;font-size:0.95rem;"
                                f"margin:0 0 4px 0;color:var(--text-color,#1a1a2e)'>"
                                f"📈 Evolução de preços — {_nome_h}</p>",
                                unsafe_allow_html=True,
                            )
                            st.plotly_chart(_fig_h, use_container_width=True)
                        _df_h = pd.DataFrame(_hist_posto).sort_values("data", ascending=False)
                        _df_h = _df_h.rename(columns={
                            "data":"Data","preco":"Preço (R$/L)",
                            "combustivel":"Combustível","municipio":"Município","uf":"UF"})
                        st.dataframe(_df_h.head(20), use_container_width=True, height=220)

                # Botão para registrar preços atuais da planilha PP
                st.markdown("---")
                _pp_df_intel = st.session_state.get("_pp_df")
                if _pp_df_intel is not None and not _pp_df_intel.empty:
                    if st.button("🔄 Registrar preços atuais da planilha PP no histórico",
                                 key="btn_intel_registrar_pp",
                                 use_container_width=True):
                        _n_reg = _hist_record_pp_df(_pp_df_intel)
                        st.success(f"✅ {_n_reg} novas observações registradas.")
                        st.rerun()
                else:
                    st.caption("Carregue a planilha de Preços PP para ativar o registro de histórico.")

            # ── Seção 2: Score de Postos ──────────────────────────────────────
            st.markdown("---")
            st.markdown("##### ⭐ Score de Postos")
            st.markdown(
                "O **Score** é calculado automaticamente na tabela de dados de cada modo. "
                "Ele combina três fatores:"
            )
            _sc1, _sc2, _sc3 = st.columns(3)
            with _sc1:
                st.markdown(
                    "<div style='background:#e3f2fd;border-radius:10px;padding:10px 12px'>"
                    "<b>💰 Preço vs ANP</b><br>"
                    "<span style='font-size:12px;color:#555'>50% do score</span><br>"
                    "<span style='font-size:11px'>Quanto mais barato que a média ANP "
                    "do município/estado, maior a pontuação.</span></div>",
                    unsafe_allow_html=True)
            with _sc2:
                st.markdown(
                    "<div style='background:#e8f5e9;border-radius:10px;padding:10px 12px'>"
                    "<b>🛒 Serviços</b><br>"
                    "<span style='font-size:12px;color:#555'>30% do score</span><br>"
                    "<span style='font-size:11px'>Quantidade de serviços disponíveis "
                    "(conveniência, ARLA 32, restaurante, etc.).</span></div>",
                    unsafe_allow_html=True)
            with _sc3:
                st.markdown(
                    "<div style='background:#fff8e1;border-radius:10px;padding:10px 12px'>"
                    "<b>📍 Distância</b><br>"
                    "<span style='font-size:12px;color:#555'>20% do score</span><br>"
                    "<span style='font-size:11px'>Proximidade ao ponto de busca "
                    "ou à rota selecionada.</span></div>",
                    unsafe_allow_html=True)
            st.markdown(
                "<div style='margin-top:10px;font-size:12px;color:#555'>"
                "🟢 <b>A</b> ≥ 75 pts &nbsp;|&nbsp; "
                "🔵 <b>B</b> 55–74 &nbsp;|&nbsp; "
                "🟡 <b>C</b> 35–54 &nbsp;|&nbsp; "
                "🔴 <b>D</b> &lt; 35"
                "</div>",
                unsafe_allow_html=True)

            # ── Seção 3: Relatório de Alertas ─────────────────────────────────
            st.markdown("---")
            st.markdown("##### ⚠️ Relatório Semanal de Alertas de Preço")

            _intel_d2    = _intel_load()
            _limiar_cfg  = _intel_d2.get("limiar", {})
            _last_report = _intel_d2.get("last_report")

            if _last_report:
                st.caption(f"📋 Último relatório gerado: {_last_report[:16].replace('T',' ')}")

            st.markdown(
                "Defina o **limiar máximo** de preço para cada combustível. "
                "Postos acima deste valor aparecerão como alertas no relatório."
            )

            # Limiares por combustível
            _COMBS_LIMIAR = [
                ("GASOLINA COMUM",    "⛽ Gasolina Comum",    5.80),
                ("GASOLINA ADITIVADA","⛽ Gasolina Aditivada", 6.20),
                ("ETANOL HIDRATADO",  "🌿 Etanol",             4.00),
                ("DIESEL S10",        "🚛 Diesel S10",         6.00),
                ("DIESEL S500",       "🚛 Diesel S500",        5.90),
            ]
            _lim_novo = {}
            _lc1, _lc2 = st.columns(2)
            for _ci, (_ck, _clbl, _cdef) in enumerate(_COMBS_LIMIAR):
                _col_lim = _lc1 if _ci % 2 == 0 else _lc2
                with _col_lim:
                    _lim_novo[_ck] = st.number_input(
                        _clbl,
                        min_value=0.0, max_value=20.0,
                        value=float(_limiar_cfg.get(_ck, _cdef)),
                        step=0.01, format="%.3f",
                        key=f"intel_lim_{_ck}",
                    )

            _col_lim_btn1, _col_lim_btn2 = st.columns([1, 1])
            with _col_lim_btn1:
                if st.button("💾 Salvar limiares", key="btn_intel_salvar_lim",
                             use_container_width=True):
                    _intel_d2["limiar"] = _lim_novo
                    _intel_save(_intel_d2)
                    st.session_state.pop("_intel_loaded", None)
                    st.success("✅ Limiares salvos.")

            with _col_lim_btn2:
                _pp_df_rep = st.session_state.get("_pp_df")
                if st.button("📄 Gerar Relatório de Alertas (.xlsx)",
                             key="btn_intel_gerar_rel",
                             use_container_width=True,
                             type="primary"):
                    if _pp_df_rep is None or _pp_df_rep.empty:
                        st.warning("⚠️ Carregue a planilha de Preços PP antes de gerar o relatório.")
                    else:
                        with st.spinner("Gerando relatório…"):
                            _bytes_rel, _fname_rel, _err_rel = _gerar_relatorio_alertas_xlsx(
                                _pp_df_rep, _lim_novo)
                        if _err_rel:
                            st.error(f"❌ Erro: {_err_rel}")
                        else:
                            st.session_state["_intel_rel_bytes"] = _bytes_rel
                            st.session_state["_intel_rel_fname"] = _fname_rel
                            st.session_state.pop("_intel_loaded", None)
                            st.rerun()

            # Botão de download aparece após geração
            _bytes_dl = st.session_state.get("_intel_rel_bytes")
            _fname_dl = st.session_state.get("_intel_rel_fname", "alertas.xlsx")
            if _bytes_dl:
                st.download_button(
                    "⬇️ Baixar relatório gerado",
                    data=_bytes_dl,
                    file_name=_fname_dl,
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True,
                    key="btn_intel_download_rel",
                )

    # ── Guia de Uso ───────────────────────────────────────────────────────────
    _col_guia_l, _col_guia_c, _col_guia_r = st.columns([1, 4, 1])
    with _col_guia_c:
        if st.button(
            "❓ Guia de Uso",
            use_container_width=True,
            type="secondary",
            key="btn_tour_sidebar",
            help="Abrir o guia interativo passo a passo",
        ):
            st.session_state["_tour_ativo"] = True
            st.rerun()

    # ── README — visualização do PDF de documentação ─────────────────────────
    _doc_bytes, _doc_nome = _carregar_doc_pdf()
    st.markdown("<div style='height:8px'></div>", unsafe_allow_html=True)
    if _doc_bytes:
        _col_doc_l, _col_doc_c, _col_doc_r = st.columns([1, 4, 1])
        with _col_doc_c:
            if st.button(
                "📄 README — Documentação",
                use_container_width=True,
                key="btn_doc_pdf",
                help="Visualizar documentação da aplicação",
            ):
                st.session_state["_doc_aberto"] = not st.session_state.get("_doc_aberto", False)
                st.rerun()

        # Exibe o PDF inline se aberto — renderizado como imagem (sem iframe, sem bloqueio)
        if st.session_state.get("_doc_aberto") and _doc_bytes:
            # Botão de download sempre disponível
            st.download_button(
                "⬇️ Baixar PDF",
                data=_doc_bytes,
                file_name=_doc_nome,
                mime="application/pdf",
                key="btn_doc_download",
            )
            try:
                import fitz  # pymupdf
                _pdf_doc   = fitz.open(stream=_doc_bytes, filetype="pdf")
                _total_pgs = len(_pdf_doc)

                if "doc_pagina_atual" not in st.session_state:
                    st.session_state["doc_pagina_atual"] = 0
                _pg_idx = st.session_state["doc_pagina_atual"]

                # Navegação
                _col_prev, _col_info, _col_next = st.columns([1, 2, 1])
                with _col_prev:
                    if st.button("◀ Anterior", use_container_width=True,
                                 disabled=_pg_idx == 0, key="doc_pg_prev"):
                        st.session_state["doc_pagina_atual"] -= 1
                        st.rerun()
                with _col_info:
                    st.markdown(
                        f"<p style='text-align:center;margin:6px 0;color:#555;font-size:13px'>"
                        f"Página <b>{_pg_idx + 1}</b> de <b>{_total_pgs}</b></p>",
                        unsafe_allow_html=True,
                    )
                with _col_next:
                    if st.button("Próxima ▶", use_container_width=True,
                                 disabled=_pg_idx >= _total_pgs - 1, key="doc_pg_next"):
                        st.session_state["doc_pagina_atual"] += 1
                        st.rerun()

                # Renderiza página atual como imagem PNG
                _page     = _pdf_doc[_pg_idx]
                _mat      = fitz.Matrix(2.0, 2.0)
                _pix      = _page.get_pixmap(matrix=_mat, alpha=False)
                _img_bytes = _pix.tobytes("png")
                _pdf_doc.close()
                st.image(_img_bytes, use_container_width=True)

            except ImportError:
                st.info("📄 Para ver o PDF aqui, instale: `pip install pymupdf`")
    else:
        _col_doc_l, _col_doc_c, _col_doc_r = st.columns([1, 4, 1])
        with _col_doc_c:
            st.markdown(
                "<div style='text-align:center;font-size:11px;color:#9E9E9E;padding:4px 0'>"
                "📄 Documentação não encontrada</div>",
                unsafe_allow_html=True,
            )

# ═══════════════════════════════════════════════════════════════════
#  TOUR DE ONBOARDING
# ═══════════════════════════════════════════════════════════════════

_TOUR_STEPS = [
    {
        "icon": "👋", "title": "Bem-vindo ao Estudo de Rede!",
        "desc": (
            "Esta plataforma permite **visualizar e analisar a rede de postos credenciados** "
            "à frota, comparar preços da ANP, planejar rotas e acompanhar a **evolução histórica "
            "de preços e score de qualidade** de cada posto. "
            "Este tour rápido mostra o essencial em menos de 2 minutos."
        ),
        "visual": [
            ("📍","Postos por estado","#e3f2fd"),
            ("🗺️","Consulta por rota","#e8f5e9"),
            ("📊","Dashboard analítico","#fff8e1"),
            ("💰","Ranking de preços","#fce4ec"),
            ("🔍","Filtros avançados","#f3e5f5"),
            ("🧠","Inteligência de dados","#ede7f6"),
            ("💾","Salve consultas","#e0f2f1"),
        ],
        "tips": ["📍 7 modos de consulta disponíveis", "🧠 Novo: histórico de preços e score por posto"],
    },
    {
        "icon": "📍", "title": "Modo 1 · Consulta por Estado (UF)",
        "desc": (
            "Selecione um **estado (UF)** na barra lateral para visualizar todos os postos "
            "credenciados. Filtre também por **município** para resultados mais precisos. "
            "Postos Gestão de Frotas aparecem em **azul e amarelo** no mapa."
        ),
        "visual": [
            ("1️⃣","Escolha o Estado na sidebar","#e3f2fd"),
            ("🏙️","Filtre por Município (opcional)","#e8f5e9"),
            ("🗺️","Mapa com todos os postos","#fff8e1"),
            ("⭐","Postos GF destacados","#fce4ec"),
        ],
        "tips": ["💡 Filtre por município para resultados mais precisos", "⭐ Postos GF têm marcadores maiores"],
    },
    {
        "icon": "🗺️", "title": "Modo 2 · Consulta por Rota",
        "desc": (
            "Defina **Origem** e **Destino** — o sistema calcula a rota e exibe todos "
            "os postos dentro do **raio configurável** (padrão: 500 m). "
            "Adicione **paradas intermediárias** para rotas mais complexas."
        ),
        "visual": [
            ("🟢","Defina ponto de Origem","#e8f5e9"),
            ("🔴","Defina ponto de Destino","#fce4ec"),
            ("🛣️","Rota calculada via OSRM","#e3f2fd"),
            ("⛽","Postos no raio da rota","#fff8e1"),
        ],
        "tips": ["🛑 Adicione paradas intermediárias", "📏 Raio de busca ajustável na sidebar"],
    },
    {
        "icon": "🔍", "title": "Filtros Avançados",
        "desc": (
            "Expanda **Filtros Avançados** na sidebar para refinar a busca. "
            "Filtre por **faixa de preço (R$/L)**, postos **abertos 24h**, "
            "e **serviços**: Pista Caminhão, ARLA 32 e Conveniência."
        ),
        "visual": [
            ("💲","Faixa de preço por combustível","#e8f5e9"),
            ("⏰","Funcionamento 24 horas","#e3f2fd"),
            ("🚛","Pista para Caminhão","#fff8e1"),
            ("🧪","ARLA 32 disponível","#fce4ec"),
        ],
        "tips": ["⚠️ Postos sem dado de serviço não são excluídos", "🔗 Filtros se combinam entre si"],
    },
    {
        "icon": "📊", "title": "Preços ANP e Tendências",
        "desc": (
            "Carregue o **.xlsx semanal da ANP** (ou o sistema busca automaticamente). "
            "Veja preços médios por combustível com indicadores de **tendência semanal ↑ ↓ ≈** "
            "e compare o preço do posto GF vs a média ANP do estado."
        ),
        "visual": [
            ("⛽","Gasolina  R$6,12  ↑","#ffebee"),
            ("🛢️","Diesel S10  R$6,48  ↓","#e8f5e9"),
            ("🧪","ARLA 32  R$3,21  ≈","#f3f4f6"),
            ("✅","GF R$0,23/L abaixo ANP","#e8f5e9"),
        ],
        "tips": ["🔄 O app busca o arquivo ANP automaticamente", "📁 Ou faça upload manual nas Configurações"],
    },
    {
        "icon": "💰", "title": "Ranking Top 5 Mais Baratos",
        "desc": (
            "Os **5 postos com menor preço** da consulta são destacados com "
            "**estrelas douradas ★** no mapa e com **cards de medalha** na aba Dados Tabulares. "
            "Selecione um combustível nos Filtros para um ranking mais preciso."
        ),
        "visual": [
            ("🥇","1º mais barato — estrela no mapa","#fff9c4"),
            ("🥈","2º mais barato","#fff9c4"),
            ("🥉","3º mais barato","#fff9c4"),
            ("🏅","4º e 5º completam o top 5","#fff9c4"),
        ],
        "tips": ["🎯 Filtre por combustível para ranking mais preciso", "⭐ Marcadores dourados no mapa"],
    },
    {
        "icon": "📊", "title": "Dashboard Analítico",
        "desc": (
            "Acesse o **Dashboard** na sidebar para ver KPIs de cobertura, "
            "penetração GF, comparativo GF vs ANP por estado e análise de preços. "
            "Exporte os dados em **CSV** para seus relatórios."
        ),
        "visual": [
            ("⛽","Total de postos na rede","#e3f2fd"),
            ("⭐","Postos Gestão de Frotas","#fff8e1"),
            ("📈","Penetração GF por estado","#e8f5e9"),
            ("💰","Economia média vs ANP","#fce4ec"),
        ],
        "tips": ["📤 Exporte os dados em CSV", "🗺️ Comparativo por estado disponível"],
    },
    {
        "icon": "🧠", "title": "Inteligência de Dados",
        "desc": (
            "O módulo **🧠 Inteligência** rastreia preços semana a semana e avalia cada posto "
            "com um **score A–D** baseado em preço (50%), serviços (30%) e distância (20%). "
            "Gere também **relatórios de alerta** com postos acima do limite de preço configurado."
        ),
        "visual": [
            ("📈","Histórico semanal de preços","#ede7f6"),
            ("🟢","Score A — posto ideal","#e8f5e9"),
            ("🟡","Score C — preço elevado","#fff8e1"),
            ("⚠️","Alerta: acima da média ANP","#fce4ec"),
        ],
        "tips": [
            "💡 Preços registrados automaticamente ao carregar a planilha PP",
            "📥 Baixe o relatório de alertas em Excel (.xlsx)",
        ],
    },
]


@st.dialog("🗺️ Guia de Uso — Estudo de Rede de Frotas", width="large")
def _tour_dialog():
    _step   = st.session_state.get("_tour_step", 0)
    _total  = len(_TOUR_STEPS)
    _s      = _TOUR_STEPS[_step]

    # ── Progresso ─────────────────────────────────────────────────────
    _dots_html = "".join(
        f"<span style='display:inline-block;width:10px;height:10px;border-radius:50%;"
        f"background:{'#1565c0' if i==_step else '#ddd'};"
        f"margin:0 3px;transition:background .2s'></span>"
        for i in range(_total)
    )
    st.markdown(
        f"<div style='text-align:center;margin:-8px 0 10px'>{_dots_html}</div>"
        f"<div style='text-align:center;font-size:11px;color:#999;margin-bottom:14px'>"
        f"Passo {_step+1} de {_total}</div>",
        unsafe_allow_html=True,
    )

    # ── Título + descrição ─────────────────────────────────────────────
    st.markdown(
        f"<div style='font-size:20px;font-weight:800;color:#1a1a1a;margin-bottom:10px'>"
        f"{_s['icon']}  {_s['title']}</div>",
        unsafe_allow_html=True,
    )
    st.markdown(_s["desc"])

    # ── Grid visual ────────────────────────────────────────────────────
    _vis = _s["visual"]
    _cols = st.columns(len(_vis))
    for _ci, (_em, _lbl, _bg) in enumerate(_vis):
        _cols[_ci].markdown(
            f"<div style='background:{_bg};border-radius:10px;padding:14px 10px;"
            f"text-align:center;border:1px solid rgba(0,0,0,.06);min-height:88px;"
            f"display:flex;flex-direction:column;align-items:center;justify-content:center'>"
            f"<div style='font-size:28px;margin-bottom:6px'>{_em}</div>"
            f"<div style='font-size:11px;color:#333;font-weight:600;line-height:1.35'>{_lbl}</div>"
            f"</div>",
            unsafe_allow_html=True,
        )

    # ── Tips ───────────────────────────────────────────────────────────
    st.markdown("<div style='margin-top:12px'>" +
        "".join(
            f"<span style='display:inline-flex;align-items:center;background:#e3f2fd;"
            f"border:1px solid #90caf9;border-radius:8px;padding:5px 11px;"
            f"font-size:11px;color:#0d47a1;margin:3px 4px 0 0'>{t}</span>"
            for t in _s["tips"]
        ) + "</div>", unsafe_allow_html=True)

    st.markdown("---")

    # ── Navegação ──────────────────────────────────────────────────────
    _c1, _c2, _c3 = st.columns([1, 1, 2])
    with _c1:
        if _step > 0:
            if st.button("← Anterior", use_container_width=True, key="tour_prev"):
                st.session_state["_tour_step"] = _step - 1
                st.rerun()
    with _c2:
        if st.button("Pular tour", use_container_width=True, key="tour_skip",
                     help="Fechar o guia sem marcar como concluído"):
            st.session_state["_tour_ativo"] = False
            st.session_state.pop("_tour_step", None)
            st.rerun()
    with _c3:
        _lbl_next = "✅ Concluir e fechar" if _step == _total - 1 else f"Próximo → ({_step+2}/{_total})"
        if st.button(_lbl_next, type="primary", use_container_width=True, key="tour_next"):
            if _step == _total - 1:
                _marcar_tour_concluido()
                st.session_state["_tour_ativo"] = False
                st.session_state.pop("_tour_step", None)
            else:
                st.session_state["_tour_step"] = _step + 1
            st.rerun()


if st.session_state.get("_tour_ativo", False):
    _tour_dialog()
    # garante step inicial
    if "_tour_step" not in st.session_state:
        st.session_state["_tour_step"] = 0


# ── BLOCO ANTIGO REMOVIDO (era st.markdown com script JS — não funciona no Streamlit) ──
if False:
    st.markdown("""
<style>
#gf-tour-overlay{position:fixed;inset:0;z-index:99990;background:rgba(10,20,40,.72);
  backdrop-filter:blur(3px);display:flex;align-items:center;justify-content:center;
  animation:gfTFadeIn .3s ease}
@keyframes gfTFadeIn{from{opacity:0}to{opacity:1}}
#gf-tour-card{background:#fff;border-radius:18px;width:min(660px,94vw);max-height:90vh;
  overflow:hidden;display:flex;flex-direction:column;
  box-shadow:0 24px 60px rgba(0,0,0,.35);
  animation:gfTSlide .35s cubic-bezier(.22,.68,0,1.2)}
@keyframes gfTSlide{from{transform:translateY(40px);opacity:0}to{transform:none;opacity:1}}
#gft-hdr{background:linear-gradient(135deg,#1565c0,#0d47a1);padding:18px 22px 14px;
  display:flex;align-items:center;justify-content:space-between;flex-shrink:0}
#gft-icon{font-size:26px;line-height:1;margin-right:12px}
#gft-title{color:#fff;font-size:15px;font-weight:700;margin:0}
#gft-sub{color:rgba(255,255,255,.72);font-size:11px;margin:2px 0 0}
#gft-x{background:rgba(255,255,255,.15);border:none;border-radius:50%;
  width:30px;height:30px;cursor:pointer;color:#fff;font-size:16px;
  display:flex;align-items:center;justify-content:center;transition:background .2s;flex-shrink:0}
#gft-x:hover{background:rgba(255,255,255,.28)}
#gft-dots{display:flex;justify-content:center;gap:7px;padding:12px 0 0;flex-shrink:0}
.gft-dot{width:8px;height:8px;border-radius:50%;background:#e0e0e0;
  transition:all .25s;cursor:pointer}
.gft-dot.on{background:#1565c0;transform:scale(1.3)}
#gft-body{padding:18px 26px;overflow-y:auto;flex:1}
.gft-step{display:none}
.gft-step.on{display:block;animation:gftStepIn .22s ease}
@keyframes gftStepIn{from{opacity:0;transform:translateX(14px)}to{opacity:1;transform:none}}
#gft-foot{padding:12px 22px 16px;display:flex;align-items:center;
  justify-content:space-between;border-top:1px solid #f0f0f0;flex-shrink:0;gap:8px;
  flex-wrap:wrap}
#gft-counter{font-size:11px;color:#999}
.gft-chk{display:flex;align-items:center;gap:5px;font-size:11px;color:#888;cursor:pointer}
.gft-chk input{accent-color:#1565c0}
.gft-btns{display:flex;gap:8px;align-items:center}
.gft-btn{border:none;border-radius:8px;padding:8px 20px;font-size:13px;
  font-weight:700;cursor:pointer;transition:all .2s}
#gft-skip{font-size:11px;color:#bbb;background:none;border:none;cursor:pointer;padding:4px 8px}
#gft-skip:hover{color:#666}
#gft-prev{background:#f0f0f0;color:#555}
#gft-prev:hover{background:#e0e0e0}
#gft-next{background:#1565c0;color:#fff;box-shadow:0 3px 10px rgba(21,101,192,.3)}
#gft-next:hover{background:#0d47a1;transform:translateY(-1px)}
</style>
<div id="gf-tour-mount"></div>
<script>
(function(){
  var STEPS = [
    { icon:'👋', title:'Bem-vindo ao Estudo de Rede!',
      desc:'Esta plataforma permite <b>visualizar e analisar a rede de postos credenciados</b> '+
           'a frota, comparar precos da ANP, planejar rotas e muito mais. '+
           'Este tour rapido vai mostrar o essencial em menos de 2 minutos.',
      visual:[
        {em:'📍',label:'Postos por estado',bg:'#e3f2fd'},
        {em:'🗺️',label:'Consulta por rota',bg:'#e8f5e9'},
        {em:'📊',label:'Dashboard analitico',bg:'#fff8e1'},
        {em:'💰',label:'Comparacao de precos',bg:'#fce4ec'},
        {em:'🔍',label:'Filtros avancados',bg:'#f3e5f5'},
        {em:'💾',label:'Salve consultas',bg:'#e0f2f1'}
      ],
      tips:['📍 Postos por estado','🗺️ Consulta por rota','📊 Dashboard analitico','💰 Ranking top 5']
    },
    { icon:'📍', title:'Modo 1 · Consulta por Estado (UF)',
      desc:'Selecione um <b>estado (UF)</b> na barra lateral para visualizar todos os postos '+
           'credenciados. Filtre por <b>municipio</b> para resultados mais precisos. '+
           'Postos Gestao de Frotas aparecem em <b>azul e amarelo</b> destacados no mapa.',
      visual:[
        {em:'1️⃣',label:'Escolha o Estado na sidebar',bg:'#e3f2fd'},
        {em:'🏙️',label:'Opcional: filtre por Municipio',bg:'#e8f5e9'},
        {em:'🗺️',label:'Mapa carrega com todos os postos',bg:'#fff8e1'},
        {em:'⭐',label:'Postos GF destacados em azul/amarelo',bg:'#fce4ec'}
      ],
      tips:['💡 Filtre por municipio para resultados mais precisos','⭐ Postos GF sao destacados visualmente']
    },
    { icon:'🗺️', title:'Modo 2 · Consulta por Rota',
      desc:'Defina <b>Origem</b> e <b>Destino</b> — o sistema calcula a rota e exibe todos '+
           'os postos dentro do <b>raio configuravel</b> (padrao: 500 m). '+
           'Adicione <b>paradas intermediarias</b> para rotas mais complexas.',
      visual:[
        {em:'🟢',label:'Defina ponto de Origem',bg:'#e8f5e9'},
        {em:'🔴',label:'Defina ponto de Destino',bg:'#fce4ec'},
        {em:'🛣️',label:'Rota calculada automaticamente',bg:'#e3f2fd'},
        {em:'⛽',label:'Postos no raio da rota',bg:'#fff8e1'}
      ],
      tips:['🛑 Adicione paradas intermediarias','📏 Raio de busca configuravel na sidebar']
    },
    { icon:'🔍', title:'Filtros Avancados',
      desc:'Na barra lateral, expanda <b>Filtros Avancados</b> para refinar a busca. '+
           'Filtre por <b>faixa de preco (R$/L)</b>, postos <b>abertos 24h</b>, '+
           'e <b>servicos</b>: Pista Caminhao, ARLA 32 e Conveniencia.',
      visual:[
        {em:'💲',label:'Faixa de preco por combustivel',bg:'#e8f5e9'},
        {em:'⏰',label:'Funcionamento 24 horas',bg:'#e3f2fd'},
        {em:'🚛',label:'Pista para Caminhao',bg:'#fff8e1'},
        {em:'🧪',label:'ARLA 32 disponivel',bg:'#fce4ec'}
      ],
      tips:['⚠️ Postos sem dado de servico nao sao excluidos','🔗 Filtros se combinam entre si']
    },
    { icon:'📊', title:'Precos ANP e Tendencias',
      desc:'Carregue o <b>.xlsx semanal da ANP</b> (ou o sistema busca automaticamente). '+
           'Veja precos medios por combustivel com indicadores de <b>tendencia semanal: ↑ ↓ ≈</b>. '+
           'Compare o preco do posto GF vs a media da ANP no estado.',
      visual:[
        {em:'⛽',label:'Gasolina  R$6,12  ↑ +0,08',bg:'#ffebee'},
        {em:'🛢️',label:'Diesel S10  R$6,48  ↓ -0,05',bg:'#e8f5e9'},
        {em:'🧪',label:'ARLA 32  R$3,21  ≈ estavel',bg:'#f3f4f6'},
        {em:'✅',label:'GF R$0,23/L abaixo ANP',bg:'#e8f5e9'}
      ],
      tips:['🔄 O app busca o arquivo ANP automaticamente','📁 Ou faca upload manual nas Configuracoes']
    },
    { icon:'💰', title:'Ranking Top 5 Mais Baratos',
      desc:'Os <b>5 postos com menor preco</b> da consulta atual sao destacados com '+
           '<b>estrelas douradas ★</b> no mapa e com <b>cards de medalha</b> na aba Dados Tabulares. '+
           'Selecione um combustivel nos Filtros para um ranking mais preciso.',
      visual:[
        {em:'🥇',label:'1o mais barato — estrela no mapa',bg:'#fff9c4'},
        {em:'🥈',label:'2o mais barato',bg:'#fff9c4'},
        {em:'🥉',label:'3o mais barato',bg:'#fff9c4'},
        {em:'🏅',label:'4o e 5o completam o top 5',bg:'#fff9c4'}
      ],
      tips:['🎯 Filtre por combustivel para ranking mais preciso','⭐ Marcadores dourados no mapa']
    },
    { icon:'📊', title:'Dashboard Analitico',
      desc:'Acesse o <b>Dashboard</b> na sidebar para ver KPIs de cobertura, '+
           'penetracao GF, comparativo GF vs ANP por estado e analise de precos. '+
           'Exporte os dados em <b>CSV</b> para seus relatorios.',
      visual:[
        {em:'⛽',label:'Total de postos na rede',bg:'#e3f2fd'},
        {em:'⭐',label:'Postos Gestao de Frotas',bg:'#fff8e1'},
        {em:'📈',label:'Penetracao GF por estado',bg:'#e8f5e9'},
        {em:'💰',label:'Economia media vs ANP',bg:'#fce4ec'}
      ],
      tips:['📤 Exporte os dados em CSV','🗺️ Comparativo por estado disponivel']
    }
  ];

  var cur = 0;
  var mount = document.getElementById('gf-tour-mount');
  if(!mount) return;

  // Constroi HTML do overlay via JS (evita sanitizacao do Streamlit)
  var ov = document.createElement('div');
  ov.id = 'gf-tour-overlay';

  var card = document.createElement('div');
  card.id = 'gf-tour-card';
  ov.appendChild(card);

  // Header
  var hdr = document.createElement('div');
  hdr.id = 'gft-hdr';
  hdr.innerHTML =
    '<div style="display:flex;align-items:center">' +
    '  <span id="gft-icon">👋</span>' +
    '  <div><div id="gft-title">Guia de Uso — Estudo de Rede de Frotas</div>' +
    '  <div id="gft-sub">Tour rapido · 7 passos · 2 minutos</div></div>' +
    '</div>' +
    '<button id="gft-x" title="Fechar">✕</button>';
  card.appendChild(hdr);

  // Dots
  var dotsWrap = document.createElement('div');
  dotsWrap.id = 'gft-dots';
  STEPS.forEach(function(_,i){
    var d = document.createElement('div');
    d.className = 'gft-dot' + (i===0?' on':'');
    d.addEventListener('click', function(){ show(i); });
    dotsWrap.appendChild(d);
  });
  card.appendChild(dotsWrap);

  // Body
  var body = document.createElement('div');
  body.id = 'gft-body';
  STEPS.forEach(function(s, i){
    var step = document.createElement('div');
    step.className = 'gft-step' + (i===0?' on':'');

    // Titulo
    var tDiv = document.createElement('div');
    tDiv.style.cssText = 'font-size:18px;font-weight:800;color:#1a1a1a;margin:0 0 10px;display:flex;align-items:center;gap:10px';
    tDiv.textContent = s.title;
    step.appendChild(tDiv);

    // Descricao
    var dDiv = document.createElement('p');
    dDiv.style.cssText = 'font-size:14px;color:#444;line-height:1.65;margin:0 0 16px';
    dDiv.innerHTML = s.desc;
    step.appendChild(dDiv);

    // Visual grid
    var grid = document.createElement('div');
    grid.style.cssText = 'display:grid;grid-template-columns:repeat(auto-fit,minmax(130px,1fr));gap:10px;margin-bottom:14px';
    s.visual.forEach(function(v){
      var cell = document.createElement('div');
      cell.style.cssText = 'background:' + v.bg + ';border-radius:10px;padding:14px 12px;text-align:center;border:1px solid rgba(0,0,0,.06)';
      cell.innerHTML = '<div style="font-size:26px;margin-bottom:6px">' + v.em + '</div>' +
                       '<div style="font-size:11px;color:#333;font-weight:600;line-height:1.4">' + v.label + '</div>';
      grid.appendChild(cell);
    });
    step.appendChild(grid);

    // Tips
    var tips = document.createElement('div');
    tips.style.cssText = 'margin-top:4px';
    s.tips.forEach(function(t){
      var tip = document.createElement('span');
      tip.style.cssText = 'display:inline-flex;align-items:center;background:#e3f2fd;border:1px solid #90caf9;'+
        'border-radius:8px;padding:5px 11px;font-size:11px;color:#0d47a1;margin:3px 3px 0 0';
      tip.textContent = t;
      tips.appendChild(tip);
    });
    step.appendChild(tips);

    body.appendChild(step);
  });
  card.appendChild(body);

  // Footer
  var foot = document.createElement('div');
  foot.id = 'gft-foot';
  foot.innerHTML =
    '<div style="display:flex;align-items:center;gap:12px">' +
    '  <span id="gft-counter">Passo 1 de 7</span>' +
    '  <label class="gft-chk"><input type="checkbox" id="gft-chk"> Nao mostrar novamente</label>' +
    '</div>' +
    '<div class="gft-btns">' +
    '  <button class="gft-btn" id="gft-skip">Pular tour</button>' +
    '  <button class="gft-btn" id="gft-prev">← Anterior</button>' +
    '  <button class="gft-btn" id="gft-next">Proximo →</button>' +
    '</div>';
  card.appendChild(foot);

  mount.appendChild(ov);

  // Logica de navegacao
  var allSteps = body.querySelectorAll('.gft-step');
  var allDots  = dotsWrap.querySelectorAll('.gft-dot');
  var counter  = document.getElementById('gft-counter');
  var btnPrev  = document.getElementById('gft-prev');
  var btnNext  = document.getElementById('gft-next');
  var btnSkip  = document.getElementById('gft-skip');
  var btnX     = document.getElementById('gft-x');
  var iconEl   = document.getElementById('gft-icon');

  function show(n){
    allSteps.forEach(function(s,i){ s.classList.toggle('on', i===n); });
    allDots.forEach(function(d,i){ d.classList.toggle('on', i===n); });
    counter.textContent = 'Passo '+(n+1)+' de '+STEPS.length;
    btnPrev.style.display = n===0 ? 'none' : '';
    btnNext.textContent   = n===STEPS.length-1 ? '✅ Concluir' : 'Proximo →';
    iconEl.textContent    = STEPS[n].icon;
    cur = n;
  }

  function fechar(gravar){
    ov.style.animation = 'gfTFadeIn .18s ease reverse';
    setTimeout(function(){ if(ov.parentNode) ov.parentNode.removeChild(ov); }, 200);
    if(gravar){
      var b = document.querySelector('.st-key-btn_tour_done_hidden button');
      if(b) b.click();
    }
  }

  btnNext.addEventListener('click', function(){
    if(cur < STEPS.length-1){ show(cur+1); }
    else { fechar(true); }
  });
  btnPrev.addEventListener('click',  function(){ if(cur>0) show(cur-1); });
  btnSkip.addEventListener('click',  function(){ fechar(document.getElementById('gft-chk').checked); });
  btnX.addEventListener('click',     function(){ fechar(document.getElementById('gft-chk').checked); });

  show(0);
})();
</script>
""", unsafe_allow_html=True)


# ═══════════════════════════════════════════════════════════════════
#  MODO 1 — Por Estado / Município
# ═══════════════════════════════════════════════════════════════════

if modo == "📍 Por UF/Município":

    if uf:
        # ── Carrega postos GF do estado diretamente da planilha (sem chamada à API ANP) ──
        if uf != st.session_state.get("_uf_carregada"):
            _pf_df_m1 = st.session_state.get("pf_coords_df", pd.DataFrame())
            if not _pf_df_m1.empty:
                df_raw_full = _pf_df_m1[
                    _pf_df_m1["uf"].fillna("").str.upper().str.strip() == uf.upper()
                ].copy().reset_index(drop=True)
            else:
                df_raw_full = pd.DataFrame()
                st.warning(
                    "⚠️ Planilha Gestão de Frotas não carregada ou sem coordenadas. "
                    "Verifique a seção **Configurações** na barra lateral."
                )

            st.session_state["df_raw_full"]  = df_raw_full
            st.session_state["_uf_carregada"] = uf
            if not df_raw_full.empty and "distribuidora" in df_raw_full.columns:
                st.session_state["distribuidoras_disponiveis"] = sorted(
                    df_raw_full["distribuidora"].dropna().unique().tolist())
            st.session_state.pop("_df_marcado",      None)
            st.session_state.pop("_df_marcado_key",  None)
            st.session_state.pop("_pf_injetados_uf", None)

        # ── Lê df_raw_full do session_state ───────────────────────────
        df_raw_full = st.session_state.get("df_raw_full", pd.DataFrame())

        # Filtra por município localmente (instantâneo, sem nova chamada à API)
        # Usa _sem_acento para que "Vitoria" encontre "VITÓRIA", "Ribeirao" → "RIBEIRÃO" etc.
        mun = municipio_input.strip()
        if mun:
            _mun_norm = _sem_acento(mun)
            df_raw = df_raw_full[
                df_raw_full["municipio"].fillna("").apply(
                    lambda x: _mun_norm in _sem_acento(x)
                )
            ].copy()
        else:
            df_raw = df_raw_full

        df_show = preparar_df(
            df_raw, distribuidoras_filtro,
            perfis_filtro=perfis_filtro_m1,
            filtro_servicos=_filtro_servicos_m1,
            filtro_24h=_filtro_24h_m1,
        )

        # ── Filtro de Preço — pós-processamento via _pp_df ───────────────
        if _preco_faixa_m1 and _fuel_sel_m1 and _fuel_sel_m1 != "— Todos —":
            _pp_m1 = st.session_state.get("_pp_df")
            if _pp_m1 is not None and "_cnpj_norm" in df_show.columns:
                _fuel_df_m1 = _pp_m1[
                    _pp_m1["combustivel_label"].str.strip() == _fuel_sel_m1
                ][["cnpj_norm","preco"]].copy()
                _lo_m1, _hi_m1 = _preco_faixa_m1
                _cnpj_ok_m1 = set(
                    _fuel_df_m1[
                        _fuel_df_m1["preco"].between(_lo_m1, _hi_m1)
                    ]["cnpj_norm"]
                )
                # Mantém postos sem preço cadastrado + postos dentro da faixa
                _sem_preco_m1 = ~df_show["_cnpj_norm"].isin(_fuel_df_m1["cnpj_norm"])
                df_show = df_show[_sem_preco_m1 | df_show["_cnpj_norm"].isin(_cnpj_ok_m1)]

        # ── Filtro Apenas Favoritos ──────────────────────────────────────────
        if _filtro_apenas_favoritos_m1 and st.session_state.get("fav_cnpjs"):
            _fav_set_m1 = st.session_state["fav_cnpjs"]
            _cnpj_col   = "_cnpj_norm" if "_cnpj_norm" in df_show.columns else "cnpj"
            df_show = df_show[
                df_show[_cnpj_col].fillna("").str.replace(r"\D", "", regex=True).isin(_fav_set_m1)
            ]

        # ── Ranking Top 5 Mais Baratos ──────────────────────────────────────
        _fuel_rank_m1 = _fuel_sel_m1 if (_fuel_sel_m1 and _fuel_sel_m1 != "— Todos —") else None
        _top5_m1 = _calcular_top5_baratos(df_show, fuel_label=_fuel_rank_m1)
        df_show  = _aplicar_rank_barato(df_show, _top5_m1)

        # ── Overlay ANP: adiciona postos do arquivo carregado pelo usuário ──
        # O mapa já contém apenas postos GF (da planilha). O overlay acrescenta
        # postos não-GF do arquivo ANP, somente quando o usuário o inseriu.
        _anp_view_m1 = st.session_state.get("_anp_df_raw")
        if _anp_view_m1 is not None:
            _anp_uf_view1 = _anp_view_m1[
                _anp_view_m1["uf"].fillna("").str.upper().str.strip() == uf.upper()
            ].copy()
            # Aplica filtro de município no overlay ANP, igual ao filtro dos postos GF
            if mun and not _anp_uf_view1.empty:
                _mun_norm_anp = _sem_acento(mun)
                _anp_uf_view1 = _anp_uf_view1[
                    _anp_uf_view1["municipio"].fillna("").apply(
                        lambda x: _mun_norm_anp in _sem_acento(x)
                    )
                ]
            if not _anp_uf_view1.empty:
                _pf_cnpjs_view1 = set(
                    df_show["cnpj"].fillna("").str.replace(r"\D", "", regex=True)
                )
                _anp_uf_view1 = _anp_uf_view1[
                    ~_anp_uf_view1["cnpj"].fillna("").str.replace(r"\D", "", regex=True)
                     .isin(_pf_cnpjs_view1)
                ]
                if not _anp_uf_view1.empty:
                    _anp_view1_prep = preparar_df(_anp_uf_view1, distribuidoras_filtro)
                    df_show = pd.concat([df_show, _anp_view1_prep], ignore_index=True)

        # ── Métricas ──────────────────────────────────────────
        _n_total_show = len(df_show)
        _n_mapa_show  = min(_n_total_show, MAX_MAPA_POSTOS)
        _mapa_label   = (f"⛽ Postos ({_n(MAX_MAPA_POSTOS)} no mapa)"
                         if _n_total_show > MAX_MAPA_POSTOS else "⛽ Postos")
        c1, c2, c3, c4 = st.columns(4)
        c1.metric(_mapa_label,          _n(_n_total_show))
        c2.metric("⭐ Credenciados PF",  _n(n_pf(df_show)))
        c3.metric("🏷️ Bandeiras",       _n(df_show['distribuidora'].nunique()) if not df_show.empty else "0")
        c4.metric("📍 Estado",          uf)

        # ── Botão Salvar (Modo 1) ─────────────────────────────────────
        _col_sv1, _col_sv2 = st.columns([3, 1])
        with _col_sv1:
            _nome_sugerido_m1 = f"Estado {uf}" + (f" — {municipio_input.strip()}" if municipio_input.strip() else "")
            _nome_salvar_m1 = st.text_input(
                "Nome da consulta",
                value=_nome_sugerido_m1,
                key="nome_salvar_m1",
                label_visibility="collapsed",
                placeholder="Nome para identificar esta consulta…",
            )
        with _col_sv2:
            if st.button("💾 Salvar", use_container_width=True, key="btn_salvar_m1",
                         help="Salvar esta consulta para acessar depois"):
                _dados_m1 = {
                    "uf": uf,
                    "municipio": municipio_input.strip(),
                    "_map_orig": st.session_state.get("_map_orig"),
                    "_map_dest": st.session_state.get("_map_dest"),
                    "_map_rota_result": (
                        {k: v for k, v in st.session_state.get("_map_rota_result", {}).items()
                         if k != "coords"}  # coords pode ser grande; recalcula ao restaurar
                        if st.session_state.get("_map_rota_result") else None
                    ),
                }
                if _salvar_rota_nova(_nome_salvar_m1 or _nome_sugerido_m1, "estado", _dados_m1):
                    st.toast(f"✅ Consulta **{_nome_salvar_m1 or _nome_sugerido_m1}** salva!", icon="💾")
                else:
                    st.error("❌ Não foi possível salvar. Verifique permissões do diretório.")

        # ── Banner de filtros ativos ──────────────────────────────────────
        _filtros_ativos_m1 = []
        if distribuidoras_filtro:
            _filtros_ativos_m1.append(f"Bandeira: {', '.join(distribuidoras_filtro)}")
        if _filtro_24h_m1:
            _filtros_ativos_m1.append("⏰ 24h")
        if _filtro_servicos_m1:
            _srv_label = {"pista_caminhao":"🚛 Pista","arla":"🧪 ARLA","conveniencia":"🛒 Conv.",
                          **{k: v for k, v in st.session_state.get("_servicos_pf_labels", {}).items()}}
            _filtros_ativos_m1 += [_srv_label.get(s,s) for s in _filtro_servicos_m1]
        if _preco_faixa_m1 and _fuel_sel_m1 and _fuel_sel_m1 != "— Todos —":
            _lo_lbl, _hi_lbl = _preco_faixa_m1
            _filtros_ativos_m1.append(f"💰 {_fuel_sel_m1}: R$ {_brl(_lo_lbl,2)}–{_brl(_hi_lbl,2)}/L")
        if _filtros_ativos_m1:
            st.info(
                "🔍 **Filtros ativos:** " + " · ".join(_filtros_ativos_m1)
                + f" &nbsp;|&nbsp; **{_n(len(df_show))} postos** exibidos",
                icon=None,
            )

        tab_mapa, tab_dados, tab_analise = st.tabs([
            "🗺️  Mapa Interativo", "📋  Dados Tabulares",
            "📊  Análise por Bandeira"])

        with tab_mapa:
            with st.spinner(f"🗺️ Carregando mapa — {_n(len(df_show))} postos…"):
                try:
                    _mapa_obj = criar_mapa(df_show)
                    _renderizar_mapa(_mapa_obj, height=660, key="mapa_m1_main")
                except Exception as _mapa_err:
                    st.error(
                        f"❌ Erro ao gerar o mapa para **{uf}**.\n\n"
                        f"**Detalhe:** `{type(_mapa_err).__name__}: {_mapa_err}`\n\n"
                        "Tente recarregar a página ou selecionar outro estado."
                    )
                    import traceback
                    st.code(traceback.format_exc(), language="python")

            # ── Exportar mapa como PNG ─────────────────────────────
            _mc1, _mc2 = st.columns([5, 1])
            with _mc2:
                if st.button("📸 Exportar mapa", use_container_width=True,
                             key="btn_exp_mapa_m1",
                             help="Baixar o mapa atual como imagem PNG"):
                    with st.spinner("🖼️ Gerando imagem…"):
                        try:
                            _titulo_png = f"Postos — {uf}" + (f" · {municipio_input.strip()}" if municipio_input.strip() else "")
                            _sub_png    = f"{len(df_show)} postos  ·  {datetime.now().strftime('%d/%m/%Y %H:%M')}"
                            _png_m1 = _exportar_mapa_postos_png(df_show, _titulo_png, _sub_png)
                            st.session_state["_mapa_png_m1"] = _png_m1
                            st.session_state["_mapa_png_m1_nome"] = (
                                f"mapa_{uf}_{datetime.now().strftime('%Y%m%d_%H%M')}.png"
                            )
                        except Exception as _emp:
                            st.error(f"Erro ao gerar imagem: {_emp}")
            if st.session_state.get("_mapa_png_m1"):
                with _mc1:
                    st.download_button(
                        "⬇️ Baixar PNG do mapa",
                        data=st.session_state["_mapa_png_m1"],
                        file_name=st.session_state.get("_mapa_png_m1_nome","mapa.png"),
                        mime="image/png",
                        use_container_width=True,
                        key="dl_mapa_png_m1",
                    )

            # ══════════════════════════════════════════════════════
            # ── PAINEL ORIGEM / DESTINO — UX redesenhado ──────────
            # ══════════════════════════════════════════════════════
            _map_o = st.session_state.get("_map_orig")
            _map_d = st.session_state.get("_map_dest")
            _rr    = st.session_state.get("_map_rota_result")

            # ── Passo 1 concluído? / Passo 2? / Passo 3?
            _p1_ok = bool(_map_o)
            _p2_ok = bool(_map_d)
            _p3_ok = bool(_rr)

            st.markdown("---")

            # ── Guia de passos ─────────────────────────────────────
            def _passo_html(num, titulo, desc, ok, ativo):
                if ok:
                    _bg, _brd, _num_bg, _num_c, _title_c = (
                        "#e8f5e9","#a5d6a7","#43a047","#fff","#2e7d32")
                    _check = "✔"
                else:
                    if ativo:
                        _bg, _brd, _num_bg, _num_c, _title_c = (
                            "#e3f2fd","#90caf9","#1565c0","#fff","#1565c0")
                    else:
                        _bg, _brd, _num_bg, _num_c, _title_c = (
                            "#fafafa","#e0e0e0","#bdbdbd","#fff","#9e9e9e")
                    _check = str(num)
                return (
                    f"<div style='display:flex;align-items:flex-start;gap:10px;"
                    f"background:{_bg};border:1px solid {_brd};border-radius:10px;"
                    f"padding:10px 14px;flex:1'>"
                    f"<div style='width:24px;height:24px;border-radius:50%;"
                    f"background:{_num_bg};color:{_num_c};font-size:12px;font-weight:700;"
                    f"display:flex;align-items:center;justify-content:center;flex-shrink:0'>"
                    f"{_check}</div>"
                    f"<div><div style='font-size:12px;font-weight:700;color:{_title_c}'>{titulo}</div>"
                    f"<div style='font-size:11px;color:#666;margin-top:1px'>{desc}</div></div>"
                    f"</div>"
                )

            st.markdown(
                f"<div style='display:flex;gap:8px;margin-bottom:14px;flex-wrap:wrap'>"
                f"{_passo_html(1,'Definir Origem','Busque e selecione o posto de partida',_p1_ok,not _p1_ok)}"
                f"{_passo_html(2,'Definir Destino','Busque e selecione o posto de chegada',_p2_ok,_p1_ok and not _p2_ok)}"
                f"{_passo_html(3,'Traçar Rota','Calcule a rota e veja os indicadores',_p3_ok,_p1_ok and _p2_ok)}"
                f"</div>",
                unsafe_allow_html=True,
            )

            # ── Cards Origem / Destino — sempre visíveis ───────────
            def _card_od(titulo, cor_brd, cor_bg, cor_txt, sel, icone_vazio, msg_vazio):
                if sel:
                    nome = sel.get("label", "?")[:55]
                    loc  = f"{sel.get('municipio','')} / {sel.get('uf','')}"
                    cnpj = sel.get("cnpj", "—")
                    return (
                        f"<div style='border:2px solid {cor_brd};border-radius:10px;"
                        f"background:{cor_bg};padding:12px 14px'>"
                        f"<div style='font-size:10px;font-weight:700;color:{cor_txt};"
                        f"letter-spacing:0.8px;text-transform:uppercase;margin-bottom:6px'>"
                        f"{titulo}</div>"
                        f"<div style='font-size:13px;font-weight:700;color:#1a1a1a;"
                        f"line-height:1.3;margin-bottom:4px'>{nome}</div>"
                        f"<div style='font-size:11px;color:#555'>📍 {loc}</div>"
                        f"<div style='font-size:11px;color:#555'>🪪 {cnpj}</div>"
                        f"</div>"
                    )
                else:
                    return (
                        f"<div style='border:2px dashed #d0d0d0;border-radius:10px;"
                        f"background:#fafafa;padding:12px 14px;text-align:center'>"
                        f"<div style='font-size:10px;font-weight:700;color:#bbb;"
                        f"letter-spacing:0.8px;text-transform:uppercase;margin-bottom:8px'>"
                        f"{titulo}</div>"
                        f"<div style='font-size:22px;margin-bottom:4px'>{icone_vazio}</div>"
                        f"<div style='font-size:11px;color:#aaa'>{msg_vazio}</div>"
                        f"</div>"
                    )

            _co_col, _cd_col = st.columns(2)
            _co_col.markdown(
                _card_od("🟢 Ponto de Origem","#43a047","#f1f8e9","#2e7d32",
                         _map_o,"📍","Busque abaixo e clique em\n'Definir como Origem'"),
                unsafe_allow_html=True,
            )
            _cd_col.markdown(
                _card_od("🔴 Ponto de Destino","#e53935","#fff8f8","#c62828",
                         _map_d,"🏁","Busque abaixo e clique em\n'Definir como Destino'"),
                unsafe_allow_html=True,
            )
            st.markdown("<div style='height:10px'></div>", unsafe_allow_html=True)

            # ── Botão Traçar Rota (destacado quando os dois estão prontos) ──
            if _map_o and _map_d:
                _mesmo_ponto_m1 = (
                    abs(_map_o["lat"] - _map_d["lat"]) < 0.0002 and
                    abs(_map_o["lon"] - _map_d["lon"]) < 0.0002
                )
                _col_btn, _col_clr = st.columns([4, 1])
                if _mesmo_ponto_m1:
                    _col_btn.warning(
                        "⚠️ Origem e Destino são o **mesmo posto**. "
                        "Busque e selecione postos diferentes abaixo."
                    )
                else:
                    if _col_btn.button(
                        f"🗺️  Traçar Rota  ·  {_map_o['label'][:22]} → {_map_d['label'][:22]}",
                        use_container_width=True, type="primary", key="btn_tracar_mapa",
                    ):
                        with st.spinner("🗺️ Calculando rota…"):
                            _cr, _dk, _dm, _lr = calcular_rota(
                                _map_o["lat"], _map_o["lon"],
                                _map_d["lat"], _map_d["lon"])
                        st.session_state["_map_rota_result"] = {
                            "coords": _cr, "dist_km": _dk, "dur_min": _dm,
                            "linha_reta": _lr, "orig": _map_o, "dest": _map_d,
                        }
                        st.rerun()
                if _col_clr.button("↺ Reiniciar", use_container_width=True,
                                   key="btn_clr_mapa_sel",
                                   help="Limpar origem, destino e rota calculada"):
                    for _k in ["_map_orig", "_map_dest", "_map_rota_result"]:
                        st.session_state.pop(_k, None)
                    st.rerun()
            elif _map_o or _map_d:
                # Um dos dois está preenchido — dica para completar
                _falta = "Destino" if _map_o else "Origem"
                st.info(f"👆 Agora busque e defina o **{_falta}** abaixo para liberar o botão de rota.")

            # ── Resultado da rota traçada ──────────────────────────
            if _rr:
                if _rr.get("linha_reta"):
                    st.warning("⚠️ OSRM indisponível — rota exibida como linha reta.")

                # Banner cabeçalho da rota
                st.markdown(
                    f"<div style='background:linear-gradient(90deg,#e8f5e9,#f1f8e9);"
                    f"border:1px solid #a5d6a7;border-radius:10px;padding:12px 16px;"
                    f"margin:6px 0 10px;display:flex;align-items:center;gap:12px;flex-wrap:wrap'>"
                    f"<span style='font-size:20px'>✅</span>"
                    f"<div style='flex:1'>"
                    f"<div style='font-size:13px;font-weight:700;color:#1b5e20'>"
                    f"{_rr['orig']['label'][:40]} → {_rr['dest']['label'][:40]}</div>"
                    f"</div></div>",
                    unsafe_allow_html=True,
                )

                # Métricas da rota — cards grandes
                _dist_rr  = _rr["dist_km"]
                _dur_rr   = _rr["dur_min"]
                _h_rr     = int(_dur_rr // 60)
                _min_rr   = int(_dur_rr % 60)
                _vel_rr   = round(_dist_rr / (_dur_rr / 60), 0) if _dur_rr > 0 else 0
                _tempo_rr = (f"{_h_rr}h {_min_rr}min" if _h_rr > 0 else f"{_min_rr} min")
                _rc1, _rc2, _rc3 = st.columns(3)
                _rc1.metric("🛣️ Distância", f"{_n(_dist_rr)} km")
                _rc2.metric("⏱️ Tempo estimado", _tempo_rr)
                _rc3.metric("🚗 Vel. média", f"{int(_vel_rr)} km/h")
                with st.spinner("🗺️ Atualizando mapa com a rota…"):
                    _mapa_rota = criar_mapa(
                        df_show, coords_rota=_rr["coords"],
                        lat_orig=_rr["orig"]["lat"], lon_orig=_rr["orig"]["lon"],
                        lat_dest=_rr["dest"]["lat"], lon_dest=_rr["dest"]["lon"],
                        label_orig=_rr["orig"]["label"], label_dest=_rr["dest"]["label"],
                    )
                    _renderizar_mapa(_mapa_rota, height=560, key="mapa_m1_rota")

            # ── Busca de postos ────────────────────────────────────
            st.markdown(
                "<div style='background:#f8f9fa;border:1px solid #e3e8f0;"
                "border-radius:10px;padding:14px 16px;margin-top:10px'>"
                "<div style='font-size:13px;font-weight:700;color:#1565c0;"
                "margin-bottom:8px'>🔍 Buscar posto para Origem ou Destino</div>"
                "<div style='font-size:11px;color:#666;margin-bottom:10px'>"
                "Digite o nome ou parte do CNPJ do posto e clique no botão correspondente.</div>"
                "</div>",
                unsafe_allow_html=True,
            )
            _col_busca_m, _col_limpa_m = st.columns([5, 1])
            _busca_txt = _col_busca_m.text_input(
                "Buscar posto",
                placeholder="Ex: Super Sol, Auto Posto Silva, 12.345…",
                key="busca_posto_mapa",
                label_visibility="collapsed",
            )
            if _col_limpa_m.button("🗑️", key="limpa_sel_mapa",
                                    help="Limpar campos de Origem e Destino"):
                for _k in ["_map_orig", "_map_dest", "_map_rota_result"]:
                    st.session_state.pop(_k, None)
                st.rerun()

            if _busca_txt and len(_busca_txt.strip()) >= 2 and not df_show.empty:
                _bt = _busca_txt.strip().upper()
                _mask_busca = df_show["razaoSocial"].fillna("").str.upper().str.contains(
                    _bt, regex=False, na=False)
                if "cnpj" in df_show.columns:
                    _btd = "".join(c for c in _busca_txt if c.isdigit())
                    if _btd:
                        _mask_cnpj = df_show["cnpj"].fillna("").apply(
                            lambda x: _btd in "".join(c for c in str(x) if c.isdigit()))
                        _mask_busca = _mask_busca | _mask_cnpj

                _res = df_show[_mask_busca].head(6)
                if not _res.empty:
                    for _idx_r, _row_r in _res.iterrows():
                        _ic  = "⭐" if bool(_row_r.get("_pro_frotas")) else "⛽"
                        _loc = f"{_row_r.get('municipio','')}/{_row_r.get('uf','')}"
                        _nm  = str(_row_r.get("razaoSocial", "?"))[:48]
                        # Card de resultado
                        st.markdown(
                            f"<div style='background:#fff;border:1px solid #e0e0e0;"
                            f"border-radius:8px;padding:8px 12px;margin-bottom:4px;"
                            f"display:flex;align-items:center;gap:8px'>"
                            f"<span style='font-size:16px'>{_ic}</span>"
                            f"<div style='flex:1'>"
                            f"<div style='font-size:12px;font-weight:600;color:#1a1a1a'>{_nm}</div>"
                            f"<div style='font-size:10px;color:#888'>📍 {_loc}</div>"
                            f"</div></div>",
                            unsafe_allow_html=True,
                        )
                        _c1r, _c2r = st.columns(2)
                        if _c1r.button(
                            "🟢 Definir como Origem",
                            key=f"set_orig_{_idx_r}",
                            use_container_width=True,
                            help="Marcar este posto como ponto de partida da rota",
                            type="primary" if not _map_o else "secondary",
                        ):
                            st.session_state["_map_orig"] = {
                                "lat":       float(_row_r["_lat"]),
                                "lon":       float(_row_r["_lon"]),
                                "label":     str(_row_r.get("razaoSocial", "Posto")),
                                "municipio": str(_row_r.get("municipio", "")),
                                "uf":        str(_row_r.get("uf", "")),
                                "cnpj":      _formatar_cnpj(str(_row_r.get("cnpj", ""))),
                            }
                            st.session_state.pop("_map_rota_result", None)
                            st.rerun()
                        if _c2r.button(
                            "🔴 Definir como Destino",
                            key=f"set_dest_{_idx_r}",
                            use_container_width=True,
                            help="Marcar este posto como ponto de chegada da rota",
                            type="primary" if _map_o and not _map_d else "secondary",
                        ):
                            st.session_state["_map_dest"] = {
                                "lat":       float(_row_r["_lat"]),
                                "lon":       float(_row_r["_lon"]),
                                "label":     str(_row_r.get("razaoSocial", "Posto")),
                                "municipio": str(_row_r.get("municipio", "")),
                                "uf":        str(_row_r.get("uf", "")),
                                "cnpj":      _formatar_cnpj(str(_row_r.get("cnpj", ""))),
                            }
                            st.session_state.pop("_map_rota_result", None)
                            st.rerun()
                else:
                    st.warning("⚠️ Nenhum posto encontrado. Tente um nome diferente ou apenas parte do CNPJ.")

            # ══════════════════════════════════════════════════════
            # ── PREÇOS E COMPARATIVO — sempre abaixo do mapa ──────
            # ══════════════════════════════════════════════════════
            st.markdown(
                "<div style='margin-top:24px;border-top:2px solid #e3e8f0;"
                "padding-top:18px'></div>",
                unsafe_allow_html=True,
            )
            _pp_df_m1    = st.session_state.get("_pp_df")
            _cnpjs_pf_m1 = st.session_state.get("cnpjs_pro_frotas", set())
            _cache_m1    = st.session_state.get("_precos_anp_cache", {})
            _sheets_m1   = _cache_m1.get("sheets")
            if _pp_df_m1 is not None and _sheets_m1 is not None and _cnpjs_pf_m1:
                _comp_m1 = _calcular_comparativo_pf_anp(
                    _pp_df_m1, _cnpjs_pf_m1, _sheets_m1, ufs=[uf] if uf else None
                )
                if _comp_m1:
                    _renderizar_comparativo_pf_anp(
                        _comp_m1,
                        subtitulo=f"Postos Gestão de Frotas vs preço médio ANP — {UF_NOME.get(uf, uf)}"
                    )
            _renderizar_precos_anp(uf, municipio_input.strip() or None)

        with tab_dados:
            # ── Mini painel: Top 5 Mais Baratos ──────────────────────────────
            if "_rank_barato" in df_show.columns and df_show["_rank_barato"].gt(0).any():
                st.markdown(
                    "<div style='font-size:14px;font-weight:700;color:#7B3F00;"
                    "margin-bottom:8px'>💰 Top 5 Mais Baratos nesta consulta</div>",
                    unsafe_allow_html=True,
                )
                _top5_rows = df_show[df_show["_rank_barato"] > 0].sort_values("_rank_barato")
                _top5_cols = st.columns(min(len(_top5_rows), 5))
                for _ti, (_, _tr5) in enumerate(zip(_top5_cols, _top5_rows.iterrows())):
                    _, _row5 = _tr5
                    _rk5   = int(_row5["_rank_barato"])
                    _em5   = _RANK_EMOJI.get(_rk5, str(_rk5))
                    _nm5   = str(_row5.get("razaoSocial","?"))[:30]
                    _pr5   = _row5.get("_preco_barato")
                    _cb5   = str(_row5.get("_comb_barato",""))
                    _mn5   = str(_row5.get("municipio",""))
                    _uf5   = str(_row5.get("uf",""))
                    _geo5  = f"{_mn5}/{_uf5}" if _mn5 and _uf5 else _mn5 or _uf5
                    _pr5_s = f"R$ {_pr5:.3f}/L" if _pr5 is not None else "—"
                    _top5_cols[_ti].markdown(
                        f"<div style='background:linear-gradient(135deg,#fff9c4,#fff3e0);"
                        f"border:2px solid #FFD700;border-radius:10px;padding:10px 12px;"
                        f"text-align:center;height:100%'>"
                        f"<div style='font-size:22px'>{_em5}</div>"
                        f"<div style='font-size:11px;font-weight:700;color:#5f3a00;"
                        f"margin:4px 0 2px;line-height:1.3'>{_nm5}</div>"
                        f"<div style='font-size:13px;font-weight:800;color:#2e7d32'>{_pr5_s}</div>"
                        f"<div style='font-size:10px;color:#888;margin-top:2px'>{_cb5}</div>"
                        f"<div style='font-size:10px;color:#aaa'>{_geo5}</div>"
                        f"</div>",
                        unsafe_allow_html=True,
                    )
                st.markdown("<div style='height:12px'></div>", unsafe_allow_html=True)

            # ── Tabela de dados ───────────────────────────────────────────────
            cols = [c for c in ["razaoSocial","cnpj","distribuidora","_pro_frotas",
                                 "endereco","bairro","municipio","uf","cep","autorizacao","statusSIGAF"]
                    if c in df_show.columns]
            df_exib = df_show[cols].copy()
            if "_pro_frotas" in df_exib.columns:
                df_exib = df_exib.rename(columns={"_pro_frotas":"Gestão de Frotas ⭐"})
            # Coluna de ranking (se existir)
            if "_rank_barato" in df_show.columns:
                df_exib.insert(0, "💰 Rank",
                    df_show["_rank_barato"].fillna(0).astype(int).map(
                        lambda v: _RANK_EMOJI.get(v, "") if v > 0 else ""
                    )
                )
            # ── Score de posto ─────────────────────────────────────────────────
            _anp_ref_m1 = None
            if st.session_state.get("_precos_anp_cache"):
                try:
                    _sh_m1 = st.session_state["_precos_anp_cache"].get("sheets", {})
                    _pr_m1 = _anp_extrair_precos(_sh_m1.get("estados"), uf=uf)
                    if _pr_m1:
                        _anp_ref_m1 = next(
                            (r["Preço Médio"] for r in _pr_m1 if "gasolina" in r.get("Combustível","").lower()),
                            None)
                except Exception:
                    pass
            _df_exib_scored = _calcular_score_df(
                df_show,
                preco_ref_anp=_anp_ref_m1,
                lat_ref=None, lon_ref=None,
            )
            if "⭐ Score" in _df_exib_scored.columns:
                df_exib.insert(0, "⭐ Score", _df_exib_scored["⭐ Score"].values)
            st.dataframe(df_exib, use_container_width=True, height=450)

            # ── Histórico de preço de posto selecionado ────────────────────────
            _hist_data = _intel_load().get("historico", {})
            if _hist_data and not df_show.empty:
                with st.expander("📈 Histórico de preços de um posto", expanded=False):
                    # Monta labels enriquecidos: "Razão Social — CNPJ • Cidade/UF"
                    _hist_opts   = ["—"]
                    _hist_cnpj_map = {}  # label → cnpj_norm
                    _hist_nome_map = {}  # label → nome limpo
                    for _i, _row_h in df_show.head(200).iterrows():
                        _nm  = str(_row_h.get("razaoSocial", "") or "").strip()
                        _cj  = re.sub(r"\D", "", str(_row_h.get("cnpj", "") or ""))
                        _mn  = str(_row_h.get("municipio", "") or "").strip().title()
                        _uf  = str(_row_h.get("uf", "") or "").strip().upper()
                        _cj_fmt = (f"{_cj[:2]}.{_cj[2:5]}.{_cj[5:8]}/{_cj[8:12]}-{_cj[12:]}"
                                   if len(_cj) == 14 else _cj)
                        _geo = f"{_mn}/{_uf}" if _mn and _uf else (_mn or _uf)
                        _label = f"{_nm}  —  {_cj_fmt}"
                        if _geo:
                            _label += f"  •  {_geo}"
                        _hist_opts.append(_label)
                        _hist_cnpj_map[_label] = _cj
                        _hist_nome_map[_label] = _nm

                    _hist_sel = st.selectbox(
                        "Selecione o posto",
                        options=_hist_opts,
                        key="hist_sel_m1",
                    )
                    if _hist_sel and _hist_sel != "—":
                        _hist_cnpj_sel = _hist_cnpj_map.get(_hist_sel, "")
                        _hist_nome_sel = _hist_nome_map.get(_hist_sel, _hist_sel)
                        if _hist_cnpj_sel and _hist_cnpj_sel in _hist_data:
                            _fig_hist = _hist_chart_posto(_hist_cnpj_sel, _hist_nome_sel)
                            if _fig_hist:
                                st.markdown(
                                    f"<p style='font-weight:600;font-size:0.95rem;"
                                    f"margin:0 0 4px 0;color:var(--text-color,#1a1a2e)'>"
                                    f"📈 Evolução de preços — {_hist_nome_sel}</p>",
                                    unsafe_allow_html=True,
                                )
                                st.plotly_chart(_fig_hist, use_container_width=True)
                        else:
                            st.info("Nenhum histórico registrado para este posto ainda.")

            st.download_button("⬇️ Baixar dados em CSV",
                               df_show.to_csv(index=False).encode("utf-8"),
                               f"postos_{uf}.csv", "text/csv", use_container_width=True)

        with tab_analise:
            if not df_show.empty and "distribuidora" in df_show.columns:
                contagem = df_show["distribuidora"].value_counts().reset_index()
                contagem.columns = ["Distribuidora","Quantidade"]
                st.markdown("**Distribuição de postos por bandeira**")
                st.bar_chart(contagem.set_index("Distribuidora"), height=400)
                if n_pf(df_show) > 0:
                    st.markdown("**Postos Gestão de Frotas por bandeira**")
                    pf_dist = df_show[df_show["_pro_frotas"]]["distribuidora"].value_counts().reset_index()
                    pf_dist.columns = ["Distribuidora","Gestão de Frotas"]
                    st.bar_chart(pf_dist.set_index("Distribuidora"), height=300)

    else:
        # Mapa vazio centrado no Brasil + instrução informativa
        _renderizar_mapa(criar_mapa(pd.DataFrame()), height=680, key="mapa_m1_vazio")
        st.info("👈 Selecione um Estado na barra lateral para carregar os postos")


# ═══════════════════════════════════════════════════════════════════
#  MODO 2 — Por Rota
# ═══════════════════════════════════════════════════════════════════

elif modo == "🗺️ Por Rota":

    _auto_rota = st.session_state.pop("_auto_buscar_rota", False)
    if buscar_rota_btn or _auto_rota:
        orig_sel = st.session_state.get("orig_sel")
        dest_sel = st.session_state.get("dest_sel")
        if buscar_rota_btn:
            _log_acesso("ROTA_CALCULAR",
                        f"{orig_sel.get('label','—') if orig_sel else '—'} → "
                        f"{dest_sel.get('label','—') if dest_sel else '—'}")
        # Coleta paradas intermediárias selecionadas (ignora as não preenchidas)
        _n_par_main = st.session_state.get("_paradas_count", 0)
        _paradas_data_main = [
            st.session_state[f"parada_sel_{_pi}"]
            for _pi in range(1, _n_par_main + 1)
            if st.session_state.get(f"parada_sel_{_pi}")
        ]
        if not orig_sel:
            st.warning("⚠️ Confirme o **ponto de Origem** selecionando uma sugestão.")
        elif not dest_sel:
            st.warning("⚠️ Confirme o **ponto de Destino** selecionando uma sugestão.")
        elif orig_sel["label"] == dest_sel["label"]:
            st.warning("⚠️ Origem e Destino não podem ser o mesmo ponto.")
        else:
            lo, ld = orig_sel, dest_sel
            _wp_ll = [[wp["lat"], wp["lon"]] for wp in _paradas_data_main]
            _n_wp  = len(_paradas_data_main)
            _spinner_txt = (
                f"🗺️ Calculando rota com {_n_wp} parada{'s' if _n_wp != 1 else ''}…"
                if _n_wp else "🗺️ Calculando rota…"
            )
            with st.spinner(_spinner_txt):
                coords_rota, dist_km, dur_min, linha_reta = calcular_rota(
                    lo["lat"], lo["lon"], ld["lat"], ld["lon"],
                    waypoints=_wp_ll if _wp_ll else None
                )
            if linha_reta:
                st.warning("⚠️ Servidor de roteamento indisponível. Usando **linha reta** como aproximação.")
            ufs_rota = ufs_ao_longo_rota(coords_rota)
            if not ufs_rota:
                st.error("❌ Não foi possível detectar estados ao longo da rota. Verifique os pontos de origem e destino.")
                ufs_rota = []

            if ufs_rota:
                st.info(f"🗺️ Estados detectados na rota: **{', '.join(ufs_rota)}**")

            # ── Carrega postos GF dos estados da rota (planilha local — sem API ANP) ──
            _pf_df_m2   = st.session_state.get("pf_coords_df", pd.DataFrame())
            _ufs_set_m2 = {u.upper() for u in ufs_rota}
            _n_pf_total_m2 = len(_pf_df_m2)

            if not _pf_df_m2.empty:
                df_todos = _pf_df_m2[
                    _pf_df_m2["uf"].fillna("").str.upper().str.strip().isin(_ufs_set_m2)
                ].copy().reset_index(drop=True)
                if df_todos.empty:
                    # pf_coords_df tem dados mas nenhum para os estados da rota
                    _ufs_disp = sorted(_pf_df_m2["uf"].fillna("").str.upper().str.strip().unique().tolist())
                    st.warning(
                        f"⚠️ A planilha Gestão de Frotas tem **{_n(_n_pf_total_m2)}** postos com coordenadas, "
                        f"mas **nenhum** nos estados **{', '.join(sorted(_ufs_set_m2))}**. "
                        f"Estados disponíveis na planilha: {', '.join(_ufs_disp) or '—'}. "
                        "Verifique se a planilha está completa ou recarregue em **Configurações**."
                    )
            else:
                df_todos = pd.DataFrame()
                st.warning(
                    "⚠️ Planilha Gestão de Frotas não carregada ou sem coordenadas. "
                    "Verifique a seção **Configurações** na barra lateral."
                )

            # Guarda diagnóstico para aviso persistente após re-render
            st.session_state["_m2_diag"] = {
                "n_pf_total":  _n_pf_total_m2,
                "n_df_todos":  len(df_todos),
                "ufs_rota":    sorted(_ufs_set_m2),
            }

            if not df_todos.empty:
                with st.spinner("📏 Calculando distâncias…"):
                    dists = dist_minima_rota_np(
                        df_todos["_lat"].values,
                        df_todos["_lon"].values,
                        coords_rota,
                    )
                    df_todos["_dist_rota"] = dists
                df_rota = df_todos[df_todos["_dist_rota"] <= raio].copy().sort_values("_dist_rota").reset_index(drop=True)
                if df_rota.empty:
                    st.warning(
                        f"⚠️ Foram encontrados **{_n(len(df_todos))}** postos Gestão de Frotas nos estados, "
                        f"mas nenhum está dentro de **{raio} m** da rota. "
                        "Tente aumentar o raio na barra lateral."
                    )
            else:
                df_rota = pd.DataFrame()

            # ── Overlay ANP: adiciona postos do arquivo carregado pelo usuário ──
            _anp_view_m2 = st.session_state.get("_anp_df_raw")
            if _anp_view_m2 is not None and not df_rota.empty:
                _anp_rota = _anp_view_m2[
                    _anp_view_m2["uf"].fillna("").str.upper().str.strip().isin(_ufs_set_m2)
                ].copy().dropna(subset=["_lat", "_lon"])
                if not _anp_rota.empty:
                    _dists_anp = dist_minima_rota_np(
                        _anp_rota["_lat"].values, _anp_rota["_lon"].values, coords_rota
                    )
                    _anp_rota["_dist_rota"] = _dists_anp
                    _anp_rota = _anp_rota[_anp_rota["_dist_rota"] <= raio]
                    if not _anp_rota.empty:
                        _pf_cnpjs_m2 = set(
                            df_rota["cnpj"].fillna("").str.replace(r"\D", "", regex=True)
                        )
                        _anp_rota = _anp_rota[
                            ~_anp_rota["cnpj"].fillna("").str.replace(r"\D", "", regex=True)
                             .isin(_pf_cnpjs_m2)
                        ]
                        if not _anp_rota.empty:
                            df_rota = pd.concat(
                                [df_rota, _marcar_df_completo(_anp_rota)], ignore_index=True
                            )

            st.session_state.update({
                "df_rota": df_rota, "coords_rota": coords_rota,
                "lat_orig": lo["lat"], "lon_orig": lo["lon"], "label_orig": lo["label"],
                "lat_dest": ld["lat"], "lon_dest": ld["lon"], "label_dest": ld["label"],
                "dist_km": dist_km, "dur_min": dur_min, "raio_usado": raio, "linha_reta": linha_reta,
                "_ufs_rota_atual": ufs_rota,        # para aba de preços
                "_paradas_data": _paradas_data_main, # paradas intermediárias (lista de dicts)
            })
            if not df_rota.empty and "distribuidora" in df_rota.columns:
                st.session_state["distribuidoras_rota"] = sorted(df_rota["distribuidora"].dropna().unique().tolist())
            else:
                st.session_state.pop("distribuidoras_rota", None)

    if "df_rota" in st.session_state:
        df_rota      = st.session_state["df_rota"]
        coords_rota  = st.session_state.get("coords_rota", [])
        lat_orig     = st.session_state.get("lat_orig"); lon_orig = st.session_state.get("lon_orig")
        lat_dest     = st.session_state.get("lat_dest"); lon_dest = st.session_state.get("lon_dest")
        label_orig   = st.session_state.get("label_orig", "Origem")
        label_dest   = st.session_state.get("label_dest", "Destino")
        dist_km      = st.session_state.get("dist_km", 0)
        dur_min      = st.session_state.get("dur_min", 0)
        raio_usado   = st.session_state.get("raio_usado", 500)
        _paradas_vis = st.session_state.get("_paradas_data", [])

        if st.session_state.get("linha_reta"):
            st.warning("⚠️ Rota exibida como **linha reta** (OSRM indisponível).")

        # ── Aviso persistente quando df_rota é vazio ─────────────────
        if df_rota.empty:
            _diag = st.session_state.get("_m2_diag", {})
            _n_tot  = _diag.get("n_pf_total", -1)
            _n_est  = _diag.get("n_df_todos",  -1)
            _ufs_r  = _diag.get("ufs_rota",    [])
            if _n_tot == 0:
                st.error(
                    "❌ Planilha Gestão de Frotas sem coordenadas (lat/lon). "
                    "Acesse **Configurações** → **🔄 Recarregar planilha** e verifique se a planilha "
                    "no repositório contém colunas **Latitude** e **Longitude**."
                )
            elif _n_est == 0 and _n_tot > 0:
                st.warning(
                    f"⚠️ A planilha tem **{_n(_n_tot)}** postos com coordenadas, "
                    f"mas **nenhum** nos estados **{', '.join(_ufs_r)}**. "
                    "Acesse **Configurações** → **🔄 Recarregar planilha** ou verifique a planilha."
                )
            elif _n_est > 0:
                st.warning(
                    f"⚠️ {_n(_n_est)} postos encontrados nos estados, "
                    f"mas nenhum dentro de **{raio_usado} m** da rota. "
                    "Aumente o **Raio da rota** na barra lateral."
                )

        df_show_r = preparar_df(
            df_rota, distribuidoras_filtro,
            perfis_filtro=perfis_filtro_m2,
            filtro_servicos=_filtro_servicos_m2,
            filtro_24h=_filtro_24h_m2,
        )

        # ── Filtro de Preço — pós-processamento via _pp_df ───────────────
        if _preco_faixa_m2 and _fuel_sel_m2 and _fuel_sel_m2 != "— Todos —":
            _pp_m2 = st.session_state.get("_pp_df")
            if _pp_m2 is not None and "_cnpj_norm" in df_show_r.columns:
                _fuel_df_m2 = _pp_m2[
                    _pp_m2["combustivel_label"].str.strip() == _fuel_sel_m2
                ][["cnpj_norm","preco"]].copy()
                _lo_m2, _hi_m2 = _preco_faixa_m2
                _cnpj_ok_m2 = set(
                    _fuel_df_m2[
                        _fuel_df_m2["preco"].between(_lo_m2, _hi_m2)
                    ]["cnpj_norm"]
                )
                _sem_preco_m2 = ~df_show_r["_cnpj_norm"].isin(_fuel_df_m2["cnpj_norm"])
                df_show_r = df_show_r[_sem_preco_m2 | df_show_r["_cnpj_norm"].isin(_cnpj_ok_m2)]

        # ── Ranking Top 5 Mais Baratos ──────────────────────────────────────
        _fuel_rank_m2 = _fuel_sel_m2 if (_fuel_sel_m2 and _fuel_sel_m2 != "— Todos —") else None
        _top5_m2 = _calcular_top5_baratos(df_show_r, fuel_label=_fuel_rank_m2)
        df_show_r = _aplicar_rank_barato(df_show_r, _top5_m2)

        c1, c2, c3, c4, c5 = st.columns(5)
        c1.metric("🛣️ Distância",      f"{_n(dist_km)} km")
        c2.metric("⏱️ Tempo estimado", f"{int(dur_min//60)}h {int(dur_min%60)}min")
        c3.metric("📍 Paradas",        str(len(_paradas_vis)) if _paradas_vis else "—")
        c4.metric("⛽ Postos na rota", _n(len(df_show_r)))
        c5.metric("⭐ Gestão de Frotas",     _n(n_pf(df_show_r)))

        # Resumo visual da rota com todas as paradas
        _all_labels = [label_orig] + [wp["label"] for wp in _paradas_vis] + [label_dest]
        _rota_str   = " → ".join(f"**{lb}**" for lb in _all_labels)
        st.success(f"✅ {_rota_str} | {_n(len(df_show_r))} postos a até {raio_usado} m")

        # ── Banner de filtros ativos (Modo 2) ────────────────────────────
        _filtros_ativos_m2 = []
        if distribuidoras_filtro:
            _filtros_ativos_m2.append(f"Bandeira: {', '.join(distribuidoras_filtro)}")
        if _filtro_24h_m2:
            _filtros_ativos_m2.append("⏰ 24h")
        if _filtro_servicos_m2:
            _srv_lbl2 = {"pista_caminhao":"🚛 Pista","arla":"🧪 ARLA","conveniencia":"🛒 Conv.",
                         **{k: v for k, v in st.session_state.get("_servicos_pf_labels", {}).items()}}
            _filtros_ativos_m2 += [_srv_lbl2.get(s,s) for s in _filtro_servicos_m2]
        if _preco_faixa_m2 and _fuel_sel_m2 and _fuel_sel_m2 != "— Todos —":
            _lo_lbl2, _hi_lbl2 = _preco_faixa_m2
            _filtros_ativos_m2.append(f"💰 {_fuel_sel_m2}: R$ {_brl(_lo_lbl2,2)}–{_brl(_hi_lbl2,2)}/L")
        if _filtros_ativos_m2:
            st.info(
                "🔍 **Filtros ativos:** " + " · ".join(_filtros_ativos_m2)
                + f" &nbsp;|&nbsp; **{_n(len(df_show_r))} postos** exibidos",
                icon=None,
            )

        # ── Botão Salvar (Modo 2) ─────────────────────────────────────
        _nome_sugerido_m2 = f"{label_orig[:30]} → {label_dest[:30]}"
        _col_sv2a, _col_sv2b = st.columns([3, 1])
        with _col_sv2a:
            _nome_salvar_m2 = st.text_input(
                "Nome da rota",
                value=_nome_sugerido_m2,
                key="nome_salvar_m2",
                label_visibility="collapsed",
                placeholder="Nome para identificar esta rota…",
            )
        with _col_sv2b:
            if st.button("💾 Salvar", use_container_width=True, key="btn_salvar_m2",
                         help="Salvar esta rota para acessar depois"):
                _dados_m2 = {
                    "orig_sel":     st.session_state.get("orig_sel"),
                    "dest_sel":     st.session_state.get("dest_sel"),
                    "paradas_data": st.session_state.get("_paradas_data", []),
                    "raio":         raio_usado,
                    "dist_km":      dist_km,
                    "dur_min":      dur_min,
                    "label_orig":   label_orig,
                    "label_dest":   label_dest,
                    "lat_orig":     lat_orig,
                    "lon_orig":     lon_orig,
                    "lat_dest":     lat_dest,
                    "lon_dest":     lon_dest,
                }
                if _salvar_rota_nova(_nome_salvar_m2 or _nome_sugerido_m2, "rota", _dados_m2):
                    st.toast(f"✅ Rota **{_nome_salvar_m2 or _nome_sugerido_m2}** salva!", icon="💾")
                else:
                    st.error("❌ Não foi possível salvar. Verifique permissões do diretório.")

        tab_m, tab_d = st.tabs([
            "🗺️  Mapa da Rota", "📋  Postos na Rota"])

        with tab_m:
            with st.spinner(f"🗺️ Carregando mapa da rota — {_n(len(df_show_r))} postos…"):
                m = criar_mapa(df_show_r, coords_rota=coords_rota,
                               lat_orig=lat_orig, lon_orig=lon_orig,
                               lat_dest=lat_dest, lon_dest=lon_dest,
                               label_orig=label_orig, label_dest=label_dest,
                               waypoints=_paradas_vis if _paradas_vis else None)
                _renderizar_mapa(m, height=660, key="mapa_m2_rota")

            # ── Exportar mapa como PNG (Modo 2) ───────────────────
            _m2c1, _m2c2 = st.columns([5, 1])
            with _m2c2:
                if st.button("📸 Exportar mapa", use_container_width=True,
                             key="btn_exp_mapa_m2",
                             help="Baixar o mapa da rota como imagem PNG"):
                    with st.spinner("🖼️ Gerando imagem…"):
                        try:
                            _titulo_m2 = f"Rota: {label_orig[:20]} → {label_dest[:20]}"
                            _sub_m2    = f"{len(df_show_r)} postos  ·  {datetime.now().strftime('%d/%m/%Y %H:%M')}"
                            _png_m2 = _exportar_mapa_postos_png(df_show_r, _titulo_m2, _sub_m2)
                            st.session_state["_mapa_png_m2"] = _png_m2
                            st.session_state["_mapa_png_m2_nome"] = (
                                f"mapa_rota_{datetime.now().strftime('%Y%m%d_%H%M')}.png"
                            )
                        except Exception as _emp2:
                            st.error(f"Erro ao gerar imagem: {_emp2}")
            if st.session_state.get("_mapa_png_m2"):
                with _m2c1:
                    st.download_button(
                        "⬇️ Baixar PNG do mapa",
                        data=st.session_state["_mapa_png_m2"],
                        file_name=st.session_state.get("_mapa_png_m2_nome","mapa_rota.png"),
                        mime="image/png",
                        use_container_width=True,
                        key="dl_mapa_png_m2",
                    )

            # ── Refinar Origem/Destino com posto da rota ──────────
            if not df_show_r.empty:
                st.markdown("---")
                st.markdown(
                    "<div style='background:#f8f9fa;border:1px solid #e3e8f0;"
                    "border-radius:10px;padding:14px 16px;margin-bottom:10px'>"
                    "<div style='font-size:13px;font-weight:700;color:#1565c0;"
                    "margin-bottom:4px'>🔄 Refinar Origem ou Destino</div>"
                    "<div style='font-size:11px;color:#666'>"
                    "Escolha um posto desta rota para redefinir o ponto de partida ou chegada.</div>"
                    "</div>",
                    unsafe_allow_html=True,
                )
                _busca_r = st.text_input(
                    "Buscar posto na rota",
                    placeholder="Ex: Auto Posto Central, Rede Sul…",
                    key="busca_posto_rota",
                    label_visibility="collapsed",
                )
                if _busca_r and len(_busca_r.strip()) >= 2:
                    _br = _busca_r.strip().upper()
                    _mask_r = df_show_r["razaoSocial"].fillna("").str.upper().str.contains(
                        _br, regex=False, na=False)
                    if "cnpj" in df_show_r.columns:
                        _brd = "".join(c for c in _busca_r if c.isdigit())
                        if _brd:
                            _mask_r_cnpj = df_show_r["cnpj"].fillna("").apply(
                                lambda x: _brd in "".join(c for c in str(x) if c.isdigit()))
                            _mask_r = _mask_r | _mask_r_cnpj

                    _res_r = df_show_r[_mask_r].head(6)
                    if not _res_r.empty:
                        for _idx_rr, _row_rr in _res_r.iterrows():
                            _ic_r   = "⭐" if bool(_row_rr.get("_pro_frotas")) else "⛽"
                            _dist_r = int(_row_rr["_dist_rota"]) if pd.notna(_row_rr.get("_dist_rota")) else 0
                            _nm_rr  = str(_row_rr.get("razaoSocial","?"))[:48]
                            _loc_rr = f"{_row_rr.get('municipio','')}/{_row_rr.get('uf','')}"
                            _maps_url_rr = (
                                f"https://maps.google.com/?q="
                                f"{float(_row_rr['_lat']):.6f},{float(_row_rr['_lon']):.6f}"
                            )
                            # Card do posto
                            st.markdown(
                                f"<div style='background:#fff;border:1px solid #e0e0e0;"
                                f"border-radius:8px;padding:8px 12px;margin-bottom:4px'>"
                                f"<div style='display:flex;align-items:center;gap:8px'>"
                                f"<span style='font-size:16px'>{_ic_r}</span>"
                                f"<div style='flex:1'>"
                                f"<div style='font-size:12px;font-weight:600;color:#1a1a1a'>{_nm_rr}</div>"
                                f"<div style='font-size:10px;color:#888'>📍 {_loc_rr} &nbsp;·&nbsp; "
                                f"🛣️ {_dist_r} m da rota</div>"
                                f"</div></div></div>",
                                unsafe_allow_html=True,
                            )
                            _sel_r = {
                                "lat":   float(_row_rr["_lat"]),
                                "lon":   float(_row_rr["_lon"]),
                                "label": str(_row_rr.get("razaoSocial", "Posto")),
                            }
                            _c_orig, _c_dest, _c_map = st.columns([2, 2, 1])
                            if _c_orig.button("🟢 Nova Origem", key=f"rota_orig_{_idx_rr}",
                                              use_container_width=True, help="Recalcular rota a partir deste posto"):
                                st.session_state["orig_sel"]          = _sel_r
                                st.session_state["_form_key"]         = st.session_state.get("_form_key", 0) + 1
                                st.session_state["_auto_buscar_rota"] = True
                                st.rerun()
                            if _c_dest.button("🔴 Novo Destino", key=f"rota_dest_{_idx_rr}",
                                              use_container_width=True, help="Recalcular rota até este posto"):
                                st.session_state["dest_sel"]          = _sel_r
                                st.session_state["_form_key"]         = st.session_state.get("_form_key", 0) + 1
                                st.session_state["_auto_buscar_rota"] = True
                                st.rerun()
                            with _c_map:
                                st.link_button("📍", _maps_url_rr,
                                               help="Ver no Google Maps",
                                               use_container_width=True)
                    else:
                        st.warning("⚠️ Nenhum posto encontrado na rota com esse nome.")

            # ══════════════════════════════════════════════════════
            # ── PREÇOS E COMPARATIVO — abaixo do mapa da rota ─────
            # ══════════════════════════════════════════════════════
            st.markdown(
                "<div style='margin-top:24px;border-top:2px solid #e3e8f0;"
                "padding-top:18px'></div>",
                unsafe_allow_html=True,
            )
            _pp_df_m2    = st.session_state.get("_pp_df")
            _cnpjs_pf_m2 = st.session_state.get("cnpjs_pro_frotas", set())
            _cache_m2    = st.session_state.get("_precos_anp_cache", {})
            _sheets_m2   = _cache_m2.get("sheets")
            _ufs_rota_m2 = list(st.session_state.get("_ufs_rota_atual", []))
            if _pp_df_m2 is not None and _sheets_m2 is not None and _cnpjs_pf_m2:
                _comp_m2 = _calcular_comparativo_pf_anp(
                    _pp_df_m2, _cnpjs_pf_m2, _sheets_m2,
                    ufs=_ufs_rota_m2 if _ufs_rota_m2 else None
                )
                if _comp_m2:
                    _renderizar_comparativo_pf_anp(
                        _comp_m2,
                        subtitulo="Postos Gestão de Frotas vs preço médio ANP — estados da rota"
                    )
            _ufs_rota = st.session_state.get("_ufs_rota_atual", [])
            _renderizar_precos_anp(None, ufs_multiplas=_ufs_rota)

        with tab_d:
            # ── Mini painel: Top 5 Mais Baratos ──────────────────────────────
            if "_rank_barato" in df_show_r.columns and df_show_r["_rank_barato"].gt(0).any():
                st.markdown(
                    "<div style='font-size:14px;font-weight:700;color:#7B3F00;"
                    "margin-bottom:8px'>💰 Top 5 Mais Baratos nesta rota</div>",
                    unsafe_allow_html=True,
                )
                _top5r_rows = df_show_r[df_show_r["_rank_barato"] > 0].sort_values("_rank_barato")
                _top5r_cols = st.columns(min(len(_top5r_rows), 5))
                for _ti2, (_, _tr5r) in enumerate(zip(_top5r_cols, _top5r_rows.iterrows())):
                    _, _row5r = _tr5r
                    _rk5r  = int(_row5r["_rank_barato"])
                    _em5r  = _RANK_EMOJI.get(_rk5r, str(_rk5r))
                    _nm5r  = str(_row5r.get("razaoSocial","?"))[:30]
                    _pr5r  = _row5r.get("_preco_barato")
                    _cb5r  = str(_row5r.get("_comb_barato",""))
                    _mn5r  = str(_row5r.get("municipio",""))
                    _uf5r  = str(_row5r.get("uf",""))
                    _geo5r = f"{_mn5r}/{_uf5r}" if _mn5r and _uf5r else _mn5r or _uf5r
                    _pr5r_s = f"R$ {_pr5r:.3f}/L" if _pr5r is not None else "—"
                    _top5r_cols[_ti2].markdown(
                        f"<div style='background:linear-gradient(135deg,#fff9c4,#fff3e0);"
                        f"border:2px solid #FFD700;border-radius:10px;padding:10px 12px;"
                        f"text-align:center;height:100%'>"
                        f"<div style='font-size:22px'>{_em5r}</div>"
                        f"<div style='font-size:11px;font-weight:700;color:#5f3a00;"
                        f"margin:4px 0 2px;line-height:1.3'>{_nm5r}</div>"
                        f"<div style='font-size:13px;font-weight:800;color:#2e7d32'>{_pr5r_s}</div>"
                        f"<div style='font-size:10px;color:#888;margin-top:2px'>{_cb5r}</div>"
                        f"<div style='font-size:10px;color:#aaa'>{_geo5r}</div>"
                        f"</div>",
                        unsafe_allow_html=True,
                    )
                st.markdown("<div style='height:12px'></div>", unsafe_allow_html=True)

            # ── Tabela de dados ───────────────────────────────────────────────
            cols_r = [c for c in ["razaoSocial","distribuidora","_pro_frotas",
                                   "municipio","uf","endereco","cep","_dist_rota"]
                      if c in df_show_r.columns]
            df_exib = df_show_r[cols_r].copy()
            if "_pro_frotas" in df_exib.columns:
                df_exib = df_exib.rename(columns={"_pro_frotas":"Gestão de Frotas ⭐"})
            if "_dist_rota" in df_exib.columns:
                df_exib = df_exib.rename(columns={"_dist_rota":"Dist. da Rota (m)"})
                df_exib["Dist. da Rota (m)"] = df_exib["Dist. da Rota (m)"].round(0).astype(int)
            # Coluna de ranking (se existir)
            if "_rank_barato" in df_show_r.columns:
                df_exib.insert(0, "💰 Rank",
                    df_show_r["_rank_barato"].map(
                        lambda v: _RANK_EMOJI.get(int(v), "") if int(v) > 0 else ""
                    )
                )

            # Coluna de link Google Maps — usa lat/lon do df_show_r original
            if "_lat" in df_show_r.columns and "_lon" in df_show_r.columns:
                df_exib["🗺️ Maps"] = df_show_r["_lat"].combine(
                    df_show_r["_lon"],
                    lambda la, lo: (
                        f"https://maps.google.com/?q={la:.6f},{lo:.6f}"
                        if pd.notna(la) and pd.notna(lo) else None
                    ),
                )
                _col_cfg = {
                    "🗺️ Maps": st.column_config.LinkColumn(
                        "🗺️ Maps",
                        display_text="Ver no Maps",
                        help="Abrir posto no Google Maps",
                    )
                }
            else:
                _col_cfg = {}

            st.dataframe(df_exib, use_container_width=True, height=450,
                         column_config=_col_cfg if _col_cfg else None)
            st.download_button("⬇️ Baixar dados em CSV",
                               df_show_r.to_csv(index=False).encode("utf-8"),
                               "postos_rota.csv","text/csv", use_container_width=True)

    else:
        # Mapa vazio centrado no Brasil + instrução informativa
        _renderizar_mapa(criar_mapa(pd.DataFrame()), height=680, key="mapa_m2_vazio")
        st.info("👈 Preencha Origem e Destino na barra lateral e clique em Traçar Rota")


# ═══════════════════════════════════════════════════════════════════
#  MODO 3 — Consulta por Posto (CNPJ / Razão Social / Nome)
# ═══════════════════════════════════════════════════════════════════

elif modo == "🔍 Consulta por Posto":

    _m3_termo  = st.session_state.get("_m3_termo", "")
    _m3_uf     = st.session_state.get("_m3_uf", "")
    _m3_map_o  = st.session_state.get("_map_orig")
    _m3_map_d  = st.session_state.get("_map_dest")
    _m3_o_ok   = bool(_m3_map_o)
    _m3_d_ok   = bool(_m3_map_d)
    _m3_rr     = st.session_state.get("_map_rota_result")

    # ── Auto-calcular rota quando Origem e Destino estão definidos ──
    if _m3_o_ok and _m3_d_ok:
        _m3_od_sig = (
            f"{_m3_map_o['lat']:.5f},{_m3_map_o['lon']:.5f};"
            f"{_m3_map_d['lat']:.5f},{_m3_map_d['lon']:.5f}"
        )
        _m3_sig_ant = st.session_state.get("_m3_rota_sig", "")
        _m3_mesmo   = (
            abs(_m3_map_o["lat"] - _m3_map_d["lat"]) < 0.0002 and
            abs(_m3_map_o["lon"] - _m3_map_d["lon"]) < 0.0002
        )
        if not _m3_mesmo and (_m3_rr is None or _m3_sig_ant != _m3_od_sig):
            with st.spinner("🗺️ Calculando rota…"):
                _cr_a, _dk_a, _dm_a, _lr_a = calcular_rota(
                    _m3_map_o["lat"], _m3_map_o["lon"],
                    _m3_map_d["lat"], _m3_map_d["lon"])
            st.session_state["_map_rota_result"] = {
                "coords": _cr_a, "dist_km": _dk_a, "dur_min": _dm_a,
                "linha_reta": _lr_a, "orig": _m3_map_o, "dest": _m3_map_d,
            }
            st.session_state["_m3_rota_sig"] = _m3_od_sig
            st.rerun()
        _m3_rr = st.session_state.get("_map_rota_result")

    # ── Cabeçalho ─────────────────────────────────────────────────
    st.markdown(
        "<h2 style='margin:0 0 4px;font-size:1.35rem;color:#1565c0'>"
        "🔍 Consulta por Posto</h2>"
        "<p style='color:#555;font-size:13px;margin:0 0 16px'>"
        "Busque por nome, razão social ou CNPJ — selecione como Origem ou Destino.</p>",
        unsafe_allow_html=True,
    )

    # ── Guia de passos ────────────────────────────────────────────
    def _passo_html_m3(num, titulo, desc, ok, ativo):
        if ok:
            _bg, _brd, _num_bg, _num_c, _title_c = (
                "#e8f5e9", "#a5d6a7", "#43a047", "#fff", "#2e7d32")
            _check = "✔"
        elif ativo:
            _bg, _brd, _num_bg, _num_c, _title_c = (
                "#e3f2fd", "#90caf9", "#1565c0", "#fff", "#1565c0")
            _check = str(num)
        else:
            _bg, _brd, _num_bg, _num_c, _title_c = (
                "#fafafa", "#e0e0e0", "#bdbdbd", "#fff", "#9e9e9e")
            _check = str(num)
        return (
            f"<div style='display:flex;align-items:flex-start;gap:10px;"
            f"background:{_bg};border:1px solid {_brd};border-radius:10px;"
            f"padding:10px 14px;flex:1'>"
            f"<div style='width:24px;height:24px;border-radius:50%;"
            f"background:{_num_bg};color:{_num_c};font-size:12px;font-weight:700;"
            f"display:flex;align-items:center;justify-content:center;flex-shrink:0'>"
            f"{_check}</div>"
            f"<div><div style='font-size:12px;font-weight:700;color:{_title_c}'>{titulo}</div>"
            f"<div style='font-size:11px;color:#666;margin-top:1px'>{desc}</div></div>"
            f"</div>"
        )
    _p1_ok_m3 = bool(_m3_termo)
    _p2_ok_m3 = _m3_o_ok and _m3_d_ok
    _p3_ok_m3 = bool(_m3_rr)
    st.markdown(
        f"<div style='display:flex;gap:8px;margin-bottom:14px;flex-wrap:wrap'>"
        f"{_passo_html_m3(1,'Buscar Posto','Pesquise por nome, razão social ou CNPJ',_p1_ok_m3,not _p1_ok_m3)}"
        f"{_passo_html_m3(2,'Definir Origem / Destino','Clique nos botões ao lado de cada resultado',_p2_ok_m3,_p1_ok_m3 and not _p2_ok_m3)}"
        f"{_passo_html_m3(3,'Traçar Rota','Com os dois pontos definidos, calcule a rota',_p3_ok_m3,_p2_ok_m3)}"
        f"</div>",
        unsafe_allow_html=True,
    )

    # ── Cards Origem / Destino — sempre visíveis ──────────────────
    def _card_od_m3(titulo, cor_brd, cor_bg, cor_txt, sel, icone_vazio, msg_vazio):
        if sel:
            nome = sel.get("label", "?")[:55]
            loc  = f"{sel.get('municipio','')} / {sel.get('uf','')}"
            cnpj = sel.get("cnpj", "—")
            return (
                f"<div style='border:2px solid {cor_brd};border-radius:10px;"
                f"background:{cor_bg};padding:12px 14px'>"
                f"<div style='font-size:10px;font-weight:700;color:{cor_txt};"
                f"letter-spacing:0.8px;text-transform:uppercase;margin-bottom:6px'>"
                f"{titulo}</div>"
                f"<div style='font-size:13px;font-weight:700;color:#1a1a1a;"
                f"line-height:1.3;margin-bottom:4px'>{nome}</div>"
                f"<div style='font-size:11px;color:#555'>📍 {loc}</div>"
                f"<div style='font-size:11px;color:#555'>🪪 {cnpj}</div>"
                f"</div>"
            )
        else:
            return (
                f"<div style='border:2px dashed #d0d0d0;border-radius:10px;"
                f"background:#fafafa;padding:12px 14px;text-align:center'>"
                f"<div style='font-size:10px;font-weight:700;color:#bbb;"
                f"letter-spacing:0.8px;text-transform:uppercase;margin-bottom:8px'>"
                f"{titulo}</div>"
                f"<div style='font-size:22px;margin-bottom:4px'>{icone_vazio}</div>"
                f"<div style='font-size:11px;color:#aaa'>{msg_vazio}</div>"
                f"</div>"
            )

    _co3, _cd3 = st.columns(2)
    _co3.markdown(
        _card_od_m3("🟢 Ponto de Origem","#43a047","#f1f8e9","#2e7d32",
                    _m3_map_o,"📍","Busque e clique em\n'Definir como Origem'"),
        unsafe_allow_html=True,
    )
    _cd3.markdown(
        _card_od_m3("🔴 Ponto de Destino","#e53935","#fff8f8","#c62828",
                    _m3_map_d,"🏁","Busque e clique em\n'Definir como Destino'"),
        unsafe_allow_html=True,
    )
    st.markdown("<div style='height:8px'></div>", unsafe_allow_html=True)

    # ── Ações da rota (recalcular / limpar) ──────────────────────
    if _m3_o_ok and _m3_d_ok:
        _m3_mesmo_ponto = (
            abs(_m3_map_o["lat"] - _m3_map_d["lat"]) < 0.0002 and
            abs(_m3_map_o["lon"] - _m3_map_d["lon"]) < 0.0002
        )
        if _m3_mesmo_ponto:
            st.warning(
                "⚠️ Origem e Destino são o **mesmo posto**. "
                "Use os botões 🟢 / 🔴 abaixo para selecionar postos diferentes."
            )
        else:
            _col_rec_m3, _col_clr_m3 = st.columns([3, 1])
            if _col_rec_m3.button(
                "🔄 Recalcular Rota",
                use_container_width=True, key="btn_tracar_m3",
                help="Recalcular a rota entre os pontos selecionados",
            ):
                st.session_state.pop("_map_rota_result", None)
                st.session_state.pop("_m3_rota_sig", None)
                st.rerun()
            if _col_clr_m3.button("↺ Limpar", use_container_width=True,
                                  key="btn_clr_m3_od",
                                  help="Limpar origem, destino e rota"):
                for _k in ["_map_orig", "_map_dest", "_map_rota_result", "_m3_rota_sig"]:
                    st.session_state.pop(_k, None)
                st.rerun()
    elif _m3_o_ok or _m3_d_ok:
        _falta_m3 = "Destino" if _m3_o_ok else "Origem"
        st.info(f"👆 Defina também o **{_falta_m3}** — a rota será calculada automaticamente.")

    # ── Resultado da rota traçada ─────────────────────────────────
    if _m3_rr:
        if _m3_rr.get("linha_reta"):
            st.warning("⚠️ OSRM indisponível — rota exibida como linha reta.")

        # Banner cabeçalho da rota
        st.markdown(
            f"<div style='background:linear-gradient(90deg,#e8f5e9,#f1f8e9);"
            f"border:1px solid #a5d6a7;border-radius:10px;padding:12px 16px;"
            f"margin:6px 0 10px;display:flex;align-items:center;gap:12px;flex-wrap:wrap'>"
            f"<span style='font-size:20px'>✅</span>"
            f"<div style='flex:1'>"
            f"<div style='font-size:13px;font-weight:700;color:#1b5e20'>"
            f"{_m3_rr['orig']['label'][:40]} → {_m3_rr['dest']['label'][:40]}</div>"
            f"</div></div>",
            unsafe_allow_html=True,
        )

        # ── Métricas da rota — cards grandes ─────────────────────
        _dist_m3  = _m3_rr["dist_km"]
        _dur_m3   = _m3_rr["dur_min"]
        _h_m3     = int(_dur_m3 // 60)
        _min_m3   = int(_dur_m3 % 60)
        _vel_m3   = round(_dist_m3 / (_dur_m3 / 60), 0) if _dur_m3 > 0 else 0
        _tempo_str = (f"{_h_m3}h {_min_m3}min" if _h_m3 > 0 else f"{_min_m3} min")

        _mc1, _mc2, _mc3 = st.columns(3)
        _mc1.metric("🛣️ Distância", f"{_n(_dist_m3)} km")
        _mc2.metric("⏱️ Tempo estimado", _tempo_str)
        _mc3.metric("🚗 Vel. média", f"{int(_vel_m3)} km/h")

    st.divider()

    if not _m3_termo:
        # Estado inicial — mapa do Brasil + dica
        if _m3_rr:
            # Já tem rota traçada — exibe com rota
            _mapa_m3_vazio = criar_mapa(
                pd.DataFrame(),
                coords_rota=_m3_rr["coords"],
                lat_orig=_m3_rr["orig"]["lat"], lon_orig=_m3_rr["orig"]["lon"],
                lat_dest=_m3_rr["dest"]["lat"], lon_dest=_m3_rr["dest"]["lon"],
                label_orig=_m3_rr["orig"]["label"], label_dest=_m3_rr["dest"]["label"],
            )
            _renderizar_mapa(_mapa_m3_vazio, height=560, key="mapa_m3_vazio_rota")
        else:
            _renderizar_mapa(criar_mapa(pd.DataFrame()), height=560, key="mapa_m3_vazio")
            st.info("👈 Digite o nome do posto, razão social ou CNPJ na barra lateral e clique em **Buscar Posto**.")

    else:
        # ── Executa busca (cacheada no session_state) ──────────────
        if st.session_state.get("_m3_resultado") is None:
            with st.spinner(f"🔍 Buscando **{_m3_termo}**…"):
                _df_m3, _fonte_m3 = _buscar_posto_completo(_m3_termo, uf=_m3_uf)
                st.session_state["_m3_resultado"] = (_df_m3, _fonte_m3)
        else:
            _df_m3, _fonte_m3 = st.session_state["_m3_resultado"]

        # ── Sem resultados ─────────────────────────────────────────
        if _df_m3 is None or _df_m3.empty:
            st.warning(
                f"⚠️ Nenhum posto encontrado para **{_m3_termo}**"
                + (f" no estado **{_m3_uf}**" if _m3_uf else "")
                + ".\n\nDicas: verifique a ortografia, tente um trecho menor do nome, "
                "ou selecione outro estado."
            )
            _renderizar_mapa(criar_mapa(pd.DataFrame()), height=500, key="mapa_m3_sem_res")

        else:
            n_res = len(_df_m3)
            n_pf  = int(_df_m3["_pro_frotas"].sum()) if "_pro_frotas" in _df_m3.columns else 0
            n_cer = int(_df_m3["_cercado"].sum())    if "_cercado"    in _df_m3.columns else 0

            # ── Métricas resumo ────────────────────────────────────
            _ca, _cb, _cc, _cd = st.columns(4)
            _ca.metric("⛽ Postos encontrados", n_res)
            _cb.metric("⭐ Gestão de Frotas",         n_pf)
            _cc.metric("⚠️ Cercados",           n_cer)
            _cd.metric("📍 Estado(s)",
                       _df_m3["uf"].nunique() if "uf" in _df_m3.columns else "—")
            st.caption(f"Fonte: {_fonte_m3}")

            # ── Botão Salvar (Modo 3) ─────────────────────────────
            _nome_sugerido_m3 = f"Busca: {_m3_termo[:40]}"
            _col_sv3a, _col_sv3b = st.columns([3, 1])
            with _col_sv3a:
                _nome_salvar_m3 = st.text_input(
                    "Nome da busca",
                    value=_nome_sugerido_m3,
                    key="nome_salvar_m3",
                    label_visibility="collapsed",
                    placeholder="Nome para identificar esta busca…",
                )
            with _col_sv3b:
                if st.button("💾 Salvar", use_container_width=True, key="btn_salvar_m3",
                             help="Salvar esta busca para acessar depois"):
                    _dados_m3 = {
                        "_m3_termo": _m3_termo,
                        "_m3_uf":    _m3_uf,
                    }
                    if _salvar_rota_nova(_nome_salvar_m3 or _nome_sugerido_m3, "busca", _dados_m3):
                        st.toast(f"✅ Busca **{_nome_salvar_m3 or _nome_sugerido_m3}** salva!", icon="💾")
                    else:
                        st.error("❌ Não foi possível salvar. Verifique permissões do diretório.")

            # ── Abas: Mapa · Detalhes · Preços ────────────────────
            _tab_map_m3, _tab_det_m3, _tab_preco_m3 = st.tabs([
                "🗺️  Mapa", "📋  Detalhes", "💰  Preços ANP"
            ])

            with _tab_map_m3:
                # Zoom adaptativo: 1 posto = zoom 15, vários estados = zoom 5
                if _m3_rr:
                    # Exibe mapa com rota traçada sobreposta aos postos encontrados
                    _mapa_m3 = criar_mapa(
                        _df_m3,
                        coords_rota=_m3_rr["coords"],
                        lat_orig=_m3_rr["orig"]["lat"], lon_orig=_m3_rr["orig"]["lon"],
                        lat_dest=_m3_rr["dest"]["lat"], lon_dest=_m3_rr["dest"]["lon"],
                        label_orig=_m3_rr["orig"]["label"], label_dest=_m3_rr["dest"]["label"],
                    )
                else:
                    _mapa_m3 = criar_mapa(_df_m3)
                    if n_res == 1:
                        # Zoom máximo para posto único
                        _mapa_m3.update_layout(
                            mapbox=dict(
                                zoom=15,
                                center=dict(
                                    lat=float(_df_m3["_lat"].iloc[0]),
                                    lon=float(_df_m3["_lon"].iloc[0]),
                                ),
                            )
                        )
                    elif "uf" in _df_m3.columns and _df_m3["uf"].nunique() == 1:
                        _mapa_m3.update_layout(mapbox=dict(zoom=8))

                _renderizar_mapa(_mapa_m3, height=560, key="mapa_m3_res")

                # ── Exportar mapa como PNG (Modo 3) ───────────────
                _m3c1, _m3c2 = st.columns([5, 1])
                with _m3c2:
                    if st.button("📸 Exportar mapa", use_container_width=True,
                                 key="btn_exp_mapa_m3",
                                 help="Baixar o mapa dos postos como imagem PNG"):
                        with st.spinner("🖼️ Gerando imagem…"):
                            try:
                                _titulo_m3 = f"Busca: {_m3_termo[:30]}" + (f" — {_m3_uf}" if _m3_uf else "")
                                _sub_m3    = f"{len(_df_m3)} postos  ·  {datetime.now().strftime('%d/%m/%Y %H:%M')}"
                                _png_m3 = _exportar_mapa_postos_png(_df_m3, _titulo_m3, _sub_m3)
                                st.session_state["_mapa_png_m3"] = _png_m3
                                st.session_state["_mapa_png_m3_nome"] = (
                                    f"busca_{_m3_termo[:15].replace(' ','_')}"
                                    f"_{datetime.now().strftime('%Y%m%d_%H%M')}.png"
                                )
                            except Exception as _emp3:
                                st.error(f"Erro ao gerar imagem: {_emp3}")
                if st.session_state.get("_mapa_png_m3"):
                    with _m3c1:
                        st.download_button(
                            "⬇️ Baixar PNG do mapa",
                            data=st.session_state["_mapa_png_m3"],
                            file_name=st.session_state.get("_mapa_png_m3_nome","mapa_busca.png"),
                            mime="image/png",
                            use_container_width=True,
                            key="dl_mapa_png_m3",
                        )

                # ── Botões de O/D para os resultados ──────────────
                _df_m3_od = _df_m3.head(8)
                st.markdown(
                    "<div style='background:#f8f9fa;border:1px solid #e3e8f0;"
                    "border-radius:10px;padding:12px 14px;margin-top:10px'>"
                    "<div style='font-size:13px;font-weight:700;color:#1565c0;"
                    "margin-bottom:6px'>📍 Usar posto como Origem ou Destino</div>"
                    "<div style='font-size:11px;color:#666'>"
                    "Clique nos botões ao lado de cada posto para defini-lo na rota.</div>"
                    "</div>",
                    unsafe_allow_html=True,
                )
                st.markdown("<div style='height:6px'></div>", unsafe_allow_html=True)
                for _idx_od, _row_od in _df_m3_od.iterrows():
                    _ic_od  = "⭐" if bool(_row_od.get("_pro_frotas")) else "⛽"
                    _nm_od  = str(_row_od.get("razaoSocial", "?"))[:50]
                    _loc_od = f"{_row_od.get('municipio','')}/{_row_od.get('uf','')}"
                    _cn_od  = str(_row_od.get("cnpj", "—"))
                    _sel_od = {
                        "lat":       float(_row_od["_lat"]),
                        "lon":       float(_row_od["_lon"]),
                        "label":     str(_row_od.get("razaoSocial", "Posto")),
                        "municipio": str(_row_od.get("municipio", "")),
                        "uf":        str(_row_od.get("uf", "")),
                        "cnpj":      _cn_od,
                    }
                    # Destaca quando já selecionado
                    _is_orig_od = (_m3_map_o or {}).get("cnpj") == _cn_od and _cn_od != "—"
                    _is_dest_od = (_m3_map_d or {}).get("cnpj") == _cn_od and _cn_od != "—"
                    _brd_od = ("#43a047" if _is_orig_od else
                               "#e53935" if _is_dest_od else "#e0e0e0")
                    _bg_od  = ("#f1f8e9" if _is_orig_od else
                               "#fff8f8" if _is_dest_od else "#fff")
                    st.markdown(
                        f"<div style='background:{_bg_od};border:1px solid {_brd_od};"
                        f"border-radius:8px;padding:8px 12px;margin-bottom:4px;"
                        f"display:flex;align-items:center;gap:8px'>"
                        f"<span style='font-size:16px'>{_ic_od}</span>"
                        f"<div style='flex:1'>"
                        f"<div style='font-size:12px;font-weight:600;color:#1a1a1a'>{_nm_od}</div>"
                        f"<div style='font-size:10px;color:#888'>📍 {_loc_od}</div>"
                        f"</div></div>",
                        unsafe_allow_html=True,
                    )
                    _c1_od, _c2_od = st.columns(2)
                    if _c1_od.button(
                        "🟢 Definir Origem" + (" ✔" if _is_orig_od else ""),
                        key=f"m3_set_orig_{_idx_od}",
                        use_container_width=True,
                        type="primary" if not _m3_o_ok else "secondary",
                        help="Marcar como ponto de partida da rota",
                    ):
                        st.session_state["_map_orig"] = _sel_od
                        st.rerun()
                    if _c2_od.button(
                        "🔴 Definir Destino" + (" ✔" if _is_dest_od else ""),
                        key=f"m3_set_dest_{_idx_od}",
                        use_container_width=True,
                        type="primary" if not _m3_d_ok else "secondary",
                        help="Marcar como ponto de chegada da rota",
                    ):
                        st.session_state["_map_dest"] = _sel_od
                        st.rerun()
                if len(_df_m3) > 8:
                    st.caption(f"Exibindo 8 de {len(_df_m3)} resultados. Refine a busca para ver mais.")

            with _tab_det_m3:
                # Card detalhado para posto único
                if n_res == 1:
                    _r = _df_m3.iloc[0]
                    _is_pf_r  = bool(_r.get("_pro_frotas"))
                    _is_cer_r = bool(_r.get("_cercado"))
                    _is_rr_r  = bool(_r.get("_rodo_rede"))
                    _badges = ""
                    if _is_pf_r:  _badges += "<span style='background:#1565c0;color:#fff;border-radius:4px;padding:2px 8px;font-size:11px;margin-right:4px'>⭐ GESTÃO DE FROTAS</span>"
                    if _is_rr_r:  _badges += "<span style='background:#FFB300;color:#333;border-radius:4px;padding:2px 8px;font-size:11px;margin-right:4px'>🚛 IPIRANGA RODOREDE</span>"
                    if _is_cer_r: _badges += "<span style='background:#FF8F00;color:#fff;border-radius:4px;padding:2px 8px;font-size:11px;margin-right:4px'>⚠️ CERCADO</span>"

                    _td = "color:#888;padding:3px 8px 3px 0"
                    _perfil_row = (
                        f"<tr><td style='{_td}'>🏷️ Perfil Venda</td>"
                        f"<td>{_r.get('_perfil_venda','')}</td></tr>"
                        if _r.get("_perfil_venda") else ""
                    )
                    _autor_row = (
                        f"<tr><td style='{_td}'>🔑 Autorização</td>"
                        f"<td>{_r.get('autorizacao','')}</td></tr>"
                        if _r.get("autorizacao") else ""
                    )
                    st.markdown(
                        f"<div style='background:#f8f9fa;border-radius:12px;padding:18px 22px;"
                        f"border:1px solid #e0e0e0;margin-bottom:12px'>"
                        f"<div style='font-size:1.1rem;font-weight:700;color:#1565c0;margin-bottom:6px'>"
                        f"{_r.get('razaoSocial','—')}</div>"
                        f"{_badges}"
                        f"<table style='margin-top:10px;font-size:13px;width:100%;border-collapse:collapse'>"
                        f"<tr><td style='{_td};width:130px'>⛽ Distribuidora</td>"
                        f"<td><b>{_r.get('distribuidora','—')}</b></td></tr>"
                        f"<tr><td style='{_td}'>📋 CNPJ</td>"
                        f"<td>{_r.get('cnpj','—')}</td></tr>"
                        f"<tr><td style='{_td}'>📍 Endereço</td>"
                        f"<td>{_r.get('endereco','—')}, {_r.get('bairro','')} — "
                        f"{_r.get('municipio','—')}/{_r.get('uf','—')} "
                        f"CEP {_r.get('cep','—')}</td></tr>"
                        f"<tr><td style='{_td}'>🗺️ Coordenadas</td>"
                        f"<td>{float(_r['_lat']):.5f}, {float(_r['_lon']):.5f}</td></tr>"
                        f"{_perfil_row}{_autor_row}"
                        f"</table></div>",
                        unsafe_allow_html=True,
                    )
                    # ── Favorito + Anotação (apenas resultado único) ───────────
                    _cnpj_det = str(_r.get("cnpj", "")).replace(".", "").replace("/", "").replace("-", "")
                    _nome_det = str(_r.get("razaoSocial", ""))
                    _mun_det  = str(_r.get("municipio", ""))
                    _uf_det   = str(_r.get("uf", ""))
                    _lat_det  = float(_r.get("_lat", 0) or 0)
                    _lon_det  = float(_r.get("_lon", 0) or 0)

                    if "fav_cnpjs" not in st.session_state:
                        st.session_state["fav_cnpjs"] = {r["cnpj"] for r in _db_favoritos()}

                    _e_fav = _cnpj_det in st.session_state["fav_cnpjs"]
                    _col_fav_a, _col_fav_b = st.columns([1, 1])
                    with _col_fav_a:
                        _fav_label = "⭐ Remover Favorito" if _e_fav else "☆ Adicionar Favorito"
                        if st.button(_fav_label, use_container_width=True, key="btn_fav_det"):
                            if _e_fav:
                                _db_remove_favorito(_cnpj_det)
                                st.session_state["fav_cnpjs"].discard(_cnpj_det)
                                st.toast("Removido dos favoritos", icon="☆")
                            else:
                                _db_add_favorito(_cnpj_det, _nome_det, _mun_det,
                                                 _uf_det, _lat_det, _lon_det)
                                st.session_state["fav_cnpjs"].add(_cnpj_det)
                                st.toast("Adicionado aos favoritos!", icon="⭐")
                            st.rerun()

                    # ── Anotação interna ──────────────────────────────────────
                    with st.expander("📝 Anotação interna", expanded=False):
                        _nota_key = f"nota_posto_{_cnpj_det}"
                        if _nota_key not in st.session_state:
                            st.session_state[_nota_key] = _db_nota_posto(_cnpj_det)
                        _nota_val = st.text_area(
                            "Notas (contato, condições, restrições…)",
                            value=st.session_state[_nota_key],
                            height=120,
                            key=f"ta_{_nota_key}",
                            placeholder="Ex: Falar com João - (11) 99999-0000 · só atende caminhão com agendamento",
                            label_visibility="collapsed",
                        )
                        if st.button("💾 Salvar anotação", key=f"btn_nota_{_cnpj_det}",
                                     use_container_width=True):
                            if _db_salvar_nota_posto(_cnpj_det, _nota_val):
                                st.session_state[_nota_key] = _nota_val
                                st.toast("✅ Anotação salva!", icon="📝")
                            else:
                                st.error("❌ Erro ao salvar. Verifique a conexão com o banco.")

                else:
                    # Tabela para múltiplos resultados
                    _cols_m3 = [c for c in [
                        "razaoSocial", "cnpj", "distribuidora",
                        "_pro_frotas", "_cercado",
                        "municipio", "uf", "endereco", "cep",
                    ] if c in _df_m3.columns]
                    _df_exib_m3 = _df_m3[_cols_m3].copy()
                    if "_pro_frotas" in _df_exib_m3.columns:
                        _df_exib_m3 = _df_exib_m3.rename(columns={"_pro_frotas": "Gestão de Frotas ⭐"})
                    if "_cercado" in _df_exib_m3.columns:
                        _df_exib_m3 = _df_exib_m3.rename(columns={"_cercado": "Cercado ⚠️"})
                    st.dataframe(_df_exib_m3, use_container_width=True, height=450)
                    st.download_button(
                        "⬇️ Baixar resultados em CSV",
                        _df_m3.to_csv(index=False).encode("utf-8"),
                        f"consulta_{_m3_termo[:20].replace(' ','_')}.csv",
                        "text/csv",
                        use_container_width=True,
                    )

            with _tab_preco_m3:
                # ── Determina UF e município do(s) posto(s) encontrado(s) ──────
                _ufs_m3   = list(_df_m3["uf"].dropna().unique()) if "uf" in _df_m3.columns else []
                _uf_m3    = _ufs_m3[0] if len(_ufs_m3) == 1 else None
                _mun_m3   = (str(_df_m3.iloc[0].get("municipio", "")) or None) if n_res == 1 else None

                # ── Seção 1: Preços reais do posto (Planilha PP) ─────────────
                _pp_df_m3      = st.session_state.get("_pp_df")
                _cnpjs_pf_m3   = st.session_state.get("cnpjs_pro_frotas", set())
                _cache_anp_m3  = st.session_state.get("_precos_anp_cache", {})
                _sheets_m3     = _cache_anp_m3.get("sheets")
                _semana_m3     = _cache_anp_m3.get("semana", "")

                if n_res == 1 and _pp_df_m3 is not None:
                    _r0       = _df_m3.iloc[0]
                    _cnpj_raw = str(_r0.get("cnpj", ""))
                    _cnpj_n   = "".join(c for c in _cnpj_raw if c.isdigit())
                    _df_pp_posto = (
                        _pp_df_m3[_pp_df_m3["cnpj_norm"] == _cnpj_n]
                        if "cnpj_norm" in _pp_df_m3.columns else pd.DataFrame()
                    )

                    if not _df_pp_posto.empty:
                        st.markdown(
                            "<div style='font-size:14px;font-weight:700;color:#1565c0;"
                            "margin-bottom:10px'>💲 Preços do Posto vs Referência ANP</div>",
                            unsafe_allow_html=True,
                        )
                        # Para cada combustível presente na planilha PP
                        _fuels_pp = (
                            _df_pp_posto["combustivel_pk"].dropna().unique().tolist()
                            if "combustivel_pk" in _df_pp_posto.columns else []
                        )
                        _cards_posto_html = ""
                        for _fp in sorted(_fuels_pp):
                            _rows_fp = _df_pp_posto[_df_pp_posto["combustivel_pk"] == _fp]
                            _p_posto = float(_rows_fp["preco"].mean()) if "preco" in _rows_fp.columns else None
                            _data_fp = ""
                            if "data_atualizacao" in _rows_fp.columns:
                                _d_fp = _rows_fp["data_atualizacao"].dropna()
                                _data_fp = str(_d_fp.iloc[0]) if not _d_fp.empty else ""

                            # Referências ANP (município → estado → Brasil)
                            _p_mun = _p_uf = _p_br = None
                            if _sheets_m3:
                                if _mun_m3:
                                    _p_mun, _, _ = _anp_preco_ponto(_sheets_m3, _mun_m3, _fp)
                                if _uf_m3:
                                    _p_uf,  _, _ = _anp_preco_ponto(_sheets_m3, _uf_m3,  _fp)
                                _p_br,  _, _ = _anp_preco_ponto(_sheets_m3, "brasil", _fp)

                            # Calcula delta vs estado
                            _delta_html = ""
                            _ref = _p_uf or _p_br
                            if _p_posto and _ref:
                                _d = _p_posto - _ref
                                _dpct = _d / _ref * 100
                                _d_cor = "#2e7d32" if _d < 0 else "#c62828"
                                _d_sinal = "▼" if _d < 0 else "▲"
                                _d_txt   = "abaixo" if _d < 0 else "acima"
                                _delta_html = (
                                    f"<div style='font-size:10px;color:{_d_cor};"
                                    f"background:{'#e8f5e9' if _d<0 else '#ffebee'};"
                                    f"border-radius:4px;padding:3px 6px;margin-top:4px;text-align:center'>"
                                    f"{_d_sinal} R$ {_brl(abs(_d),3)} ({abs(_dpct):.1f}%) "
                                    f"{_d_txt} do estado</div>"
                                )

                            _nome_comb = PRODUTO_CURTO.get(_fp, _fp)
                            _data_html = f"<div style='font-size:10px;color:#aaa;margin-top:2px'>{_data_fp}</div>" if _data_fp else ""
                            _refs_html = ""
                            if _p_mun:
                                _refs_html += f"<tr><td style='font-size:11px;color:#555;padding:2px 0'>Município</td><td style='font-size:11px;font-weight:600;text-align:right'>R$ {_brl(_p_mun,3)}</td></tr>"
                            if _p_uf:
                                _refs_html += f"<tr><td style='font-size:11px;color:#555;padding:2px 0'>Estado ({_uf_m3})</td><td style='font-size:11px;font-weight:600;text-align:right'>R$ {_brl(_p_uf,3)}</td></tr>"
                            if _p_br:
                                _refs_html += f"<tr><td style='font-size:11px;color:#555;padding:2px 0'>Brasil</td><td style='font-size:11px;font-weight:600;text-align:right'>R$ {_brl(_p_br,3)}</td></tr>"
                            if _refs_html:
                                _refs_html = (
                                    "<div style='border-top:1px solid #eee;margin-top:8px;padding-top:6px'>"
                                    "<div style='font-size:10px;color:#888;margin-bottom:4px'>Referência ANP</div>"
                                    f"<table style='width:100%;border-collapse:collapse'>{_refs_html}</table></div>"
                                )

                            _p_str = f"R$ {_brl(_p_posto, 3)}/L" if _p_posto else "—"
                            _cards_posto_html += (
                                f"<div style='border-radius:10px;border:1px solid #e0e0e0;"
                                f"padding:14px 16px;background:#fff;box-shadow:0 1px 6px rgba(0,0,0,.07)'>"
                                f"<div style='font-size:12px;font-weight:700;color:#1565c0;margin-bottom:2px'>{_nome_comb}</div>"
                                f"<div style='font-size:20px;font-weight:800;color:#212121'>{_p_str}</div>"
                                f"{_data_html}{_delta_html}{_refs_html}"
                                f"</div>"
                            )

                        if _cards_posto_html:
                            st.markdown(
                                f"<div style='display:grid;grid-template-columns:repeat(auto-fill,minmax(200px,1fr));"
                                f"gap:12px;margin-bottom:20px'>{_cards_posto_html}</div>",
                                unsafe_allow_html=True,
                            )
                            if _semana_m3:
                                st.caption(f"Referência ANP: {_semana_m3}")
                        st.divider()

                # ── Seção 2: Tabela ANP completa (Modo 1 — sem calculadora) ──
                if _uf_m3:
                    _renderizar_precos_anp(_uf_m3, municipio=_mun_m3)
                elif _ufs_m3:
                    # Múltiplos estados: selectbox para evitar chaves duplicadas
                    _nomes_m3 = [f"{_u} — {UF_NOME.get(_u, _u)}" for _u in _ufs_m3]
                    _uf_sel_m3 = st.selectbox(
                        "Estado", _nomes_m3,
                        key="preco_m3_uf_sel",
                        label_visibility="collapsed",
                    )
                    _uf_code_m3 = _uf_sel_m3.split(" — ")[0] if _uf_sel_m3 else _ufs_m3[0]
                    _renderizar_precos_anp(_uf_code_m3)
                else:
                    st.info("ℹ️ Estado não identificado para este posto. Preços ANP indisponíveis.")


# ═══════════════════════════════════════════════════════════════════
#  MODO 4 — Rotas Salvas
# ═══════════════════════════════════════════════════════════════════

elif modo == "📋 Rotas Salvas":

    st.markdown(
        "<h2 style='margin:0 0 4px;font-size:1.35rem;color:#2E7D32'>"
        "📋 Rotas Salvas</h2>"
        "<p style='color:#555;font-size:13px;margin:0 0 16px'>"
        "Consultas e rotas salvas anteriormente. Clique em <b>Restaurar</b> para "
        "recarregar a visualização completa.</p>",
        unsafe_allow_html=True,
    )

    _rotas_list = _carregar_rotas_salvas()

    if not _rotas_list:
        st.info(
            "📭 Nenhuma consulta salva ainda.\n\n"
            "Use o botão **💾 Salvar** que aparece ao visualizar resultados em qualquer modo "
            "(Estado, Rota ou Busca) para guardar uma consulta aqui."
        )
        _renderizar_mapa(criar_mapa(pd.DataFrame()), height=480, key="mapa_m4_vazio")

    else:
        # ── Filtros rápidos ───────────────────────────────────────────
        _tipos_disp = sorted({r.get("tipo", "outro") for r in _rotas_list})
        _tipo_labels = {"estado": "📍 Estado", "rota": "🗺️ Rota", "busca": "🔍 Busca", "roteirizacao": "🛣️ Roteirização"}
        _filtro_tipo = st.multiselect(
            "Filtrar por tipo",
            options=_tipos_disp,
            format_func=lambda t: _tipo_labels.get(t, t),
            default=_tipos_disp,
            key="filtro_tipo_rotas",
            label_visibility="collapsed",
        )
        _rotas_filtradas = [r for r in _rotas_list if r.get("tipo") in _filtro_tipo]

        if not _rotas_filtradas:
            st.warning("⚠️ Nenhuma consulta corresponde ao filtro selecionado.")
        else:
            st.caption(f"{len(_rotas_filtradas)} consulta(s) encontrada(s) — mais recente primeiro")

        # ── Cards de rotas (ordem reversa — mais recente primeiro) ────
        for _rv in reversed(_rotas_filtradas):
            _rv_id    = _rv.get("id", "")
            _rv_nome  = _rv.get("nome", "—")
            _rv_tipo  = _rv.get("tipo", "outro")
            _rv_data  = _rv.get("criado_em", "")
            _rv_ic    = _icone_tipo(_rv_tipo)
            _tipo_lbl = _tipo_labels.get(_rv_tipo, _rv_tipo)

            # Subtítulo específico por tipo
            if _rv_tipo == "estado":
                _rv_uf  = _rv.get("uf", "")
                _rv_mun = _rv.get("municipio", "")
                _rv_sub = f"{UF_NOME.get(_rv_uf, _rv_uf)}" + (f" — {_rv_mun}" if _rv_mun else "")
                _rv_tag = "📍 Estado"
                _rv_cor = "#1565c0"
                _rv_bg  = "#e3f2fd"
            elif _rv_tipo == "rota":
                _rv_orig = _rv.get("label_orig", "?")[:35]
                _rv_dest = _rv.get("label_dest", "?")[:35]
                _rv_km   = _rv.get("dist_km", 0)
                _rv_sub  = f"{_rv_orig} → {_rv_dest}"
                _rv_sub += f" · {_n(_rv_km)} km" if _rv_km else ""
                _rv_tag  = "🗺️ Rota"
                _rv_cor  = "#2E7D32"
                _rv_bg   = "#e8f5e9"
            elif _rv_tipo == "roteirizacao":
                _rv_orig = _rv.get("label_orig", "?")[:30]
                _rv_dest = _rv.get("label_dest", "?")[:30]
                _rv_km   = _rv.get("dist_km", 0)
                _rv_placa = _rv.get("placa", "")
                _rv_comb  = _rv.get("combustivel", "")
                _rv_sub   = f"{_rv_orig} → {_rv_dest}"
                _rv_sub  += f" · {_n(_rv_km)} km" if _rv_km else ""
                if _rv_placa: _rv_sub += f" · 🚛 {_rv_placa}"
                if _rv_comb:  _rv_sub += f" · ⛽ {_rv_comb}"
                _rv_tag   = "🛣️ Roteirização"
                _rv_cor   = "#004D40"
                _rv_bg    = "#e0f7fa"
            else:  # busca
                _rv_term = _rv.get("_m3_termo", "")
                _rv_uf_b = _rv.get("_m3_uf", "")
                _rv_sub  = f'"{_rv_term}"' + (f" — {_rv_uf_b}" if _rv_uf_b else "")
                _rv_tag  = "🔍 Busca"
                _rv_cor  = "#6A1B9A"
                _rv_bg   = "#f3e5f5"

            # Card visual
            st.markdown(
                f"<div style='border-left:4px solid {_rv_cor};background:{_rv_bg};"
                f"border-radius:0 10px 10px 0;padding:12px 16px;margin-bottom:4px'>"
                f"<div style='display:flex;align-items:center;gap:8px;flex-wrap:wrap'>"
                f"<span style='font-size:1rem;font-weight:700;color:#1a1a1a'>{_rv_ic} {_rv_nome}</span>"
                f"<span style='background:{_rv_cor};color:#fff;border-radius:4px;padding:1px 7px;"
                f"font-size:10px;font-weight:600'>{_rv_tag}</span>"
                f"</div>"
                f"<div style='font-size:12px;color:#555;margin-top:4px'>{_rv_sub}</div>"
                f"<div style='font-size:10px;color:#999;margin-top:3px'>🕐 {_rv_data}</div>"
                f"</div>",
                unsafe_allow_html=True,
            )

            # Botões de ação
            _btn_key_rest = f"rest_{_rv_id}"
            _btn_key_del  = f"del_{_rv_id}"
            _col_r, _col_d = st.columns([4, 1])

            with _col_r:
                if st.button(
                    f"🔄 Restaurar — {_rv_nome[:40]}",
                    key=_btn_key_rest,
                    use_container_width=True,
                    type="primary",
                    help="Restaurar esta consulta e visualizar os resultados",
                ):
                    # ── Restaura o session_state conforme o tipo ──────────────
                    if _rv_tipo == "estado":
                        # Limpa estado anterior
                        for _k in ["df_raw_full", "_uf_carregada", "_map_orig",
                                   "_map_dest", "_map_rota_result",
                                   "distribuidoras_disponiveis"]:
                            st.session_state.pop(_k, None)
                        st.session_state["_form_key_m1"] = st.session_state.get("_form_key_m1", 0) + 1
                        # Restaura campos específicos do modo estado via flags
                        st.session_state["_restore_uf"]  = _rv.get("uf", "")
                        st.session_state["_restore_mun"] = _rv.get("municipio", "")
                        st.session_state["_map_orig"]    = _rv.get("_map_orig")
                        st.session_state["_map_dest"]    = _rv.get("_map_dest")
                        if _rv.get("_map_orig") and _rv.get("_map_dest"):
                            # Recalcula a rota ao restaurar
                            st.session_state["_restore_recalc_rota_m1"] = True
                        st.session_state["modo_selecionado"] = "📍 Por UF/Município"

                    elif _rv_tipo == "rota":
                        # Limpa rotas anteriores
                        for _k in ["df_rota", "coords_rota",
                                   "lat_orig", "lon_orig", "label_orig",
                                   "lat_dest", "lon_dest", "label_dest",
                                   "dist_km", "dur_min", "raio_usado", "linha_reta",
                                   "distribuidoras_rota", "_ufs_rota_atual", "_paradas_data"]:
                            st.session_state.pop(_k, None)
                        _n_p_old = st.session_state.get("_paradas_count", 0)
                        for _pi in range(1, _n_p_old + 1):
                            st.session_state.pop(f"parada_sel_{_pi}", None)
                            st.session_state.pop(f"txt_parada_{_pi}", None)
                        # Restaura origem, destino, paradas
                        st.session_state["orig_sel"]      = _rv.get("orig_sel")
                        st.session_state["dest_sel"]      = _rv.get("dest_sel")
                        _paradas_rest = _rv.get("paradas_data", [])
                        st.session_state["_paradas_count"] = len(_paradas_rest)
                        for _pi, _pw in enumerate(_paradas_rest, 1):
                            st.session_state[f"parada_sel_{_pi}"] = _pw
                        # Dispara recálculo automático
                        st.session_state["_form_key"] = st.session_state.get("_form_key", 0) + 1
                        st.session_state["_auto_buscar_rota"] = True
                        st.session_state["modo_selecionado"]  = "🗺️ Por Rota"

                    elif _rv_tipo == "roteirizacao":
                        # 1. Limpa estado de roteirização anterior
                        _old_np = st.session_state.get("_rot_np", 0)
                        for _pi in range(1, _old_np + 1):
                            st.session_state.pop(f"rot_parada_sel_{_pi}", None)
                            st.session_state.pop(f"rot_txt_parada_{_pi}", None)
                        st.session_state.pop("_rot_result", None)
                        # 2. Restaura dados do veículo na sidebar
                        if _rv.get("placa"):
                            st.session_state["rot_placa"]      = _rv["placa"]
                        if _rv.get("combustivel"):
                            st.session_state["rot_combustivel"] = _rv["combustivel"]
                        if _rv.get("capacidade") is not None:
                            st.session_state["rot_capacidade"] = float(_rv["capacidade"])
                        if _rv.get("autonomia") is not None:
                            st.session_state["rot_autonomia"]  = float(_rv["autonomia"])
                        # 3. Restaura origem e destino
                        _orig_r = _rv.get("orig") or {
                            "label": _rv.get("label_orig", ""),
                            "lat":   _rv.get("lat_orig"),
                            "lon":   _rv.get("lon_orig"),
                            "tipo":  "cidade",
                        }
                        _dest_r = _rv.get("dest") or {
                            "label": _rv.get("label_dest", ""),
                            "lat":   _rv.get("lat_dest"),
                            "lon":   _rv.get("lon_dest"),
                            "tipo":  "cidade",
                        }
                        st.session_state["rot_orig_sel"] = _orig_r
                        st.session_state["rot_dest_sel"] = _dest_r
                        st.session_state["rot_txt_orig"] = _orig_r.get("label", "")
                        st.session_state["rot_txt_dest"] = _dest_r.get("label", "")
                        # 4. Restaura paradas (waypoints)
                        _par_r = _rv.get("paradas", [])
                        st.session_state["_rot_np"] = len(_par_r)
                        for _pi, _pw in enumerate(_par_r, 1):
                            st.session_state[f"rot_parada_sel_{_pi}"] = _pw
                            st.session_state[f"rot_txt_parada_{_pi}"] = _pw.get("label", "")
                        # 5. Restaura resultado da rota diretamente (sem recalcular)
                        _rr_saved = _rv.get("rot_result")
                        if _rr_saved:
                            st.session_state["_rot_result"] = _rr_saved
                        # 6. Força re-render do formulário e navega para o modo
                        st.session_state["_rot_fk"] = st.session_state.get("_rot_fk", 0) + 1
                        st.session_state["modo_selecionado"] = "🛣️ Roteirização"

                    else:  # busca
                        for _k in ["_m3_termo", "_m3_uf", "_m3_resultado"]:
                            st.session_state.pop(_k, None)
                        st.session_state["_m3_termo"]         = _rv.get("_m3_termo", "")
                        st.session_state["_m3_uf"]            = _rv.get("_m3_uf", "")
                        st.session_state["_form_key_m3"]      = st.session_state.get("_form_key_m3", 0) + 1
                        st.session_state["modo_selecionado"]  = "🔍 Consulta por Posto"

                    st.toast(f"✅ Restaurando **{_rv_nome}**…", icon="🔄")
                    st.rerun()

            with _col_d:
                if st.button("🗑️", key=_btn_key_del, use_container_width=True,
                             help=f"Excluir '{_rv_nome}'"):
                    if _deletar_rota(_rv_id):
                        st.toast(f"🗑️ **{_rv_nome}** excluída.", icon="✅")
                        st.rerun()
                    else:
                        st.error("❌ Erro ao excluir.")

            st.markdown("<div style='height:8px'></div>", unsafe_allow_html=True)

        # ── Ação global ───────────────────────────────────────────────
        st.markdown("---")
        _col_clr_all, _ = st.columns([2, 3])
        with _col_clr_all:
            if st.button("🗑️ Excluir todas as consultas", use_container_width=True,
                         key="btn_del_todas",
                         help="Remove permanentemente todas as rotas salvas"):
                if _gravar_rotas_salvas([]):
                    st.toast("🗑️ Todas as consultas foram excluídas.", icon="✅")
                    st.rerun()
                else:
                    st.error("❌ Erro ao excluir.")


# ═══════════════════════════════════════════════════════════════════
#  MODO 6 — DASHBOARD ANALÍTICO
# ═══════════════════════════════════════════════════════════════════

# Referência ANP: total estimado de postos por UF (fonte SIMP/ANP 2023)
_ANP_REF_UF = {
    "AC": 320,  "AL": 980,  "AM": 780,  "AP": 330,  "BA": 4100,
    "CE": 2600, "DF": 1300, "ES": 1550, "GO": 2900, "MA": 1400,
    "MG": 8500, "MS": 1650, "MT": 1850, "PA": 1500, "PB": 1150,
    "PE": 2500, "PI": 890,  "PR": 4300, "RJ": 5000, "RN": 1050,
    "RO": 620,  "RR": 270,  "RS": 4700, "SC": 3600, "SE": 780,
    "SP":13000, "TO": 570,
}

_UF_NOME_DASH = {
    "AC":"Acre","AL":"Alagoas","AM":"Amazonas","AP":"Amapá","BA":"Bahia",
    "CE":"Ceará","DF":"Distrito Federal","ES":"Espírito Santo","GO":"Goiás",
    "MA":"Maranhão","MG":"Minas Gerais","MS":"Mato Grosso do Sul",
    "MT":"Mato Grosso","PA":"Pará","PB":"Paraíba","PE":"Pernambuco",
    "PI":"Piauí","PR":"Paraná","RJ":"Rio de Janeiro","RN":"Rio Grande do Norte",
    "RO":"Rondônia","RR":"Roraima","RS":"Rio Grande do Sul",
    "SC":"Santa Catarina","SE":"Sergipe","SP":"São Paulo","TO":"Tocantins",
}

if modo == "📊 Dashboard":

    _pf_dash  = st.session_state.get("pf_coords_df", pd.DataFrame())
    _pp_dash  = st.session_state.get("_pp_df")

    st.markdown(
        "<h2 style='margin:0 0 4px;font-size:1.35rem;"
        "background:linear-gradient(135deg,#0D47A1,#E65100);"
        "-webkit-background-clip:text;-webkit-text-fill-color:transparent'>"
        "📊 Dashboard Analítico GF</h2>"
        "<p style='color:#555;font-size:13px;margin:0 0 14px'>"
        "KPIs de cobertura geográfica e penetração da rede GF nos estados brasileiros.</p>",
        unsafe_allow_html=True,
    )

    if _pf_dash.empty:
        st.warning(
            "⚠️ Nenhum dado GF carregado. "
            "Importe a planilha de postos em **Configurações** para visualizar o dashboard."
        )
    else:
        # ── Pré-processamento ─────────────────────────────────────────────
        _df = _pf_dash.copy()
        _df["uf"] = _df["uf"].fillna("").str.strip().str.upper()
        _df["municipio"] = _df["municipio"].fillna("").str.strip()
        _df_valid = _df[_df["uf"].isin(_ANP_REF_UF.keys())]

        _total_gf    = len(_df)
        _valid_coord = int(_df[pd.notna(_df["_lat"]) & pd.notna(_df["_lon"])].shape[0])
        _total_ufs   = int(_df_valid["uf"].nunique())
        _total_mun   = int(_df_valid["municipio"].replace("", pd.NA).dropna().nunique())
        _cobertura_br= round(_total_ufs / 27 * 100, 1)

        # ── KPIs — linha 1 ───────────────────────────────────────────────
        _k1, _k2, _k3, _k4, _k5 = st.columns(5)
        for _col, _lbl, _val, _delta in [
            (_k1, "⛽ Postos GF",          _fmt_int(_total_gf), None),
            (_k2, "📍 Com Coordenadas",    _fmt_int(_valid_coord),
             f"{_valid_coord/_total_gf*100:.0f}% do total"),
            (_k3, "🗺️ Estados Cobertos",  f"{_total_ufs} / 27",
             f"{_cobertura_br:.0f}% do Brasil"),
            (_k4, "🏙️ Municípios",        _fmt_int(_total_mun), None),
            (_k5, "📊 Média por UF",
             _fmt_int(_total_gf / _total_ufs) if _total_ufs else "—", None),
        ]:
            _col.metric(_lbl, _val, _delta)

        st.markdown("<div style='height:6px'></div>", unsafe_allow_html=True)

        # ── Tabs do dashboard ─────────────────────────────────────────────
        _dt1, _dt2, _dt3, _dt4, _dt5, _dt6 = st.tabs([
            "📊 Cobertura por Estado",
            "🎯 Penetração vs ANP",
            "🗺️ Mapa de Densidade",
            "⛽ Combustíveis",
            "⚠️ Alertas de Preço",
            "⚖️ Modo Comparativo",
        ])

        # ──────────────────────────────────────────────────────────────────
        # TAB 1 — Cobertura por Estado
        # ──────────────────────────────────────────────────────────────────
        with _dt1:
            _uf_cnt = (
                _df_valid.groupby("uf").size().reset_index(name="postos_gf")
                .sort_values("postos_gf", ascending=False)
            )
            _uf_cnt["uf_nome"] = _uf_cnt["uf"].map(_UF_NOME_DASH).fillna(_uf_cnt["uf"])

            # Gráfico de barras horizontal — ranking
            _fig_bar = go.Figure()
            _colors_bar = [
                "#0D47A1" if v >= _uf_cnt["postos_gf"].quantile(0.75) else
                "#1976D2" if v >= _uf_cnt["postos_gf"].median() else
                "#90CAF9"
                for v in _uf_cnt["postos_gf"]
            ]
            _fig_bar.add_trace(go.Bar(
                y=_uf_cnt["uf_nome"],
                x=_uf_cnt["postos_gf"],
                orientation="h",
                marker_color=_colors_bar,
                text=_uf_cnt["postos_gf"].astype(str),
                textposition="outside",
                hovertemplate="<b>%{y}</b><br>Postos GF: %{x}<extra></extra>",
            ))
            _fig_bar.update_layout(
                title="Ranking de Postos GF por Estado",
                xaxis_title="Quantidade de Postos GF",
                yaxis=dict(autorange="reversed"),
                height=max(400, len(_uf_cnt) * 22 + 80),
                margin=dict(l=10, r=60, t=45, b=30),
                plot_bgcolor="rgba(0,0,0,0)",
                paper_bgcolor="rgba(0,0,0,0)",
                font=dict(size=11),
            )
            _fig_bar.update_xaxes(showgrid=True, gridcolor="#E3F2FD", gridwidth=0.5)
            st.plotly_chart(_fig_bar, use_container_width=True)

            # Tabela detalhada
            st.markdown("##### Detalhamento por Estado")
            _uf_det = _uf_cnt.copy()
            _uf_det["% do Total GF"] = (_uf_det["postos_gf"] / _total_gf * 100).round(1).astype(str) + "%"
            _uf_det["Municípios GF"] = _uf_det["uf"].apply(
                lambda u: int(_df_valid[_df_valid["uf"]==u]["municipio"]
                              .replace("", pd.NA).dropna().nunique())
            )
            _uf_det = _uf_det.rename(columns={
                "uf": "UF", "uf_nome": "Estado", "postos_gf": "Postos GF"
            })[["UF","Estado","Postos GF","% do Total GF","Municípios GF"]]
            st.dataframe(_uf_det.reset_index(drop=True), use_container_width=True, height=350)

        # ──────────────────────────────────────────────────────────────────
        # TAB 2 — Penetração vs ANP
        # ──────────────────────────────────────────────────────────────────
        with _dt2:
            _uf_pen = _uf_cnt.copy()
            _uf_pen["anp_total"] = _uf_pen["uf"].map(_ANP_REF_UF).fillna(0)
            _uf_pen["penetracao_pct"] = (
                _uf_pen["postos_gf"] / _uf_pen["anp_total"] * 100
            ).round(2)
            _uf_pen = _uf_pen.sort_values("penetracao_pct", ascending=False)

            # Gráfico de penetração
            _pen_max = float(_uf_pen["penetracao_pct"].max())
            _pen_colors = [
                "#2E7D32" if v >= _pen_max * 0.66 else
                "#F57F17" if v >= _pen_max * 0.33 else
                "#B71C1C"
                for v in _uf_pen["penetracao_pct"]
            ]
            _fig_pen = go.Figure()
            _fig_pen.add_trace(go.Bar(
                y=_uf_pen["uf_nome"],
                x=_uf_pen["penetracao_pct"],
                orientation="h",
                marker_color=_pen_colors,
                text=_uf_pen["penetracao_pct"].apply(lambda v: f"{v:.2f}%"),
                textposition="outside",
                hovertemplate=(
                    "<b>%{y}</b><br>"
                    "Penetração GF: %{x:.2f}%<br>"
                    "<extra></extra>"
                ),
            ))
            _fig_pen.update_layout(
                title="Penetração GF (% dos postos ANP por estado)",
                xaxis_title="Penetração GF (%)",
                yaxis=dict(autorange="reversed"),
                height=max(400, len(_uf_pen) * 22 + 80),
                margin=dict(l=10, r=80, t=45, b=30),
                plot_bgcolor="rgba(0,0,0,0)",
                paper_bgcolor="rgba(0,0,0,0)",
                font=dict(size=11),
            )
            _fig_pen.update_xaxes(showgrid=True, gridcolor="#E8F5E9", gridwidth=0.5)

            # Linha de média
            _pen_media = float(_uf_pen["penetracao_pct"].mean())
            _fig_pen.add_vline(
                x=_pen_media, line_dash="dash", line_color="#888",
                annotation_text=f"Média: {_pen_media:.2f}%",
                annotation_position="top right",
                annotation_font_size=10,
            )
            st.plotly_chart(_fig_pen, use_container_width=True)

            # Legenda cores
            _lc1, _lc2, _lc3 = st.columns(3)
            _lc1.markdown("<span style='color:#2E7D32;font-weight:700'>🟢 Alta penetração</span> — top 33%", unsafe_allow_html=True)
            _lc2.markdown("<span style='color:#F57F17;font-weight:700'>🟡 Média penetração</span> — meio 33%", unsafe_allow_html=True)
            _lc3.markdown("<span style='color:#B71C1C;font-weight:700'>🔴 Baixa penetração</span> — fundo 33%", unsafe_allow_html=True)

            st.markdown("##### Tabela de Penetração por Estado")
            _tbl_pen = _uf_pen[["uf","uf_nome","postos_gf","anp_total","penetracao_pct"]].copy()
            _tbl_pen.columns = ["UF","Estado","Postos GF","Total ANP (ref.)","Penetração (%)"]
            _tbl_pen["Total ANP (ref.)"] = _tbl_pen["Total ANP (ref.)"].astype(int)
            st.dataframe(
                _tbl_pen.reset_index(drop=True),
                use_container_width=True, height=350,
                column_config={
                    "Penetração (%)": st.column_config.ProgressColumn(
                        "Penetração (%)", min_value=0,
                        max_value=float(_tbl_pen["Penetração (%)"].max()),
                        format="%.2f%%",
                    )
                }
            )
            st.caption(
                "⚠️ Os totais ANP são estimativas de referência (SIMP/ANP 2023). "
                "Para análise precisa importe os dados oficiais mais recentes."
            )

        # ──────────────────────────────────────────────────────────────────
        # TAB 3 — Mapa de Densidade
        # ──────────────────────────────────────────────────────────────────
        with _dt3:
            _df_map = _df[pd.notna(_df["_lat"]) & pd.notna(_df["_lon"])].copy()
            if _df_map.empty:
                st.warning("Nenhum posto com coordenadas válidas para exibir no mapa.")
            else:
                # Mapa de scatter com densidade
                _fig_map_d = go.Figure()
                _fig_map_d.add_trace(go.Scattermapbox(
                    lat=_df_map["_lat"].tolist(),
                    lon=_df_map["_lon"].tolist(),
                    mode="markers",
                    marker=dict(
                        size=6,
                        color="#1565C0",
                        opacity=0.65,
                    ),
                    text=_df_map.apply(
                        lambda r: f"{r.get('razaoSocial','Posto GF')}<br>"
                                  f"{r.get('municipio','')} / {r.get('uf','')}",
                        axis=1,
                    ).tolist(),
                    hoverinfo="text",
                    name="Postos GF",
                ))
                # Centroide do Brasil
                _clat = float(_df_map["_lat"].mean())
                _clon = float(_df_map["_lon"].mean())
                _fig_map_d.update_layout(
                    mapbox=dict(
                        style="carto-positron",
                        center=dict(lat=_clat, lon=_clon),
                        zoom=3.8,
                    ),
                    height=520,
                    margin=dict(l=0, r=0, t=30, b=0),
                    title="Distribuição Geográfica dos Postos GF",
                )
                st.plotly_chart(_fig_map_d, use_container_width=True)

                # Top 10 municípios
                st.markdown("##### Top 10 Municípios com Mais Postos GF")
                _top_mun = (
                    _df_valid.groupby(["municipio","uf"]).size()
                    .reset_index(name="postos")
                    .sort_values("postos", ascending=False)
                    .head(10)
                )
                _top_mun.columns = ["Município","UF","Postos GF"]
                _fig_top = go.Figure(go.Bar(
                    x=_top_mun["Município"] + " / " + _top_mun["UF"],
                    y=_top_mun["Postos GF"],
                    marker_color="#1565C0",
                    text=_top_mun["Postos GF"],
                    textposition="outside",
                ))
                _fig_top.update_layout(
                    xaxis_tickangle=-35,
                    height=320,
                    margin=dict(l=10, r=10, t=20, b=80),
                    plot_bgcolor="rgba(0,0,0,0)",
                    paper_bgcolor="rgba(0,0,0,0)",
                )
                _fig_top.update_yaxes(showgrid=True, gridcolor="#E3F2FD")
                st.plotly_chart(_fig_top, use_container_width=True)

        # ──────────────────────────────────────────────────────────────────
        # TAB 4 — Combustíveis
        # ──────────────────────────────────────────────────────────────────
        with _dt4:
            if _pp_dash is None or _pp_dash.empty:
                st.info("📋 Importe a planilha de preços em **Configurações** para ver os dados de combustíveis.")
            else:
                # Combustíveis disponíveis e preços
                _comb_df = (
                    _pp_dash.groupby("combustivel_label")["preco"]
                    .agg(["count","mean","min","max"])
                    .reset_index()
                )
                _comb_df.columns = ["Combustível","Qtd Registros","Preço Médio (R$/L)","Mín (R$/L)","Máx (R$/L)"]
                _comb_df = _comb_df.sort_values("Qtd Registros", ascending=False)

                # KPIs de combustíveis
                _cf1, _cf2, _cf3 = st.columns(3)
                _cf1.metric("⛽ Combustíveis cadastrados", str(len(_comb_df)))
                _cf2.metric("📋 Total de registros de preço", _fmt_int(len(_pp_dash)))
                if "cnpj_norm" in _pp_dash.columns:
                    _cf3.metric("🏪 Postos com preço cadastrado",
                                _fmt_int(_pp_dash["cnpj_norm"].nunique()))

                # Gráfico de preços médios
                _fig_comb = go.Figure()
                _fig_comb.add_trace(go.Bar(
                    x=_comb_df["Combustível"],
                    y=_comb_df["Preço Médio (R$/L)"],
                    marker_color="#E65100",
                    text=_comb_df["Preço Médio (R$/L)"].apply(lambda v: f"R$ {v:.3f}".replace(".",",")),
                    textposition="outside",
                    name="Preço Médio",
                    error_y=dict(
                        type="data",
                        array=(_comb_df["Máx (R$/L)"] - _comb_df["Preço Médio (R$/L)"]).tolist(),
                        arrayminus=(_comb_df["Preço Médio (R$/L)"] - _comb_df["Mín (R$/L)"]).tolist(),
                        visible=True,
                        color="#B0BEC5",
                    ),
                ))
                _fig_comb.update_layout(
                    title="Preço Médio por Combustível (R$/L) — base GF",
                    yaxis_title="R$/L",
                    height=350,
                    margin=dict(l=10, r=10, t=45, b=60),
                    plot_bgcolor="rgba(0,0,0,0)",
                    paper_bgcolor="rgba(0,0,0,0)",
                    xaxis_tickangle=-20,
                )
                _fig_comb.update_yaxes(showgrid=True, gridcolor="#FBE9E7")
                st.plotly_chart(_fig_comb, use_container_width=True)

                # Tabela de combustíveis
                st.markdown("##### Detalhamento por Combustível")
                for col in ["Preço Médio (R$/L)","Mín (R$/L)","Máx (R$/L)"]:
                    _comb_df[col] = _comb_df[col].apply(lambda v: f"R$ {v:.3f}".replace(".",","))
                st.dataframe(_comb_df.reset_index(drop=True), use_container_width=True)

        # ──────────────────────────────────────────────────────────────────
        # TAB 5 — Alertas de Preço  (comparação por município via ANP)
        # ──────────────────────────────────────────────────────────────────
        with _dt5:
            # ── Referência ANP por UF (fallback hardcoded — mantido para compatibilidade) ──
            _ANP_REF_COMB = {
                "GASOLINA COMUM":    {"SP": 6.29, "RJ": 6.48, "MG": 6.22, "RS": 6.15,
                                      "PR": 6.08, "SC": 6.05, "BA": 6.35, "GO": 6.18,
                                      "MT": 6.40, "MS": 6.25, "PA": 6.52, "AM": 6.70,
                                      "PE": 6.42, "CE": 6.38, "MA": 6.55, "PI": 6.48,
                                      "AL": 6.43, "SE": 6.40, "RN": 6.37, "PB": 6.41,
                                      "ES": 6.28, "TO": 6.50, "RO": 6.58, "AC": 6.72,
                                      "RR": 6.75, "AP": 6.65, "DF": 6.20},
                "DIESEL S10":        {"SP": 6.05, "RJ": 6.18, "MG": 5.98, "RS": 5.92,
                                      "PR": 5.88, "SC": 5.85, "BA": 6.12, "GO": 6.00,
                                      "MT": 6.22, "MS": 6.08, "PA": 6.28, "AM": 6.45,
                                      "PE": 6.15, "CE": 6.10, "MA": 6.30, "PI": 6.22,
                                      "AL": 6.18, "SE": 6.15, "RN": 6.12, "PB": 6.16,
                                      "ES": 6.02, "TO": 6.25, "RO": 6.35, "AC": 6.48,
                                      "RR": 6.50, "AP": 6.40, "DF": 5.95},
                "DIESEL COMUM":      {"SP": 5.98, "RJ": 6.10, "MG": 5.90, "RS": 5.85,
                                      "PR": 5.80, "SC": 5.78, "BA": 6.05, "GO": 5.93,
                                      "MT": 6.15, "MS": 6.00, "PA": 6.20, "AM": 6.38,
                                      "PE": 6.08, "CE": 6.02, "MA": 6.22, "PI": 6.15,
                                      "AL": 6.10, "SE": 6.08, "RN": 6.05, "PB": 6.09,
                                      "ES": 5.95, "TO": 6.18, "RO": 6.28, "AC": 6.40,
                                      "RR": 6.42, "AP": 6.32, "DF": 5.88},
                "ETANOL":            {"SP": 3.80, "RJ": 4.20, "MG": 3.95, "RS": 4.10,
                                      "PR": 3.88, "SC": 4.05, "BA": 4.28, "GO": 3.72,
                                      "MT": 3.68, "MS": 3.75, "PA": 4.50, "AM": 4.85,
                                      "PE": 4.32, "CE": 4.38, "MA": 4.45, "PI": 4.40,
                                      "AL": 3.90, "SE": 4.25, "RN": 4.30, "PB": 4.35,
                                      "ES": 4.05, "TO": 4.20, "RO": 4.55, "AC": 4.90,
                                      "RR": 4.95, "AP": 4.70, "DF": 4.00},
                "GNV":               {"SP": 4.20, "RJ": 4.35, "MG": 4.15, "RS": 4.10,
                                      "PR": 4.08, "SC": 4.05, "BA": 4.38, "GO": 4.22,
                                      "MT": 4.45, "MS": 4.30, "PA": 4.55, "AM": 4.70,
                                      "PE": 4.40, "CE": 4.35, "MA": 4.55, "PI": 4.48,
                                      "AL": 4.43, "SE": 4.40, "RN": 4.37, "PB": 4.41,
                                      "ES": 4.18, "TO": 4.50, "RO": 4.58, "AC": 4.72,
                                      "RR": 4.75, "AP": 4.65, "DF": 4.10},
            }
            _ALERT_THRESH = 0.05  # 5% acima da média ANP

            # ── Fontes de dados necessárias ────────────────────────────────
            _pp_alert   = st.session_state.get("_pp_df")
            _anp_cache  = st.session_state.get("_precos_anp_cache", {})
            _sheets_alert = _anp_cache.get("sheets")

            _sem_pp   = _pp_alert is None or _pp_alert.empty
            _sem_anp  = _sheets_alert is None

            if _sem_pp and _sem_anp:
                st.info(
                    "ℹ️ Para ativar os alertas de preço, carregue:\n\n"
                    "- **Planilha de Preços por Posto** (em Configurações → Preços dos Postos GF)\n"
                    "- **Planilha ANP de Preços Semanais** (em Configurações → Preços ANP)"
                )
            elif _sem_pp:
                st.info(
                    "ℹ️ Carregue a **Planilha de Preços por Posto** em "
                    "Configurações → Preços dos Postos GF para ativar os alertas."
                )
            else:
                # ── 1. Cruzar preços GF com dados do posto (municipio/uf) ──
                # _pp_alert colunas: cnpj_norm, combustivel_pk, combustivel_label, preco
                # _pf_dash   colunas: cnpj, municipio, uf, razaoSocial, distribuidora
                _pf_info = _pf_dash[["cnpj", "municipio", "uf", "razaoSocial"]].copy() \
                    if "cnpj" in _pf_dash.columns else pd.DataFrame()

                if _pf_info.empty:
                    st.warning("⚠️ Sem dados de localização dos postos GF (municipio/uf). "
                               "Reimporte a planilha de postos.")
                else:
                    _merged = _pp_alert.merge(
                        _pf_info, left_on="cnpj_norm", right_on="cnpj", how="inner"
                    )
                    _merged["municipio"] = _merged["municipio"].fillna("").str.strip()
                    _merged["uf"]        = _merged["uf"].fillna("").str.strip().str.upper()
                    _merged = _merged[_merged["preco"] > 0]

                    if _merged.empty:
                        st.warning("⚠️ Nenhum posto GF com preço e localização encontrado. "
                                   "Verifique se os CNPJs da planilha de preços correspondem "
                                   "aos CNPJs da planilha de postos.")
                    else:
                        # ── 2. Construir lookup ANP por (uf_norm, mun_norm, pk) ──
                        # Nível 1: municipios   → {(uf_n, mun_n, pk): preco}
                        # Nível 2: estados      → {(uf_n, pk): preco}
                        # Nível 3: hardcoded    → _ANP_REF_COMB
                        _anp_by_mun   = {}   # (uf_n, mun_n, pk) → float
                        _anp_by_state = {}   # (uf_n, pk) → float

                        def _build_anp_lookup(sheets):
                            """Extrai lookup ANP de municipios e estados."""
                            def _extract_sheet(df):
                                c_est  = _anp_col(df, "estado", "estados")
                                c_mun  = _anp_col(df, "munic")
                                c_prod = _anp_col(df, "produto")
                                c_med  = _anp_col(df, "medio revenda", "media revenda", "preco medio")
                                return c_est, c_mun, c_prod, c_med

                            # Municipios
                            if "municipios" in sheets:
                                _df_m = sheets["municipios"]
                                _ce, _cm, _cp, _cmed = _extract_sheet(_df_m)
                                if _ce and _cp and _cmed and _cm:
                                    for _, _r in _df_m.iterrows():
                                        _uf_n  = _anp_norm(str(_r.get(_ce, "")))
                                        _mn_n  = _anp_norm(str(_r.get(_cm, "")))
                                        _pk    = _anp_norm(str(_r.get(_cp, "")))
                                        try:
                                            _v = float(str(_r.get(_cmed, "")).replace(",", "."))
                                            if _v > 0:
                                                _anp_by_mun[(_uf_n, _mn_n, _pk)] = _v
                                        except (ValueError, TypeError):
                                            pass

                            # Estados
                            if "estados" in sheets:
                                _df_e = sheets["estados"]
                                _ce, _, _cp, _cmed = _extract_sheet(_df_e)
                                if _ce and _cp and _cmed:
                                    for _, _r in _df_e.iterrows():
                                        _uf_n = _anp_norm(str(_r.get(_ce, "")))
                                        _pk   = _anp_norm(str(_r.get(_cp, "")))
                                        try:
                                            _v = float(str(_r.get(_cmed, "")).replace(",", "."))
                                            if _v > 0:
                                                _anp_by_state[(_uf_n, _pk)] = _v
                                        except (ValueError, TypeError):
                                            pass

                        if _sheets_alert:
                            _build_anp_lookup(_sheets_alert)

                        # ── 3. Mapeamento UF sigla → nome normalizado ANP ──
                        _UF_NOME_ANP = {k: _anp_norm(v) for k, v in UF_NOME.items()}

                        def _get_anp_by_municipio(row):
                            """Busca preço ANP: município → estado → hardcoded."""
                            _pk_raw = str(row.get("combustivel_pk", ""))
                            _pk_can = _PP_PARA_ANP_PK.get(_pk_raw, _pk_raw)
                            _uf_sig = str(row.get("uf", "")).upper().strip()
                            _uf_n   = _UF_NOME_ANP.get(_uf_sig, _anp_norm(_uf_sig))
                            _mn_n   = _anp_norm(str(row.get("municipio", "")))

                            # Nível 1 — municipio (match exato)
                            for _pk_try in [_pk_raw, _pk_can]:
                                _v = _anp_by_mun.get((_uf_n, _mn_n, _pk_try))
                                if _v: return _v, "Município"
                            # Nível 1 — municipio (match por substring)
                            if _mn_n:
                                for (_u, _m, _p), _v in _anp_by_mun.items():
                                    if _u == _uf_n and (_mn_n in _m or _m in _mn_n):
                                        if _pk_raw in _p or _p in _pk_raw or \
                                           _pk_can in _p or _p in _pk_can:
                                            return _v, "Município (aprox.)"

                            # Nível 2 — estado
                            for _pk_try in [_pk_raw, _pk_can]:
                                _v = _anp_by_state.get((_uf_n, _pk_try))
                                if _v: return _v, "Estado"
                            # Nível 2 — estado por substring de produto
                            for (_u, _p), _v in _anp_by_state.items():
                                if _u == _uf_n:
                                    if _pk_raw in _p or _p in _pk_raw or \
                                       _pk_can in _p or _p in _pk_can:
                                        return _v, "Estado (aprox.)"

                            # Nível 3 — fallback hardcoded
                            for _key, _refs in _ANP_REF_COMB.items():
                                if _anp_norm(_key) in _pk_raw or _pk_raw in _anp_norm(_key) or \
                                   _anp_norm(_key) in _pk_can or _pk_can in _anp_norm(_key):
                                    _v = _refs.get(_uf_sig)
                                    if _v: return _v, "Referência (fixo)"
                            return None, None

                        # ── 4. Calcular média GF por posto e obter ref ANP ──
                        # Agrupa por (cnpj_norm, combustivel_pk) → preco médio GF
                        _grp = (
                            _merged.groupby(["cnpj_norm", "combustivel_pk", "combustivel_label",
                                             "municipio", "uf", "razaoSocial"])
                            ["preco"].mean()
                            .reset_index()
                            .rename(columns={"preco": "preco_gf"})
                        )

                        _refs   = _grp.apply(_get_anp_by_municipio, axis=1, result_type="expand")
                        _grp["_anp_ref"]   = _refs[0]
                        _grp["_nivel_anp"] = _refs[1]
                        _grp = _grp.dropna(subset=["_anp_ref"])
                        _grp["_diff_pct"] = (_grp["preco_gf"] - _grp["_anp_ref"]) / _grp["_anp_ref"]
                        _grp["_diff_rs"]  = _grp["preco_gf"] - _grp["_anp_ref"]
                        _grp["_alerta"]   = _grp["_diff_pct"] > _ALERT_THRESH

                        _alert_df = _grp  # alias para compatibilidade com restante do bloco

                        # Colunas para exibição posterior
                        _price_col = "preco_gf"
                        _comb_col  = "combustivel_label"

                        def _get_anp_ref(row):  # mantido para compatibilidade de reutilização
                            return row.get("_anp_ref")

                        # ── 5. KPIs e visualizações ────────────────────────
                        _n_total    = len(_alert_df)
                        _n_alertas  = int(_alert_df["_alerta"].sum())
                        _n_ok       = _n_total - _n_alertas
                        _pct_alert  = (_n_alertas / _n_total * 100) if _n_total > 0 else 0
                        _pior_diff  = _alert_df.loc[_alert_df["_alerta"], "_diff_pct"].max() if _n_alertas > 0 else 0
                        _media_diff = _alert_df.loc[_alert_df["_alerta"], "_diff_pct"].mean() if _n_alertas > 0 else 0
                        _nivel_info = (
                            _alert_df["_nivel_anp"].value_counts().idxmax()
                            if "_nivel_anp" in _alert_df.columns and not _alert_df["_nivel_anp"].isna().all()
                            else "Referência (fixo)"
                        )

                        # KPIs
                        _ac1, _ac2, _ac3, _ac4 = st.columns(4)
                        _ac1.metric("⚠️ Postos em Alerta", _fmt_int(_n_alertas),
                                    delta=f"{_pct_alert:.1f}% da base",
                                    delta_color="inverse")
                        _ac2.metric("✅ Dentro da Média", _fmt_int(_n_ok),
                                    delta=f"{100-_pct_alert:.1f}% da base")
                        _ac3.metric("📈 Pior Desvio", f"+{_pior_diff*100:.1f}%" if _n_alertas > 0 else "—")
                        _ac4.metric("📊 Desvio Médio", f"+{_media_diff*100:.1f}%" if _n_alertas > 0 else "—")

                        if not _sem_anp:
                            st.caption(f"🔍 Referência ANP utilizada principalmente: **{_nivel_info}**")
                        else:
                            st.caption("ℹ️ Usando referência fixa (ANP não carregada). "
                                       "Carregue a planilha ANP em Configurações para comparação por município.")

                        st.markdown("---")

                        if _n_alertas == 0:
                            st.success("✅ Nenhum posto GF com preço acima de 5% da média ANP. "
                                       "Todos os preços estão dentro do parâmetro de referência.")
                        else:
                            # ── Alertas por UF ────────────────────────────────────
                            _alerta_uf = (
                                _alert_df[_alert_df["_alerta"]]
                                .groupby("uf")
                                .agg(
                                    postos_alerta=(_price_col, "count"),
                                    preco_medio=(_price_col, "mean"),
                                    pior_desvio=("_diff_pct", "max"),
                                )
                                .reset_index()
                                .sort_values("postos_alerta", ascending=False)
                            )
                            _alerta_uf["uf_nome"] = _alerta_uf["uf"].map(_UF_NOME_DASH).fillna(_alerta_uf["uf"])

                            st.markdown("#### ⚠️ Alertas por Estado")
                            _col_g1, _col_g2 = st.columns([2, 1])

                            with _col_g1:
                                _fig_alerta = go.Figure()
                                _color_alert = [
                                    "#B71C1C" if v > 0.10 else
                                    "#E53935" if v > 0.07 else
                                    "#EF9A9A"
                                    for v in _alerta_uf["pior_desvio"]
                                ]
                                _fig_alerta.add_trace(go.Bar(
                                    y=_alerta_uf["uf_nome"],
                                    x=_alerta_uf["postos_alerta"],
                                    orientation="h",
                                    marker_color=_color_alert,
                                    text=_alerta_uf["postos_alerta"].astype(str),
                                    textposition="outside",
                                    hovertemplate=(
                                        "<b>%{y}</b><br>"
                                        "Postos em alerta: %{x}<br>"
                                        "<extra></extra>"
                                    ),
                                ))
                                _fig_alerta.update_layout(
                                    title="Postos em Alerta por Estado (preço > ANP + 5%)",
                                    xaxis_title="Quantidade de Postos",
                                    yaxis=dict(autorange="reversed"),
                                    height=max(300, len(_alerta_uf) * 24 + 80),
                                    margin=dict(l=10, r=60, t=45, b=30),
                                    plot_bgcolor="rgba(0,0,0,0)",
                                    paper_bgcolor="rgba(0,0,0,0)",
                                    font=dict(size=11),
                                )
                                _fig_alerta.update_xaxes(showgrid=True, gridcolor="#FFEBEE")
                                st.plotly_chart(_fig_alerta, use_container_width=True)

                            with _col_g2:
                                st.markdown("##### Resumo por Estado")
                                _alerta_uf_disp = _alerta_uf[["uf_nome", "postos_alerta", "pior_desvio"]].copy()
                                _alerta_uf_disp.columns = ["Estado", "Postos", "Pior Desvio"]
                                _alerta_uf_disp["Pior Desvio"] = _alerta_uf_disp["Pior Desvio"].apply(
                                    lambda v: f"+{v*100:.1f}%"
                                )
                                st.dataframe(_alerta_uf_disp, use_container_width=True, hide_index=True)

                            # ── Top piores postos ─────────────────────────────────
                            st.markdown("#### 🏆 Top 20 Postos com Maior Desvio")
                            _top_piores = (
                                _alert_df[_alert_df["_alerta"]]
                                .nlargest(20, "_diff_pct")
                                .copy()
                            )
                            _disp_cols_tp = []
                            for _c in ["razaoSocial", "municipio", "uf",
                                       "combustivel_label", _price_col,
                                       "_anp_ref", "_nivel_anp", "_diff_pct", "_diff_rs"]:
                                if _c in _top_piores.columns:
                                    _disp_cols_tp.append(_c)
                            _top_piores_disp = _top_piores[_disp_cols_tp].copy()

                            _rename_tp = {
                                "razaoSocial":       "Posto",
                                "municipio":         "Município",
                                "uf":                "UF",
                                "combustivel_label": "Combustível",
                                _price_col:          "Preço GF (R$/L)",
                                "_anp_ref":          "Ref. ANP (R$/L)",
                                "_nivel_anp":        "Base ANP",
                                "_diff_pct":         "Desvio %",
                                "_diff_rs":          "Desvio R$/L",
                            }
                            _top_piores_disp = _top_piores_disp.rename(columns=_rename_tp)

                            if "Desvio %" in _top_piores_disp.columns:
                                _top_piores_disp["Desvio %"] = _top_piores_disp["Desvio %"].apply(
                                    lambda v: f"+{v*100:.1f}%" if pd.notna(v) else "—"
                                )
                            if "Desvio R$/L" in _top_piores_disp.columns:
                                _top_piores_disp["Desvio R$/L"] = _top_piores_disp["Desvio R$/L"].apply(
                                    lambda v: f"+R$ {v:.3f}".replace(".", ",") if pd.notna(v) and v > 0 else "—"
                                )
                            for _fc in ["Preço GF (R$/L)", "Ref. ANP (R$/L)"]:
                                if _fc in _top_piores_disp.columns:
                                    _top_piores_disp[_fc] = _top_piores_disp[_fc].apply(
                                        lambda v: f"R$ {v:.3f}".replace(".", ",") if pd.notna(v) else "—"
                                    )

                            st.dataframe(
                                _top_piores_disp.reset_index(drop=True),
                                use_container_width=True,
                                hide_index=True,
                            )

                            # ── Exportar lista de alertas ─────────────────────────
                            st.markdown("---")
                            _exp_alert_cols = st.columns([3, 1])
                            with _exp_alert_cols[1]:
                                _all_alertas = _alert_df[_alert_df["_alerta"]].copy()
                                _all_alertas["Desvio_pct"] = (_all_alertas["_diff_pct"] * 100).round(2)
                                _all_alertas["Desvio_RS"]  = _all_alertas["_diff_rs"].round(3)
                                _all_alertas = _all_alertas.drop(
                                    columns=["_anp_ref","_diff_pct","_diff_rs","_alerta","_nivel_anp"],
                                    errors="ignore"
                                )
                                _csv_alertas = _all_alertas.to_csv(index=False).encode("utf-8-sig")
                                st.download_button(
                                    label="📥 Exportar Alertas (CSV)",
                                    data=_csv_alertas,
                                    file_name=f"alertas_preco_gf_{pd.Timestamp.now().strftime('%Y%m%d_%H%M')}.csv",
                                    mime="text/csv",
                                    use_container_width=True,
                                )
                            with _exp_alert_cols[0]:
                                st.caption(
                                    f"⚠️ **{_n_alertas} postos** com preço acima de 5% da referência ANP. "
                                    f"Exporte a lista completa para análise detalhada."
                                )


        # ══════════════════════════════════════════════════════════════════
        # TAB 6 — MODO COMPARATIVO
        # ══════════════════════════════════════════════════════════════════
        with _dt6:
            st.markdown(
                "<p style='color:#555;font-size:13px;margin:0 0 14px'>"
                "Compare dois estados ou duas regiões: cobertura, distribuidoras "
                "e preços médios lado a lado.</p>",
                unsafe_allow_html=True,
            )

            # ── Dicionários de suporte ────────────────────────────────────────
            _REGIOES_BR = {
                "Norte":        ["AM","PA","AC","RO","RR","AP","TO"],
                "Nordeste":     ["MA","PI","CE","RN","PB","PE","AL","SE","BA"],
                "Centro-Oeste": ["MT","MS","GO","DF"],
                "Sudeste":      ["SP","RJ","MG","ES"],
                "Sul":          ["PR","SC","RS"],
            }
            # Total de municípios por UF (IBGE 2023) — usado para cobertura %
            _TOTAL_MUNS_UF = {
                "AC":22,"AL":102,"AP":16,"AM":62,"BA":417,"CE":184,"DF":1,
                "ES":78,"GO":246,"MA":217,"MT":141,"MS":79,"MG":853,"PA":144,
                "PB":223,"PR":399,"PE":185,"PI":224,"RJ":92,"RN":167,"RS":497,
                "RO":52,"RR":15,"SC":295,"SP":645,"SE":75,"TO":139,
            }

            # ── Seletor de modo ───────────────────────────────────────────────
            _cmp_modo = st.radio(
                "Comparar por:",
                ["🗺️ Estados", "🌎 Regiões"],
                horizontal=True,
                key="dash_cmp_modo",
            )

            _ufs_disponiveis = sorted(_df_valid["uf"].unique().tolist())

            if _cmp_modo == "🗺️ Estados":
                _sc1, _sc2 = st.columns(2)
                with _sc1:
                    _uf_a = st.selectbox(
                        "Estado A", _ufs_disponiveis,
                        index=0,
                        format_func=lambda u: f"{u} — {_UF_NOME_DASH.get(u, u)}",
                        key="dash_cmp_uf_a",
                    )
                with _sc2:
                    _default_b_idx = 1 if len(_ufs_disponiveis) > 1 else 0
                    _uf_b = st.selectbox(
                        "Estado B", _ufs_disponiveis,
                        index=_default_b_idx,
                        format_func=lambda u: f"{u} — {_UF_NOME_DASH.get(u, u)}",
                        key="dash_cmp_uf_b",
                    )
                _label_a  = f"{_uf_a} — {_UF_NOME_DASH.get(_uf_a, _uf_a)}"
                _label_b  = f"{_uf_b} — {_UF_NOME_DASH.get(_uf_b, _uf_b)}"
                _ufs_a    = [_uf_a]
                _ufs_b    = [_uf_b]
            else:
                _regioes_disp = sorted(_REGIOES_BR.keys())
                _sc1, _sc2 = st.columns(2)
                with _sc1:
                    _reg_a = st.selectbox("Região A", _regioes_disp,
                                          index=0, key="dash_cmp_reg_a")
                with _sc2:
                    _reg_b = st.selectbox("Região B", _regioes_disp,
                                          index=min(1, len(_regioes_disp)-1),
                                          key="dash_cmp_reg_b")
                _label_a = f"🌎 {_reg_a}"
                _label_b = f"🌎 {_reg_b}"
                _ufs_a   = _REGIOES_BR[_reg_a]
                _ufs_b   = _REGIOES_BR[_reg_b]

            # ── Função: calcula métricas para um conjunto de UFs ──────────────
            def _cmp_metricas(ufs, df_v, pp):
                """Retorna dict com KPIs e dados detalhados para um grupo de UFs."""
                _sub = df_v[df_v["uf"].isin(ufs)].copy()
                _n_postos   = len(_sub)
                _n_muns     = int(_sub["municipio"].replace("", pd.NA).dropna().nunique())
                _n_distrib  = int(
                    _sub["distribuidora"].replace("", pd.NA).dropna().nunique()
                    if "distribuidora" in _sub.columns else 0
                )
                _n_coord    = int(_sub[pd.notna(_sub["_lat"]) & pd.notna(_sub["_lon"])].shape[0])
                _total_muns_reg = sum(_TOTAL_MUNS_UF.get(u, 0) for u in ufs)
                _cob_pct    = round(_n_muns / _total_muns_reg * 100, 1) if _total_muns_reg else 0
                _media_mun  = round(_n_postos / _n_muns, 1) if _n_muns else 0

                # Distribuidoras ranking
                _distrib_cnt = pd.Series(dtype=int)
                if "distribuidora" in _sub.columns:
                    _distrib_cnt = (
                        _sub["distribuidora"]
                        .replace("", pd.NA).dropna()
                        .value_counts()
                        .head(10)
                    )

                # Preços médios por combustível (se disponível)
                _precos = {}
                if pp is not None and not pp.empty and "cnpj" in _sub.columns:
                    _cnpjs = set(_sub["cnpj"].dropna().astype(str)
                                 .str.replace(r"\D", "", regex=True))
                    _pp_sub = pp[pp["cnpj_norm"].isin(_cnpjs)]
                    if not _pp_sub.empty:
                        _precos = (
                            _pp_sub.groupby("combustivel_label")["preco"]
                            .mean().round(3).to_dict()
                        )

                return {
                    "n_postos":    _n_postos,
                    "n_muns":      _n_muns,
                    "n_distrib":   _n_distrib,
                    "n_coord":     _n_coord,
                    "cob_pct":     _cob_pct,
                    "media_mun":   _media_mun,
                    "distrib_cnt": _distrib_cnt,
                    "precos":      _precos,
                    "df_sub":      _sub,
                }

            _ma = _cmp_metricas(_ufs_a, _df_valid, _pp_dash)
            _mb = _cmp_metricas(_ufs_b, _df_valid, _pp_dash)

            # ── Helper visual ─────────────────────────────────────────────────
            def _badge_cmp(txt, cor_bg, cor_txt="#fff"):
                return (
                    f"<span style='background:{cor_bg};color:{cor_txt};"
                    f"border-radius:5px;padding:1px 8px;font-size:11px;"
                    f"font-weight:700'>{txt}</span>"
                )

            def _winner_cmp(val_a, val_b, higher_is_better=True):
                """Retorna tuple (badge_a, badge_b) indicando quem é melhor."""
                if val_a == val_b or (val_a == 0 and val_b == 0):
                    return (_badge_cmp("=", "#607D8B"),
                            _badge_cmp("=", "#607D8B"))
                if higher_is_better:
                    _wa = "#2E7D32" if val_a > val_b else "#C62828"
                    _wb = "#2E7D32" if val_b > val_a else "#C62828"
                else:
                    _wa = "#2E7D32" if val_a < val_b else "#C62828"
                    _wb = "#2E7D32" if val_b < val_a else "#C62828"
                _sym_a = "▲" if _wa == "#2E7D32" else "▼"
                _sym_b = "▲" if _wb == "#2E7D32" else "▼"
                return _badge_cmp(_sym_a, _wa), _badge_cmp(_sym_b, _wb)

            st.markdown("---")

            # ── KPIs lado a lado ──────────────────────────────────────────────
            st.markdown("#### 📊 Comparativo de Cobertura")
            _kpi_rows = [
                ("⛽ Postos GF",      _ma["n_postos"],  _mb["n_postos"],  True),
                ("🏙️ Municípios GF", _ma["n_muns"],    _mb["n_muns"],    True),
                ("📈 Cobertura %",    _ma["cob_pct"],   _mb["cob_pct"],   True),
                ("🏢 Distribuidoras", _ma["n_distrib"], _mb["n_distrib"], True),
                ("📍 Com Coord.",     _ma["n_coord"],   _mb["n_coord"],   True),
                ("📊 Média GF/Mun.", _ma["media_mun"], _mb["media_mun"], True),
            ]

            # Cabeçalho da tabela comparativa
            _hc0, _hca, _hcw, _hcb = st.columns([2, 2, 1, 2])
            _hc0.markdown(
                "<div style='font-size:12px;color:#888;font-weight:600'>Indicador</div>",
                unsafe_allow_html=True)
            _hca.markdown(
                f"<div style='font-size:13px;font-weight:700;color:#0D47A1'>{_label_a}</div>",
                unsafe_allow_html=True)
            _hcw.markdown("")
            _hcb.markdown(
                f"<div style='font-size:13px;font-weight:700;color:#B71C1C'>{_label_b}</div>",
                unsafe_allow_html=True)
            st.markdown("<div style='height:4px'></div>", unsafe_allow_html=True)

            def _fmt_num(v):
                if isinstance(v, float):
                    return f"{v:,.1f}".replace(",", "X").replace(".", ",").replace("X", ".")
                return f"{v:,}".replace(",", ".")

            for _kpi_lbl, _va, _vb, _hib in _kpi_rows:
                _ba, _bb = _winner_cmp(_va, _vb, higher_is_better=_hib)
                _c0, _ca, _cw, _cb = st.columns([2, 2, 1, 2])
                _c0.markdown(
                    f"<div style='font-size:12px;color:#555;padding:4px 0'>{_kpi_lbl}</div>",
                    unsafe_allow_html=True)
                _ca.markdown(
                    f"<div style='font-size:14px;font-weight:700;color:#0D47A1;padding:2px 0'>"
                    f"{_fmt_num(_va)}&nbsp;{_ba}</div>",
                    unsafe_allow_html=True)
                _cw.markdown(
                    "<div style='text-align:center;font-size:12px;"
                    "color:#aaa;padding:4px 0'>vs</div>",
                    unsafe_allow_html=True)
                _cb.markdown(
                    f"<div style='font-size:14px;font-weight:700;color:#B71C1C;padding:2px 0'>"
                    f"{_fmt_num(_vb)}&nbsp;{_bb}</div>",
                    unsafe_allow_html=True)

            # ── Distribuidoras ────────────────────────────────────────────────
            st.markdown("---")
            st.markdown("#### 🏢 Distribuidoras — Top 10")
            _gc_a, _gc_b = st.columns(2)

            def _chart_distrib_cmp(title, distrib_cnt, cor_base):
                if distrib_cnt.empty:
                    return None
                _nc  = len(distrib_cnt)
                _r   = int(cor_base[1:3], 16)
                _g   = int(cor_base[3:5], 16)
                _b_v = int(cor_base[5:7], 16)
                _cors = [
                    cor_base if i == 0 else
                    f"rgba({_r},{_g},{_b_v},{max(0.3, 1 - i * 0.08):.2f})"
                    for i in range(_nc)
                ]
                _fig = go.Figure()
                _fig.add_trace(go.Bar(
                    y=distrib_cnt.index.tolist(),
                    x=distrib_cnt.values.tolist(),
                    orientation="h",
                    marker_color=_cors,
                    text=distrib_cnt.values.tolist(),
                    textposition="outside",
                    hovertemplate="<b>%{y}</b><br>Postos: %{x}<extra></extra>",
                ))
                _fig.update_layout(
                    title=title,
                    xaxis_title="Postos GF",
                    yaxis=dict(autorange="reversed"),
                    height=max(260, _nc * 30 + 80),
                    margin=dict(l=10, r=50, t=45, b=20),
                    plot_bgcolor="rgba(0,0,0,0)",
                    paper_bgcolor="rgba(0,0,0,0)",
                    font=dict(size=10),
                )
                _fig.update_xaxes(showgrid=True, gridcolor="#ECEFF1")
                return _fig

            with _gc_a:
                _fig_da = _chart_distrib_cmp(_label_a, _ma["distrib_cnt"], "#0D47A1")
                if _fig_da:
                    st.plotly_chart(_fig_da, use_container_width=True)
                else:
                    st.info("Sem dados de distribuidora para este grupo.")

            with _gc_b:
                _fig_db = _chart_distrib_cmp(_label_b, _mb["distrib_cnt"], "#B71C1C")
                if _fig_db:
                    st.plotly_chart(_fig_db, use_container_width=True)
                else:
                    st.info("Sem dados de distribuidora para este grupo.")

            # ── Preços médios por combustível ─────────────────────────────────
            _precos_a = _ma["precos"]
            _precos_b = _mb["precos"]

            if _precos_a or _precos_b:
                st.markdown("---")
                st.markdown("#### 💲 Preços Médios por Combustível (R$/L)")
                _all_combs = sorted(set(list(_precos_a.keys()) + list(_precos_b.keys())))

                if _all_combs:
                    _fig_preco = go.Figure()
                    _fig_preco.add_trace(go.Bar(
                        name=_label_a,
                        x=_all_combs,
                        y=[_precos_a.get(c) for c in _all_combs],
                        marker_color="#1565C0",
                        text=[f"R$ {_precos_a[c]:.3f}".replace(".", ",")
                              if c in _precos_a else "" for c in _all_combs],
                        textposition="outside",
                        hovertemplate="<b>%{x}</b><br>%{fullData.name}: R$ %{y:.3f}<extra></extra>",
                    ))
                    _fig_preco.add_trace(go.Bar(
                        name=_label_b,
                        x=_all_combs,
                        y=[_precos_b.get(c) for c in _all_combs],
                        marker_color="#C62828",
                        text=[f"R$ {_precos_b[c]:.3f}".replace(".", ",")
                              if c in _precos_b else "" for c in _all_combs],
                        textposition="outside",
                        hovertemplate="<b>%{x}</b><br>%{fullData.name}: R$ %{y:.3f}<extra></extra>",
                    ))
                    _fig_preco.update_layout(
                        barmode="group",
                        yaxis_title="R$/L",
                        height=380,
                        margin=dict(l=10, r=10, t=30, b=60),
                        plot_bgcolor="rgba(0,0,0,0)",
                        paper_bgcolor="rgba(0,0,0,0)",
                        legend=dict(orientation="h", y=-0.20, x=0.5, xanchor="center"),
                        font=dict(size=11),
                    )
                    _fig_preco.update_yaxes(showgrid=True, gridcolor="#ECEFF1")
                    st.plotly_chart(_fig_preco, use_container_width=True)

                    # Tabela delta
                    st.markdown("##### Δ Diferença de Preços (A − B)")
                    _delta_rows = []
                    for _c in _all_combs:
                        _pa = _precos_a.get(_c)
                        _pb = _precos_b.get(_c)
                        if _pa is not None and _pb is not None:
                            _diff     = _pa - _pb
                            _diff_pct = (_diff / _pb * 100) if _pb else 0
                            _delta_rows.append({
                                "Combustível": _c,
                                f"Preço A":    f"R$ {_pa:.3f}".replace(".", ","),
                                f"Preço B":    f"R$ {_pb:.3f}".replace(".", ","),
                                "Δ R$/L":      f"{'+' if _diff >= 0 else ''}{_diff:.3f}".replace(".", ","),
                                "Δ %":         f"{'+' if _diff_pct >= 0 else ''}{_diff_pct:.1f}%",
                                "Mais barato": (
                                    _label_a[:20] if _pa < _pb else
                                    (_label_b[:20] if _pb < _pa else "Igual")
                                ),
                            })
                    if _delta_rows:
                        st.dataframe(pd.DataFrame(_delta_rows),
                                     use_container_width=True, hide_index=True)
            else:
                st.markdown("---")
                st.info(
                    "ℹ️ Para comparar preços médios, carregue a planilha de Preços "
                    "em **Configurações → Preços dos Postos GF**."
                )

            # ── Mini mapas geográficos ─────────────────────────────────────────
            st.markdown("---")
            st.markdown("#### 🗺️ Distribuição Geográfica")
            _gm_a, _gm_b = st.columns(2)

            def _mini_mapa_cmp(df_sub, titulo, cor_marker):
                _sc = df_sub[pd.notna(df_sub["_lat"]) & pd.notna(df_sub["_lon"])].copy()
                if _sc.empty:
                    return None
                _sc = _sc.sample(min(len(_sc), 800), random_state=42)
                _fig_m = go.Figure()
                _cd = (
                    _sc[["razaoSocial", "municipio", "uf"]].values
                    if all(c in _sc.columns for c in ["razaoSocial", "municipio", "uf"])
                    else None
                )
                _fig_m.add_trace(go.Scattergeo(
                    lat=_sc["_lat"], lon=_sc["_lon"],
                    mode="markers",
                    marker=dict(size=5, color=cor_marker, opacity=0.75,
                                line=dict(color="white", width=0.4)),
                    hovertemplate=(
                        "<b>%{customdata[0]}</b><br>"
                        "%{customdata[1]} — %{customdata[2]}<extra></extra>"
                    ),
                    customdata=_cd,
                ))
                _lat_c = (_sc["_lat"].max() + _sc["_lat"].min()) / 2
                _lon_c = (_sc["_lon"].max() + _sc["_lon"].min()) / 2
                _fig_m.update_layout(
                    title=dict(text=titulo, font=dict(size=12)),
                    geo=dict(
                        scope="south america",
                        center=dict(lat=_lat_c, lon=_lon_c),
                        projection_type="mercator",
                        showland=True, landcolor="#F5F5F5",
                        showcoastlines=True, coastlinecolor="#BDBDBD",
                        showframe=False,
                        lataxis=dict(range=[_sc["_lat"].min() - 2, _sc["_lat"].max() + 2]),
                        lonaxis=dict(range=[_sc["_lon"].min() - 2, _sc["_lon"].max() + 2]),
                    ),
                    height=380,
                    margin=dict(l=0, r=0, t=40, b=0),
                    paper_bgcolor="rgba(0,0,0,0)",
                )
                return _fig_m

            with _gm_a:
                _fm_a = _mini_mapa_cmp(_ma["df_sub"], _label_a, "#1565C0")
                if _fm_a:
                    st.plotly_chart(_fm_a, use_container_width=True)
                else:
                    st.info(f"Sem coordenadas para {_label_a}.")

            with _gm_b:
                _fm_b = _mini_mapa_cmp(_mb["df_sub"], _label_b, "#C62828")
                if _fm_b:
                    st.plotly_chart(_fm_b, use_container_width=True)
                else:
                    st.info(f"Sem coordenadas para {_label_b}.")

            # ── Resumo executivo ──────────────────────────────────────────────
            st.markdown("---")
            st.markdown("#### 📋 Resumo Executivo")
            _re_a, _re_b = st.columns(2)
            for _re_col, _lbl, _m, _cor in [
                (_re_a, _label_a, _ma, "#0D47A1"),
                (_re_b, _label_b, _mb, "#B71C1C"),
            ]:
                with _re_col:
                    _preco_item = ""
                    if _m["precos"]:
                        _best_c = min(_m["precos"], key=_m["precos"].get)
                        _best_v = _m["precos"][_best_c]
                        _preco_item = (
                            f"<li>Combustível mais barato: <b>{_best_c}</b> "
                            f"@ R$ {_best_v:.3f}".replace(".", ",") + "/L</li>"
                        )
                    st.markdown(
                        f"<div style='border:2px solid {_cor};"
                        f"border-radius:10px;padding:14px 16px;background:#fafafa'>"
                        f"<div style='font-size:14px;font-weight:700;color:{_cor};"
                        f"margin-bottom:8px'>{_lbl}</div>"
                        f"<ul style='font-size:12px;color:#333;margin:0;padding-left:16px'>"
                        f"<li><b>{_fmt_int(_m['n_postos'])}</b> postos GF credenciados</li>"
                        f"<li><b>{_fmt_int(_m['n_muns'])}</b> municípios atendidos "
                        f"(<b>{_m['cob_pct']:.1f}%</b> de cobertura)</li>"
                        f"<li><b>{_m['n_distrib']}</b> distribuidoras presentes</li>"
                        f"<li>Média de <b>{_m['media_mun']:.1f}</b> posto(s)/município</li>"
                        + _preco_item +
                        f"</ul></div>",
                        unsafe_allow_html=True,
                    )


# ── Restauração pós-rerun: recalcula rota do Modo 1 se solicitado ──────────
# Este bloco roda APÓS o rerun causado pelo botão Restaurar (Modo 1 com rota).
# Nesse momento o modo já está em "📍 Por UF/Município" e os flags existem.
if (
    st.session_state.get("modo_selecionado") == "📍 Por UF/Município"
    and st.session_state.get("_restore_recalc_rota_m1")
):
    _o_rest = st.session_state.get("_map_orig")
    _d_rest = st.session_state.get("_map_dest")
    if _o_rest and _d_rest:
        with st.spinner("🗺️ Recalculando rota restaurada…"):
            _cr_rest, _dk_rest, _dm_rest, _lr_rest = calcular_rota(
                _o_rest["lat"], _o_rest["lon"],
                _d_rest["lat"], _d_rest["lon"]
            )
        st.session_state["_map_rota_result"] = {
            "coords":   _cr_rest,
            "dist_km":  _dk_rest,
            "dur_min":  _dm_rest,
            "linha_reta": _lr_rest,
            "orig":     _o_rest,
            "dest":     _d_rest,
        }
    st.session_state.pop("_restore_recalc_rota_m1", None)
    st.rerun()


# ═══════════════════════════════════════════════════════════════════
#  MODO — Inteligência de Dados
# ═══════════════════════════════════════════════════════════════════

elif modo == "🧠 Inteligência":

    st.markdown("""
    <style>
    .intel-hero {
        background: linear-gradient(135deg, #1a0533 0%, #0d1b4b 60%, #061840 100%);
        border-radius: 18px;
        padding: 2rem 2.2rem 1.6rem;
        margin-bottom: 1.5rem;
        position: relative;
        overflow: hidden;
    }
    .intel-hero::before {
        content: "";
        position: absolute;
        top: -40px; right: -40px;
        width: 200px; height: 200px;
        background: radial-gradient(circle, rgba(106,27,154,0.3) 0%, transparent 70%);
    }
    .intel-hero-title {
        font-size: 1.6rem; font-weight: 800;
        background: linear-gradient(135deg, #ffffff 0%, #ce93d8 100%);
        -webkit-background-clip: text; -webkit-text-fill-color: transparent;
        background-clip: text; margin: 0 0 0.3rem;
    }
    .intel-hero-sub { font-size: 0.9rem; color: rgba(255,255,255,0.55); margin: 0; }
    .intel-kpi-card {
        background: linear-gradient(135deg, #f3e5f5, #fce4ec);
        border: 1px solid #ce93d8;
        border-radius: 14px;
        padding: 1rem 1.2rem;
        text-align: center;
    }
    .intel-kpi-num { font-size: 1.8rem; font-weight: 800; color: #4a148c; line-height: 1; }
    .intel-kpi-lbl { font-size: 0.75rem; color: #888; margin-top: 4px; }
    </style>
    """, unsafe_allow_html=True)

    # ── Hero banner ────────────────────────────────────────────────
    st.markdown("""
    <div class='intel-hero'>
      <div class='intel-hero-title'>🧠 Inteligência de Dados</div>
      <div class='intel-hero-sub'>
        Histórico de preços por posto · Score composto · Relatório de alertas
      </div>
    </div>
    """, unsafe_allow_html=True)

    _intel_d_pg = _intel_load()
    _hist_all_pg = _intel_d_pg.get("historico", {})
    _n_cnpjs_pg  = len(_hist_all_pg)
    _n_obs_pg    = sum(len(v) for v in _hist_all_pg.values())
    _datas_pg    = sorted({e["data"] for v in _hist_all_pg.values() for e in v})
    _semanas_pg  = len(_datas_pg)

    # ── KPIs ──────────────────────────────────────────────────────
    _ki1, _ki2, _ki3, _ki4 = st.columns(4)
    with _ki1:
        st.markdown(f"<div class='intel-kpi-card'>"
                    f"<div class='intel-kpi-num'>{_fmt_int(_n_cnpjs_pg)}</div>"
                    f"<div class='intel-kpi-lbl'>📍 Postos rastreados</div></div>",
                    unsafe_allow_html=True)
    with _ki2:
        st.markdown(f"<div class='intel-kpi-card'>"
                    f"<div class='intel-kpi-num'>{_fmt_int(_n_obs_pg)}</div>"
                    f"<div class='intel-kpi-lbl'>📊 Observações</div></div>",
                    unsafe_allow_html=True)
    with _ki3:
        st.markdown(f"<div class='intel-kpi-card'>"
                    f"<div class='intel-kpi-num'>{_semanas_pg}</div>"
                    f"<div class='intel-kpi-lbl'>📅 Semanas de histórico</div></div>",
                    unsafe_allow_html=True)
    with _ki4:
        _last_rpt = _intel_d_pg.get("last_report","")
        _last_rpt_fmt = _last_rpt[:10] if _last_rpt else "Nunca"
        st.markdown(f"<div class='intel-kpi-card'>"
                    f"<div class='intel-kpi-num' style='font-size:1.1rem'>{_last_rpt_fmt}</div>"
                    f"<div class='intel-kpi-lbl'>📋 Último relatório</div></div>",
                    unsafe_allow_html=True)

    st.markdown("<div style='height:1.2rem'></div>", unsafe_allow_html=True)

    # ── Abas principais ───────────────────────────────────────────
    _tab_hist, _tab_score, _tab_alertas = st.tabs([
        "📈 Histórico de Preços",
        "⭐ Score de Postos",
        "⚠️ Relatório de Alertas",
    ])

    # ══ ABA 1: Histórico ══════════════════════════════════════════
    with _tab_hist:
        if _n_obs_pg == 0:
            st.info(
                "**Nenhum histórico registrado ainda.**\n\n"
                "Os preços são acumulados automaticamente toda vez que você carrega "
                "a planilha de **Preços PP** em Configurações → 💲 Preços PP.\n\n"
                "Após algumas semanas de uso, o gráfico de evolução de preços "
                "estará disponível aqui."
            )
        else:
            st.markdown("#### Visualizar histórico de um posto")
            _col_h1, _col_h2 = st.columns([2, 1])
            with _col_h1:
                _cnpj_hist_pg = st.text_input(
                    "CNPJ do posto (somente números)",
                    key="intel_cnpj_pg",
                    placeholder="Ex: 12345678000199",
                    max_chars=18,
                )
            with _col_h2:
                _comb_hist_pg = st.selectbox(
                    "Filtrar por combustível",
                    options=["Todos"] + list({
                        e["combustivel"]
                        for v in _hist_all_pg.values()
                        for e in v
                        if e.get("combustivel")
                    }),
                    key="intel_comb_hist_pg",
                )

            if _cnpj_hist_pg:
                _cnpj_h_n2 = re.sub(r"\D", "", _cnpj_hist_pg)
                _hist_posto2 = _hist_all_pg.get(_cnpj_h_n2, [])
                if not _hist_posto2:
                    st.warning("Nenhum histórico encontrado para este CNPJ.")
                else:
                    _nome_h2 = _hist_posto2[0].get("nome", f"Posto {_cnpj_h_n2}")
                    _comb_f2 = None if _comb_hist_pg == "Todos" else _comb_hist_pg
                    _fig_h2  = _hist_chart_posto(_cnpj_h_n2, _nome_h2, combustivel=_comb_f2)
                    if _fig_h2:
                        _titulo_h2 = f"📈 Evolução de preços — {_nome_h2}"
                        if _comb_f2:
                            _titulo_h2 += f" ({_comb_f2.title()})"
                        st.markdown(
                            f"<p style='font-weight:600;font-size:0.95rem;"
                            f"margin:0 0 4px 0;color:var(--text-color,#1a1a2e)'>"
                            f"{_titulo_h2}</p>",
                            unsafe_allow_html=True,
                        )
                        st.plotly_chart(_fig_h2, use_container_width=True)

                    _df_h2 = pd.DataFrame(_hist_posto2)
                    if _comb_f2:
                        _df_h2 = _df_h2[_df_h2["combustivel"] == _comb_f2]
                    _df_h2 = _df_h2.sort_values("data", ascending=False)
                    _df_h2 = _df_h2.rename(columns={
                        "data":"Data","preco":"Preço R$/L",
                        "combustivel":"Combustível","municipio":"Município","uf":"UF"})
                    st.dataframe(_df_h2.head(52), use_container_width=True, height=250)

            # Registrar preços da planilha PP
            st.markdown("---")
            st.markdown("##### Registrar preços no histórico")
            _pp_df_pg = st.session_state.get("_pp_df")
            if _pp_df_pg is not None and not _pp_df_pg.empty:
                st.caption(f"Planilha PP carregada: {_fmt_int(len(_pp_df_pg))} registros")
                if st.button("🔄 Registrar preços atuais da planilha PP",
                             key="btn_hist_reg_pg", use_container_width=True):
                    _n_reg_pg = _hist_record_pp_df(_pp_df_pg)
                    st.success(f"✅ {_n_reg_pg} novas observações registradas.")
                    st.rerun()
            else:
                st.info("Carregue a planilha de Preços PP em **Configurações → 💲 Preços PP** para ativar o registro automático de histórico.")

    # ══ ABA 2: Score ══════════════════════════════════════════════
    with _tab_score:
        st.markdown(
            "O **Score** é calculado automaticamente na tabela de dados de cada modo.\n\n"
            "Ele aparece como `A 82`, `B 61`, `C 43` ou `D 28` — combinando três fatores:"
        )
        _sc_c1, _sc_c2, _sc_c3 = st.columns(3)
        with _sc_c1:
            st.markdown(
                "<div style='background:linear-gradient(135deg,#e3f2fd,#bbdefb);"
                "border-radius:14px;padding:1.2rem;text-align:center'>"
                "<div style='font-size:2rem'>💰</div>"
                "<div style='font-weight:800;color:#0d47a1;font-size:1rem'>Preço vs ANP</div>"
                "<div style='font-size:0.75rem;color:#555;margin-top:6px'><b>50%</b> do score</div>"
                "<div style='font-size:0.78rem;color:#666;margin-top:8px'>"
                "Quanto mais barato que a média ANP regional, maior a pontuação.</div>"
                "</div>", unsafe_allow_html=True)
        with _sc_c2:
            st.markdown(
                "<div style='background:linear-gradient(135deg,#e8f5e9,#c8e6c9);"
                "border-radius:14px;padding:1.2rem;text-align:center'>"
                "<div style='font-size:2rem'>🛒</div>"
                "<div style='font-weight:800;color:#2e7d32;font-size:1rem'>Serviços</div>"
                "<div style='font-size:0.75rem;color:#555;margin-top:6px'><b>30%</b> do score</div>"
                "<div style='font-size:0.78rem;color:#666;margin-top:8px'>"
                "Conveniência, ARLA 32, restaurante, banheiro e outros serviços disponíveis.</div>"
                "</div>", unsafe_allow_html=True)
        with _sc_c3:
            st.markdown(
                "<div style='background:linear-gradient(135deg,#fff8e1,#ffe082);"
                "border-radius:14px;padding:1.2rem;text-align:center'>"
                "<div style='font-size:2rem'>📍</div>"
                "<div style='font-weight:800;color:#f57f17;font-size:1rem'>Distância</div>"
                "<div style='font-size:0.75rem;color:#555;margin-top:6px'><b>20%</b> do score</div>"
                "<div style='font-size:0.78rem;color:#666;margin-top:8px'>"
                "Proximidade ao ponto de busca ou à rota selecionada.</div>"
                "</div>", unsafe_allow_html=True)

        st.markdown("<div style='height:1rem'></div>", unsafe_allow_html=True)
        st.markdown("""
| Grau | Pontuação | Significado |
|------|-----------|-------------|
| 🟢 **A** | ≥ 75 pts | Excelente — preço competitivo, muitos serviços, bem localizado |
| 🔵 **B** | 55–74 pts | Bom — acima da média em pelo menos dois fatores |
| 🟡 **C** | 35–54 pts | Regular — desempenho mediano |
| 🔴 **D** | < 35 pts | Abaixo da média — preço elevado ou sem serviços |
        """)
        st.info("💡 O Score aparece como primeira coluna na tabela de dados do Modo **📍 Por UF/Município**. "
                "Quanto mais dados da planilha Pró-Frotas e ANP estiverem carregados, mais preciso ele fica.")

    # ══ ABA 3: Alertas ════════════════════════════════════════════
    with _tab_alertas:
        st.markdown("#### ⚠️ Configurar Limiares e Gerar Relatório")
        st.markdown(
            "Defina o **preço máximo aceitável** para cada combustível. "
            "Ao gerar o relatório, todos os postos da planilha GF que estiverem "
            "**acima do limiar** serão listados no arquivo Excel."
        )

        _intel_d3    = _intel_load()
        _limiar_cfg3 = _intel_d3.get("limiar", {})

        _COMBS_LIM3 = [
            ("GASOLINA COMUM",    "⛽ Gasolina Comum",     5.80),
            ("GASOLINA ADITIVADA","⛽ Gasolina Aditivada",  6.20),
            ("ETANOL HIDRATADO",  "🌿 Etanol Hidratado",    4.00),
            ("DIESEL S10",        "🚛 Diesel S10",          6.00),
            ("DIESEL S500",       "🚛 Diesel S500",         5.90),
        ]
        _lim_novo3 = {}
        _lca, _lcb = st.columns(2)
        for _ci3, (_ck3, _clbl3, _cdef3) in enumerate(_COMBS_LIM3):
            with (_lca if _ci3 % 2 == 0 else _lcb):
                _lim_novo3[_ck3] = st.number_input(
                    _clbl3,
                    min_value=0.0, max_value=20.0,
                    value=float(_limiar_cfg3.get(_ck3, _cdef3)),
                    step=0.01, format="%.3f",
                    key=f"intel_lim3_{_ck3}",
                )

        _col_btn_a, _col_btn_b = st.columns([1, 1])
        with _col_btn_a:
            if st.button("💾 Salvar limiares", key="btn_intel3_salvar",
                         use_container_width=True):
                _intel_d3["limiar"] = _lim_novo3
                _intel_save(_intel_d3)
                st.session_state.pop("_intel_loaded", None)
                st.success("✅ Limiares salvos.")

        with _col_btn_b:
            _pp_df3 = st.session_state.get("_pp_df")
            if st.button("📄 Gerar Relatório Excel",
                         key="btn_intel3_gerar",
                         use_container_width=True,
                         type="primary"):
                if _pp_df3 is None or _pp_df3.empty:
                    st.warning("⚠️ Carregue a planilha de Preços PP em Configurações antes de gerar o relatório.")
                else:
                    with st.spinner("Gerando relatório…"):
                        _bytes3, _fname3, _err3 = _gerar_relatorio_alertas_xlsx(
                            _pp_df3, _lim_novo3)
                    if _err3:
                        st.error(f"❌ Erro: {_err3}")
                    else:
                        st.session_state["_intel_rel_bytes"] = _bytes3
                        st.session_state["_intel_rel_fname"] = _fname3
                        st.session_state.pop("_intel_loaded", None)
                        st.rerun()

        _bytes_dl3 = st.session_state.get("_intel_rel_bytes")
        _fname_dl3 = st.session_state.get("_intel_rel_fname", "alertas.xlsx")
        if _bytes_dl3:
            st.markdown("---")
            st.download_button(
                "⬇️ Baixar relatório gerado",
                data=_bytes_dl3,
                file_name=_fname_dl3,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
                key="btn_intel3_download",
            )

        if not (st.session_state.get("_pp_df") is not None):
            st.info("ℹ️ Para gerar o relatório, carregue a planilha de **Preços PP** em "
                    "**Configurações → 💲 Preços PP**.")


# ═══════════════════════════════════════════════════════════════════
#  MODO ADMIN — Controle de Acesso de Usuários
# ═══════════════════════════════════════════════════════════════════

elif modo == "🔐 Admin":
    _email_admin_check = (st.session_state.get("_auth_user") or {}).get("email", "")
    if _email_admin_check.lower() != _ADMIN_EMAIL.lower():
        st.error("🚫 Acesso restrito ao administrador.")
        st.stop()

    st.markdown("## 🔐 Painel de Administração — Controle de Acesso")

    # ── Modo de acesso global ──────────────────────────────────────
    _modo_atual_db = _db_modo_acesso()
    st.markdown("### ⚙️ Modo de Acesso Global")
    _col_m1, _col_m2 = st.columns(2)
    with _col_m1:
        st.info(f"**Modo atual:** {'🔓 Blacklist (aberto)' if _modo_atual_db == 'blacklist' else '🔒 Allowlist (restrito)'}")
    with _col_m2:
        if _modo_atual_db == "blacklist":
            if st.button("🔒 Mudar para Allowlist (restrito)", use_container_width=True):
                _db_set_modo_acesso("allowlist")
                st.toast("✅ Modo alterado para Allowlist.", icon="🔒")
                st.rerun()
        else:
            if st.button("🔓 Mudar para Blacklist (aberto)", use_container_width=True):
                _db_set_modo_acesso("blacklist")
                st.toast("✅ Modo alterado para Blacklist.", icon="🔓")
                st.rerun()

    st.caption("**Blacklist:** todos entram, exceto os bloqueados. **Allowlist:** só entram os aprovados.")
    st.divider()

    # ── Adicionar e-mail manualmente ──────────────────────────────
    st.markdown("### ➕ Adicionar Usuário")
    _ca1, _ca2, _ca3, _ca4 = st.columns([3, 2, 2, 1])
    with _ca1:
        _novo_email = st.text_input("E-mail", placeholder="usuario@empresa.com", key="admin_novo_email")
    with _ca2:
        _novo_status = st.selectbox("Status", ["permitido", "bloqueado"], key="admin_novo_status")
    with _ca3:
        _novo_motivo = st.text_input("Motivo (opcional)", key="admin_novo_motivo")
    with _ca4:
        st.markdown("<div style='height:28px'></div>", unsafe_allow_html=True)
        if st.button("Adicionar", use_container_width=True, key="admin_add_btn", type="primary"):
            if _novo_email and "@" in _novo_email:
                if _db_atualizar_status_acesso(_novo_email.strip(), _novo_status,
                                               _novo_motivo, _ADMIN_EMAIL):
                    st.toast(f"✅ {_novo_email} → {_novo_status}", icon="✅")
                    st.rerun()
            else:
                st.warning("Digite um e-mail válido.")

    st.divider()

    # ── Lista de usuários controlados ─────────────────────────────
    st.markdown("### 👥 Usuários Gerenciados")

    _registros = _db_listar_controle_acesso()

    # Também mostra usuários dos logs que ainda não estão na lista
    _logs_emails = set()
    try:
        _logs_rec = _db_ler_logs(limite=5000)
        _logs_emails = {r.get("user_email","") for r in _logs_rec
                        if r.get("user_email") and r.get("user_email") != "—"}
    except Exception:
        pass
    _emails_gerenc = {r["email"] for r in _registros}
    _emails_novos  = _logs_emails - _emails_gerenc - {_ADMIN_EMAIL}

    if _emails_novos:
        st.info(f"💡 **{len(_emails_novos)} e-mail(s)** nos logs ainda não gerenciados: "
                f"{', '.join(sorted(_emails_novos)[:5])}{'…' if len(_emails_novos)>5 else ''}")

    if not _registros:
        st.caption("Nenhum usuário gerenciado ainda. Use o formulário acima para adicionar.")
    else:
        # Filtro de status
        _filtro_status = st.pills("Filtrar por status",
                                   ["Todos", "✅ Permitido", "🚫 Bloqueado", "⏳ Pendente"],
                                   default="Todos", key="admin_filtro_status")

        _mapa_filtro = {"Todos": None, "✅ Permitido": "permitido",
                        "🚫 Bloqueado": "bloqueado", "⏳ Pendente": "pendente"}
        _filtro_val = _mapa_filtro.get(_filtro_status)
        _reg_filtrados = [r for r in _registros
                          if _filtro_val is None or r.get("status") == _filtro_val]

        for _reg in _reg_filtrados:
            _em   = _reg.get("email", "")
            _st   = _reg.get("status", "")
            _mot  = _reg.get("motivo", "") or ""
            _nom  = _reg.get("nome", "") or ""
            _ult  = _reg.get("ultimo_acesso", "") or ""
            _icone = {"permitido": "✅", "bloqueado": "🚫", "pendente": "⏳"}.get(_st, "❔")

            with st.container(border=True):
                _rc1, _rc2, _rc3, _rc4 = st.columns([4, 2, 2, 2])
                with _rc1:
                    st.markdown(f"**{_em}**")
                    if _nom:
                        st.caption(f"👤 {_nom}")
                    if _ult:
                        st.caption(f"🕐 Último acesso: {_ult[:16]}")
                with _rc2:
                    st.markdown(f"{_icone} **{_st.title()}**")
                    if _mot:
                        st.caption(f"_{_mot}_")
                with _rc3:
                    _nova_acao = "bloqueado" if _st != "bloqueado" else "permitido"
                    _btn_label = "🚫 Bloquear" if _nova_acao == "bloqueado" else "✅ Permitir"
                    if st.button(_btn_label, key=f"admin_toggle_{_em}",
                                 use_container_width=True):
                        _db_atualizar_status_acesso(_em, _nova_acao, "", _ADMIN_EMAIL)
                        st.toast(f"✅ {_em} → {_nova_acao}", icon="🔄")
                        st.rerun()
                with _rc4:
                    _motivo_blq = st.text_input("Motivo", value=_mot,
                                                key=f"admin_mot_{_em}",
                                                placeholder="opcional",
                                                label_visibility="collapsed")
                    if st.button("💾", key=f"admin_save_mot_{_em}",
                                 help="Salvar motivo"):
                        _db_atualizar_status_acesso(_em, _st, _motivo_blq, _ADMIN_EMAIL)
                        st.toast("Motivo salvo!", icon="💾")
                        st.rerun()

    st.divider()

    # ── Resumo de acessos recentes (dos logs) ─────────────────────
    st.markdown("### 📊 Acessos Recentes")
    if _logs_rec:
        _login_logs = [r for r in _logs_rec if r.get("acao") == "LOGIN"][:20]
        if _login_logs:
            _df_log = pd.DataFrame(_login_logs)[
                ["timestamp","user_email","user_name","ip","auth_provider"]
            ].rename(columns={
                "timestamp": "Data/Hora", "user_email": "E-mail",
                "user_name": "Nome", "ip": "IP", "auth_provider": "Provider"
            })
            st.dataframe(_df_log, use_container_width=True, hide_index=True)
        else:
            st.caption("Nenhum login registrado ainda.")


# ═══════════════════════════════════════════════════════════════════
#  MODO 5 — Roteirização
# ═══════════════════════════════════════════════════════════════════

elif modo == "🛣️ Roteirização":

    # ── Dados do veículo (definidos no sidebar) ─────────────────────
    _rot_placa = str(st.session_state.get("rot_placa") or "")
    _rot_comb  = str(st.session_state.get("rot_combustivel") or "")
    _rot_cap   = float(st.session_state.get("rot_capacidade") or 80.0)
    _rot_aut   = float(st.session_state.get("rot_autonomia")  or 10.0)
    _rot_min   = _rot_cap * 0.25
    _rot_fk    = st.session_state.get("_rot_fk", 0)
    _rot_np    = int(st.session_state.get("_rot_np", 0))

    # ── Cabeçalho ────────────────────────────────────────────────────
    st.markdown(
        "<h2 style='margin:0 0 4px;font-size:1.35rem;"
        "background:linear-gradient(135deg,#004D40,#00796B);"
        "-webkit-background-clip:text;-webkit-text-fill-color:transparent'>"
        "🛣️ Roteirização</h2>"
        "<p style='color:#555;font-size:13px;margin:0 0 12px'>"
        "Informe os pontos da rota — a aplicação sugere os melhores postos GF.</p>",
        unsafe_allow_html=True,
    )

    # ── Badge do veículo ─────────────────────────────────────────────
    if _rot_placa or _rot_comb:
        _v_parts = []
        if _rot_placa: _v_parts.append(f"🚛 <b>{_rot_placa.upper()}</b>")
        if _rot_comb:  _v_parts.append(f"⛽ {_rot_comb}")
        if _rot_cap:   _v_parts.append(f"🛢 {_rot_cap:.0f} L")
        if _rot_aut:   _v_parts.append(f"📏 {_rot_aut:.1f} km/L")
        _v_parts.append(f"🔋 alcance ~{(_rot_cap - _rot_min) * _rot_aut:.0f} km")
        st.markdown(
            "<div style='background:linear-gradient(90deg,#e0f7fa,#f1f8e9);"
            "border:1px solid #80deea;border-radius:8px;padding:7px 14px;"
            "font-size:12px;color:#004D40;margin-bottom:14px'>"
            + " · ".join(_v_parts) + "</div>",
            unsafe_allow_html=True,
        )
    else:
        st.info("💡 Configure os dados do veículo no **menu lateral** antes de calcular.")

    # ════════════════════════════════════════════════════════════════
    #  Helper — autocomplete de ponto (inline, sem st.rerun em loop)
    # ════════════════════════════════════════════════════════════════
    def _ponto_rot(label, key_sel, key_txt, icon="📍", cor="#1565C0",
                   deletavel=False, del_key=None):
        """Retorna True se botão de deleção foi clicado."""
        sel = st.session_state.get(key_sel)
        if sel:
            if deletavel and del_key:
                _c1, _c2 = st.columns([10, 1])
            else:
                _c1 = st.container(); _c2 = None
            with _c1:
                _tipo_s = sel.get("tipo", "")
                _ico_s  = {"estado": "🗺️", "cidade": "📍", "posto": "⛽"}.get(_tipo_s, "📍")
                st.markdown(
                    f"<div style='background:linear-gradient(90deg,{cor}18,transparent);"
                    f"border-left:3px solid {cor};border-radius:0 8px 8px 0;"
                    f"padding:6px 10px;font-size:12px;margin-bottom:2px'>"
                    f"<span style='color:{cor};font-weight:700;font-size:10px;"
                    f"text-transform:uppercase;letter-spacing:.6px'>{icon} {label}</span><br>"
                    f"<span style='color:#1a1a1a'>{_ico_s} {sel['label'][:55]}</span></div>",
                    unsafe_allow_html=True,
                )
            if _c2:
                with _c2:
                    if st.button("✕", key=del_key, use_container_width=True):
                        st.session_state.pop(key_sel, None)
                        return True
            else:
                if st.button(f"✕ Limpar {label}",
                             key=f"_rot_clr_{key_sel}_{_rot_fk}", help=f"Limpar {label}"):
                    st.session_state.pop(key_sel, None)
                    st.rerun()
            return False

        # ── Digitando ───────────────────────────────────────────────
        if deletavel and del_key:
            _ci, _cd = st.columns([10, 1])
        else:
            _ci = st.container(); _cd = None

        with _ci:
            _txt = st.text_input(
                f"{icon} {label}", placeholder="Cidade, UF, CNPJ ou Nome do Posto",
                key=f"{key_txt}_{_rot_fk}",
            )
        if _cd:
            with _cd:
                if st.button("✕", key=del_key, use_container_width=True):
                    return True

        _ant = st.session_state.get(f"_ra_{key_sel}", "")
        if _txt != _ant:
            st.session_state[f"_ra_{key_sel}"] = _txt

        if len(_txt.strip()) >= 2:
            _ts = _txt.strip(); _tu = _ts.upper()
            _sg: list = []
            if _tu in UFS:
                _bb = BBOX_UFS.get(_tu, (-15.8, -47.9, -15.7, -47.8))
                _sg = [{"label": f"🗺️ Estado {_tu}",
                        "lat": (_bb[0]+_bb[2])/2, "lon": (_bb[1]+_bb[3])/2,
                        "tipo": "estado"}]
            elif len(_ts) >= 3:
                _sc = _buscar_cidades_cache(_ts) or [
                    dict(s, tipo="cidade") for s in sugestoes_nominatim(_ts)
                ]
                _sp = buscar_posto_por_texto(_ts)
                _sg = _sc[:4] + _sp[:3]
            if _sg:
                _lbs = [s["label"] for s in _sg]
                _ix = st.selectbox("", range(len(_lbs)),
                    format_func=lambda i: _lbs[i],
                    key=f"_rs_{key_sel}_{_rot_fk}",
                    index=None, placeholder="↑ selecione…",
                    label_visibility="collapsed")
                if _ix is not None:
                    st.session_state[key_sel] = _sg[_ix]
                    st.rerun()
        return False

    def _rail_r(cor="#90CAF9", h=8):
        st.markdown(
            f"<div style='margin:2px 0 2px 8px;border-left:2px dashed {cor};"
            f"height:{h}px'></div>", unsafe_allow_html=True)

    # ════════════════════════════════════════════════════════════════
    #  FORMULÁRIO DA ROTA
    # ════════════════════════════════════════════════════════════════
    _rot_orig = st.session_state.get("rot_orig_sel")
    _rot_dest = st.session_state.get("rot_dest_sel")
    _rot_res  = st.session_state.get("_rot_result")

    # Se há resultado, colapsa o formulário num expander
    _form_expanded = (_rot_res is None)

    with st.expander("📍 Pontos da Rota", expanded=_form_expanded):

        st.markdown(
            "<div style='font-size:11px;color:#555;margin-bottom:8px'>"
            "💡 Digite cidade, UF, CNPJ ou nome do posto e selecione a sugestão.</div>",
            unsafe_allow_html=True)

        _ponto_rot("Ponto de Origem", "rot_orig_sel", "rot_txt_orig", "🟢", "#2E7D32")
        _rot_orig = st.session_state.get("rot_orig_sel")

        _dels_rot = []
        for _ri in range(1, _rot_np + 1):
            _rail_r("#FF8F00", 6)
            _del = _ponto_rot(
                f"Parada {_ri}", f"rot_parada_sel_{_ri}", f"rot_txt_parada_{_ri}",
                "🟠", "#E65100", deletavel=True, del_key=f"_rd_{_ri}_{_rot_fk}")
            if _del:
                _dels_rot.append(_ri)

        if _dels_rot:
            for _di in sorted(_dels_rot, reverse=True):
                for _dj in range(_di, _rot_np):
                    _nxt = st.session_state.get(f"rot_parada_sel_{_dj+1}")
                    st.session_state[f"rot_parada_sel_{_dj}"] = _nxt if _nxt else None
                    if not _nxt:
                        st.session_state.pop(f"rot_parada_sel_{_dj}", None)
                st.session_state.pop(f"rot_parada_sel_{_rot_np}", None)
            st.session_state["_rot_np"] = max(0, _rot_np - len(_dels_rot))
            st.rerun()

        _rail_r("#BDBDBD", 5)
        _ca, _cb = st.columns([2, 3])
        with _ca:
            if _rot_np < 10:
                if st.button("➕ Adicionar Parada", key=f"rot_add_{_rot_fk}",
                             use_container_width=True):
                    st.session_state["_rot_np"] = _rot_np + 1
                    st.rerun()
        with _cb:
            if _rot_np:
                st.caption(f"{_rot_np}/10 parada(s)")

        _rail_r("#C62828", 8)
        _ponto_rot("Ponto de Destino", "rot_dest_sel", "rot_txt_dest", "🔴", "#C62828")
        _rot_dest = st.session_state.get("rot_dest_sel")

        st.markdown("<div style='height:8px'></div>", unsafe_allow_html=True)

        _rot_pronto = bool(_rot_orig) and bool(_rot_dest)
        if not _rot_pronto:
            st.markdown(
                "<div style='background:#fff3e0;border-radius:8px;padding:6px 10px;"
                "font-size:11px;color:#e65100;text-align:center'>"
                "⚠️ Informe <b>Origem</b> e <b>Destino</b></div>",
                unsafe_allow_html=True)

        _cc, _ccl = st.columns([4, 1])
        with _cc:
            _rot_calcular = st.button(
                "🛣️ Calcular Roteirização",
                use_container_width=True,
                disabled=not _rot_pronto,
                key=f"rot_calc_{_rot_fk}",
                type="primary",
            )
        with _ccl:
            if st.button("🗑️", use_container_width=True,
                         key=f"rot_lmp_{_rot_fk}", help="Limpar tudo"):
                for _k in [k for k in list(st.session_state.keys())
                           if k.startswith("rot_") or k.startswith("_rot_")
                           or k.startswith("_ra_") or k.startswith("_rs_")
                           or k.startswith("_rd_")]:
                    st.session_state.pop(_k, None)
                st.rerun()

    # ════════════════════════════════════════════════════════════════
    #  CÁLCULO (roda no mesmo pass — sem st.rerun após armazenar)
    # ════════════════════════════════════════════════════════════════
    if _rot_calcular and _rot_pronto:
        _log_acesso("ROTEIRIZACAO_CALCULAR",
                    f"{st.session_state.get('rot_placa','—')} | "
                    f"orig={st.session_state.get('_rot_orig_sel',{}).get('label','—')} → "
                    f"dest={st.session_state.get('_rot_dest_sel',{}).get('label','—')}")
        _wps_calc = []
        for _ri in range(1, _rot_np + 1):
            _ws = st.session_state.get(f"rot_parada_sel_{_ri}")
            if _ws and "lat" in _ws:
                _wps_calc.append([float(_ws["lat"]), float(_ws["lon"])])

        with st.spinner("🗺️ Calculando rota…"):
            try:
                _rc_c, _rd_c, _rm_c, _rlr_c = calcular_rota(
                    float(_rot_orig["lat"]), float(_rot_orig["lon"]),
                    float(_rot_dest["lat"]), float(_rot_dest["lon"]),
                    waypoints=_wps_calc or None,
                )
            except Exception as _e_calc:
                st.error(f"❌ Erro ao calcular rota: {_e_calc}")
                st.stop()

        _paradas_salvas = [
            st.session_state.get(f"rot_parada_sel_{_ri}")
            for _ri in range(1, _rot_np + 1)
            if st.session_state.get(f"rot_parada_sel_{_ri}")
        ]
        st.session_state["_rot_result"] = {
            "coords":     _rc_c,
            "dist_km":    _rd_c,
            "dur_min":    _rm_c,
            "linha_reta": _rlr_c,
            "orig":       _rot_orig,
            "dest":       _rot_dest,
            "paradas":    _paradas_salvas,
            "placa":      _rot_placa,
            "combustivel": _rot_comb,
            "capacidade": _rot_cap,
            "autonomia":  _rot_aut,
        }
        _rot_res = st.session_state["_rot_result"]   # usa na mesma passagem

    # ════════════════════════════════════════════════════════════════
    #  RESULTADOS
    # ════════════════════════════════════════════════════════════════
    if _rot_res:
        _rc    = _rot_res["coords"]
        _rd    = float(_rot_res["dist_km"])
        _rm    = float(_rot_res["dur_min"])
        _rlr   = bool(_rot_res.get("linha_reta", False))
        _ro    = _rot_res["orig"]
        _rt    = _rot_res["dest"]
        _rp    = _rot_res.get("paradas", [])
        _rcomb = _rot_res.get("combustivel", _rot_comb)
        _rcap  = float(_rot_res.get("capacidade", _rot_cap))
        _raut  = float(_rot_res.get("autonomia",  _rot_aut))
        _rmin  = _rcap * 0.25

        # ── Métricas ──────────────────────────────────────────────
        _mc1, _mc2, _mc3, _mc4 = st.columns(4)
        _mc1.metric("📏 Distância",     f"{_rd:,.0f} km".replace(",", "."))
        _mc2.metric("⏱️ Tempo est.",    f"{int(_rm//60)}h {int(_rm%60):02d}min")
        _mc3.metric("🛢 Consumo",       f"{_rd/_raut:.0f} L" if _raut else "—")
        _mc4.metric("📍 Pontos rota",   str(len(_rp) + 2))
        if _rlr:
            st.caption("⚠️ Rota calculada em linha reta (OSRM indisponível).")

        # ── Otimização de abastecimento ───────────────────────────
        _pp_df_r  = st.session_state.get("_pp_df")
        _pf_df_r  = st.session_state.get("pf_coords_df", pd.DataFrame())
        _sugest: list = []
        _cands:  list = []
        _range_avail = (_rcap - _rmin) * _raut

        if _pp_df_r is not None and not _pf_df_r.empty and _rcomb:
            _cm = _pp_df_r[
                _pp_df_r["combustivel_label"].str.upper().str.strip()
                == _rcomb.upper().strip()
            ]
            if not _cm.empty:
                _cpk = _cm["combustivel_pk"].iloc[0]
                _pr  = (_pp_df_r[_pp_df_r["combustivel_pk"] == _cpk]
                        [["cnpj_norm", "preco"]].copy())
                _pfc = _pf_df_r.copy()
                _pfc["_cn"] = _pfc["cnpj"].fillna("").str.replace(r"\D","",regex=True)
                _mg  = _pfc.merge(_pr, left_on="_cn", right_on="cnpj_norm", how="inner")
                # ── Vetorizado: sem iterrows() ─────────────────────
                _mg_valid = _mg[pd.notna(_mg["_lat"]) & pd.notna(_mg["_lon"])].copy()
                if not _mg_valid.empty:
                    _cands = [
                        {
                            "label":     str(r.get("razaoSocial") or "Posto GF")[:45],
                            "cnpj":      str(r["_cn"]),
                            "lat":       float(r["_lat"]),
                            "lon":       float(r["_lon"]),
                            "preco":     float(r["preco"]),
                            "municipio": str(r.get("municipio", "")),
                            "uf":        str(r.get("uf", "")),
                        }
                        for r in _mg_valid.to_dict("records")
                    ]

        if _cands and _rc and _raut > 0:
            _MAX_DEV = 5.0   # km do corredor

            # ── Vetorizar projeção: NumPy broadcasting O(candidatos + pontos_rota) ──
            # Subsamplea a rota para no máximo 300 pontos (suficiente para 5 km de precisão)
            _rc_arr = np.array(_rc, dtype=np.float64)   # (N, 2) → [lat, lon]
            if len(_rc_arr) > 300:
                _idx = np.round(np.linspace(0, len(_rc_arr) - 1, 300)).astype(int)
                _rc_arr = _rc_arr[_idx]

            # Distâncias acumuladas ao longo da rota (km)
            _R = 6371.0
            _dlat = np.radians(np.diff(_rc_arr[:, 0]))
            _dlon = np.radians(np.diff(_rc_arr[:, 1]))
            _lat1r = np.radians(_rc_arr[:-1, 0])
            _lat2r = np.radians(_rc_arr[1:, 0])
            _a_seg = np.sin(_dlat/2)**2 + np.cos(_lat1r)*np.cos(_lat2r)*np.sin(_dlon/2)**2
            _seg_km = 2 * _R * np.arcsin(np.sqrt(_a_seg))
            _cum_km = np.concatenate([[0.0], np.cumsum(_seg_km)])  # (N,)

            # Para cada candidato: distância haversine a cada ponto da rota (vetorizado)
            _c_lats = np.radians(np.array([c["lat"] for c in _cands]))   # (M,)
            _c_lons = np.radians(np.array([c["lon"] for c in _cands]))   # (M,)
            _r_lats = np.radians(_rc_arr[:, 0])                          # (N,)
            _r_lons = np.radians(_rc_arr[:, 1])                          # (N,)

            # Broadcasting (M, N)
            _dlat_c = _c_lats[:, None] - _r_lats[None, :]
            _dlon_c = _c_lons[:, None] - _r_lons[None, :]
            _aa = (np.sin(_dlat_c/2)**2
                   + np.cos(_c_lats[:, None]) * np.cos(_r_lats[None, :])
                   * np.sin(_dlon_c/2)**2)
            _dist_mat = 2 * _R * np.arcsin(np.sqrt(np.clip(_aa, 0, 1)))  # (M, N) km

            _best_idx = np.argmin(_dist_mat, axis=1)   # (M,) — índice do ponto mais próximo
            _perp_km  = _dist_mat[np.arange(len(_cands)), _best_idx]   # (M,)
            _km_along = _cum_km[_best_idx]                              # (M,)

            _ests = []
            for _i, _pc in enumerate(_cands):
                _perp = float(_perp_km[_i])
                _kma  = float(_km_along[_i])
                if _perp <= _MAX_DEV and 0 <= _kma <= _rd:
                    _ests.append({**_pc, "_km": _kma, "_dev": _perp})
            _ests.sort(key=lambda x: x["_km"])

            # ════════════════════════════════════════════════════════
            # ALGORITMO INTELIGENTE DE PARADAS — v2
            # ────────────────────────────────────────────────────────
            # Princípios:
            #   1. Só para quando necessário (combustível abaixo de
            #      65%) OU quando há vantagem real de preço.
            #   2. Look-ahead: se há posto >3% mais barato além da
            #      janela obrigatória, abastece o mínimo no posto
            #      atual para "voar" até o mais barato.
            #   3. Na última parada, abastece só o necessário para
            #      o destino (sem encher o tanque desnecessariamente).
            #   4. Evita paradas minúsculas (< 5 L) — descarta e
            #      avança posição sem registrar parada.
            # ════════════════════════════════════════════════════════
            _alcance_efetivo = (_rcap - _rmin) * _raut   # km por ciclo
            _PCT_BAIXO       = 0.65   # abaixo deste % considera parar voluntariamente
            _PRECO_VANT      = 0.03   # vantagem mínima de preço para look-ahead (3%)
            _LITROS_MIN_STOP = 5      # descarta paradas com < 5 L

            _pos          = 0.0
            _fuel         = float(_rcap)
            _seen: set    = set()
            _ultimo_preco = None      # último preço pago (para comparação)

            for _ in range(30):       # no máximo 30 paradas
                if _pos >= _rd:
                    break

                # Até onde posso ir antes de atingir o mínimo?
                _can_go = (_fuel - _rmin) * _raut
                _must   = _pos + _can_go

                if _must >= _rd:
                    break             # alcança o destino sem parar

                # ── Janela obrigatória: [_pos, _must] ──────────────
                _janela = [e for e in _ests
                           if _pos < e["_km"] <= _must
                           and e["cnpj"] not in _seen]

                # ── Janela estendida: além de _must, alcançável com
                #    tanque cheio a partir de qualquer posto da janela
                _janela_ext = [e for e in _ests
                               if _must < e["_km"] <= _pos + _alcance_efetivo * 1.85
                               and e["cnpj"] not in _seen]

                if not _janela:
                    # Emergência: pega o mais próximo disponível
                    _alem = [e for e in _ests
                             if e["_km"] > _pos and e["cnpj"] not in _seen]
                    if not _alem:
                        break
                    _best = dict(min(_alem, key=lambda x: x["_km"]))
                    _best["motivo"]    = "emergencia"
                    _fill_target_km   = None
                else:
                    _best_obrig  = min(_janela, key=lambda x: x.get("preco", 9999))
                    _preco_obrig = _best_obrig.get("preco", 9999)

                    # ── Look-ahead: posto mais barato além da janela?
                    _fill_target_km = None
                    if _janela_ext:
                        _best_ext  = min(_janela_ext, key=lambda x: x.get("preco", 9999))
                        _preco_ext = _best_ext.get("preco", 9999)
                        if _preco_ext < _preco_obrig * (1 - _PRECO_VANT):
                            _fill_target_km = _best_ext["_km"]

                    _best = dict(_best_obrig)
                    _best["motivo"] = "estrategico" if _fill_target_km else "mais_barato"

                # ── Combustível ao chegar no posto ─────────────────
                _km_ate       = _best["_km"] - _pos
                _fuel_chegada = max(0.0, _fuel - (_km_ate / _raut))
                _pct_chegada  = (_fuel_chegada / _rcap * 100) if _rcap else 0.0

                # ── Pula parada se tanque ainda alto e preço não compensa
                if (_best.get("motivo") != "emergencia"
                        and _pct_chegada >= _PCT_BAIXO * 100
                        and _ultimo_preco is not None
                        and _best.get("preco", 9999) >= _ultimo_preco * (1 - _PRECO_VANT)
                        and not _fill_target_km):
                    # Avança sem registrar parada
                    _pos  = _best["_km"]
                    _fuel = _fuel_chegada
                    _seen.add(_best["cnpj"])   # evita revisitar
                    continue

                # ── Quantos litros abastecer? ───────────────────────
                _dist_restante = _rd - _best["_km"]

                if _fill_target_km:
                    # Abastece mínimo para chegar ao posto mais barato à frente
                    _dist_ate_target    = _fill_target_km - _best["_km"]
                    _litros_necessarios = (_dist_ate_target / _raut) * 1.10 + _rmin
                elif _dist_restante <= _alcance_efetivo:
                    # Última parada: só o necessário para o destino + margem
                    _litros_necessarios = (_dist_restante / _raut) * 1.15 + _rmin
                else:
                    # Parada intermediária: enche o tanque (máximo alcance)
                    _litros_necessarios = _rcap

                _litros_fill = max(0.0, _litros_necessarios - _fuel_chegada)
                _litros_fill = min(_litros_fill, _rcap - _fuel_chegada)
                _litros_fill = math.ceil(_litros_fill)

                # Descarta paradas com abastecimento insignificante
                if _litros_fill < _LITROS_MIN_STOP:
                    _pos  = _best["_km"]
                    _fuel = _fuel_chegada
                    _seen.add(_best["cnpj"])
                    continue

                _fuel_apos = min(_fuel_chegada + _litros_fill, _rcap)
                _pct_apos  = (_fuel_apos / _rcap * 100) if _rcap else 0.0
                _custo_ab  = round(_litros_fill * _best.get("preco", 0.0), 2)

                _best["fuel_chegada"]     = round(_fuel_chegada, 1)
                _best["pct_chegada"]      = round(_pct_chegada, 1)
                _best["litros_sugeridos"] = int(_litros_fill)
                _best["custo_abast"]      = _custo_ab
                _best["fuel_apos"]        = round(_fuel_apos, 1)
                _best["pct_apos"]         = round(_pct_apos, 1)

                _seen.add(_best["cnpj"])
                _sugest.append(_best)
                _ultimo_preco = _best.get("preco")

                _fuel = _fuel_apos
                _pos  = _best["_km"]

        # ── Tabs ─────────────────────────────────────────────────
        _t_mapa, _t_abast, _t_custo, _t_res = st.tabs(
            ["🗺️  Mapa da Rota", "⛽  Abastecimento", "💰  Custo da Viagem", "📋  Resumo"]
        )

        with _t_mapa:
            try:
                _wps_mapa = [
                    {"lat": float(p["lat"]), "lon": float(p["lon"]),
                     "label": str(p.get("label", f"Parada {_wi+1}"))}
                    for _wi, p in enumerate(_rp) if "lat" in p and "lon" in p
                ] or None

                _cnpjs_sg = {s["cnpj"] for s in _sugest}
                if _cnpjs_sg and not _pf_df_r.empty:
                    _dfm = _pf_df_r.copy()
                    _dfm["_cn2"] = _dfm["cnpj"].fillna("").str.replace(r"\D","",regex=True)
                    _dfm = preparar_df(_dfm[_dfm["_cn2"].isin(_cnpjs_sg)], [])
                else:
                    _dfm = pd.DataFrame()

                _fig_r = criar_mapa(
                    _dfm,
                    coords_rota=_rc,
                    lat_orig=float(_ro["lat"]),  lon_orig=float(_ro["lon"]),
                    lat_dest=float(_rt["lat"]),  lon_dest=float(_rt["lon"]),
                    label_orig=str(_ro.get("label","Origem"))[:30],
                    label_dest=str(_rt.get("label","Destino"))[:30],
                    waypoints=_wps_mapa,
                )
                _renderizar_mapa(_fig_r, height=560, key="mapa_rot_v2")
            except Exception as _e_mapa:
                st.error(f"❌ Erro ao renderizar mapa: {_e_mapa}")

        with _t_abast:
            if not _rcomb:
                st.info("Configure o **combustível** no menu lateral.")
            elif not _cands:
                st.warning(
                    f"⚠️ Nenhum posto GF com preço de **{_rcomb}** encontrado. "
                    "Verifique se a planilha de preços foi carregada em **Configurações**.")
            elif not _sugest:
                st.success(
                    f"✅ Nenhuma parada necessária — alcance efetivo "
                    f"(~{_range_avail:.0f} km) é suficiente para os {_rd:.0f} km desta rota.")
            else:
                _n_ab         = len(_sugest)
                _custo_total  = sum(s.get("custo_abast", 0) for s in _sugest)
                _litros_total = sum(s.get("litros_sugeridos", 0) for s in _sugest)
                # Banner resumo
                st.markdown(
                    f"<div style='background:linear-gradient(90deg,#e0f7fa,#f1f8e9);"
                    f"border-radius:8px;padding:9px 14px;margin-bottom:10px;"
                    f"font-size:12px;color:#004D40'>"
                    f"⛽ <b>{_n_ab} parada(s) sugerida(s)</b> &nbsp;·&nbsp; "
                    f"🛢 Total a abastecer: <b>{_litros_total} L</b> &nbsp;·&nbsp; "
                    f"💰 Custo total estimado: <b>R$ {_brl(_custo_total, 2)}</b>"
                    f"</div>",
                    unsafe_allow_html=True)

                for _ia, _ab in enumerate(_sugest, 1):
                    _motivo = _ab.get("motivo", "mais_barato")
                    if _motivo == "mais_barato":
                        _cor_ab = "#1B5E20"; _bg_ab = "#f1f8e9"
                        _tag_ab = "🏆 Melhor preço"
                    elif _motivo == "estrategico":
                        _cor_ab = "#E65100"; _bg_ab = "#fff3e0"
                        _tag_ab = "🎯 Estratégico — posto mais barato à frente"
                    else:
                        _cor_ab = "#B71C1C"; _bg_ab = "#fce4ec"
                        _tag_ab = "⚠️ Emergência — posto mais próximo"

                    _litros  = _ab.get("litros_sugeridos", 0)
                    _custo_p = _ab.get("custo_abast", 0)
                    _f_ch    = _ab.get("fuel_chegada", 0)
                    _p_ch    = _ab.get("pct_chegada", 0)
                    _f_ap    = _ab.get("fuel_apos", 0)
                    _p_ap    = _ab.get("pct_apos", 0)
                    _preco_l = _ab.get("preco", 0)

                    # Barra de nível do tanque (chegada → saída)
                    _bar_ch  = max(2, int(_p_ch))
                    _bar_ap  = max(2, int(_p_ap))
                    _bar_cor = "#4CAF50" if _p_ap >= 50 else ("#FF9800" if _p_ap >= 25 else "#F44336")

                    st.markdown(
                        f"<div style='border-left:4px solid {_cor_ab};background:{_bg_ab};"
                        f"border-radius:0 10px 10px 0;padding:10px 14px;margin-bottom:8px'>"

                        # Linha 1: nome + tag
                        f"<div style='display:flex;align-items:center;gap:8px;margin-bottom:4px'>"
                        f"<b style='color:{_cor_ab};font-size:13px'>#{_ia} {_ab['label']}</b>"
                        f"<span style='background:{_cor_ab};color:#fff;border-radius:4px;"
                        f"padding:1px 7px;font-size:10px;white-space:nowrap'>{_tag_ab}</span>"
                        f"</div>"

                        # Linha 2: localização + km + preço
                        f"<div style='font-size:11px;color:#555;margin-bottom:5px'>"
                        f"📍 {_ab.get('municipio','')} / {_ab.get('uf','')} &nbsp;·&nbsp; "
                        f"🛣 <b>{_ab.get('_km',0):.0f} km</b> da origem &nbsp;·&nbsp; "
                        f"💰 <b>R$ {_preco_l:.3f}/L</b>"
                        f"</div>"

                        # Linha 3: nível tanque — chegada e saída
                        f"<div style='font-size:11px;color:#444;margin-bottom:5px'>"
                        f"🔋 Chega com <b>{_f_ch:.0f} L ({_p_ch:.0f}%)</b> &nbsp;→&nbsp; "
                        f"⛽ Abastece <b style='color:{_cor_ab}'>{_litros} L de {_rcomb}</b> &nbsp;→&nbsp; "
                        f"🔋 Sai com <b style='color:{_bar_cor}'>{_f_ap:.0f} L ({_p_ap:.0f}%)</b>"
                        f"</div>"

                        # Linha 4: custo + barra visual do tanque
                        f"<div style='display:flex;align-items:center;gap:10px'>"
                        f"<span style='font-size:12px;font-weight:700;color:{_cor_ab}'>"
                        f"💵 Custo: R$ {_custo_p:.2f}</span>"
                        f"<div style='flex:1;background:#e0e0e0;border-radius:4px;height:8px;overflow:hidden'>"
                        f"<div style='width:{_bar_ap}%;background:{_bar_cor};height:100%;border-radius:4px'></div>"
                        f"</div>"
                        f"<span style='font-size:10px;color:#888'>{_p_ap:.0f}%</span>"
                        f"</div>"

                        f"</div>",
                        unsafe_allow_html=True)

        # ══════════════════════════════════════════════════════════
        # TAB CUSTO DA VIAGEM
        # ══════════════════════════════════════════════════════════
        with _t_custo:
            # ── Sem planilha de preços ────────────────────────────
            if _pp_df_r is None or _pp_df_r.empty:
                st.info(
                    "ℹ️ Para calcular o custo da viagem, carregue a **planilha de preços** "
                    "em Configurações → 💲 Preços PP."
                )
            else:
                # ── Dados base ────────────────────────────────────
                _consumo_total_l = _rd / _raut if _raut > 0 else 0

                # Preço médio dos candidatos GF para esse combustível
                _preco_medio_gf = (
                    sum(c["preco"] for c in _cands) / len(_cands) if _cands else None
                )

                # Referência ANP (se disponível via planilha carregada)
                _anp_cache_r   = st.session_state.get("_precos_anp_cache", {})
                _sheets_r      = _anp_cache_r.get("sheets")
                _preco_anp_ref = None
                if _sheets_r and "estados" in _sheets_r:
                    _df_est_r = _sheets_r["estados"]
                    _ce_r = _anp_col(_df_est_r, "estado", "estados")
                    _cp_r = _anp_col(_df_est_r, "produto")
                    _cm_r = _anp_col(_df_est_r, "medio revenda", "media revenda", "preco medio")
                    if _ce_r and _cp_r and _cm_r:
                        _uf_origem = _ro.get("uf", "") or ""
                        _pk_r = _anp_norm(_rcomb)
                        _anp_vals = []
                        for _, _rr in _df_est_r.iterrows():
                            _uf_n = _anp_norm(str(_rr.get(_ce_r, "")))
                            _pk_n = _anp_norm(str(_rr.get(_cp_r, "")))
                            if _uf_n == _anp_norm(_uf_origem) or _uf_n in _anp_norm(_uf_origem):
                                if _pk_r in _pk_n or _pk_n in _pk_r:
                                    try:
                                        _v = float(str(_rr.get(_cm_r, "")).replace(",", "."))
                                        if _v > 0:
                                            _anp_vals.append(_v)
                                    except (ValueError, TypeError):
                                        pass
                        if _anp_vals:
                            _preco_anp_ref = sum(_anp_vals) / len(_anp_vals)

                # Custo total dos abastecimentos sugeridos
                _custo_sugest  = sum(s.get("custo_abast", 0) for s in _sugest)
                _litros_sugest = sum(s.get("litros_sugeridos", 0) for s in _sugest)

                # Custo estimado total (inclui combustível já no tanque consumido)
                _custo_km      = _custo_sugest / _rd if _rd > 0 and _custo_sugest > 0 else 0
                _custo_100km   = _custo_km * 100

                # Custo se fosse ao preço médio GF (referência de mercado)
                _custo_medio_gf = (
                    _consumo_total_l * _preco_medio_gf if _preco_medio_gf else None
                )
                # Custo se fosse ao preço ANP
                _custo_anp = (
                    _consumo_total_l * _preco_anp_ref if _preco_anp_ref else None
                )

                # ── Header ────────────────────────────────────────
                st.markdown(
                    "<div style='font-size:13px;font-weight:700;color:#1B5E20;"
                    "margin-bottom:12px'>💰 Custo Estimado da Viagem</div>",
                    unsafe_allow_html=True,
                )

                # ── Caso sem paradas necessárias ──────────────────
                if not _sugest and _preco_medio_gf:
                    _custo_sem_par = _consumo_total_l * _preco_medio_gf
                    _ck1, _ck2, _ck3 = st.columns(3)
                    _ck1.metric("💰 Custo estimado",
                                f"R$ {_custo_sem_par:,.2f}".replace(",","X").replace(".",",").replace("X","."))
                    _ck2.metric("🛢 Consumo total",
                                f"{_consumo_total_l:.1f} L")
                    _ck3.metric("📏 Custo/km",
                                f"R$ {_custo_sem_par/_rd:.3f}".replace(".",",") if _rd > 0 else "—")
                    st.info(
                        f"✅ O alcance efetivo do veículo (~{_range_avail:.0f} km) cobre toda a rota "
                        f"({_rd:.0f} km). O custo é calculado com base no preço médio dos postos GF "
                        f"para **{_rcomb}**."
                    )

                elif _sugest:
                    # ── KPIs ──────────────────────────────────────
                    _kc1, _kc2, _kc3, _kc4 = st.columns(4)
                    _kc1.metric(
                        "💰 Total Abastecimentos",
                        f"R$ {_custo_sugest:,.2f}".replace(",","X").replace(".",",").replace("X","."),
                        help="Custo total das paradas de abastecimento sugeridas",
                    )
                    _kc2.metric(
                        "🛢 Total Abastecido",
                        f"{_litros_sugest:.0f} L",
                        delta=f"{_litros_sugest/_consumo_total_l*100:.0f}% do consumo" if _consumo_total_l > 0 else None,
                    )
                    _kc3.metric(
                        "📏 Custo por km",
                        f"R$ {_custo_km:.4f}".replace(".",",") if _custo_km > 0 else "—",
                        help="Custo de abastecimento dividido pela distância total",
                    )
                    _kc4.metric(
                        "⛽ Preço médio pago",
                        f"R$ {_custo_sugest/_litros_sugest:.3f}/L".replace(".",",") if _litros_sugest > 0 else "—",
                        help="Preço médio ponderado pelos volumes abastecidos",
                    )

                    st.markdown("<div style='height:6px'></div>", unsafe_allow_html=True)

                    # ── Gráfico custo acumulado ao longo da rota ──
                    if len(_sugest) >= 1:
                        st.markdown("#### 📈 Custo Acumulado ao Longo da Rota")
                        _km_pts  = [0.0] + [s.get("_km", 0) for s in _sugest] + [float(_rd)]
                        _custo_pts = [0.0]
                        _acc = 0.0
                        for _s in _sugest:
                            _acc += _s.get("custo_abast", 0)
                            _custo_pts.append(_acc)
                        _custo_pts.append(_acc)  # linha reta até o destino

                        _lbl_pts = (
                            [f"Origem: {_ro.get('label','')[:25]}"]
                            + [s["label"][:30] for s in _sugest]
                            + [f"Destino: {_rt.get('label','')[:25]}"]
                        )

                        _fig_acc = go.Figure()
                        # Área preenchida
                        _fig_acc.add_trace(go.Scatter(
                            x=_km_pts, y=_custo_pts,
                            mode="lines+markers",
                            name="Custo acumulado",
                            line=dict(color="#1B5E20", width=2.5),
                            marker=dict(size=10, color="#2E7D32",
                                       line=dict(color="white", width=2)),
                            fill="tozeroy",
                            fillcolor="rgba(27,94,32,0.10)",
                            text=_lbl_pts,
                            hovertemplate=(
                                "<b>%{text}</b><br>"
                                "Km: %{x:.0f} km<br>"
                                "Acumulado: R$ %{y:,.2f}<extra></extra>"
                            ),
                        ))
                        # Marcadores de cada parada
                        for _is, _sp in enumerate(_sugest):
                            _acc_s = sum(_sugest[j].get("custo_abast", 0) for j in range(_is + 1))
                            _fig_acc.add_annotation(
                                x=_sp.get("_km", 0),
                                y=_acc_s,
                                text=f"R$ {_sp.get('custo_abast',0):.0f}",
                                showarrow=True, arrowhead=2,
                                arrowcolor="#E65100", arrowsize=1,
                                ax=0, ay=-32,
                                font=dict(size=9, color="#E65100"),
                                bgcolor="white",
                                bordercolor="#E65100",
                                borderwidth=1,
                                borderpad=2,
                            )
                        _fig_acc.update_layout(
                            xaxis_title="Distância (km)",
                            yaxis_title="Custo acumulado (R$)",
                            height=280,
                            margin=dict(l=10, r=10, t=20, b=40),
                            plot_bgcolor="rgba(0,0,0,0)",
                            paper_bgcolor="rgba(0,0,0,0)",
                            showlegend=False,
                            font=dict(size=11),
                        )
                        _fig_acc.update_xaxes(showgrid=True, gridcolor="#E8F5E9")
                        _fig_acc.update_yaxes(showgrid=True, gridcolor="#E8F5E9",
                                              tickprefix="R$ ")
                        st.plotly_chart(_fig_acc, use_container_width=True)

                    # ── Breakdown por posto ───────────────────────
                    st.markdown("#### 🏢 Custo por Posto de Abastecimento")
                    _fig_bk = go.Figure()
                    _nomes_bk = [f"#{_i+1} {s['label'][:28]}" for _i, s in enumerate(_sugest)]
                    _custos_bk = [s.get("custo_abast", 0) for s in _sugest]
                    _precos_bk = [s.get("preco", 0) for s in _sugest]
                    _litros_bk = [s.get("litros_sugeridos", 0) for s in _sugest]
                    _cores_bk  = [
                        "#1B5E20" if p == min(_precos_bk) else
                        "#E65100" if p == max(_precos_bk) else
                        "#2E7D32"
                        for p in _precos_bk
                    ]
                    _fig_bk.add_trace(go.Bar(
                        y=_nomes_bk,
                        x=_custos_bk,
                        orientation="h",
                        marker_color=_cores_bk,
                        text=[
                            f"R$ {c:.2f}  ({l}L @ R$ {p:.3f}/L)"
                            .replace(".",",")
                            for c, l, p in zip(_custos_bk, _litros_bk, _precos_bk)
                        ],
                        textposition="outside",
                        hovertemplate=(
                            "<b>%{y}</b><br>"
                            "Custo: R$ %{x:,.2f}<br>"
                            "<extra></extra>"
                        ),
                    ))
                    _fig_bk.update_layout(
                        xaxis_title="Custo (R$)",
                        yaxis=dict(autorange="reversed"),
                        height=max(200, len(_sugest) * 48 + 60),
                        margin=dict(l=10, r=120, t=20, b=30),
                        plot_bgcolor="rgba(0,0,0,0)",
                        paper_bgcolor="rgba(0,0,0,0)",
                        font=dict(size=10),
                    )
                    _fig_bk.update_xaxes(showgrid=True, gridcolor="#E8F5E9",
                                         tickprefix="R$ ")
                    st.plotly_chart(_fig_bk, use_container_width=True)

                    # ── Comparativo de preços ─────────────────────
                    _preco_pago_medio = (
                        _custo_sugest / _litros_sugest if _litros_sugest > 0 else None
                    )
                    _tem_comparativo = (
                        _preco_medio_gf is not None or _preco_anp_ref is not None
                    )
                    if _tem_comparativo and _preco_pago_medio:
                        st.markdown("#### 📊 Comparativo de Preços")
                        _comp_nomes, _comp_vals, _comp_cores = [], [], []
                        _comp_nomes.append("Preço pago (postos GF selecionados)")
                        _comp_vals.append(_preco_pago_medio)
                        _comp_cores.append("#1B5E20")
                        if _preco_medio_gf and abs(_preco_medio_gf - _preco_pago_medio) > 0.001:
                            _comp_nomes.append("Preço médio rede GF (rota)")
                            _comp_vals.append(_preco_medio_gf)
                            _comp_cores.append("#42A5F5")
                        if _preco_anp_ref:
                            _comp_nomes.append(f"Referência ANP ({_ro.get('uf','UF')})")
                            _comp_vals.append(_preco_anp_ref)
                            _comp_cores.append("#E65100")

                        _fig_cmp = go.Figure()
                        _fig_cmp.add_trace(go.Bar(
                            x=_comp_nomes,
                            y=_comp_vals,
                            marker_color=_comp_cores,
                            text=[f"R$ {v:.3f}".replace(".", ",") for v in _comp_vals],
                            textposition="outside",
                            hovertemplate="<b>%{x}</b><br>R$ %{y:.3f}/L<extra></extra>",
                        ))
                        _fig_cmp.update_layout(
                            yaxis_title="R$/L",
                            height=260,
                            margin=dict(l=10, r=10, t=20, b=60),
                            plot_bgcolor="rgba(0,0,0,0)",
                            paper_bgcolor="rgba(0,0,0,0)",
                            font=dict(size=10),
                            showlegend=False,
                        )
                        _fig_cmp.update_yaxes(
                            showgrid=True, gridcolor="#E8F5E9",
                            tickprefix="R$ ",
                            range=[0, max(_comp_vals) * 1.2],
                        )
                        st.plotly_chart(_fig_cmp, use_container_width=True)

                    # ── Projeção de economia ──────────────────────
                    _economia_rows = []
                    if _preco_medio_gf and _preco_pago_medio and _litros_sugest > 0:
                        _custo_se_medio = _litros_sugest * _preco_medio_gf
                        _eco_medio = _custo_se_medio - _custo_sugest
                        _economia_rows.append({
                            "Cenário": "vs. Preço médio GF da rota",
                            "Custo (R$)": f"R$ {_custo_se_medio:,.2f}".replace(",","X").replace(".",",").replace("X","."),
                            "Economia (R$)": f"{'+ R$ ' if _eco_medio >= 0 else '- R$ '}{abs(_eco_medio):,.2f}".replace(",","X").replace(".",",").replace("X","."),
                            "Economia (%)": f"{'▼' if _eco_medio >= 0 else '▲'} {abs(_eco_medio/_custo_se_medio*100):.1f}%",
                        })
                    if _custo_anp and _litros_sugest > 0:
                        _eco_anp = _custo_anp - _custo_sugest
                        # Custo ANP proporcional aos litros abastecidos
                        _custo_anp_prop = _litros_sugest * _preco_anp_ref
                        _eco_anp_prop = _custo_anp_prop - _custo_sugest
                        _economia_rows.append({
                            "Cenário": f"vs. Referência ANP ({_ro.get('uf','UF')})",
                            "Custo (R$)": f"R$ {_custo_anp_prop:,.2f}".replace(",","X").replace(".",",").replace("X","."),
                            "Economia (R$)": f"{'+ R$ ' if _eco_anp_prop >= 0 else '- R$ '}{abs(_eco_anp_prop):,.2f}".replace(",","X").replace(".",",").replace("X","."),
                            "Economia (%)": f"{'▼' if _eco_anp_prop >= 0 else '▲'} {abs(_eco_anp_prop/_custo_anp_prop*100):.1f}%" if _custo_anp_prop > 0 else "—",
                        })

                    if _economia_rows:
                        st.markdown("#### 💡 Projeção de Economia")
                        _eco_df = pd.DataFrame(_economia_rows)
                        st.dataframe(_eco_df, use_container_width=True, hide_index=True)
                        st.caption(
                            "🟢 Economia positiva = postos GF selecionados têm preço abaixo da referência. "
                            "Valores calculados com base nos litros efetivamente abastecidos."
                        )

                    # ── Tabela detalhada ──────────────────────────
                    st.markdown("#### 📋 Tabela Detalhada de Abastecimentos")
                    _det_rows = []
                    for _id, _sp in enumerate(_sugest, 1):
                        _det_rows.append({
                            "#":         _id,
                            "Posto":     _sp["label"][:35],
                            "Município": f"{_sp.get('municipio','')} / {_sp.get('uf','')}",
                            "Km na rota":f"{_sp.get('_km',0):.0f} km",
                            "Preço R$/L": f"R$ {_sp.get('preco',0):.3f}".replace(".",","),
                            "Litros":    f"{_sp.get('litros_sugeridos',0):.0f} L",
                            "Custo":     f"R$ {_sp.get('custo_abast',0):,.2f}".replace(",","X").replace(".",",").replace("X","."),
                            "Nível após":f"{_sp.get('pct_apos',0):.0f}% ({_sp.get('fuel_apos',0):.0f} L)",
                        })
                    if _det_rows:
                        _df_det = pd.DataFrame(_det_rows)
                        st.dataframe(_df_det, use_container_width=True, hide_index=True)

                        # Rodapé totais
                        st.markdown(
                            f"<div style='background:#E8F5E9;border-radius:8px;"
                            f"padding:10px 14px;font-size:12px;color:#1B5E20;"
                            f"display:flex;gap:24px;flex-wrap:wrap;margin-top:4px'>"
                            f"<span>🛢 <b>Total abastecido:</b> {_litros_sugest:.0f} L</span>"
                            f"<span>💰 <b>Custo total:</b> R$ {_custo_sugest:,.2f}</span>".replace(",","X").replace(".",",").replace("X",".")
                            + f"<span>📏 <b>Custo/100 km:</b> R$ {_custo_100km:.2f}</span>".replace(".",",")
                            + f"<span>⛽ <b>Preço médio pago:</b> R$ {_preco_pago_medio:.3f}/L</span>".replace(".",",")
                            + f"</div>",
                            unsafe_allow_html=True,
                        )
                else:
                    st.warning(
                        f"⚠️ Nenhum posto GF com preço de **{_rcomb}** encontrado na rota. "
                        "Verifique a planilha de preços em Configurações."
                    )

        with _t_res:
            st.markdown(
                "<div style='font-size:12px;font-weight:700;color:#004D40;"
                "margin-bottom:10px'>📋 Resumo da Roteirização</div>",
                unsafe_allow_html=True)
            _pontos_r = [_ro] + _rp + [_rt]
            _cores_r  = ["#2E7D32"] + ["#E65100"]*len(_rp) + ["#C62828"]
            _icons_r  = ["🟢"] + ["🟠"]*len(_rp) + ["🔴"]
            _lbls_r   = ["Origem"] + [f"Parada {i+1}" for i in range(len(_rp))] + ["Destino"]
            for _pt, _cr, _ic, _lb in zip(_pontos_r, _cores_r, _icons_r, _lbls_r):
                _ico_t = {"estado":"🗺️","cidade":"📍","posto":"⛽"}.get(_pt.get("tipo",""),"📍")
                st.markdown(
                    f"<div style='border-left:3px solid {_cr};padding:5px 10px;"
                    f"margin-bottom:3px;font-size:12px'>"
                    f"<b style='color:{_cr}'>{_ic} {_lb}:</b> {_ico_t} {_pt.get('label','')[:55]}"
                    f"</div>", unsafe_allow_html=True)
            st.markdown("<div style='height:8px'></div>", unsafe_allow_html=True)
            _custo_r   = sum(s.get("custo_abast", 0) for s in _sugest)
            _litros_r  = sum(s.get("litros_sugeridos", 0) for s in _sugest)
            for _lbl, _val in [
                ("📏 Distância",      f"{_rd:,.0f} km".replace(",",".")),
                ("⏱️ Tempo",         f"{int(_rm//60)}h {int(_rm%60):02d}min"),
                ("⛽ Combustível",    _rcomb or "—"),
                ("🚛 Placa",         _rot_res.get("placa","") or "—"),
                ("🛢 Consumo total", f"{_rd/_raut:.0f} L" if _raut else "—"),
                ("🛢 Total abastec.",f"{_litros_r} L" if _sugest else "—"),
                ("💰 Custo abast.",  f"R$ {_custo_r:.2f}".replace(".",",") if _custo_r else "—"),
                ("⛽ Paradas abast.", str(len(_sugest)) if _sugest else "Nenhuma"),
            ]:
                _cr1, _cr2 = st.columns([3, 4])
                _cr1.caption(_lbl); _cr2.markdown(f"**{_val}**")

        # ── Exportações (Card · PDF · GPX) ────────────────────────
        st.markdown("---")
        _exp_col1, _exp_col2, _exp_col3, _exp_col4 = st.columns([2.5, 1, 1, 1])
        with _exp_col1:
            st.markdown(
                "<div style='font-size:11px;color:#555;padding:6px 0'>"
                "📤 <b>Card</b> para WhatsApp/e-mail &nbsp;·&nbsp; "
                "📄 <b>PDF</b> para impressão &nbsp;·&nbsp; "
                "🗺️ <b>GPX</b> para GPS</div>",
                unsafe_allow_html=True)

        with _exp_col4:
            if st.button("📤 Card PNG", use_container_width=True,
                         key="rot_gerar_card"):
                with st.spinner("🖼️ Gerando card de compartilhamento…"):
                    try:
                        _card_bytes = _gerar_card_rota_png(
                            {
                                "placa":      _rot_res.get("placa",""),
                                "combustivel":_rcomb,
                                "capacidade": _rcap,
                                "autonomia":  _raut,
                                "dist_km":    _rd,
                                "dur_min":    _rm,
                                "coords":     _rc,
                                "orig":       _ro,
                                "dest":       _rt,
                                "paradas":    _rp,
                            },
                            _sugest,
                        )
                        _nome_card = (
                            f"rota_{(_ro.get('label','orig'))[:15].replace(' ','_')}"
                            f"_{(_rt.get('label','dest'))[:15].replace(' ','_')}"
                            f"_{datetime.now().strftime('%Y%m%d_%H%M')}.png"
                        )
                        st.session_state["_rot_card_bytes"] = _card_bytes
                        st.session_state["_rot_card_nome"]  = _nome_card
                    except Exception as _ecard:
                        st.error(f"❌ Erro ao gerar card: {_ecard}")

        if st.session_state.get("_rot_card_bytes"):
            st.download_button(
                label="⬇️ Baixar Card PNG",
                data=st.session_state["_rot_card_bytes"],
                file_name=st.session_state.get("_rot_card_nome", "card_rota.png"),
                mime="image/png",
                use_container_width=True,
                key="rot_download_card",
            )

        with _exp_col3:
            # ── GPX: gerado imediatamente ao clicar ──────────────
            if st.button("🗺️ Exportar GPX", use_container_width=True,
                         key="rot_gerar_gpx"):
                try:
                    _gpx_bytes = gerar_gpx_roteirizacao(
                        {
                            "placa":    _rot_res.get("placa",""),
                            "dist_km":  _rd,
                            "coords":   _rc,
                            "orig":     _ro,
                            "dest":     _rt,
                            "paradas":  _rp,
                        },
                        _sugest,
                    )
                    _nome_gpx = (
                        f"rota_{(_ro.get('label','orig'))[:15].replace(' ','_')}"
                        f"_{(_rt.get('label','dest'))[:15].replace(' ','_')}"
                        f"_{datetime.now().strftime('%Y%m%d_%H%M')}.gpx"
                    )
                    st.session_state["_rot_gpx_bytes"] = _gpx_bytes
                    st.session_state["_rot_gpx_nome"]  = _nome_gpx
                except Exception as _egpx:
                    st.error(f"❌ Erro ao gerar GPX: {_egpx}")

        if st.session_state.get("_rot_gpx_bytes"):
            st.download_button(
                label="⬇️ Baixar arquivo GPX",
                data=st.session_state["_rot_gpx_bytes"],
                file_name=st.session_state.get("_rot_gpx_nome", "rota.gpx"),
                mime="application/gpx+xml",
                use_container_width=True,
                key="rot_download_gpx",
            )

        with _exp_col2:
            if st.button("📄 Gerar PDF", use_container_width=True,
                         key="rot_gerar_pdf", type="primary"):
                with st.spinner("📄 Gerando relatório PDF…"):
                    try:
                        _pdf_bytes = gerar_pdf_roteirizacao(
                            {
                                "placa":      _rot_res.get("placa",""),
                                "combustivel":_rcomb,
                                "capacidade": _rcap,
                                "autonomia":  _raut,
                                "dist_km":    _rd,
                                "dur_min":    _rm,
                                "coords":     _rc,
                                "orig":       _ro,
                                "dest":       _rt,
                                "paradas":    _rp,
                            },
                            _sugest,
                            logo_b64=_LOGO_B64,
                        )
                        if _pdf_bytes:
                            _nome_pdf = (
                                f"rota_{(_ro.get('label','orig'))[:15].replace(' ','_')}"
                                f"_{(_rt.get('label','dest'))[:15].replace(' ','_')}"
                                f"_{datetime.now().strftime('%Y%m%d_%H%M')}.pdf"
                            )
                            st.session_state["_rot_pdf_bytes"] = _pdf_bytes
                            st.session_state["_rot_pdf_nome"]  = _nome_pdf
                        else:
                            st.error("❌ Não foi possível gerar o PDF. Verifique se reportlab está instalado.")
                    except Exception as _epdf:
                        st.error(f"❌ Erro ao gerar PDF: {_epdf}")

        if st.session_state.get("_rot_pdf_bytes"):
            st.download_button(
                label="⬇️ Baixar Relatório PDF",
                data=st.session_state["_rot_pdf_bytes"],
                file_name=st.session_state.get("_rot_pdf_nome", "relatorio_rota.pdf"),
                mime="application/pdf",
                use_container_width=True,
                key="rot_download_pdf",
            )

        # ── Salvar ────────────────────────────────────────────────
        st.markdown("---")
        _nome_sug = (f"{_ro.get('label','Origem')[:20]} → {_rt.get('label','Destino')[:20]}")
        _cn1, _cn2 = st.columns([4, 1])
        with _cn1:
            _nome_in = st.text_input("Nome para salvar", value=_nome_sug,
                                     key="rot_nome_salvar")
        with _cn2:
            st.markdown("<div style='height:28px'></div>", unsafe_allow_html=True)
            if st.button("💾 Salvar", key="rot_salvar"):
                if _salvar_rota_nova(_nome_in or _nome_sug, "roteirizacao", {
                    # ── Campos de exibição (lista de Rotas Salvas) ──
                    "label_orig":  _ro.get("label", ""),
                    "label_dest":  _rt.get("label", ""),
                    "lat_orig":    _ro.get("lat"),
                    "lon_orig":    _ro.get("lon"),
                    "lat_dest":    _rt.get("lat"),
                    "lon_dest":    _rt.get("lon"),
                    "dist_km":     _rd,
                    "dur_min":     _rm,
                    "placa":       _rot_placa,
                    "combustivel": _rcomb,
                    "capacidade":  _rcap,
                    "autonomia":   _raut,
                    # ── Objetos completos para restauração do formulário ──
                    "orig":        _ro,   # dict completo → rot_orig_sel
                    "dest":        _rt,   # dict completo → rot_dest_sel
                    "paradas":     _rp,   # lista de waypoints → rot_parada_sel_*
                    # ── Resultado calculado (restaura sem recalcular) ──
                    "rot_result": {
                        "coords":      _rc,
                        "dist_km":     _rd,
                        "dur_min":     _rm,
                        "linha_reta":  _rlr,
                        "orig":        _ro,
                        "dest":        _rt,
                        "paradas":     _rp,
                        "placa":       _rot_placa,
                        "combustivel": _rcomb,
                        "capacidade":  _rcap,
                        "autonomia":   _raut,
                    },
                    # ── Sugestões de abastecimento ──
                    "abastecimentos": _sugest,
                }):
                    st.toast("✅ Roteirização salva!", icon="💾")
                else:
                    st.error("❌ Erro ao salvar.")

# ── Restauração pós-rerun: recalcula rota do Modo 1 se solicitado ──────────
if (
    st.session_state.get("modo_selecionado") == "📍 Por UF/Município"
    and st.session_state.get("_restore_recalc_rota_m1")
):
    _o_rest = st.session_state.get("_map_orig")
    _d_rest = st.session_state.get("_map_dest")
    if _o_rest and _d_rest:
        with st.spinner("🗺️ Recalculando rota restaurada…"):
            _cr_rest, _dk_rest, _dm_rest, _lr_rest = calcular_rota(
                _o_rest["lat"], _o_rest["lon"],
                _d_rest["lat"], _d_rest["lon"]
            )
        st.session_state["_map_rota_result"] = {
            "coords":   _cr_rest,
            "dist_km":  _dk_rest,
            "dur_min":  _dm_rest,
            "linha_reta": _lr_rest,
            "orig":     _o_rest,
            "dest":     _d_rest,
        }
    st.session_state.pop("_restore_recalc_rota_m1", None)
    st.rerun()
