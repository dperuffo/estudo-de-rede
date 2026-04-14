#!/usr/bin/env python3
"""
═══════════════════════════════════════════════════════════════
  IMPORTAR POSTOS – Relatório Ponto de Venda.xlsx → gestao_frota.db
═══════════════════════════════════════════════════════════════
  Como usar:
    python3 importar_postos.py
    python3 importar_postos.py "Relatório Ponto de Venda.xlsx"
═══════════════════════════════════════════════════════════════
"""
import sys, os, json, sqlite3

# ── Dependência: openpyxl ──────────────────────────────────────
try:
    import openpyxl
except ImportError:
    print("Instalando openpyxl...")
    os.system(f"{sys.executable} -m pip install openpyxl --quiet")
    import openpyxl

# ── Caminhos ──────────────────────────────────────────────────
BASE_DIR  = os.path.dirname(os.path.abspath(__file__))
DB_FILE   = os.path.join(BASE_DIR, 'gestao_frota.db')
XLSX_FILE = sys.argv[1] if len(sys.argv) > 1 else os.path.join(BASE_DIR, 'Relatório Ponto de Venda.xlsx')

# ── Mapeamentos ───────────────────────────────────────────────
UF_MAP = {
    'Acre':'AC','Alagoas':'AL','Amapá':'AP','Amazonas':'AM',
    'Bahia':'BA','Ceará':'CE','Distrito Federal':'DF',
    'Espírito Santo':'ES','Goiás':'GO','Maranhão':'MA',
    'Mato Grosso':'MT','Mato Grosso do Sul':'MS','Minas Gerais':'MG',
    'Pará':'PA','Paraíba':'PB','Paraná':'PR','Pernambuco':'PE',
    'Piauí':'PI','Rio de Janeiro':'RJ','Rio Grande do Norte':'RN',
    'Rio Grande do Sul':'RS','Rondônia':'RO','Roraima':'RR',
    'Santa Catarina':'SC','São Paulo':'SP','Sergipe':'SE','Tocantins':'TO',
}

BANDEIRA_MAP = {
    'Ipiranga':'Ipiranga','Shell':'Shell','Br':'BR Petrobras',
    'Vibra':'Vibra','Bandeira Branca':'Bandeira Branca','Ale':'Ale',
    'Raizen':'Raízen','Raízen':'Raízen',
    'Petrox':'Outros','Rodoil':'Outros','-':'Outros',
}

# ── Helpers ───────────────────────────────────────────────────
def sim(val):
    return str(val or '').strip().lower() == 'sim'

def txt(val):
    return str(val or '').strip()

def latlon(val):
    try:
        return str(val).replace(',', '.').strip()
    except Exception:
        return ''

def normalizar_bandeira(val):
    return BANDEIRA_MAP.get(txt(val), 'Outros')

def normalizar_uf(val):
    v = txt(val)
    if len(v) <= 2:
        return v.upper()
    return UF_MAP.get(v, v[:2].upper() if v else '')

# ── Leitor de linha por cabeçalho ─────────────────────────────
class Linha:
    """Acessa colunas de uma linha pelo nome do cabeçalho, sem depender da posição."""
    def __init__(self, row_tuple, col_map):
        # Padding fixo de 50 posições — evita index out of range em linhas curtas
        self._row = list(row_tuple) + [None] * 50
        # Exclui a chave especial '_all' (valor dict) para não confundir índices
        self._map = {k: v for k, v in col_map.items() if k != '_all'}

    def get(self, *nomes):
        """Retorna o valor da primeira coluna encontrada dentre os nomes passados."""
        for nome in nomes:
            idx = self._map.get(nome)
            if idx is not None:
                return self._row[idx]
        return None

def build_col_map(header_row):
    """Cria mapa {nome_coluna_normalizado: índice} a partir da linha de cabeçalho.
    Armazena TODAS as ocorrências de nomes duplicados na lista _all_indices."""
    m = {}
    all_indices = {}   # nome → [lista de índices em ordem]
    for i, h in enumerate(header_row):
        if h is not None:
            k = str(h).strip()
            m[k] = i                                    # última ocorrência (padrão)
            all_indices.setdefault(k, []).append(i)
    m['_all'] = all_indices
    return m

def parse_servicos(linha):
    servicos = []
    if sim(linha.get('Possui restaurante?')):                       servicos.append('Restaurante')
    if sim(linha.get('Possui banheiro?')):                          servicos.append('Banheiro')
    if sim(linha.get('Possui estacionamento?')):                    servicos.append('Estacionamento')
    if (sim(linha.get('Possui troca de óleo?')) or
        sim(linha.get('Possui óleo a granel?'))):                   servicos.append('Lubrificação')

    outros_raw = txt(linha.get('Outros serviços (conteúdo separado por vírgula)') or '')
    if outros_raw and outros_raw != '-':
        for item in outros_raw.split(','):
            item = item.strip()
            if not item or item == '-': continue
            if 'Lavagem' in item:
                if 'Lavagem' not in servicos: servicos.append('Lavagem')
            else:
                if 'Outros' not in servicos:  servicos.append('Outros')

    if sim(linha.get('Possui conveniência?')) and 'Outros' not in servicos:
        servicos.append('Outros')

    # Planilha simplificada: sem colunas de serviço → lista vazia é ok
    return servicos

# ── Importação ────────────────────────────────────────────────
def main():
    print()
    print('  ╔══════════════════════════════════════════╗')
    print('  ║   Importar Postos – Ponto de Venda       ║')
    print('  ╚══════════════════════════════════════════╝')
    print()

    if not os.path.isfile(XLSX_FILE):
        print(f'  ✕  Arquivo não encontrado: {XLSX_FILE}')
        print(f'     Coloque o arquivo na mesma pasta que este script e tente novamente.')
        sys.exit(1)

    if not os.path.isfile(DB_FILE):
        print(f'  ✕  Banco de dados não encontrado: {DB_FILE}')
        print(f'     Execute "python3 server.py" ao menos uma vez para criar o banco.')
        sys.exit(1)

    print(f'  📂  Excel:  {os.path.basename(XLSX_FILE)}')
    print(f'  🗄️   Banco:  {os.path.basename(DB_FILE)}')
    print()

    # ── Abre a planilha ──
    print('  Lendo planilha...')
    wb = openpyxl.load_workbook(XLSX_FILE, data_only=True, read_only=True)

    NOME_ABA = 'Ponto de Venda'
    if NOME_ABA in wb.sheetnames:
        ws = wb[NOME_ABA]
    else:
        ws = wb.worksheets[0]
        print(f'  ℹ  Usando aba: "{ws.title}"  (abas disponíveis: {wb.sheetnames})')

    rows_iter = ws.iter_rows(values_only=True)

    # ── Lê cabeçalho e monta mapa de colunas ──
    header_row = next(rows_iter, None)
    if header_row is None:
        print('  ✕  A planilha está vazia.')
        sys.exit(1)

    col_map = build_col_map(header_row)

    # ── Detecta formato da planilha e exibe diagnóstico ──
    all_cols = [k for k in col_map if k != '_all']
    print(f'  ℹ  {len(all_cols)} colunas detectadas: {all_cols}')

    # ── Processa linhas ──
    registros = []
    erros     = 0
    all_idx   = col_map.get('_all', {})

    for i, row_tuple in enumerate(rows_iter):
        if not any(v for v in row_tuple if v is not None):
            continue
        try:
            linha    = Linha(row_tuple, col_map)
            row_list = list(row_tuple) + [None] * 50  # padding seguro

            # ── CNPJ: aceita 'CNPJ', 'CNPJ POSTO', 'Cnpj' ──
            cnpj = txt(linha.get('CNPJ', 'CNPJ POSTO', 'Cnpj'))
            # ── Razão Social: aceita 'Nome', 'RAZÃO SOCIAL', 'RAZAO SOCIAL', 'Razão Social' ──
            razao = txt(linha.get('Nome', 'RAZÃO SOCIAL', 'RAZAO SOCIAL', 'Razão Social', 'Razao Social'))
            if not cnpj or not razao:
                continue

            # ── Bandeira ──
            bandeira_raw = linha.get('Bandeira', 'BANDEIRA', 'bandeira')

            # ── Endereço ──
            cep         = txt(linha.get('CEP', 'cep'))
            logradouro  = txt(linha.get('Logradouro', 'LOGRADOURO', 'logradouro'))
            numero      = txt(linha.get('Número', 'Numero', 'NUMERO', 'NÚMERO', 'numero'))
            complemento = txt(linha.get('Complemento', 'COMPLEMENTO', 'complemento'))
            bairro      = txt(linha.get('Bairro', 'BAIRRO', 'bairro'))
            cidade      = txt(linha.get('Cidade', 'CIDADE', 'cidade'))
            uf_raw      = linha.get('UF', 'uf')

            # ── GPS ──
            lat = latlon(linha.get('Latitude', 'LATITUDE', 'latitude'))
            lon = latlon(linha.get('Longitude', 'LONGITUDE', 'longitude'))

            # ── Contato / Gestor ──
            gestor = txt(linha.get('Nome Contato', 'NOME DO GESTOR', 'Nome do Gestor',
                                   'Gestor', 'GESTOR', 'Nome Responsável'))

            # ── Telefone: primeira ocorrência de 'Telefone' ou 'TELEFONE' ──
            tel_indices = all_idx.get('Telefone', all_idx.get('TELEFONE', []))
            tel_contato = txt(row_list[tel_indices[0]]) if tel_indices else ''

            # ── E-mail responsável ──
            email_resp = txt(linha.get(
                'E-MAIL DO RESPONSAVEL', 'E-mail do Responsável',
                'E-mail Responsável', 'Email Responsável'
            ))
            # Fallback: primeira ocorrência de 'E-mail' ou 'E-MAIL'
            if not email_resp:
                em_idx = all_idx.get('E-mail', all_idx.get('E-MAIL', []))
                email_resp = txt(row_list[em_idx[0]]) if em_idx else ''

            # ── E-mail NF ──
            email_nf = txt(linha.get(
                'E-MAIL PARA RECEBIMENTO DE NF', 'E-mail para NF',
                'E-mail NF', 'Email NF'
            ))
            # Fallback: segunda ocorrência de 'E-mail'
            if not email_nf:
                em_idx = all_idx.get('E-mail', all_idx.get('E-MAIL', []))
                email_nf = txt(row_list[em_idx[1]]) if len(em_idx) > 1 else ''

            # ── Dados bancários (presentes na planilha do usuário) ──
            banco   = txt(linha.get('BANCO',    'Banco',    'banco'))
            agencia = txt(linha.get('AGENCIA',  'Agência',  'Agencia', 'agencia'))
            conta   = txt(linha.get('CONTA CORRENTE', 'Conta Corrente', 'CONTA', 'Conta', 'conta'))

            servicos = parse_servicos(linha)

            registros.append({
                'cnpj':        cnpj,
                'razao':       razao,
                'bandeira':    normalizar_bandeira(bandeira_raw),
                'cep':         cep,
                'logradouro':  logradouro,
                'numero':      numero,
                'complemento': complemento,
                'bairro':      bairro,
                'cidade':      cidade,
                'uf':          normalizar_uf(uf_raw),
                'lat':         lat,
                'lon':         lon,
                'gestor':      gestor,
                'telefone':    tel_contato,
                'email_resp':  email_resp,
                'email_nf':    email_nf,
                'banco':       banco,
                'agencia':     agencia,
                'conta':       conta,
                'servicos':    json.dumps(servicos, ensure_ascii=False),
                'combustiveis':json.dumps({},       ensure_ascii=False),
                'fotos':       json.dumps([],       ensure_ascii=False),
            })
        except Exception as e:
            erros += 1
            if erros <= 5:
                print(f'  ⚠  Linha {i+2} ignorada: {e}')

    wb.close()
    print(f'  ✓  {len(registros)} registros lidos ({erros} com erro)')

    if not registros:
        print('  Nenhum registro válido encontrado. Verifique se o arquivo é o correto.')
        sys.exit(1)

    # ── Confirma ──
    print()
    resp = input(f'  Importar {len(registros)} postos? '
                 f'(registros existentes com mesmo CNPJ serão ignorados) [S/n]: ').strip().lower()
    if resp not in ('', 's', 'sim', 'y', 'yes'):
        print('  Importação cancelada.')
        sys.exit(0)

    # ── Insere no banco ──
    conn = sqlite3.connect(DB_FILE)
    conn.execute("PRAGMA journal_mode=WAL")
    inseridos = 0
    duplicatas = 0

    for r in registros:
        try:
            conn.execute('''INSERT OR IGNORE INTO postos
                (cnpj,razao,bandeira,cep,logradouro,numero,complemento,
                 bairro,cidade,uf,lat,lon,
                 gestor,telefone,email_resp,email_nf,
                 banco,agencia,conta,
                 servicos,combustiveis,fotos)
                VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)''', [
                r['cnpj'],r['razao'],r['bandeira'],
                r['cep'],r['logradouro'],r['numero'],r['complemento'],
                r['bairro'],r['cidade'],r['uf'],
                r['lat'],r['lon'],
                r['gestor'],r['telefone'],r['email_resp'],r['email_nf'],
                r['banco'], r['agencia'], r['conta'],
                r['servicos'],r['combustiveis'],r['fotos'],
            ])
            if conn.execute('SELECT changes()').fetchone()[0]:
                inseridos += 1
            else:
                duplicatas += 1
        except Exception as e:
            duplicatas += 1
            if duplicatas <= 3:
                print(f'  ⚠  CNPJ {r["cnpj"]}: {e}')

    conn.commit()
    conn.close()

    print()
    print(f'  ✅  Concluído!')
    print(f'      {inseridos} postos inseridos')
    if duplicatas:
        print(f'      {duplicatas} ignorados (CNPJ já existente)')
    print()


if __name__ == '__main__':
    main()
